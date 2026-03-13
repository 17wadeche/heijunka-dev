import re
import sys
import argparse
import time
import glob
from pathlib import Path
from datetime import datetime as _dt, date as _date, timedelta
from dateutil import parser as dateparser
import pandas as pd
import csv
import numpy as np
import shutil
import subprocess
from openpyxl import load_workbook
import tempfile, uuid
import json
from functools import lru_cache
def _file_sig(path: str | Path) -> tuple[str, int, int]:
    p = Path(path)
    try:
        st = p.stat()
        return (str(p), st.st_mtime_ns, st.st_size)
    except Exception:
        return (str(p), 0, 0)

from openpyxl import load_workbook as _load_workbook_orig
@lru_cache(maxsize=128)
def _load_workbook_cached(path_str: str, data_only: bool, read_only: bool, mtime_ns: int, size: int):
    return _load_workbook_orig(Path(path_str), data_only=data_only, read_only=read_only)
def load_workbook_fast(path: Path, *, data_only: bool = True, read_only: bool = True):
    sig_path, sig_mtime, sig_size = _file_sig(path)
    return _load_workbook_cached(sig_path, data_only, read_only, sig_mtime, sig_size)
from functools import lru_cache
@lru_cache(maxsize=256)
def _read_excel_cached(path_str: str, sheet_name: str, engine: str | None,
                       usecols, nrows, mtime_ns: int, size: int):
    from pandas import read_excel
    return read_excel(
        path_str,
        sheet_name=sheet_name,
        engine=engine,
        header=None,
        usecols=usecols,
        nrows=nrows,
    )
def read_excel_fast(path: Path, *, sheet_name: str, engine: str | None = None,
                    usecols=None, nrows: int | None = None):
    usecols_key = tuple(usecols) if isinstance(usecols, list) else usecols
    sig_path, sig_mtime, sig_size = _file_sig(path)
    return _read_excel_cached(sig_path, sheet_name, engine, usecols_key, nrows, sig_mtime, sig_size)
REPO_DIR = Path(r"C:\heijunka-dev")
REPO_CSV = REPO_DIR / "metrics_aggregate_dev.csv"
REPO_XLSX = REPO_DIR / "metrics_aggregate_dev.xlsx"
GIT_BRANCH = "main"
TIMELINESS_CSV = REPO_DIR / "timeliness.csv"
EXCLUDED_SOURCE_FILES = {
    r"C:\Users\wadec8\Medtronic PLC\SVT PXM Team - Archived Heijunka\SVT Future Heijunka.xlsm",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Heijunka Population & SW.xlsx",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\PVH Future Heijunka (UPDATE) Template.xlsm",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Versatility Matrix_June 2024.xlsx",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - Heijunka\CRDN Heijunka Template.xlsm",
    r"c:\Users\wadec8\Medtronic PLC\CAS Virtual VMB - PA Board\Scheduling Assistant.xlsx",
    r"c:\Users\wadec8\Medtronic PLC\CAS Virtual VMB - PA Board\Aging FACs 26 09.xlsx",
    r"c:\Users\wadec8\Medtronic PLC\CAS Virtual VMB - PA Board\Tier 1 Escalations and Recognition.pptx",
    r"c:\Users\wadec8\Medtronic PLC\CAS Virtual VMB - PA Board\Time Studies.docx",
    r"c:\Users\wadec8\Medtronic PLC\CAS Virtual VMB - PA Board\Updated Cryo Electronic data sheet.xlsx",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - Aortic - Heijunka\Saved Heijunkas\Aortic Heijunka Template 2.0.xlsm",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - Aortic - Heijunka\Saved Heijunkas\Aortic Heijunka Template 3.0.xlsm",
    r"c:\Users\wadec8\Medtronic PLC\Doran, Elaine - Heijunka Production Analysis\Archived Heijunka\ECT Future Heijunka 04 August 2025.xlsm",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Archive_Production Analysis\PVEV Heijunka_2020_2022_Archived.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Archive_Production Analysis\Prod Analysis Drafts for upcoming weeks\PVH Production Analysis & Heijunka_16 Oct 2023.xlsm"
}
EXCLUDED_DIRS = {
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\Weekly Heijunka Archived",
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\Clinical",
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\Commercial",
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\Remediation",
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\WIP Blitz Power Hour",
    r"C:\Users\wadec8\Medtronic PLC\SVT PXM Team - Archived Heijunka",
    r"c:\Users\wadec8\Medtronic PLC\Doran, Elaine - Heijunka Production Analysis\Heijunka Template",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Archive",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Finding Work Tool",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Archive_Production Analysis\2023",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Archive_Production Analysis\2022",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Archive_Production Analysis\2021",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Archive_Production Analysis\2020",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Archive_Production Analysis\2019",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Archive_Production Analysis\Prod Analysis Drafts for upcoming weeks",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\PVH Smartsheet Gameboard",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Standard Works",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Upcoming Weeks Heijunka Drafts",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - Aortic - Heijunka\Saved Heijunkas\Templates",
    r"c:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - Heijunka\Archived\Archived Heijunka 2024",
    r"C:\Users\wadec8\Medtronic PLC\CQXM - IV Resource Site - COS Supportive Materials\Archive_Production Analysis\2024\8. Aug"
}
EXCLUDED_DIRS = {s.lower().rstrip("\\").replace("/", "\\") for s in EXCLUDED_DIRS}
A1_RE = re.compile(r"([A-Za-z]+)(\d+)")
EXCLUDED_SOURCE_FILES = {s.lower().replace("/", "\\") for s in EXCLUDED_SOURCE_FILES}
def _is_excluded_path(p: Path) -> bool:
    try:
        sp = str(p).lower().replace("/", "\\").strip().rstrip("\\")
        if sp in EXCLUDED_SOURCE_FILES:
            return True
        for d in EXCLUDED_DIRS:
            d_norm = d.strip().rstrip("\\")
            if sp == d_norm or sp.startswith(d_norm + "\\"):
                return True
            if ("\\" + d_norm + "\\") in ("\\" + sp + "\\"):
                return True
        if p.name.lower().strip() == "svt future heijunka.xlsm":
            return True
    except Exception:
        pass
    return False
TEAM_CONFIG = [
    {
        "name": "ECT",
        "root": r"C:\Users\wadec8\Medtronic PLC\Doran, Elaine - Heijunka Production Analysis",
        "pattern": "*.xls*",
        "period": {"sheet": "#12 Production Analysis", "cell": "C4"},
        "cells": {
            "Individual (WIP-Non WIP)": {
                "Total Available Hours": "I39",
                "Completed Hours":       "I40",
            }
        },
        "sum_columns": {
            "#12 Production Analysis": {
                "Target Output": {"col": "F", "row_start": 7, "row_end": 199},
                "Actual Output": {"col": "I", "row_start": 7, "row_end": 199},
                "Completed Hours Detail": {"col": "G", "row_start": 7, "row_end": 199, "divide": 60, "skip_hidden": True},
            }
        },
        "unique_key": ["team", "period_date"],
    },
    {
        "name": "ECT",
        "root": r"C:\Users\wadec8\Medtronic PLC\Doran, Elaine - Heijunka Production Analysis\Archived Heijunka",
        "pattern": "*.xls*",
        "period": {"sheet": "#12 Production Analysis", "cell": "C4"},
        "cells": {
            "Individual (WIP-Non WIP)": {
                "Total Available Hours": "I39",
                "Completed Hours":       "I40",
            }
        },
        "sum_columns": {
            "#12 Production Analysis": {
                "Target Output": {"col": "F", "row_start": 7, "row_end": 199},
                "Actual Output": {"col": "I", "row_start": 7, "row_end": 199},
                "Completed Hours Detail": {"col": "G", "row_start": 7, "row_end": 199, "divide": 60, "skip_hidden": True},
            }
        },
        "unique_key": ["team", "period_date"],
    }
]
SKIP_PATTERNS = [r"~\$", r"\.tmp$"]
DATE_REGEXES = [
    r"\b\d{4}[-_]\d{2}[-_]\d{2}\b",
    r"\b\d{8}\b",
    r"\b\d{1,2}[-_]\d{1,2}[-_]\d{2,4}\b",
    r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*[-_ ]\d{1,2}[-_, ]\d{2,4}\b",
]
USE_FILE_MTIME_IF_NO_DATE = True
OUT_XLSX = Path.cwd() / "metrics_aggregate_dev.xlsx"
OUT_CSV  = Path.cwd() / "metrics_aggregate_dev.csv"
def looks_like_temp(name: str) -> bool:
    return any(re.search(pat, name, flags=re.IGNORECASE) for pat in SKIP_PATTERNS)
def parse_date_from_text(text: str):
    for rx in DATE_REGEXES:
        m = re.search(rx, text, flags=re.IGNORECASE)
        if m:
            try:
                return dateparser.parse(m.group(0), dayfirst=False, yearfirst=False).date()
            except Exception:
                pass
    return None
def _filter_future_periods(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "period_date" not in df.columns:
        return df
    today = pd.Timestamp.today().normalize()
    d = pd.to_datetime(df["period_date"], errors="coerce").dt.normalize()
    src = (
        df.get("source_file", "")
          .astype(str)
          .str.casefold()
          .str.replace("/", "\\", regex=False)
    )
    pvh_archive_2025_token = "\\cqxm - iv resource site - cos supportive materials\\archive_production analysis\\2025"
    is_pvh_archive_2025 = (
        df.get("team", "").astype(str).str.casefold().eq("pvh")
        & src.str.contains(pvh_archive_2025_token, regex=False)
    )
    keep = d.isna() | (d <= today) | is_pvh_archive_2025
    return df.loc[keep].copy()
def _filter_ect_min_year(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "team" not in df.columns or "period_date" not in df.columns:
        return df
    cutoff = pd.Timestamp("2024-08-19")
    d = pd.to_datetime(df["period_date"], errors="coerce")
    keep = ~((df["team"] == "ECT") & (d < cutoff))
    return df.loc[keep].copy()
def _excel_serial_to_date(n) -> _date | None:
    try:
        return (_dt(1899, 12, 30) + timedelta(days=float(n))).date()
    except Exception:
        return None
YEAR_RX = re.compile(r"\b(?:19|20)\d{2}\b")
def _coerce_to_date_for_filter2(v, require_explicit_year: bool = False) -> _date | None:
    if isinstance(v, _dt):
        return v.date()
    if isinstance(v, _date):
        return v
    if isinstance(v, (int, float)):
        d = _excel_serial_to_date(v)
        if d and _dt(1900, 1, 1).date() <= d <= _dt(2100, 1, 1).date():
            return d
        return None
    s = str(v)
    if require_explicit_year and not YEAR_RX.search(s):
        return None
    try:
        return dateparser.parse(s, dayfirst=False, yearfirst=False).date()
    except Exception:
        return None
def _coerce_to_date_for_filter(v) -> _date | None:
    if isinstance(v, _dt):
        return v.date()
    if isinstance(v, _date):
        return v
    if isinstance(v, (int, float)):
        d = _excel_serial_to_date(v)
        if d and _dt(1900, 1, 1).date() <= d <= _dt(2100, 1, 1).date():
            return d
        return None
    try:
        d = dateparser.parse(str(v)).date()
        return d
    except Exception:
        return None
def _git(args: list[str], cwd: Path) -> tuple[int, str, str]:
    p = subprocess.Popen(args, cwd=str(cwd), stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    out, err = p.communicate()
    cmd = " ".join(args)
    if p.returncode != 0:
        print(f"[git] {cmd} -> code={p.returncode}\nSTDERR: {err}\nSTDOUT: {out}")
    else:
        if err.strip():
            print(f"[git] {cmd} (ok) stderr: {err.strip()}")
    return p.returncode, out.strip(), err.strip()
def _ensure_git_identity(repo_dir: Path):
    code, out, _ = _git(["git", "config", "--get", "user.email"], repo_dir)
    if code != 0 or not out:
        _git(["git", "config", "user.email", "heijunka-bot@example.com"], repo_dir)
    code, out, _ = _git(["git", "config", "--get", "user.name"], repo_dir)
    if code != 0 or not out:
        _git(["git", "config", "user.name", "Heijunka Bot"], repo_dir)
def git_autocommit_and_push_many(repo_dir: Path, files: list[Path], branch: str = "main"):
    if not (repo_dir / ".git").exists():
        print(f"[WARN] {repo_dir} is not a git repo; skipping auto-commit.")
        return
    _git(["git", "config", "--global", "--add", "safe.directory", str(repo_dir)], repo_dir)
    _ensure_git_identity(repo_dir)
    code, remotes, _ = _git(["git", "remote"], repo_dir)
    if code != 0 or not remotes.strip():
        print("[WARN] No git remote configured (e.g., 'origin'); cannot push.")
        return
    remote = "origin" if "origin" in remotes.split() else remotes.split()[0]
    code, heads, _ = _git(["git", "branch", "--list", branch], repo_dir)
    _git(["git", "checkout", branch], repo_dir) if heads.strip() else _git(["git", "checkout", "-B", branch], repo_dir)
    _, dirty, _ = _git(["git", "status", "--porcelain"], repo_dir)
    if dirty.strip():
        print("[git] Working tree has local changes; using --autostash during pull.")
    _git(["git", "fetch", remote], repo_dir)
    code, _, _ = _git(["git", "pull", "--rebase", "--autostash", remote, branch], repo_dir)
    if code != 0:
        print("[WARN] 'git pull --rebase --autostash' failed; skipping commit to avoid conflicts.")
        return
    staged_any = False
    for fp in files:
        if fp and fp.exists():
            try:
                rel = fp.relative_to(repo_dir).as_posix()
            except Exception:
                rel = str(fp)
            _git(["git", "add", "--", rel], repo_dir)
            staged_any = True
    if not staged_any:
        print("[git] Nothing to stage; skipping push.")
        return
    code, status, _ = _git(["git", "status", "--porcelain"], repo_dir)
    if code != 0 or not status.strip():
        print("[git] No changes detected after add; skipping push.")
        return
    msg = f"Auto-update metrics at {_dt.now().isoformat(timespec='seconds')}"
    code, _, _ = _git(["git", "commit", "-m", msg], repo_dir)
    if code != 0:
        print("[WARN] git commit failed; see logs above.")
        return
    code, _, _ = _git(["git", "push", "-u", remote, branch], repo_dir)
    if code != 0:
        print("[WARN] git push failed; check credentials / PAT / VPN.")
    else:
        print("[git] Pushed:", ", ".join([str(f.name) for f in files if f.exists()]))
def run_apply_closures():
    try:
        script = Path(__file__).with_name("apply_closures.py")
        if not script.exists():
            print(f"[apply_closures] Not found: {script}")
            return
        print(f"[apply_closures] Running: {script}")
        res = subprocess.run([sys.executable, str(script)], capture_output=True, text=True)
        if res.stdout.strip():
            print("[apply_closures][stdout]\n" + res.stdout)
        if res.stderr.strip():
            print("[apply_closures][stderr]\n" + res.stderr)
        if res.returncode != 0:
            print(f"[apply_closures] Exit code {res.returncode}")
    except Exception as e:
        print(f"[apply_closures] Failed: {e}")
def git_pull_repo(repo_dir: Path, branch: str = "main"):
    if not (repo_dir / ".git").exists():
        print(f"[WARN] {repo_dir} is not a git repo; skipping pre-pull.")
        return
    _git(["git", "config", "--global", "--add", "safe.directory", str(repo_dir)], repo_dir)
    _ensure_git_identity(repo_dir)
    code, remotes, _ = _git(["git", "remote"], repo_dir)
    if code != 0 or not remotes.strip():
        print("[WARN] No git remote configured; skipping pre-pull.")
        return
    remote = "origin" if "origin" in remotes.split() else remotes.split()[0]
    heads_code, heads, _ = _git(["git", "branch", "--list", branch], repo_dir)
    if heads.strip():
        _git(["git", "checkout", branch], repo_dir)
    else:
        _git(["git", "checkout", "-B", branch], repo_dir)
    _git(["git", "fetch", remote], repo_dir)
    _git(["git", "pull", "--rebase", "--autostash", remote, branch], repo_dir)
def git_autocommit_and_push(repo_dir: Path, file_path: Path, branch: str = "main"):
    if not (repo_dir / ".git").exists():
        print(f"[WARN] {repo_dir} is not a git repo; skipping auto-commit.")
        return
    if not file_path.exists():
        print(f"[WARN] {file_path} not found to commit.")
        return
    _git(["git", "config", "--global", "--add", "safe.directory", str(repo_dir)], repo_dir)
    _ensure_git_identity(repo_dir)
    code, _, _ = _git(["git", "--version"], repo_dir)
    if code != 0:
        print("[WARN] Git not available on PATH. Install Git or adjust PATH.")
        return
    code, remotes, _ = _git(["git", "remote"], repo_dir)
    if code != 0 or not remotes.strip():
        print("[WARN] No git remote configured (e.g., 'origin'); cannot push.")
        return
    remote = "origin" if "origin" in remotes.split() else remotes.split()[0]
    code, heads, _ = _git(["git", "branch", "--list", branch], repo_dir)
    if heads.strip():
        _git(["git", "checkout", branch], repo_dir)
    else:
        _git(["git", "checkout", "-B", branch], repo_dir)
    _, dirty, _ = _git(["git", "status", "--porcelain"], repo_dir)
    if dirty.strip():
        print("[git] Working tree has local changes; using --autostash during pull.")
    _git(["git", "fetch", remote], repo_dir)
    code, _, _ = _git(["git", "rev-parse", "--abbrev-ref", "--symbolic-full-name", "@{u}"], repo_dir)
    if code != 0:
        _git(["git", "branch", "--set-upstream-to", f"{remote}/{branch}", branch], repo_dir)
    code, _, _ = _git(["git", "pull", "--rebase", "--autostash", remote, branch], repo_dir)
    if code != 0:
        print("[WARN] 'git pull --rebase --autostash' failed; skipping commit to avoid conflicts. Resolve repo state manually.")
        return
    try:
        rel = file_path.relative_to(repo_dir).as_posix()
    except Exception:
        rel = str(file_path)
    _git(["git", "add", "--", rel], repo_dir)
    code, status, _ = _git(["git", "status", "--porcelain", "--", rel], repo_dir)
    if code != 0:
        print("[WARN] Could not check git status; aborting commit.")
        return
    if not status.strip():
        print("[git] No changes in CSV; skipping push.")
        return
    msg = f"Auto-update {rel} at {_dt.now().isoformat(timespec='seconds')}"
    code, _, _ = _git(["git", "commit", "-m", msg], repo_dir)
    if code != 0:
        print("[WARN] git commit failed; see logs above.")
        return
    code, _, _ = _git(["git", "push", "-u", remote, branch], repo_dir)
    if code != 0:
        print("[WARN] git push failed; see logs above. Check credentials / PAT / VPN.")
    else:
        print("[git] Pushed CSV update to", f"{remote}/{branch}")
def _ph_values_by_person(ws, col_end: str) -> tuple[dict, float | None, float | None]:
    def _to_float(v):
        if v is None: 
            return None
        try:
            return float(str(v).replace(",", "").strip())
        except Exception:
            return None
    rng_names = ws.Range(f"B30:{col_end}30").Value
    rng_actual = ws.Range(f"B50:{col_end}50").Value
    rng_avail  = ws.Range(f"B59:{col_end}59").Value
    if isinstance(rng_names, (tuple, list)) and isinstance(rng_names[0], (tuple, list)):
        names = list(rng_names[0])
    else:
        names = list(rng_names if isinstance(rng_names, (tuple, list)) else [rng_names])
    if isinstance(rng_actual, (tuple, list)) and isinstance(rng_actual[0], (tuple, list)):
        actuals = list(rng_actual[0])
    else:
        actuals = list(rng_actual if isinstance(rng_actual, (tuple, list)) else [rng_actual])
    if isinstance(rng_avail, (tuple, list)) and isinstance(rng_avail[0], (tuple, list)):
        avails = list(rng_avail[0])
    else:
        avails = list(rng_avail if isinstance(rng_avail, (tuple, list)) else [rng_avail])
    m = max(len(names), len(actuals), len(avails))
    names  += [None] * (m - len(names))
    actuals += [None] * (m - len(actuals))
    avails  += [None] * (m - len(avails))
    per = {}
    tot_actual = 0.0
    tot_avail  = 0.0
    any_actual = False
    any_avail  = False
    for nm, a, t in zip(names, actuals, avails):
        nm_str = (str(nm).strip() if nm is not None else "")
        if not nm_str:
            continue
        a_f = _to_float(a)
        t_f = _to_float(t)
        if a_f is not None:
            any_actual = True
            tot_actual += a_f
        if t_f is not None:
            any_avail = True
            tot_avail += t_f
        per[nm_str] = {"actual": (a_f if a_f is not None else 0.0),
                       "available": (t_f if t_f is not None else 0.0)}
    return per, (tot_actual if any_actual else None), (tot_avail if any_avail else None)
def _to_excel_com_value(v):
    if isinstance(v, (int, float, str)):
        return v
    if isinstance(v, _dt):
        return v
    if isinstance(v, _date):
        return _dt(v.year, v.month, v.day)
    return str(v)
def read_one_cell_openpyxl(ws, a1: str):
    r, c = a1_to_rowcol(a1)
    vals = list(ws.iter_rows(min_row=r, max_row=r, min_col=c, max_col=c, values_only=True))
    return vals[0][0] if vals else None
def read_one_cell_xlsb(file_path: Path, sheet: str, a1: str):
    df = read_excel_fast(file_path, sheet_name=sheet, engine="pyxlsb")
    m = A1_RE.fullmatch(a1.strip())
    if not m:
        return None
    col_letters, row_str = m.groups()
    r = int(row_str) - 1
    c = col_letter_to_index(col_letters) - 1
    try:
        return df.iat[r, c]
    except Exception:
        return None
def infer_period_date(path: Path):
    d = parse_date_from_text(str(path))
    if d:
        return d
    if USE_FILE_MTIME_IF_NO_DATE:
        try:
            return _dt.fromtimestamp(path.stat().st_mtime).date()
        except Exception:
            return None
    return None
def safe_div(n, d):
    try:
        n = float(str(n).replace(",", "").strip())
        d = float(str(d).replace(",", "").strip())
        if d == 0:
            return None
        return n / d
    except Exception:
        return None
def col_letter_to_index(letter: str) -> int:
    letter = letter.strip().upper()
    num = 0
    for ch in letter:
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f"Invalid column letter: {letter}")
        num = num * 26 + (ord(ch) - ord('A') + 1)
    return num
def a1_to_rowcol(a1: str) -> tuple[int, int]:
    m = A1_RE.fullmatch(a1.strip())
    if not m:
        raise ValueError(f"Bad cell address: {a1}")
    letters, row = m.groups()
    col = 0
    for ch in letters.upper():
        col = col * 26 + (ord(ch) - ord('A') + 1)
    return int(row), col
import re
def sum_column_openpyxl_filtered(ws, target_col: str,
                                 include_contains: dict | None = None,
                                 exclude_regex: dict | None = None,
                                 row_start: int | None = None,
                                 row_end: int | None = None,
                                 skip_hidden: bool = False) -> float | None:
    t_idx = col_letter_to_index(target_col)            # 1-based
    need_cols = {t_idx}
    if include_contains:
        need_cols |= {col_letter_to_index(c) for c in include_contains.keys()}
    if exclude_regex:
        need_cols |= {col_letter_to_index(c) for c in exclude_regex.keys()}
    min_c, max_c = min(need_cols), max(need_cols)
    row_start = row_start or 1
    row_end = row_end or ws.max_row
    total, any_vals = 0.0, False
    for r_idx, row_vals in enumerate(
        ws.iter_rows(min_row=row_start, max_row=row_end,
                     min_col=min_c, max_col=max_c, values_only=True),
        start=row_start
    ):
        if skip_hidden and hasattr(ws, "row_dimensions"):
            rd = ws.row_dimensions.get(r_idx) if hasattr(ws.row_dimensions, "get") else None
            if rd is not None and getattr(rd, "hidden", False):
                continue
        row_map_by_idx = {min_c + off: v for off, v in enumerate(row_vals)}
        if include_contains:
            ok = True
            for col_letter, needle in include_contains.items():
                c_idx = col_letter_to_index(col_letter)
                v = row_map_by_idx.get(c_idx)
                if not (isinstance(v, str) and needle.lower() in v.lower()):
                    ok = False
                    break
            if not ok:
                continue
        if exclude_regex:
            bad = False
            for col_letter, rx in exclude_regex.items():
                c_idx = col_letter_to_index(col_letter)
                v = row_map_by_idx.get(c_idx)
                if isinstance(v, str) and re.search(rx, v, flags=re.IGNORECASE):
                    bad = True
                    break
            if bad:
                continue
        val = row_map_by_idx.get(t_idx)
        try:
            if val is None:
                continue
            if isinstance(val, (int, float)):
                total += float(val); any_vals = True
            else:
                total += float(str(val).replace(",", "").strip()); any_vals = True
        except Exception:
            continue
    return total if any_vals else None
def _people_available_openpyxl_generic(ws,
                                       name_col: str = "A",
                                       avail_col: str = "I",
                                       start_row: int = 6,
                                       end_row: int = 32,
                                       step: int = 3) -> list[tuple[str, float | None]]:
    out: list[tuple[str, float | None]] = []
    BAD_NAMES = {"", "#REF!", "-", "–", "—", "0"}
    for r in range(start_row, end_row + 1, step):
        nm = ws[f"{name_col}{r}"].value
        nm = (str(nm).strip() if nm is not None else "")
        if nm in BAD_NAMES:
            continue
        avail = safe_numeric(ws[f"{avail_col}{r}"].value)
        out.append((nm, avail))
    return out
def _svt_completed_hours_by_person_openpyxl(ws_pa,
                                            people: list[str],
                                            name_col: str = "C",
                                            minutes_col: str = "G",
                                            row_start: int = 1,
                                            row_end:   int = 200,
                                            skip_hidden: bool = True) -> dict[str, float]:
    kc = col_letter_to_index(name_col)
    mc = col_letter_to_index(minutes_col)
    want = {p.strip().casefold(): p.strip() for p in people if str(p).strip()}
    totals_min = {canonical: 0.0 for canonical in want.keys()}
    min_c = min(kc, mc)
    max_c = max(kc, mc)
    for r_idx, row_vals in enumerate(
        ws_pa.iter_rows(min_row=row_start, max_row=row_end,
                        min_col=min_c, max_col=max_c, values_only=True),
        start=row_start
    ):
        if skip_hidden and hasattr(ws_pa, "row_dimensions"):
            rd = ws_pa.row_dimensions.get(r_idx) if hasattr(ws_pa.row_dimensions, "get") else None
            if rd is not None and getattr(rd, "hidden", False):
                continue
        name_val = row_vals[kc - min_c]
        mins_val = row_vals[mc - min_c]
        if name_val is None or mins_val is None:
            continue
        name_key = str(name_val).strip().casefold()
        if name_key not in totals_min:
            continue
        try:
            mins = float(str(mins_val).replace(",", "").strip())
            if mins > 0:
                totals_min[name_key] += mins
        except Exception:
            continue
    out = {}
    for key, total_m in totals_min.items():
        out[want[key]] = round(total_m / 60.0, 2)
    return out
def _hc_in_wip_from_file(file_path: Path, sheet_name: str,
                         key_col_letter: str = "C", cond_col_letter: str = "G",
                         row_start: int = 7, row_end: int = 200) -> int | None:
    ext = file_path.suffix.lower()
    try:
        if ext in (".xlsx", ".xlsm"):
            wb = load_workbook(file_path, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                return None
            ws = wb[sheet_name]
            return _svt_hc_in_wip_openpyxl(ws, key_col_letter, cond_col_letter, row_start, row_end)
        elif ext == ".xlsb":
            from pandas import read_excel
            df = read_excel(file_path, sheet_name=sheet_name, engine="pyxlsb", header=None)
            rs = row_start - 1
            re_ = row_end 
            k = col_letter_to_index(key_col_letter) - 1
            c = col_letter_to_index(cond_col_letter) - 1
            cond_series = pd.to_numeric(df.iloc[rs:re_, c], errors="coerce")
            mask = cond_series > 0
            keys_series = df.iloc[rs:re_, k].where(mask)
            uniq = set()
            for v in keys_series.dropna().tolist():
                s = str(v).strip()
                if s and s.lower() != "nan":
                    uniq.add(s)
            return len(uniq)
        else:
            return None
    except Exception:
        return None
_TIMEISH_RE = re.compile(
    r"""
    (                       
      \b\d{1,2}\s*[:]\s*\d{2}\s*(?:AM|PM)\b      
     |\b\d{1,2}\s*(?:AM|PM)\b                   
     |\b\d{1,2}\s*[-–]\s*\d{1,2}\s*(?:AM|PM)?\b  
     |\b(?:AM|PM)\b
    )
    (?:\s*[A-Z]{2,4})?      
    """,
    re.IGNORECASE | re.VERBOSE,
)
PAREN_TRIM_RE = re.compile(r"\s*\((?:audit|reverse\s*shadowing|flex)\)\s*", re.IGNORECASE)
AUDIT_WORD_RE = re.compile(r"\b(?:audit)\b", re.IGNORECASE)
TRAIL_PRACTICE_RE = re.compile(r"\s+practice\b", re.IGNORECASE)
SPLIT_RE = re.compile(r"\s*(?:&|,|-|/|\+|\band\b|reverse\s*shadowing\s*with)\s*", re.IGNORECASE)
ALIAS_MAP = {
    "jerlie": "Jerlie", "jerile": "Jerlie",
    "natalie": "Natalie", "natlalie": "Natalie",
    "sean": "Sean", "sh": "Sean",
    "orla": "Orla",
}
def _looks_like_timeish(s: str) -> bool:
    return bool(_TIMEISH_RE.search(s or ""))
def _should_exclude_name(name: str) -> bool:
    n = (name or "").strip()
    if not n:
        return True
    lowered = n.casefold()
    if lowered == "team member" or lowered.startswith("team member "):
        return True
    if lowered in {"nan", "audit", "team", "everyone"}:
        return True
    if "power hour" in lowered:
        return True
    if "affera" in n.lower():
        return True
    if "ftq" in n.lower():
        return True
    if "capa" in n.lower():
        return True
    if "training" in n.lower():
        return True
    if "practice" in n.lower():
        return True
    if "problem solving" in n.lower():
        return True
    if "timing studies" in n.lower():
        return True
    if "training tasks" in n.lower():
        return True
    if "training station 2" in n.lower():
        return True
    if "prism 2" in n.lower():
        return True
    if "eu" in n.lower():
        return True
    if "0" in n:
        return True
    if _looks_like_timeish(n):
        return True
    return False
def _clean_person_token(s: str) -> str:
    if s is None:
        return ""
    s = s.strip().strip('"').strip("'")
    s = re.sub(r"[?!]+", "", s)          
    s = PAREN_TRIM_RE.sub(" ", s)         
    s = AUDIT_WORD_RE.sub(" ", s) 
    s = TRAIL_PRACTICE_RE.sub("", s)    
    s = re.sub(r"\s{2,}", " ", s)        
    return s.strip()
def _alias_canonicalize(name: str) -> str:
    k = (name or "").strip().casefold()
    return ALIAS_MAP.get(k, (name or "").strip())
def _split_people(name: str) -> list[str]:
    if not name:
        return []
    parts = SPLIT_RE.split(name)       
    cleaned = []
    for p in parts:
        px = _clean_person_token(p)
        if not px:
            continue
        px = _alias_canonicalize(px)  
        if not _should_exclude_name(px):
            cleaned.append(px)
    return cleaned
def _normalize_cell_station_hours(ch: dict) -> dict:
    out: dict[str, float] = {}
    for raw_key, v in (ch or {}).items():
        key = (raw_key or "").strip()
        if not key or key.casefold() == "nan":
            continue
        try:
            out[key] = round(float(v or 0.0), 2)
        except Exception:
            continue
    return out
def _normalize_outputs_by_person(op: dict) -> dict:
    out: dict[str, dict] = {}
    for raw_name, vals in (op or {}).items():
        vals = vals or {}
        output_val = round(float(vals.get("output", 0) or 0.0), 2)
        target_val = round(float(vals.get("target", 0) or 0.0), 2)
        if output_val == 0.0 and target_val == 0.0:
            continue
        name_clean = _clean_person_token(raw_name)
        if _should_exclude_name(name_clean):
            continue
        people = _split_people(name_clean)
        if len(people) >= 2:
            for person in people:
                person = _alias_canonicalize(person)
                cur = out.get(person, {"output": 0.0, "target": 0.0})
                cur["output"] = round(cur["output"] + output_val, 2)
                cur["target"] = round(cur["target"] + target_val, 2)
                out[person] = cur
            continue
        person = _alias_canonicalize(people[0] if people else name_clean)
        if _should_exclude_name(person):
            continue
        cur = out.get(person, {"output": 0.0, "target": 0.0})
        cur["output"] = round(cur["output"] + output_val, 2)
        cur["target"] = round(cur["target"] + target_val, 2)
        out[person] = cur
    out2: dict[str, dict] = {}
    for k, v in out.items():
        ov = round(float(v.get("output", 0.0)), 2)
        tv = round(float(v.get("target", 0.0)), 2)
        if ov == 0.0 and tv == 0.0:
            continue
        out2[k] = {"output": ov, "target": tv}
    return out2
def _normalize_outputs_by_cell(op: dict) -> dict:
    out: dict[str, dict] = {}
    for raw_key, vals in (op or {}).items():
        key = (raw_key or "").strip()
        if not key or key.casefold() == "nan":
            continue
        vals = vals or {}
        ov = round(float(vals.get("output", 0) or 0.0), 2)
        tv = round(float(vals.get("target", 0) or 0.0), 2)
        if ov == 0.0 and tv == 0.0:
            continue
        out[key] = {"output": ov, "target": tv}
    return out
def _normalize_person_hours(ph: dict) -> dict:
    out: dict[str, dict] = {}
    for raw_name, vals in (ph or {}).items():
        name = (raw_name or "").strip()
        vals = vals or {}
        actual = float(vals.get("actual", 0) or 0.0)
        avail  = vals.get("available", None)
        if _should_exclude_name(name):
            continue
        people = _split_people(name)
        if len(people) >= 2:
            for person in people:
                cur = out.get(person, {"actual": 0.0, "available": 0.0})
                cur["actual"] = round(cur["actual"] + actual, 2)
                try:
                    if avail is not None and float(avail) > 0:
                        cur["available"] = max(cur.get("available", 0.0), float(avail))
                    else:
                        cur["available"] = max(cur.get("available", 0.0), 25.0)
                except Exception:
                    cur["available"] = max(cur.get("available", 0.0), 25.0)
                out[person] = cur
            continue
        person = _alias_canonicalize(_clean_person_token(people[0] if people else name))
        if _should_exclude_name(person):
            continue
        cur = out.get(person, {"actual": 0.0, "available": 0.0})
        cur["actual"] = round(cur["actual"] + actual, 2)
        try:
            if avail is not None and float(avail) > 0:
                cur["available"] = max(cur.get("available", 0.0), float(avail))
        except Exception:
            pass
        out[person] = cur
    for k, v in out.items():
        v["actual"] = round(float(v.get("actual", 0.0)), 2)
        v["available"] = round(float(v.get("available", 0.0)), 2)
    return out
def _hc_in_wip_from_person_hours(ph: dict) -> int:
    return sum(
        1
        for _name, vals in (ph or {}).items()
        if float((vals or {}).get("actual", 0.0) or 0.0) > 0.0
    )
def sum_column_pyxlsb_filtered(file_path: Path, sheet: str, target_col: str,
                               include_contains: dict | None = None,
                               exclude_regex: dict | None = None,
                               row_start: int | None = None,
                               row_end: int | None = None) -> float | None:
    try:
        df = read_excel_fast(file_path, sheet_name=sheet, engine="pyxlsb")
        t = col_letter_to_index(target_col) - 1
        s = pd.to_numeric(df.iloc[:, t], errors="coerce")
        mask = pd.Series(True, index=df.index)
        if include_contains:
            for col_letter, needle in include_contains.items():
                c = col_letter_to_index(col_letter) - 1
                mask &= df.iloc[:, c].astype(str).str.contains(needle, case=False, na=False)
        if exclude_regex:
            for col_letter, rx in exclude_regex.items():
                c = col_letter_to_index(col_letter) - 1
                mask &= ~df.iloc[:, c].astype(str).str.contains(rx, case=False, na=False, regex=True)
        if row_start or row_end:
            rs = (row_start - 1) if row_start else 0
            re_ = (row_end - 1) if row_end else len(df) - 1
            idx = df.index[rs:re_ + 1]
            mask = mask.loc[idx]
        s = s.where(mask)
        tot = s.dropna().sum()
        return float(tot) if pd.notna(tot) else None
    except Exception:
        return None
def sum_column_from_file(file_path: Path, sheet: str, col_letter: str,
                         include_contains: dict | None = None,
                         exclude_regex: dict | None = None,
                         row_start: int | None = None,
                         row_end: int | None = None,
                         skip_hidden: bool = False) -> float | None:
    ext = file_path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        wb = load_workbook(file_path, data_only=True, read_only=not skip_hidden)
        if sheet not in wb.sheetnames:
            return None
        ws = wb[sheet]
        return sum_column_openpyxl_filtered(ws, col_letter,
                                            include_contains, exclude_regex,
                                            row_start, row_end, skip_hidden)
    elif ext == ".xlsb":
        return sum_column_pyxlsb_filtered(file_path, sheet, col_letter,
                                          include_contains, exclude_regex,
                                          row_start, row_end)
    else:
        return None
def sum_column(ws, col_letter):
    col = 0
    for ch in col_letter.strip().upper():
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f"Invalid column letter: {col_letter}")
        col = col * 26 + (ord(ch) - ord('A') + 1)
    total = 0.0
    any_vals = False
    for (val,) in ws.iter_rows(min_col=col, max_col=col, values_only=True):
        try:
            if val is None: 
                continue
            if isinstance(val, (int, float)):
                total += float(val); any_vals = True
            else:
                v = float(str(val).replace(",", "").strip())
                total += v; any_vals = True
        except Exception:
            continue
    return total if any_vals else None
def safe_numeric(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    try:
        s = str(x).strip().replace(",", "")
        return float(s)
    except Exception:
        return None
def _nest_hours_person_by_station(pairs: pd.DataFrame) -> dict:
    if pairs.empty:
        return {}
    pairs = pairs.copy()
    pairs["station"] = pairs["station"].astype(str).str.strip()
    pairs["person"]  = pairs["person"].astype(str).str.strip()
    pairs["hours"]   = pd.to_numeric(pairs["hours"], errors="coerce")
    _bad = {"", "-", "–", "—", "nan"}
    pairs = pairs[
        pairs["hours"].notna() &
        ~pairs["station"].str.casefold().isin(_bad) &
        ~pairs["person"].str.casefold().isin(_bad)
    ]
    agg = pairs.groupby(["station","person"])["hours"].sum(min_count=1)
    out = {}
    for (st, pe), v in agg.items():
        if pd.isna(v):
            continue
        out.setdefault(st, {})[pe] = round(float(v), 2)
    return out
def _nest_outputs_person_by_station(pairs: pd.DataFrame) -> dict:
    if pairs.empty:
        return {}
    pairs = pairs.copy()
    pairs["station"] = pairs["station"].astype(str).str.strip()
    pairs["person"]  = pairs["person"].astype(str).str.strip()
    for c in ["actual_output","target_output"]:
        pairs[c] = pd.to_numeric(pairs[c], errors="coerce").fillna(0.0)
    agg = (pairs
           .groupby(["station","person"])[["actual_output","target_output"]]
           .sum(min_count=1)
           .replace({np.nan: 0.0}))
    out = {}
    for (st, pe), row in agg.iterrows():
        if not st or not pe:
            continue
        out.setdefault(st, {})[pe] = {
            "output": round(float(row["actual_output"]), 2),
            "target": round(float(row["target_output"]), 2),
        }
    for st in list(out.keys()):
        for pe in list(out[st].keys()):
            v = out[st][pe]
            if (v.get("output",0)==0) and (v.get("target",0)==0):
                del out[st][pe]
        if not out[st]:
            del out[st]
    return out
def _uplh_by_person_by_station(hours_map: dict, outputs_map: dict) -> dict:
    out: dict[str, dict[str, dict]] = {}
    for station, people_hours in (hours_map or {}).items():
        for person, hours in (people_hours or {}).items():
            try:
                h = float(hours)
                if h <= 0:
                    continue
            except Exception:
                continue
            op_station = (outputs_map or {}).get(station, {})
            op_vals = op_station.get(person, {}) if isinstance(op_station, dict) else {}
            ao = float(op_vals.get("output", 0) or 0.0)
            to = float(op_vals.get("target", 0) or 0.0)
            actual_uplh = (ao / h) if h else None
            target_uplh = (to / h) if h else None
            if actual_uplh is None and target_uplh is None:
                continue
            out.setdefault(station, {})
            out[station][person] = {
                "actual": round(actual_uplh, 2) if actual_uplh is not None else None,
                "target": round(target_uplh, 2) if target_uplh is not None else None,
            }
    return out
def normalize_period_date(df: pd.DataFrame) -> pd.DataFrame:
    if "period_date" in df.columns:
        df["period_date"] = pd.to_datetime(df["period_date"], errors="coerce").dt.normalize()
    return df
def read_with_openpyxl(file_path: Path, cells_cfg: dict, sumcols_cfg: dict) -> dict:
    need_visible_rows = any(
        isinstance(spec, dict) and spec.get("skip_hidden", False)
        for mapping in (sumcols_cfg or {}).values()
        for spec in mapping.values()
    )
    wb = load_workbook_fast(file_path, data_only=True, read_only=not need_visible_rows)
    out = {}
    for sheet_name, mapping in (cells_cfg or {}).items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        for out_name, addr in mapping.items():
            try:
                r, c = a1_to_rowcol(addr)
                vals = list(ws.iter_rows(min_row=r, max_row=r, min_col=c, max_col=c, values_only=True))
                out[out_name] = vals[0][0] if vals else None
            except Exception:
                out[out_name] = None
    for sheet_name, mapping in (sumcols_cfg or {}).items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        for out_name, spec in mapping.items():
            try:
                if isinstance(spec, str):
                    out[out_name] = sum_column(ws, spec)
                elif isinstance(spec, dict):
                    col_letter   = spec.get("col")
                    row_start    = spec.get("row_start")
                    row_end      = spec.get("row_end")
                    include_filt = spec.get("include_contains")
                    exclude_rx   = spec.get("exclude_regex")
                    skip_hidden  = spec.get("skip_hidden", False)
                    divide       = spec.get("divide")
                    val = sum_column_openpyxl_filtered(
                        ws,
                        col_letter,
                        include_contains=include_filt,
                        exclude_regex=exclude_rx,
                        row_start=row_start,
                        row_end=row_end,
                        skip_hidden=skip_hidden
                    )
                    if val is not None and divide:
                        try:
                            val = float(val) / float(divide)
                        except Exception:
                            pass
                    out[out_name] = val
                else:
                    out[out_name] = None
            except Exception:
                out[out_name] = None
    return out
def read_with_pyxlsb(file_path: Path, cells_cfg: dict, sumcols_cfg: dict) -> dict:
    from pandas import read_excel
    out = {}
    for sheet_name, mapping in (cells_cfg or {}).items():
        df = read_excel(file_path, sheet_name=sheet_name, engine="pyxlsb", header=None)
        for out_name, addr in mapping.items():
            m = re.fullmatch(r"([A-Za-z]+)(\d+)", addr.strip())
            if not m:
                out[out_name] = None
                continue
            col_letters, row_str = m.groups()
            r = int(row_str) - 1
            c = col_letter_to_index(col_letters) - 1
            try:
                val = df.iat[r, c]
            except Exception:
                val = None
            out[out_name] = val
    for sheet_name, mapping in (sumcols_cfg or {}).items():
        df = read_excel(file_path, sheet_name=sheet_name, engine="pyxlsb", header=None)
        for out_name, spec in mapping.items():
            try:
                if isinstance(spec, str):
                    c = col_letter_to_index(spec) - 1
                    series = pd.to_numeric(df.iloc[:, c], errors="coerce")
                    total = float(series.dropna().sum())
                    out[out_name] = total
                elif isinstance(spec, dict):
                    col_letter   = spec.get("col")
                    row_start    = spec.get("row_start")
                    row_end      = spec.get("row_end")
                    include_filt = spec.get("include_contains")
                    exclude_rx   = spec.get("exclude_regex")
                    divide       = spec.get("divide")
                    val = sum_column_pyxlsb_filtered(
                        file_path,
                        sheet_name,
                        col_letter,
                        include_contains=include_filt,
                        exclude_regex=exclude_rx,
                        row_start=row_start,
                        row_end=row_end
                    )
                    if val is not None and divide:
                        try:
                            val = float(val) / float(divide)
                        except Exception:
                            pass
                    out[out_name] = val
                else:
                    out[out_name] = None
            except Exception:
                out[out_name] = None
    return out
def _read_sheet_as_df(file_path: Path, sheet_name: str):
    ext = file_path.suffix.lower()
    engine = "pyxlsb" if ext == ".xlsb" else None
    try:
        return read_excel_fast(file_path, sheet_name=sheet_name, engine=engine)
    except Exception:
        return None
def _sum_output_target_by(df: pd.DataFrame, key_col_idx: int, out_col_idx: int, tgt_col_idx: int) -> dict:
    if df is None or df.empty:
        return {}
    cols = [key_col_idx, out_col_idx, tgt_col_idx]
    n = df.shape[1]
    if any(i >= n for i in cols):
        return {}
    sub = df.iloc[:, cols].copy()
    sub.columns = ["key", "output", "target"]
    sub["key"] = sub["key"].astype(str).str.strip()
    bad_keys = {"", "-", "–", "—", "nan"}  
    sub = sub[~sub["key"].isin(bad_keys)]
    sub = sub[sub["key"].str.len() > 0] 
    sub["output"] = pd.to_numeric(sub["output"], errors="coerce")
    sub["target"] = pd.to_numeric(sub["target"], errors="coerce")
    sub = sub[sub[["output", "target"]].notna().any(axis=1)]
    if sub.empty:
        return {}
    grp = sub.groupby("key", dropna=False, sort=True).agg({"output": "sum", "target": "sum"})
    out = {}
    for k, row in grp.iterrows():
        out[k] = {
            "output": float(round(row.get("output", 0.0) if pd.notna(row.get("output")) else 0.0, 2)),
            "target": float(round(row.get("target", 0.0) if pd.notna(row.get("target")) else 0.0, 2)),
        }
    return out
def _person_cell_hours_outputs_for_team(file_path: Path, team_name: str) -> tuple[dict, dict]:
    team = (team_name or "").strip().casefold()
    if team in ("svt","ect","pvh","crdn","aortic"):
        sheet = "#12 Production Analysis"
        person_idx = 2  
        cell_idx   = 3 
        mins_idx   = 6
        tgt_idx    = 5 if team != "aortic" else 4 
        out_idx    = 8 
    elif team == "tct clinical":
        sheet = "Clinical #12 Prod Analysis"
        person_idx = 2; cell_idx = 3; mins_idx = 7; tgt_idx = 6; out_idx = 9
    elif team == "tct commercial":
        sheet = "Commercial #12 Prod Analysis"
        person_idx = 2; cell_idx = 3; mins_idx = 7; tgt_idx = 6; out_idx = 9
    else:
        return {}, {}
    df = _read_sheet_as_df(file_path, sheet)
    if df is None or df.empty:
        return {}, {}
    n = df.shape[1]
    need = [person_idx, cell_idx, mins_idx, tgt_idx, out_idx]
    if any(i >= n for i in need):
        return {}, {}
    sub = df.iloc[:, need].copy()
    sub.columns = ["person","station","mins","target","actual"]
    sub["person"]  = sub["person"].astype(str).str.strip()
    sub["station"] = sub["station"].astype(str).str.strip()
    sub["mins"]    = pd.to_numeric(sub["mins"], errors="coerce")
    sub["target"]  = pd.to_numeric(sub["target"], errors="coerce")
    sub["actual"]  = pd.to_numeric(sub["actual"], errors="coerce")
    _bad = {"", "-", "–", "—", "nan"}
    sub = sub[
        ~sub["person"].str.casefold().isin(_bad) &
        ~sub["station"].str.casefold().isin(_bad)
    ]
    sub["person"] = sub["person"].apply(_clean_person_token)
    sub = sub[~sub["person"].apply(_should_exclude_name)]
    sub = sub[(sub[["mins","target","actual"]].notna().any(axis=1))]
    if team_name.strip().casefold() == "pvh":
        sub = sub[sub["station"].str.casefold() != "non-wip"]
    pairs_hours = (
        sub.loc[sub["mins"].notna(), ["station","person","mins"]]
           .rename(columns={"mins":"hours"})
           .assign(hours=lambda d: d["hours"] / 60.0)
    )
    hours_pc = _nest_hours_person_by_station(pairs_hours)
    pairs_outs = sub.loc[:, ["station","person","actual","target"]] \
                    .rename(columns={"actual":"actual_output","target":"target_output"})
    outs_pc = _nest_outputs_person_by_station(pairs_outs)
    return hours_pc, outs_pc
def _outputs_person_and_cell_for_team(file_path: Path, team_name: str) -> tuple[dict, dict]:
    team = (team_name or "").strip().casefold()
    if team in ("svt", "ect", "pvh", "crdn", "aortic"):
        sheet = "#12 Production Analysis"
        person_idx = 2  
        cell_idx   = 3  
        target_idx = 4  
        output_idx = 8  
        if team not in ("aortic",):
            target_idx = 5
    elif team == "tct clinical":
        sheet = "Clinical #12 Prod Analysis"
        person_idx = 2; cell_idx = 3; target_idx = 6; output_idx = 9
    elif team == "tct commercial":
        sheet = "Commercial #12 Prod Analysis"
        person_idx = 2; cell_idx = 3; target_idx = 6; output_idx = 9
    else:
        return {}, {}
    df = _read_sheet_as_df(file_path, sheet)
    if df is None:
        return {}, {}
    by_person = _sum_output_target_by(df, key_col_idx=person_idx, out_col_idx=output_idx, tgt_col_idx=target_idx)
    by_cell   = _sum_output_target_by(df, key_col_idx=cell_idx,   out_col_idx=output_idx, tgt_col_idx=target_idx)
    return by_person, by_cell
def _cell_station_hours_for_team(file_path: Path, team_name: str) -> dict:
    team = (team_name or "").strip().casefold()
    if team in ("svt", "ect", "pvh", "crdn", "aortic"):
        sheet = "#12 Production Analysis"
        key_idx = 3
        mins_idx = 6
    elif team == "tct clinical":
        sheet = "Clinical #12 Prod Analysis"; key_idx = 3; mins_idx = 7
    elif team == "tct commercial":
        sheet = "Commercial #12 Prod Analysis"; key_idx = 3; mins_idx = 7
    else:
        return {}
    df = _read_sheet_as_df(file_path, sheet)
    if df is None or df.empty: return {}
    n = df.shape[1]
    if key_idx >= n or mins_idx >= n: return {}
    sub = df.iloc[:, [key_idx, mins_idx]].copy()
    sub.columns = ["cell_station", "mins"]
    sub["cell_station"] = sub["cell_station"].astype(str).str.strip()
    bad = {"", "-", "–", "—", "nan"}
    sub = sub[~sub["cell_station"].isin(bad)]
    sub["mins"] = pd.to_numeric(sub["mins"], errors="coerce")
    sub = sub.dropna(subset=["mins"])
    if sub.empty: return {}
    agg = sub.groupby("cell_station", dropna=False)["mins"].sum()
    return {k: round(float(v) / 60.0, 2) for k, v in agg.items() if pd.notna(v) and float(v) > 0}
def _hours_by_cs_by_person_for_team(file_path: Path, team_name: str) -> dict:
    team = (team_name or "").strip().casefold()
    ext = file_path.suffix.lower()
    def _norm_station(s: str) -> str | None:
        if s is None:
            return None
        k = str(s).strip()
        if not k:
            return None
        if k in {"-", "–", "—"}:
            return None
        if k.casefold() == "nan":
            return None
        return k
    out: dict[str, dict[str, float]] = {}
    if team == "aortic":
        sheet = "#12 Production Analysis"
        try:
            if ext in (".xlsx", ".xlsm"):
                wb = load_workbook_fast(file_path, data_only=True, read_only=True)
                if sheet not in wb.sheetnames:
                    return {}
                ws = wb[sheet]
                c_idx = col_letter_to_index("C")
                d_idx = col_letter_to_index("D")
                for r in ws.iter_rows(min_row=7, max_row=199,
                                      min_col=min(c_idx, d_idx),
                                      max_col=max(c_idx, d_idx),
                                      values_only=True):
                    person_raw = r[c_idx - min(c_idx, d_idx)]
                    station_raw = r[d_idx - min(c_idx, d_idx)]
                    person = _clean_person_token(str(person_raw) if person_raw is not None else "")
                    station = _norm_station(station_raw)
                    if not person or _should_exclude_name(person) or not station:
                        continue
                    out.setdefault(station, {})
                    out[station][person] = round(out[station].get(person, 0.0) + 2.0, 2)
                return out
            elif ext == ".xlsb":
                df = read_excel_fast(file_path, sheet_name=sheet, engine="pyxlsb")
                C = col_letter_to_index("C") - 1
                D = col_letter_to_index("D") - 1
                sub = df.iloc[6:199, [C, D]].copy()
                sub.columns = ["person", "station"]
                for _, row in sub.iterrows():
                    person = _clean_person_token(str(row["person"]) if pd.notna(row["person"]) else "")
                    station = _norm_station(row["station"])
                    if not person or _should_exclude_name(person) or not station:
                        continue
                    out.setdefault(station, {})
                    out[station][person] = round(out[station].get(person, 0.0) + 2.0, 2)
                return out
        except Exception:
            return {}
    try:
        if team in {"svt", "ect", "pvh", "crdn"}:
            sheet = "#12 Production Analysis"
            person_col = "C"; station_col = "D"; minutes_col = "G"
        elif team == "tct clinical":
            sheet = "Clinical #12 Prod Analysis"
            person_col = "C"; station_col = "D"; minutes_col = "H"
        elif team == "tct commercial":
            sheet = "Commercial #12 Prod Analysis"
            person_col = "C"; station_col = "D"; minutes_col = "H"
        else:
            return {}
        if ext == ".xlsb":
            df = read_excel_fast(file_path, sheet_name=sheet, engine="pyxlsb")
        else:
            df = read_excel_fast(file_path, sheet_name=sheet, engine=None)
        p_i = col_letter_to_index(person_col) - 1
        s_i = col_letter_to_index(station_col) - 1
        m_i = col_letter_to_index(minutes_col) - 1
        nrows = df.shape[0]
        sub = df.iloc[0:200, [p_i, s_i, m_i]].copy()
        sub.columns = ["person", "station", "mins"]
        sub["person"] = sub["person"].astype(str).str.strip()
        sub["station"] = sub["station"].apply(_norm_station)
        if team == "pvh":
            sub = sub[sub["station"].astype(str).str.casefold() != "non-wip"]
        sub["mins"] = pd.to_numeric(sub["mins"], errors="coerce")
        sub = sub.dropna(subset=["station", "mins"])
        sub = sub[sub["mins"] > 0]
        for _, row in sub.iterrows():
            person = _clean_person_token(row["person"])
            if not person or _should_exclude_name(person):
                continue
            station = row["station"]
            hours = float(row["mins"]) / 60.0
            out.setdefault(station, {})
            out[station][person] = round(out[station].get(person, 0.0) + hours, 2)
        return out
    except Exception:
        return {}
def read_metrics_from_file(file_path: Path, cells_cfg: dict, sumcols_cfg: dict) -> dict:
    ext = file_path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return read_with_openpyxl(file_path, cells_cfg, sumcols_cfg)
    elif ext == ".xlsb":
        return read_with_pyxlsb(file_path, cells_cfg, sumcols_cfg)
    else:
        raise ValueError(f"Unsupported file type: {ext}")
def collect_for_team(team_cfg: dict) -> list[dict]:
    team_name = team_cfg["name"]
    pattern = team_cfg.get("pattern", "*.xlsx")
    cells_cfg = team_cfg.get("cells", {})
    sumcols_cfg = team_cfg.get("sum_columns", {})
    files: list[Path] = []
    if "root" in team_cfg:
        root = Path(team_cfg["root"])
        if not root.exists():
            print(f"[WARN] Root not found for {team_name}: {root}", file=sys.stderr)
            return []
        files = [p for p in root.rglob(pattern) if p.is_file()]
    elif "file_glob" in team_cfg:
        files = [Path(p) for p in glob.glob(team_cfg["file_glob"]) if Path(p).is_file()]
    elif "file" in team_cfg:
        p = Path(team_cfg["file"])
        if not p.exists():
            print(f"[WARN] File not found for {team_name}: {p}", file=sys.stderr)
            return []
        files = [p]
    else:
        print(f"[WARN] No 'root', 'file_glob', or 'file' in TEAM_CONFIG for {team_name}", file=sys.stderr)
        return []
    rows = []
    for p in files:
        if team_name == "PVH" and r"\Archive_Production Analysis\2025" in str(p):
            print(f"[PVH-2025][scan] considering: {p}")
        if p.is_dir():
            if team_name == "PVH" and r"\Archive_Production Analysis\2025" in str(p):
                print(f"[PVH-2025][skip] is_dir: {p}")
            continue
        if looks_like_temp(p.name):
            if team_name == "PVH" and r"\Archive_Production Analysis\2025" in str(p):
                print(f"[PVH-2025][skip] temp/lock: {p}")
            continue
        if _is_excluded_path(p):
            if team_name == "PVH" and r"\Archive_Production Analysis\2025" in str(p):
                print(f"[PVH-2025][skip] excluded by path rules: {p}")
            continue
        if p.is_dir() or looks_like_temp(p.name) or _is_excluded_path(p):
            continue
        try:
            period = None
            per_cfg = team_cfg.get("period")
            if per_cfg and isinstance(per_cfg, dict):
                sheet = per_cfg.get("sheet")
                cell  = per_cfg.get("cell")
                if sheet and cell:
                    ext = p.suffix.lower()
                    try:
                        if ext == ".xlsb":
                            val = read_one_cell_xlsb(p, sheet, cell)
                        else:
                            wb = load_workbook(p, data_only=True, read_only=True)
                            if sheet in wb.sheetnames:
                                ws = wb[sheet]
                                val = read_one_cell_openpyxl(ws, cell)
                            else:
                                val = None
                        if isinstance(val, (_dt, _date)):
                            period = val.date() if isinstance(val, _dt) else val
                        elif isinstance(val, (int, float)):
                            period = _excel_serial_to_date(val)
                        elif val is not None:
                            period = dateparser.parse(str(val)).date()
                    except Exception:
                        period = None
            if period is None:
                period = infer_period_date(p)
                if team_name == "PVH" and r"\Archive_Production Analysis\2025" in str(p):
                    print(f"[PVH-2025][warn] No period in C4 / parse failed; falling back to filename/mtime for {p.name}")
            if team_name == "ECT":
                cutoff = _dt(2024, 8, 19).date()
                if isinstance(period, _dt):
                    period = period.date()
                if isinstance(period, _date) and period < cutoff:
                    continue
            if team_name.lower().startswith("tct"):
                today = _dt.today().date()
                if isinstance(period, _dt):
                    period = period.date()
                if isinstance(period, _date) and period > today:
                    print(f"[skip] TCT future period {period} -> {p}")
                    continue
            if team_name == "PVH" and r"\Archive_Production Analysis\2025" in str(p):
                print(f"[PVH-2025][period] {p.name} -> period={period}")
            values = read_metrics_from_file(p, cells_cfg, sumcols_cfg)
            cbs = team_cfg.get("cells_by_sheet") or {}
            if cbs:
                try:
                    wb = load_workbook_fast(p, data_only=True, read_only=True)
                    for sheet_name, mapping in cbs.items():
                        if sheet_name not in wb.sheetnames:
                            continue
                        ws = wb[sheet_name]
                        for out_name, addr in mapping.items():
                            if isinstance(addr, list):
                                values[out_name] = sum_cells_openpyxl(ws, addr)
                            else:
                                values[out_name] = read_one_cell_openpyxl(ws, addr)
                except Exception:
                    pass
            if team_name in ("ECT", "PVH", "CRDN", "Aortic"):
                try:
                    if team_name == "Aortic":
                        values["HC in WIP"] = _aortic_hc_in_wip_from_file(
                            p, "#12 Production Analysis", col_c="C", row_start=7, row_end=199, skip_hidden=True
                        )
                    else:
                        values["HC in WIP"] = _hc_in_wip_from_file(p, "#12 Production Analysis")
                except Exception:
                    values["HC in WIP"] = None
                try:
                    by_person, by_cell = _outputs_person_and_cell_for_team(p, team_name)
                    if by_person:
                        values["Outputs by Person"] = json.dumps(by_person, ensure_ascii=False)
                    if by_cell:
                        values["Outputs by Cell/Station"] = json.dumps(by_cell, ensure_ascii=False)
                except Exception:
                    pass
                try:
                    cs_hours = _cell_station_hours_for_team(p, team_name)
                    if cs_hours:
                        values["Cell/Station Hours"] = json.dumps(cs_hours, ensure_ascii=False)
                except Exception:
                    pass
                try:
                    wb_tmp = load_workbook(p, data_only=True, read_only=True)
                    per_person = {}
                    if "Individual (WIP-Non WIP)" in wb_tmp.sheetnames:
                        ws_ind = wb_tmp["Individual (WIP-Non WIP)"]
                        people_avail = _people_available_openpyxl_generic(
                            ws_ind, name_col="A", avail_col="I", start_row=6, end_row=32, step=3
                        )
                        names = [n for n, _ in people_avail]
                        if "#12 Production Analysis" in wb_tmp.sheetnames and names:
                            ws_pa = wb_tmp["#12 Production Analysis"]
                            completed_hours = _svt_completed_hours_by_person_openpyxl(
                                ws_pa=ws_pa, people=names, name_col="C", minutes_col="G",
                                row_start=7, row_end=199, skip_hidden=True
                            )
                            for name, avail in people_avail:
                                a = completed_hours.get(name, 0.0)
                                per_person[name] = {"actual": round(float(a or 0.0), 2),
                                                    "available": round(float(avail or 0.0), 2)}
                    if per_person:
                        values["Person Hours"] = json.dumps(per_person, ensure_ascii=False)
                except Exception:
                    pass
            if team_name == "Aortic":
                try:
                    per_actual = _aortic_hours_by_col(p, sheet="#12 Production Analysis", col_letter="C")
                    if per_actual:
                        existing_pp = {}
                        if "Person Hours" in values and values["Person Hours"]:
                            try:
                                existing_pp = json.loads(values["Person Hours"])
                            except Exception:
                                existing_pp = {}
                        BAD_NAMES = {"", "#REF!", "nan", "0", "-", "–", "—"}
                        existing_pp = {k: v for k, v in existing_pp.items() if str(k).strip() not in BAD_NAMES}
                        merged = {}
                        for name, hours in per_actual.items():
                            prev = existing_pp.get(name, {"actual": 0.0, "available": 0.0})
                            prev["actual"] = float(hours)
                            merged[name] = prev
                        for name, prev in existing_pp.items():
                            merged.setdefault(name, prev)
                        values["Person Hours"] = json.dumps(merged, ensure_ascii=False)
                except Exception:
                    pass
                try:
                    cs_hours = _aortic_hours_by_col(p, sheet="#12 Production Analysis", col_letter="D")
                    if cs_hours:
                        values["Cell/Station Hours"] = json.dumps(cs_hours, ensure_ascii=False)
                except Exception:
                    pass
                try:
                    ph_raw = json.loads(values.get("Person Hours") or "{}")
                except Exception:
                    ph_raw = {}
                try:
                    ph_norm = _normalize_person_hours(ph_raw)
                except Exception:
                    ph_norm = ph_raw or {}
                completed_total = 0.0
                hc_in_wip = 0
                for rec in (ph_norm or {}).values():
                    try:
                        a = float((rec or {}).get("actual") or 0.0)
                        completed_total += a
                        if a > 0.0:
                            hc_in_wip += 1
                    except Exception:
                        pass
                values["Completed Hours"] = round(completed_total, 2)
                values["HC in WIP"] = int(hc_in_wip)
                try:
                    values["Person Hours"] = json.dumps(ph_norm, ensure_ascii=False)
                except Exception:
                    pass
            if team_name in ("SVT", "TCT Clinical", "TCT Commercial"):
                try:
                    by_person, by_cell = _outputs_person_and_cell_for_team(p, team_name)
                    if by_person:
                        values["Outputs by Person"] = json.dumps(by_person, ensure_ascii=False)
                    if by_cell:
                        values["Outputs by Cell/Station"] = json.dumps(by_cell, ensure_ascii=False)
                except Exception:
                    pass
                try:
                    cs_hours = _cell_station_hours_for_team(p, team_name)
                    if cs_hours:
                        values["Cell/Station Hours"] = json.dumps(cs_hours, ensure_ascii=False)
                except Exception:
                    pass
            try:
                hours_pc2, outs_pc2 = _person_cell_hours_outputs_for_team(p, team_name)
                if hours_pc2 and not values.get("Hours by Cell/Station - by person"):
                    values["Hours by Cell/Station - by person"] = json.dumps(hours_pc2, ensure_ascii=False)
                if outs_pc2:
                    values["Output by Cell/Station - by person"] = json.dumps(outs_pc2, ensure_ascii=False)
                uplh_hours_map = None
                if hours_pc2:
                    uplh_hours_map = hours_pc2
                elif values.get("Hours by Cell/Station - by person"):
                    try:
                        uplh_hours_map = json.loads(values["Hours by Cell/Station - by person"])
                    except Exception:
                        uplh_hours_map = None
                if uplh_hours_map and (outs_pc2 or values.get("Output by Cell/Station - by person")):
                    if not outs_pc2:
                        try:
                            outs_pc2 = json.loads(values["Output by Cell/Station - by person"])
                        except Exception:
                            outs_pc2 = {}
                    uplh_pc2 = _uplh_by_person_by_station(uplh_hours_map or {}, outs_pc2 or {})
                    if uplh_pc2:
                        values["UPLH by Cell/Station - by person"] = json.dumps(uplh_pc2, ensure_ascii=False)
            except Exception:
                pass
            try:
                cs_by_person = _hours_by_cs_by_person_for_team(p, team_name)
                if cs_by_person and not values.get("Hours by Cell/Station - by person"):
                    values["Hours by Cell/Station - by person"] = json.dumps(cs_by_person, ensure_ascii=False)
            except Exception:
                pass
            try:
                if not values.get("UPLH by Cell/Station - by person"):
                    hours_map = cs_by_person or json.loads(values.get("Hours by Cell/Station - by person", "{}"))
                    try:
                        outputs_map = outs_pc2
                    except NameError:
                        outputs_map = None
                    if not outputs_map and values.get("Output by Cell/Station - by person"):
                        outputs_map = json.loads(values["Output by Cell/Station - by person"])
                    uplh_pc = _uplh_by_person_by_station(hours_map or {}, outputs_map or {})
                    if uplh_pc:
                        values["UPLH by Cell/Station - by person"] = json.dumps(uplh_pc, ensure_ascii=False)
            except Exception:
                pass
            sheet_for_hc = None
            if team_name == "TCT Commercial":
                sheet_for_hc = "Commercial #12 Prod Analysis"
            elif team_name == "TCT Clinical":
                sheet_for_hc = "Clinical #12 Prod Analysis"
            if sheet_for_hc:
                values["HC in WIP"] = _hc_in_wip_from_file(p, sheet_for_hc)
            if team_name == "SVT":
                try:
                    wb_tmp = load_workbook(p, data_only=True, read_only=True)
                    if "#12 Production Analysis" in wb_tmp.sheetnames:
                        ws_hc = wb_tmp["#12 Production Analysis"]
                        values["HC in WIP"] = _svt_hc_in_wip_openpyxl(ws_hc)
                except Exception:
                    values["HC in WIP"] = None
            if team_name == "SVT":
                try:
                    wb_vis = load_workbook_fast(p, data_only=True, read_only=False)
                    if "Individual" in wb_vis.sheetnames and "#12 Production Analysis" in wb_vis.sheetnames:
                        ws_ind = wb_vis["Individual"]
                        ws_pa  = wb_vis["#12 Production Analysis"]
                        people_avail = _people_available_openpyxl_generic(
                            ws_ind, name_col="A", avail_col="I", start_row=6, end_row=50, step=3
                        )
                        _exclude = {
                            "team member 1",
                            "team member 2",
                            "team member 3",
                            "team member 4",
                            "total available hours",
                            "total pitches",
                        }
                        _norm = lambda s: str(s).strip().casefold()
                        people_avail = [(n, a) for (n, a) in (people_avail or []) if _norm(n) not in _exclude]
                        names = [n for n, _ in people_avail] if people_avail else []
                        completed_hours = _svt_completed_hours_by_person_openpyxl(
                            ws_pa=ws_pa,
                            people=names,
                            name_col="C",
                            minutes_col="G",
                            row_start=1,
                            row_end=200,
                            skip_hidden=True
                        ) if names else {}
                        per_person = {}
                        for name, avail in (people_avail or []):
                            a = float(completed_hours.get(name, 0.0) or 0.0)
                            try:
                                avail_f = float(avail) if avail is not None else 0.0
                            except Exception:
                                avail_f = 0.0
                            per_person[str(name).strip()] = {"actual": round(a, 2), "available": round(avail_f, 2)}
                        if per_person:
                            values["Person Hours"] = json.dumps(per_person, ensure_ascii=False)
                except Exception:
                    pass
            if p.suffix.lower() == ".xlsb" and team_name in ("TCT Commercial", "TCT Clinical"):
                try:
                    ind_sheet = "Individual(WIP-Non WIP)"
                    pa_sheet  = ("Commercial #12 Prod Analysis" if team_name == "TCT Commercial"
                                else "Clinical #12 Prod Analysis")
                    if team_name == "TCT Commercial":
                        people_avail = _tct_people_available_xlsb(
                            p, ind_sheet, name_col="A", avail_col="I", start_row=24, end_row=68, step=3
                        )
                    else:
                        people_avail = _tct_people_available_xlsb(
                            p, ind_sheet, name_col="Z", avail_col="AG", start_row=24, end_row=68, step=3
                        )
                    names = [n for n, _ in people_avail]
                    completed_hours = {}
                    if names:
                        completed_hours = _completed_hours_by_person_pyxlsb(
                            p, pa_sheet, names, name_col="C", minutes_col="H", row_start=1, row_end=200
                        )
                    if people_avail:
                        per_person = {}
                        for name, avail in people_avail:
                            a = completed_hours.get(name, 0.0)
                            per_person[str(name).strip()] = {
                                "actual":    round(float(a or 0.0), 2),
                                "available": round(float(avail or 0.0), 2),
                            }
                        values["Person Hours"] = json.dumps(per_person, ensure_ascii=False)
                except Exception:
                    pass
                try:
                    hours_pc, outs_pc = _person_cell_hours_outputs_for_team(p, team_name)
                    if hours_pc:
                        values["Hours by Cell/Station - by person"] = json.dumps(hours_pc, ensure_ascii=False)
                    if outs_pc:
                        values["Output by Cell/Station - by person"] = json.dumps(outs_pc, ensure_ascii=False)
                    if hours_pc or outs_pc:
                        uplh_pc = _uplh_by_person_by_station(hours_pc or {}, outs_pc or {})
                        if uplh_pc:
                            values["UPLH by Cell/Station - by person"] = json.dumps(uplh_pc, ensure_ascii=False)
                except Exception:
                    pass
            rows.append({
                "team": team_name,
                "period_date": period,
                "source_file": str(p),
                **values
            })
        except Exception as e:
            error_cols = {}
            for mapping in cells_cfg.values():
                for out_name in mapping.keys():
                    error_cols.setdefault(out_name, None)
            for mapping in sumcols_cfg.values():
                for out_name in mapping.keys():
                    error_cols.setdefault(out_name, None)
            rows.append({
                "team": team_name,
                "period_date": infer_period_date(p),
                "source_file": str(p),
                "error": str(e),
                **error_cols
            })
    if team_name == "PVH":
        kept = sum(1 for r in rows if r.get("period_date") is not None)
        print(f"[PVH] rows collected (incl. 2025): {kept} (total={len(rows)})")
    rows = apply_fallbacks_for_team(rows, team_cfg)
    return rows
def apply_fallbacks_for_team(rows: list[dict], team_cfg: dict) -> list[dict]:
    fb = team_cfg.get("fallback_total_available_hours")
    if not fb:
        return rows
    by_date = {}
    dated_rows = []
    for i, r in enumerate(rows):
        d = r.get("period_date")
        if d:
            dated_rows.append((d, i))
            by_date.setdefault(d, []).append(i)
    dated_rows.sort(key=lambda t: t[0])
    def find_nearest_earlier(target_date):
        lo, hi = 0, len(dated_rows) - 1
        best_idx = None
        while lo <= hi:
            mid = (lo + hi) // 2
            d_mid, idx_mid = dated_rows[mid]
            if d_mid < target_date:
                best_idx = idx_mid
                lo = mid + 1
            else:
                hi = mid - 1
        return best_idx
    for cur_date, row_idx in dated_rows:
        row = rows[row_idx]
        tah_val = row.get("Total Available Hours")
        try:
            is_zero = float(str(tah_val).replace(",", "").strip()) == 0.0
        except Exception:
            is_zero = False
        if not is_zero:
            continue
        target_prev = cur_date - timedelta(days=7)
        prev_indices = by_date.get(target_prev)
        candidate_idx = None
        if prev_indices:
            candidate_idx = prev_indices[0]
        else:
            candidate_idx = find_nearest_earlier(cur_date)
        if candidate_idx is None:
            continue
        prev_file = Path(rows[candidate_idx].get("source_file", ""))
        if not prev_file.exists():
            continue
        repl = sum_column_from_file(
            prev_file,
            fb["sheet"],
            fb["column"],
            include_contains=fb.get("include_contains"),
            exclude_regex=fb.get("exclude_regex"),
            row_start=fb.get("row_start"),
            row_end=fb.get("row_end"),
            skip_hidden=fb.get("skip_hidden", False),
        )
        if repl is not None:
            row["Total Available Hours"] = repl
            row["fallback_used"] = f"{fb['sheet']}!{fb['column']} from {prev_file.name}"
    return rows
def build_master(rows: list[dict]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df = normalize_period_date(df)
    base_cols = ["team", "period_date", "source_file"]
    metric_cols = [c for c in df.columns if c not in base_cols + ["error"]]
    cols = base_cols + metric_cols + (["error"] if "error" in df.columns else [])
    df = df.reindex(columns=cols)
    if "period_date" in df.columns:
        df = df.sort_values(["team", "period_date", "source_file"], ascending=[True, True, True])
    return df
def save_outputs(df: pd.DataFrame):
    if df.empty:
        print("No rows collected. Check paths/sheets/cells.")
        return
    df = _dedupe_by_team_unique_key(df)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as xlw:
        df.to_excel(xlw, index=False, sheet_name="All Metrics")
    preferred_cols = [
        "team", "period_date", "source_file",
        "Total Available Hours", "Completed Hours",
        "Target Output", "Actual Output",
        "Target UPLH", "Actual UPLH",
        "UPLH WP1", "UPLH WP2",
        "HC in WIP", "Actual HC Used",
        "People in WIP", "Person Hours",
        "Outputs by Person", "Outputs by Cell/Station",
        "Cell/Station Hours",
        "Hours by Cell/Station - by person",
        "Output by Cell/Station - by person",
        "UPLH by Cell/Station - by person",
        "Open Complaint Timeliness",
        "fallback_used", "error",
    ]
    cols = [c for c in preferred_cols if c in df.columns]
    out = df.loc[:, cols].copy()
    if "period_date" in out.columns:
        out["period_date"] = pd.to_datetime(out["period_date"], errors="coerce").dt.normalize()
    numeric_cols = {
        "Total Available Hours", "Completed Hours", "Target Output", "Actual Output",
        "Target UPLH", "Actual UPLH", "Actual HC Used", "UPLH WP1", "UPLH WP2"
    } & set(out.columns)
    for c in numeric_cols:
        out[c] = pd.to_numeric(out[c], errors="coerce")
    sort_cols = [c for c in ["team", "period_date", "source_file"] if c in out.columns]
    if sort_cols:
        out = out.sort_values(sort_cols, na_position="last")
    dedupe_subset = [c for c in ["team", "period_date", "source_file"] if c in out.columns]
    if dedupe_subset:
        out = out.drop_duplicates(subset=dedupe_subset, keep="last")
    if "period_date" in out.columns:
        out["period_date"] = pd.to_datetime(out["period_date"], errors="coerce").dt.strftime("%Y-%m-%d")
    out = out.replace({np.nan: ""})
    out.to_csv(
        OUT_CSV,
        index=False,
        sep=",",
        encoding="utf-8",
        lineterminator="\n",
        quoting=csv.QUOTE_MINIMAL,
        date_format="%Y-%m-%d",
    )
    print(f"Saved Excel: {OUT_XLSX.resolve()}")
    print(f"Saved CSV:   {OUT_CSV.resolve()}")
    REPO_DIR.mkdir(parents=True, exist_ok=True)
    try:
        shutil.copyfile(OUT_CSV,  REPO_CSV)
        print(f"Copied CSV  -> {REPO_CSV.resolve()}")
    except Exception as e:
        print(f"[WARN] Copy CSV failed: {e}")
    try:
        shutil.copyfile(OUT_XLSX, REPO_XLSX)
        print(f"Copied XLSX -> {REPO_XLSX.resolve()}")
    except Exception as e:
        print(f"[WARN] Copy XLSX failed: {e}")
    run_apply_closures()
    try:
        git_autocommit_and_push_many(REPO_DIR, [REPO_CSV, REPO_XLSX], branch=GIT_BRANCH)
    except Exception as e:
        print(f"[WARN] Git push (many) failed: {e}")
    def _samefile(a: Path, b: Path) -> bool:
        try:
            return a.resolve().samefile(b.resolve())
        except Exception:
            return str(a.resolve()).lower() == str(b.resolve()).lower()
    if not _samefile(OUT_CSV, REPO_CSV):
        shutil.copyfile(OUT_CSV, REPO_CSV)
        print(f"Copied CSV  -> {REPO_CSV}")
    else:
        print("[info] OUT_CSV and REPO_CSV are the same file; skipping copy.")
    if not _samefile(OUT_XLSX, REPO_XLSX):
        shutil.copyfile(OUT_XLSX, REPO_XLSX)
        print(f"Copied XLSX -> {REPO_XLSX}")
    else:
        print("[info] OUT_XLSX and REPO_XLSX are the same file; skipping copy.")
def merge_with_existing(new_df: pd.DataFrame) -> pd.DataFrame:
    new_df = normalize_period_date(new_df)
    old_frames = []
    if OUT_XLSX.exists():
        try:
            old_xlsx = pd.read_excel(OUT_XLSX, sheet_name="All Metrics")
            old_frames.append(old_xlsx)
        except Exception:
            pass
    if REPO_CSV.exists():
        try:
            old_csv = pd.read_csv(REPO_CSV, dtype=str, keep_default_na=False)
            for col in old_csv.columns:
                if col.lower() in {
                    "total available hours", "completed hours", "target output", "actual output",
                    "target uplh", "actual uplh", "actual hc used", "hc in wip",
                    "open complaint timeliness"
                }:
                    old_csv[col] = pd.to_numeric(old_csv[col], errors="coerce")
            old_frames.append(old_csv)
        except Exception:
            pass
    if OUT_CSV.exists():
        try:
            old_local = pd.read_csv(OUT_CSV, dtype=str, keep_default_na=False)
            for col in old_local.columns:
                if col.lower() in {
                    "total available hours", "completed hours", "target output", "actual output",
                    "target uplh", "actual uplh", "actual hc used", "hc in wip",
                    "open complaint timeliness"
                }:
                    old_local[col] = pd.to_numeric(old_local[col], errors="coerce")
            old_frames.append(old_local)
        except Exception:
            pass
    if old_frames:
        old = pd.concat(old_frames, ignore_index=True)
    else:
        return new_df
    old = normalize_period_date(old)
    team_keys: dict[str, list[str]] = {}
    for cfg in TEAM_CONFIG:
        key = cfg.get("unique_key", ["team", "period_date", "source_file"])
        team_keys[cfg["name"]] = key
    def make_key_row(r) -> tuple:
        kcols = team_keys.get(r.get("team"), ["team", "period_date", "source_file"])
        parts = []
        for c in kcols:
            if c == "period_date":
                ts = pd.to_datetime(r.get(c), errors="coerce")
                parts.append(ts.normalize().date().isoformat() if pd.notna(ts) else None)
            else:
                parts.append(r.get(c))
        return tuple(parts)
    old = old.copy()
    old["_origin"] = "old"
    old["_key"] = old.apply(make_key_row, axis=1)
    new_df = new_df.copy()
    new_df["_origin"] = "new"
    new_df["_key"] = new_df.apply(make_key_row, axis=1)
    combined = pd.concat([old, new_df], ignore_index=True)
    combined = normalize_period_date(combined)
    if "source_file" in combined.columns:
        norm = (combined["source_file"]
                .astype(str)
                .str.lower()
                .str.replace("/", "\\", regex=False))
        is_old = combined["_origin"] == "old"
        pvh_allow_token = "\\cqxm - iv resource site - cos supportive materials\\archive_production analysis\\2025"
        is_pvh = combined["team"].astype(str).str.casefold().eq("pvh")
        allow_pvh_2025 = is_pvh & norm.str.contains(pvh_allow_token, regex=False)
        keep = pd.Series(True, index=combined.index)
        if EXCLUDED_SOURCE_FILES:
            keep &= is_old | allow_pvh_2025 | (~norm.isin(EXCLUDED_SOURCE_FILES))
        if EXCLUDED_DIRS:
            keep &= is_old | allow_pvh_2025 | (~norm.str.startswith(tuple(EXCLUDED_DIRS)))
        combined = combined.loc[keep].copy()
    combined["_origin_rank"] = combined["_origin"].map({"old": 0, "new": 1}).fillna(1)
    combined = combined.sort_values(
        ["team", "period_date", "_origin_rank", "source_file"],
        ascending=[True, True, True, True]
    ).drop_duplicates(subset=["_key"], keep="last")
    combined = _filter_future_periods(combined)
    combined = _filter_ect_min_year(combined)
    combined = combined.drop(columns=[c for c in ["_key", "_origin", "_origin_rank"] if c in combined.columns])
    base_cols = ["team", "period_date", "source_file"]
    metric_cols = [c for c in combined.columns if c not in base_cols + ["error"]]
    cols = base_cols + metric_cols + (["error"] if "error" in combined.columns else [])
    combined = combined.reindex(columns=cols)
    if "period_date" in combined.columns:
        combined = combined.sort_values(["team", "period_date", "source_file"], ascending=[True, True, True])
    dbg = combined.copy()
    pvh = dbg[dbg["team"].astype(str).str.upper().eq("PVH")]
    print(
        "[debug] PVH rows:", len(pvh),
        "with period_date NA:", pvh["period_date"].isna().sum(),
        "unique periods:", pvh["period_date"].dropna().dt.date.nunique()
    )
    return combined
def _dedupe_by_team_unique_key(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    team_keys = {}
    for cfg in TEAM_CONFIG:
        key = cfg.get("unique_key", ["team", "period_date", "source_file"])
        team_keys[cfg["name"]] = key
    def _make_key_row(r) -> tuple:
        kcols = team_keys.get(r.get("team"), ["team", "period_date", "source_file"])
        parts = []
        for c in kcols:
            if c == "period_date":
                ts = pd.to_datetime(r.get(c), errors="coerce")
                parts.append(ts.normalize().date().isoformat() if pd.notna(ts) else None)
            else:
                parts.append(r.get(c))
        return tuple(parts)
    df = df.copy()
    
    df["_dedupe_key"] = df.apply(_make_key_row, axis=1)
    df = (df
          .sort_values(["team", "period_date", "source_file"], na_position="last")
          .drop_duplicates(subset=["_dedupe_key"], keep="last")
          .drop(columns=["_dedupe_key"]))
    return df
def _read_timeliness_csv_standardized() -> pd.DataFrame:
    p = TIMELINESS_CSV
    if not p.exists():
        return pd.DataFrame(columns=["team", "period_date", "Open Complaint Timeliness"])
    try:
        t = pd.read_csv(p, dtype=str, keep_default_na=False)
    except Exception as e:
        print(f"[timeliness] Failed to read {p}: {e}")
        return pd.DataFrame(columns=["team", "period_date", "Open Complaint Timeliness"])
    if t.shape[1] < 3:
        t = t.iloc[:, :3].copy()
        t.columns = ["team", "period_date", "Open Complaint Timeliness"]
    else:
        lower_cols = [str(c).strip().lower() for c in t.columns]
        def _first_match(names):
            for want in names:
                if want in lower_cols:
                    return t.columns[lower_cols.index(want)]
            return None
        team_col = _first_match(["team"])
        date_col = _first_match(["period_date", "period", "date"])
        val_col  = _first_match(["open complaint timeliness","timeliness","value","metric"])
        if not (team_col and date_col and val_col):
            t = t.iloc[:, :3].copy()
            t.columns = ["team", "period_date", "Open Complaint Timeliness"]
        else:
            t = t.rename(columns={
                team_col: "team", date_col: "period_date", val_col: "Open Complaint Timeliness"
            })
    t["team"] = t["team"].astype(str).str.strip()
    t["period_date"] = pd.to_datetime(t["period_date"], errors="coerce").dt.normalize()
    s = t["Open Complaint Timeliness"].astype(str).str.strip().str.replace("%","", regex=False)
    t["Open Complaint Timeliness"] = pd.to_numeric(s.str.extract(r'([+-]?\d+(?:\.\d+)?)', expand=False), errors="coerce")
    t = t.drop_duplicates(subset=["team","period_date"], keep="last")
    return t
def ensure_timeliness_placeholders(metrics_df: pd.DataFrame):
    if metrics_df.empty or "team" not in metrics_df.columns or "period_date" not in metrics_df.columns:
        return
    pairs = (
        metrics_df[["team","period_date"]]
        .dropna()
        .copy()
    )
    pairs["team"] = pairs["team"].astype(str).str.strip()
    pairs["period_date"] = pd.to_datetime(pairs["period_date"], errors="coerce").dt.normalize()
    pairs = pairs.dropna().drop_duplicates()
    t = _read_timeliness_csv_standardized()
    pairs["_team_key"] = pairs["team"].str.casefold()
    t["_team_key"] = t["team"].str.casefold()
    missing = (
        pairs.merge(t[["_team_key","period_date"]], on=["_team_key","period_date"], how="left", indicator=True)
             .loc[lambda d: d["_merge"] == "left_only", ["team","period_date"]]
             .drop_duplicates()
    )
    if missing.empty:
        print("[timeliness] No missing timeliness rows to add.")
        return
    placeholders = missing.copy()
    placeholders["Open Complaint Timeliness"] = np.nan
    out = pd.concat([t[["team","period_date","Open Complaint Timeliness"]], placeholders], ignore_index=True)
    out = out.drop_duplicates(subset=["team","period_date"], keep="last")
    out = out.sort_values(["team","period_date"]).reset_index(drop=True)
    TIMELINESS_CSV.parent.mkdir(parents=True, exist_ok=True)
    out_to_csv = out.copy()
    out_to_csv["period_date"] = pd.to_datetime(out_to_csv["period_date"], errors="coerce").dt.strftime("%Y-%m-%d")
    out_to_csv.to_csv(TIMELINESS_CSV, index=False)
    print(f"[timeliness] Added {len(placeholders)} placeholder row(s) to {TIMELINESS_CSV}.")
    git_autocommit_and_push(REPO_DIR, TIMELINESS_CSV, branch=GIT_BRANCH)
def add_open_complaint_timeliness(df: pd.DataFrame) -> pd.DataFrame:
    try:
        p = TIMELINESS_CSV
    except NameError:
        p = Path(r"C:\heijunka-dev") / "timeliness.csv"
    if not p.exists():
        print(f"[timeliness] {p} not found; skipping join.")
        return df
    try:
        t = pd.read_csv(p, dtype=str, keep_default_na=False)
    except Exception as e:
        print(f"[timeliness] Failed to read {p}: {e}")
        return df
    if t.shape[1] < 3:
        print("[timeliness] Expected at least 3 columns (team, date/period_date, value). Skipping join.")
        return df
    lower_cols = [str(c).strip().lower() for c in t.columns]
    def _first_match(names):
        for want in names:
            if want in lower_cols:
                return t.columns[lower_cols.index(want)]
        return None
    team_col = _first_match(["team"])
    date_col = _first_match(["period_date", "period", "date"])
    val_col  = _first_match(["open complaint timeliness", "timeliness", "value", "metric"])
    if not (team_col and date_col and val_col):
        t = t.iloc[:, :3].copy()
        t.columns = ["team", "period_date", "Open Complaint Timeliness"]
    else:
        t = t.rename(columns={
            team_col: "team",
            date_col: "period_date",
            val_col:  "Open Complaint Timeliness"
        })
    def _team_key(s: pd.Series) -> pd.Series:
        return s.astype(str).str.strip()
    def _to_num(series: pd.Series) -> pd.Series:
        s = series.astype(str).str.strip()
        s = s.str.replace("%", "", regex=False)
        extracted = s.str.extract(r'([+-]?\d+(?:\.\d+)?)', expand=False)
        return pd.to_numeric(extracted, errors="coerce")
    t["team"] = t["team"].astype(str).str.strip()
    t["_team_key"] = _team_key(t["team"])
    t["period_date"] = pd.to_datetime(t["period_date"], errors="coerce").dt.normalize()
    t["Open Complaint Timeliness"] = _to_num(t["Open Complaint Timeliness"])
    t = t.dropna(subset=["_team_key", "period_date"]).drop_duplicates(subset=["_team_key", "period_date"], keep="last")
    out = df.copy()
    if "team" not in out.columns or "period_date" not in out.columns:
        print("[timeliness] 'team'/'period_date' not found in metrics df; skipping join.")
        return df
    out["team"] = out["team"].astype(str).str.strip()
    out["_team_key"] = _team_key(out["team"])
    out["period_date"] = pd.to_datetime(out["period_date"], errors="coerce").dt.normalize()
    exact = out.merge(
        t[["_team_key", "period_date", "Open Complaint Timeliness"]],
        on=["_team_key", "period_date"],
        how="left",
        suffixes=("", "_t")
    )
    if "Open Complaint Timeliness" in out.columns:
        left_vals  = pd.to_numeric(exact["Open Complaint Timeliness"], errors="coerce")
        right_vals = pd.to_numeric(exact.get("Open Complaint Timeliness_t"), errors="coerce")
        exact["Open Complaint Timeliness"] = right_vals.combine_first(left_vals)
        if "Open Complaint Timeliness_t" in exact.columns:
            exact = exact.drop(columns=["Open Complaint Timeliness_t"])
    else:
        if "Open Complaint Timeliness_t" in exact.columns and "Open Complaint Timeliness" not in exact.columns:
            exact = exact.rename(columns={"Open Complaint Timeliness_t": "Open Complaint Timeliness"})
    mask_missing = exact["Open Complaint Timeliness"].isna() if "Open Complaint Timeliness" in exact.columns else pd.Series(False, index=exact.index)
    def _week_monday(s: pd.Series) -> pd.Series:
        d = pd.to_datetime(s, errors="coerce").dt.normalize()
        return d - pd.to_timedelta(d.dt.weekday, unit="D")
    if mask_missing.any():
        tmp = exact.loc[mask_missing].copy()
        tmp["_week_key"] = _week_monday(tmp["period_date"])
        t2 = t.copy()
        t2["_week_key"] = _week_monday(t2["period_date"])
        t2_week = (
            t2[["_team_key", "_week_key", "period_date", "Open Complaint Timeliness"]]
            .sort_values(["_team_key", "_week_key", "period_date"])
            .drop_duplicates(["_team_key", "_week_key"], keep="last")
        )
        wk = tmp.merge(
            t2_week,
            on=["_team_key", "_week_key"],
            how="left",
            suffixes=("", "_wk")
        )
        fill = pd.to_numeric(wk["Open Complaint Timeliness"], errors="coerce")
        exact.loc[mask_missing, "Open Complaint Timeliness"] = fill.values
        mask_missing = exact["Open Complaint Timeliness"].isna()
    if mask_missing.any():
        tol = pd.Timedelta(days=6)
        left_near = exact.loc[mask_missing, ["_team_key", "period_date"]].copy()
        left_near = left_near.sort_values(["_team_key", "period_date"])
        right_near = t[["_team_key", "period_date", "Open Complaint Timeliness"]].copy()
        right_near = right_near.sort_values(["_team_key", "period_date"])
        filled_vals = []
        for team_key, left_g in left_near.groupby("_team_key"):
            r_g = right_near[right_near["_team_key"] == team_key]
            if r_g.empty:
                filled_vals.extend([np.nan] * len(left_g))
                continue
            merged_g = pd.merge_asof(
                left_g.sort_values("period_date"),
                r_g.sort_values("period_date"),
                on="period_date",
                direction="nearest",
                tolerance=tol
            )
            filled_vals.extend(merged_g["Open Complaint Timeliness"].tolist())
        exact.loc[mask_missing, "Open Complaint Timeliness"] = pd.to_numeric(filled_vals, errors="coerce")
    exact = exact.drop(columns=[c for c in ["_team_key"] if c in exact.columns])
    return exact
def _available_team_names() -> list[str]:
    seen = []
    for cfg in TEAM_CONFIG:
        nm = cfg.get("name", "").strip()
        if nm and nm not in seen:
            seen.append(nm)
    return seen
def run_once(selected_teams: set[str] | None = None):
    if selected_teams:
        wanted = {t.strip().casefold() for t in selected_teams if str(t).strip()}
        cfg_iter = [c for c in TEAM_CONFIG if c.get("name","").strip().casefold() in wanted]
        if not cfg_iter:
            print(f"[info] No TEAM_CONFIG entries matched {sorted(selected_teams)}")
            return
        print("[info] Running for teams:", ", ".join(sorted({c['name'] for c in cfg_iter})))
    else:
        cfg_iter = TEAM_CONFIG
    all_rows = []
    for cfg in cfg_iter:
        else:
            all_rows.extend(collect_for_team(cfg))     
    df = build_master(all_rows)
    df = _filter_future_periods(df)
    df = _filter_ect_min_year(df)
    df["Target UPLH"] = df.apply(lambda r: safe_div(r.get("Target Output"), r.get("Total Available Hours")), axis=1)
    df["Actual UPLH"] = df.apply(lambda r: safe_div(r.get("Actual Output"), r.get("Completed Hours")), axis=1)
    df["Target UPLH"] = df["Target UPLH"].round(2)
    df["Actual UPLH"] = df["Actual UPLH"].round(2)
    if "UPLH WP1" in df.columns: df["UPLH WP1"] = pd.to_numeric(df["UPLH WP1"], errors="coerce").round(2)
    if "UPLH WP2" in df.columns: df["UPLH WP2"] = pd.to_numeric(df["UPLH WP2"], errors="coerce").round(2)
    df["Actual HC Used"] = pd.to_numeric(df.get("Completed Hours"), errors="coerce") / 30
    df["Actual HC Used"] = df["Actual HC Used"].round(2)
    df = merge_with_existing(df)
    try:
        if REPO_CSV.exists():
            repo = pd.read_csv(REPO_CSV, dtype=str, keep_default_na=False)
        else:
            repo = pd.DataFrame(columns=df.columns)
        local_hist = pd.read_csv(OUT_CSV, dtype=str, keep_default_na=False) if OUT_CSV.exists() else pd.DataFrame(columns=df.columns)
        def _norm_dates(s): return pd.to_datetime(s, errors="coerce").dt.normalize()
        repo_svt  = repo[repo.get("team","") == "SVT"].copy()
        local_svt = local_hist[local_hist.get("team","") == "SVT"].copy()
        curr_svt  = df[df["team"] == "SVT"].copy()
        r_dates = _norm_dates(repo_svt["period_date"])
        l_dates = _norm_dates(local_svt["period_date"])
        c_dates = _norm_dates(curr_svt["period_date"])
        r_n, l_n, c_n = r_dates.nunique(), l_dates.nunique(), c_dates.nunique()
        if (c_n + 1 < r_n) or (c_n + 1 < l_n):
            print("[safety] SVT history shrank; restoring union of repo/local/current.")
            df = merge_with_existing(df)
    except Exception:
        pass
    try:
        if REPO_CSV.exists():
            repo = pd.read_csv(REPO_CSV, dtype=str, keep_default_na=False)
            repo_svt = repo[repo["team"] == "SVT"]
            curr_svt = df[df["team"] == "SVT"]
            r_n = pd.to_datetime(repo_svt["period_date"], errors="coerce").dt.normalize().nunique()
            c_n = pd.to_datetime(curr_svt["period_date"], errors="coerce").dt.normalize().nunique()
            if c_n + 3 < r_n:
                print("[safety] SVT periods dropped unexpectedly; restoring repo history union.")
                df = merge_with_existing(df)
    except Exception:
        pass
    ensure_timeliness_placeholders(df)
    df = add_open_complaint_timeliness(df)
    df = _dedupe_by_team_unique_key(df)
    save_outputs(df)
    try:
        counts = df.groupby("team", dropna=False)["period_date"].nunique().sort_values(ascending=False)
        print("\n[summary] weeks per team (unique period_date):")
        print(counts.to_string())
    except Exception:
        pass
def watch_mode(selected_teams: set[str] | None = None):
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
    class NewFileHandler(FileSystemEventHandler):
        def on_created(self, event):
            if event.is_directory:
                return
            name = event.src_path.lower()
            if (name.endswith(".xlsx") or name.endswith(".xlsm") or name.endswith(".xlsb")) and not looks_like_temp(name):
                print(f"[watch] New file: {event.src_path}")
                time.sleep(2)
                run_once()
    if selected_teams:
        wanted = {t.strip().casefold() for t in selected_teams if str(t).strip()}
        cfg_iter = [c for c in TEAM_CONFIG if c.get("name","").strip().casefold() in wanted]
        if not cfg_iter:
            print(f"[watch][info] No TEAM_CONFIG entries matched {sorted(selected_teams)}")
            return
        print("[watch] Watching only teams:", ", ".join(sorted({c['name'] for c in cfg_iter})))
    else:
        cfg_iter = TEAM_CONFIG
    roots = []
    for cfg in TEAM_CONFIG:
        r = cfg.get("root")
        if r:
            roots.append(r)
    run_once(selected_teams)
    obs = Observer()
    for r in roots:
        rp = Path(r)
        if rp.exists():
            obs.schedule(NewFileHandler(), str(rp), recursive=True)
            print(f"[watch] Watching: {rp}")
        else:
            print(f"[watch][WARN] Missing root: {rp}")
    obs.start()
    print("[watch] Running. Ctrl+C to stop.")
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        obs.stop()
    obs.join()
def main():
    parser = argparse.ArgumentParser(description="Aggregate metrics from synced SharePoint Excel files (.xlsx/.xlsm/.xlsb)")
    parser.add_argument("--watch", action="store_true", help="Watch folders and refresh on new files")
    parser.add_argument("--team", action="append", default=None,
                        help="Run only for this team name (repeatable). Examples: --team SVT --team Aortic")
    parser.add_argument("--list-teams", action="store_true", help="List available team names and exit")
    args = parser.parse_args()
    if args.list_teams:
        print("Available teams:")
        for nm in _available_team_names():
            print(" -", nm)
        return
    selected = set(args.team) if args.team else None
    if args.watch:
        watch_mode(selected)
    else:
        run_once(selected)
if __name__ == "__main__":
    main()