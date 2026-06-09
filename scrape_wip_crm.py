from __future__ import annotations
import argparse
import csv
import datetime as _dt
import json
import os
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from zipfile import BadZipFile
LIT_LETTERS_TEAM = "Lit & Letters"
def _is_lit_letters_path(path: str) -> bool:
    base = os.path.basename(_norm_path(path)).lower()
    return (
        "pab for letters" in base
        and "lit" in base
        and "principals" in base
    )
TEAM_BY_SOURCE: Dict[str, str] = {
    r"C:\Users\wadec8\Medtronic PLC\MCS COS Transformation - VMB Scheduling\Heijunka Current.xlsm": "MCS",
    r"C:\Users\wadec8\Medtronic PLC\Diagnostics MDR - Heijunka and Production Analysis": "CDS",
    r"C:\Users\wadec8\Medtronic PLC\Defibrillation Solutions - Schedule and PAB": "DS",
    r"C:\Users\wadec8\Medtronic PLC\Cardiac Pacing Therapies CQXM - Heijunka & PAB": "CPT",
    r"C:\Users\wadec8\Medtronic PLC\Tier1 PXM - Non Implantables - Heijunka": "NI",
    r"C:\Users\wadec8\Medtronic PLC\CRM CQXM Reports - 1.9 Heijunka Tracker": "NI & PM MEIC",
    r"C:\Users\wadec8\Medtronic PLC\Tier1 PXM - Non Implantables - Heijunka\PM-CTS PAB":"PM-CTS",
}
EXCLUDED_FILES = {
    os.path.normpath(r"C:\Users\wadec8\Medtronic PLC\Cardiac Pacing Therapies CQXM - Heijunka & PAB\Archive\2026\4. April 2026\Not USED Week 20 Apr 2026 Heijunka & PAB.xlsm"),
    os.path.normpath(r"C:\Users\wadec8\Medtronic PLC\Defibrillation Solutions - Schedule and PAB\Archive\Assigned DS COS PEs for 2025.xlsx"),
    os.path.normpath(r"C:\Users\wadec8\Medtronic PLC\Defibrillation Solutions - Schedule and PAB\Archive\(will be archived) DS_Schedule_PAS 6.5 V1.xlsx"),
    os.path.normpath(r"C:\Users\wadec8\Medtronic PLC\Defibrillation Solutions - Schedule and PAB\Archive\CPT Event Support.xlsx"),
    os.path.normpath(r"C:\Users\wadec8\Medtronic PLC\Defibrillation Solutions - Schedule and PAB\Archive\DS Production Analysis Sheet and Schedule.xlsx"),
    os.path.normpath(r"C:\Users\wadec8\Medtronic PLC\Tier1 PXM - Non Implantables - Heijunka\PM-CTS PAB\Revised PM-CTS Template.xlsm"),
    
}
TEAM_BY_BASENAME: Dict[str, str] = {
    "Heijunka Current.xlsm": "MCS",
}
CDS_NEW_HOURS_START = _dt.date(2026, 4, 24)
CDS_NEW_HOURS_START = _dt.date(2026, 4, 24)
AF_ACTUAL_HOURS_START = _dt.date(2026, 4, 24)
def _use_af_actual_hours(period: Optional[_dt.date]) -> bool:
    return isinstance(period, _dt.date) and period > AF_ACTUAL_HOURS_START
def _cds_use_new_hours_layout(period: Optional[_dt.date]) -> bool:
    return isinstance(period, _dt.date) and period >= CDS_NEW_HOURS_START
MCS_ROOT_HINT = _norm_mcs = os.path.normpath(r"C:\Users\wadec8\Medtronic PLC\MCS COS Transformation - VMB Scheduling")
CDS_ROOT_HINT = _norm_cds = os.path.normpath(r"C:\Users\wadec8\Medtronic PLC\Diagnostics MDR - Heijunka and Production Analysis")
DS_ROOT_HINT = _norm_ds = os.path.normpath(r"C:\Users\wadec8\Medtronic PLC\Defibrillation Solutions - Schedule and PAB")
CPT_ROOT_HINT = _norm_cpt = os.path.normpath(r"C:\Users\wadec8\Medtronic PLC\Cardiac Pacing Therapies CQXM - Heijunka & PAB")
NI_ROOT_HINT = _norm_ni = os.path.normpath(r"C:\Users\wadec8\Medtronic PLC\Tier1 PXM - Non Implantables - Heijunka")
MEIC_ROOT_HINT = _norm_meic = os.path.normpath(r"C:\Users\wadec8\Medtronic PLC\CRM CQXM Reports - 1.9 Heijunka Tracker")
PM_CTS_ROOT_HINT = norm_pm_cts = os.path.normpath(r"C:\Users\wadec8\Medtronic PLC\Tier1 PXM - Non Implantables - Heijunka\PM-CTS PAB")
_AVAIL_PAT = re.compile(r"\bavailability\b", re.IGNORECASE)
_PROD_PAT = re.compile(r"\b(production|product)\s+analysis\b", re.IGNORECASE)
_MONTH_MAP = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}
def _cell_date(v: Any, *, default_year: Optional[int] = None) -> Optional[_dt.date]:
    if v is None:
        return None
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    if isinstance(v, (int, float)):
        try:
            return (_dt.datetime(1899, 12, 30) + _dt.timedelta(days=float(v))).date()
        except Exception:
            return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%m-%d-%y"):
            try:
                return _dt.datetime.strptime(s, fmt).date()
            except ValueError:
                pass
        return parse_period_date_from_text(s, default_year=default_year)
    return None
def _col_idx(col: str) -> int:
    n = 0
    for ch in col.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n
def _norm_ws_text(v: Any) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip().lower())
def _find_col_by_header(ws: Worksheet, header_text: str, *, max_header_row: int = 10) -> Optional[int]:
    want = _norm_ws_text(header_text)
    for r in range(1, min(ws.max_row, max_header_row) + 1):
        for c in range(1, ws.max_column + 1):
            if _norm_ws_text(ws.cell(r, c).value) == want:
                return c
    return None
def _find_number_right_of_label(ws: Worksheet, label_text: str, *, lookahead: int = 8) -> Optional[float]:
    want = _norm_ws_text(label_text)
    for row in ws.iter_rows():
        for cell in row:
            if want in _norm_ws_text(cell.value):
                for c in range(cell.column + 1, min(ws.max_column, cell.column + lookahead) + 1):
                    n = _cell_number(ws.cell(cell.row, c).value)
                    if n is not None:
                        return n
    return None
def _is_excluded_station_lit(area: Any) -> bool:
    s = str(area).strip().lower() if area is not None else ""
    return s in {"non-wip", "essential non-wip"}
def _lit_cell_station(area: Any, subarea: Any) -> str:
    a = str(area).strip() if area is not None else ""
    s = str(subarea).strip() if subarea is not None else ""
    return s if s and s.lower() != "no subareas" else a
def _iter_rows_lit_pab(
    ws_pab: Worksheet, start_row: int = 2
) -> Iterable[Tuple[int, str, str, str, Optional[float], Optional[float]]]:
    for r in range(start_row, ws_pab.max_row + 1):
        person = ws_pab[f"C{r}"].value
        area = ws_pab[f"D{r}"].value
        subarea = ws_pab[f"E{r}"].value
        mins = _cell_number(ws_pab[f"H{r}"].value)
        output = _cell_number(ws_pab[f"J{r}"].value)
        p = str(person).strip() if person is not None else ""
        a = str(area).strip() if area is not None else ""
        s = str(subarea).strip() if subarea is not None else ""
        if p == "" and a == "" and s == "" and mins is None and output is None:
            continue
        hours = mins / 60.0 if mins is not None else None
        yield (r, p, a, s, hours, output)
def compute_total_available_hours_lit(ws_perf: Worksheet) -> Optional[float]:
    return _find_number_right_of_label(ws_perf, "Total Team workable hours")
def compute_completed_hours_lit(
    ws_perf: Worksheet,
) -> Tuple[Optional[float], Dict[str, float], List[str]]:
    actual_col = _find_col_by_header(ws_perf, "Overall VIP hours") or _col_idx("AI")
    total = _find_number_right_of_label(ws_perf, "Total WIP Hours")
    actual_by_person: Dict[str, float] = {}
    people_in_wip: List[str] = []
    seen = set()
    for r in range(5, ws_perf.max_row + 1):
        person = ws_perf[f"A{r}"].value
        p = str(person).strip() if person is not None else ""
        if not p or is_excluded_person(p):
            continue
        actual = _cell_number(ws_perf.cell(r, actual_col).value)
        if actual is None or actual == 0:
            continue
        actual_by_person[p] = actual_by_person.get(p, 0.0) + actual
        if p not in seen:
            seen.add(p)
            people_in_wip.append(p)
    if total is None:
        total = sum(actual_by_person.values())
    return total, actual_by_person, people_in_wip
def compute_person_available_hours_lit(ws_perf: Worksheet) -> Dict[str, float]:
    available_col = _find_col_by_header(ws_perf, "Total Workable Hours") or _col_idx("AM")
    out: Dict[str, float] = {}
    for r in range(5, ws_perf.max_row + 1):
        person = ws_perf[f"A{r}"].value
        p = str(person).strip() if person is not None else ""
        if not p or is_excluded_person(p):
            continue
        available = _cell_number(ws_perf.cell(r, available_col).value)
        if available is not None:
            out[p] = out.get(p, 0.0) + available
    return out
def compute_target_actual_output_lit(ws_pab: Worksheet) -> Tuple[float, float]:
    target = 0.0
    actual = 0.0
    for _, _, area, _, _, output in _iter_rows_lit_pab(ws_pab):
        if _is_excluded_station_lit(area):
            continue
        if output is not None:
            actual += output
    return target, actual
def compute_outputs_by_person_lit(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, person, area, _, _, output in _iter_rows_lit_pab(ws_pab):
        if not person or is_excluded_person(person) or _is_excluded_station_lit(area):
            continue
        out.setdefault(person, {"output": 0.0, "target": 0.0})
        if output is not None:
            out[person]["output"] += output
    return out
def compute_outputs_by_station_lit(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, _, area, subarea, _, output in _iter_rows_lit_pab(ws_pab):
        if _is_excluded_station_lit(area):
            continue
        station = _lit_cell_station(area, subarea)
        if not station:
            continue
        out.setdefault(station, {"output": 0.0, "target": 0.0})
        if output is not None:
            out[station]["output"] += output
    return out
def compute_station_hours_lit(ws_pab: Worksheet) -> Tuple[Dict[str, float], Dict[str, Dict[str, float]]]:
    station_hours: Dict[str, float] = {}
    station_hours_by_person: Dict[str, Dict[str, float]] = {}
    for _, person, area, subarea, hours, _ in _iter_rows_lit_pab(ws_pab):
        if _is_excluded_station_lit(area):
            continue
        station = _lit_cell_station(area, subarea)
        if not station or hours is None:
            continue
        station_hours[station] = station_hours.get(station, 0.0) + hours
        if person and not is_excluded_person(person):
            station_hours_by_person.setdefault(station, {})
            station_hours_by_person[station][person] = (
                station_hours_by_person[station].get(person, 0.0) + hours
            )
    return station_hours, station_hours_by_person
def compute_output_by_station_by_person_lit(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, person, area, subarea, _, output in _iter_rows_lit_pab(ws_pab):
        if not person or is_excluded_person(person) or _is_excluded_station_lit(area):
            continue
        station = _lit_cell_station(area, subarea)
        if not station or output is None:
            continue
        out.setdefault(station, {})
        out[station][person] = out[station].get(person, 0.0) + output
    return out
def scrape_one_workbook_lit_letters(path: str) -> List[Dict[str, Any]]:
    team = LIT_LETTERS_TEAM
    wb = load_workbook(path, data_only=True)
    err_msgs: List[str] = []
    ws_pab = wb[_sheet_ci(wb, "#3 PAB")] if _sheet_ci(wb, "#3 PAB") else None
    ws_perf = wb[_sheet_ci(wb, "#6 Performance WIP Time")] if _sheet_ci(wb, "#6 Performance WIP Time") else None
    if ws_pab is None:
        err_msgs.append("missing_#3_pab_sheet")
    if ws_perf is None:
        err_msgs.append("missing_#6_performance_wip_time_sheet")
    period = parse_period_date_from_filename(path, default_year=2026)
    if period is None:
        err_msgs.append("missing_period_date_from_filename")
    total_available = None
    completed_hours = None
    actual_hours_by_person: Dict[str, float] = {}
    people: List[str] = []
    person_avail: Dict[str, float] = {}
    target_output = None
    actual_output = None
    outputs_by_person: Dict[str, Dict[str, float]] = {}
    outputs_by_station: Dict[str, Dict[str, float]] = {}
    station_hours: Dict[str, float] = {}
    station_hours_by_person: Dict[str, Dict[str, float]] = {}
    output_by_station_by_person: Dict[str, Dict[str, float]] = {}
    uplh_by_station_by_person: Dict[str, Dict[str, float]] = {}
    try:
        if ws_perf is not None:
            total_available = compute_total_available_hours_lit(ws_perf)
            completed_hours, actual_hours_by_person, people = compute_completed_hours_lit(ws_perf)
            person_avail = compute_person_available_hours_lit(ws_perf)
        if ws_pab is not None:
            target_output, actual_output = compute_target_actual_output_lit(ws_pab)
            outputs_by_person = compute_outputs_by_person_lit(ws_pab)
            outputs_by_station = compute_outputs_by_station_lit(ws_pab)
            station_hours, station_hours_by_person = compute_station_hours_lit(ws_pab)
            output_by_station_by_person = compute_output_by_station_by_person_lit(ws_pab)
            uplh_by_station_by_person = compute_uplh_by_station_by_person(
                output_by_station_by_person,
                station_hours_by_person,
            )
    except Exception as e:
        err_msgs.append(f"lit_letters_parse_error: {e!r}")
    target_uplh = safe_div(float(target_output or 0.0), float(completed_hours or 0.0))
    actual_uplh = safe_div(float(actual_output or 0.0), float(completed_hours or 0.0))
    hc_in_wip = len(people) if people else 0
    actual_hc_used = safe_div(float(completed_hours or 0.0), 32.5)
    return [{
        "team": team,
        "period_date": iso_date(period),
        "source_file": path,
        "Total Available Hours": float(total_available) if total_available is not None else "",
        "Completed Hours": float(completed_hours) if completed_hours is not None else "",
        "Target Output": float(target_output) if target_output is not None else "",
        "Actual Output": float(actual_output) if actual_output is not None else "",
        "Target UPLH": float(target_uplh) if target_uplh is not None else "",
        "Actual UPLH": float(actual_uplh) if actual_uplh is not None else "",
        "UPLH WP1": "",
        "UPLH WP2": "",
        "HC in WIP": hc_in_wip,
        "Actual HC Used": float(actual_hc_used) if actual_hc_used is not None else "",
        "People in WIP": dumps_json(people) if ws_perf is not None else "",
        "Person Hours": build_person_hours_json(person_avail, actual_hours_by_person) if ws_perf is not None else "",
        "Outputs by Person": dumps_json(outputs_by_person) if ws_pab is not None else "",
        "Outputs by Cell/Station": dumps_json(outputs_by_station) if ws_pab is not None else "",
        "Cell/Station Hours": dumps_json(station_hours) if ws_pab is not None else "",
        "Hours by Cell/Station - by person": dumps_json(station_hours_by_person) if ws_pab is not None else "",
        "Output by Cell/Station - by person": dumps_json(output_by_station_by_person) if ws_pab is not None else "",
        "UPLH by Cell/Station - by person": dumps_json(uplh_by_station_by_person) if ws_pab is not None else "",
        "error": "; ".join(err_msgs) if err_msgs else "",
        "Closures": "",
        "Opened": "",
    }]
def _is_excluded_station_cds(v: Any) -> bool:
    s = str(v).strip().lower() if v is not None else ""
    return s in {"non-wip", "essential non-wip"}
def _iter_rows_cds_pab(
    ws_pab: Worksheet, start_row: int = 2
) -> Iterable[Tuple[int, str, str, Optional[float], Optional[float], Optional[float]]]:
    for r in range(start_row, ws_pab.max_row + 1):
        person = ws_pab[f"C{r}"].value
        cell_station = ws_pab[f"D{r}"].value
        target_g = _cell_number(ws_pab[f"G{r}"].value)
        hours_i = _cell_number(ws_pab[f"I{r}"].value)
        actual_j = _cell_number(ws_pab[f"J{r}"].value)
        p = str(person).strip() if person is not None else ""
        cs = str(cell_station).strip() if cell_station is not None else ""
        if p == "" and cs == "" and target_g is None and hours_i is None and actual_j is None:
            continue
        yield (r, p, cs, target_g, hours_i, actual_j)
def compute_period_date_cds(ws_metrics: Worksheet) -> Optional[_dt.date]:
    d = _cell_date(ws_metrics["B3"].value, default_year=2026)
    if d is None:
        return None
    return d - _dt.timedelta(days=4)
def compute_total_available_hours_cds(ws_wip_plan: Worksheet) -> Optional[float]:
    return _cell_number(ws_wip_plan["CW3"].value)
def _cds_use_r_layout(ws_perf: Worksheet) -> bool:
    for r in range(5, 11):
        if _cell_number(ws_perf[f"AA{r}"].value) is not None:
            return False
    return True
def compute_completed_hours_cds(
    ws_perf: Worksheet,
    period: Optional[_dt.date] = None,
) -> Tuple[Optional[float], Dict[str, float], List[str]]:
    if _use_af_actual_hours(period):
        total = _cell_number(ws_perf["AF10"].value)
        actual_col = "AF"
    elif _cds_use_new_hours_layout(period):
        total = _cell_number(ws_perf["W10"].value)
        actual_col = "W"
    else:
        use_r_layout = _cds_use_r_layout(ws_perf)
        total_col = "AB" if not use_r_layout else "R"
        actual_col = "AB" if not use_r_layout else "R"
        total = _cell_number(ws_perf[f"{total_col}11"].value)
    actual_by_person: Dict[str, float] = {}
    people_in_wip: List[str] = []
    seen = set()
    summed_actual = 0.0
    for r in range(5, 11):
        person = ws_perf[f"A{r}"].value
        actual = _cell_number(ws_perf[f"{actual_col}{r}"].value)
        p = str(person).strip() if person is not None else ""
        if not p or is_excluded_person(p) or actual is None or actual == 0:
            continue
        actual_by_person[p] = actual_by_person.get(p, 0.0) + actual
        summed_actual += actual
        if p not in seen:
            seen.add(p)
            people_in_wip.append(p)
    if total is None and _use_af_actual_hours(period):
        total = summed_actual
    return total, actual_by_person, people_in_wip
def compute_person_available_hours_cds(
    ws_perf: Worksheet,
    period: Optional[_dt.date] = None,
) -> Dict[str, float]:
    if _cds_use_new_hours_layout(period):
        available_col = "AA"
    else:
        use_r_layout = _cds_use_r_layout(ws_perf)
        available_col = "AA" if not use_r_layout else "R"
    out: Dict[str, float] = {}
    for r in range(5, 11):
        person = ws_perf[f"A{r}"].value
        available = _cell_number(ws_perf[f"{available_col}{r}"].value)
        p = str(person).strip() if person is not None else ""
        if not p or is_excluded_person(p) or available is None:
            continue
        out[p] = out.get(p, 0.0) + available
    return out
def compute_target_actual_output_cds(ws_pab: Worksheet) -> Tuple[float, float]:
    targ = 0.0
    act = 0.0
    for _, _, _, target_g, _, actual_j in _iter_rows_cds_pab(ws_pab):
        if target_g is not None:
            targ += target_g
        if actual_j is not None:
            act += actual_j
    return targ, act
def compute_outputs_by_person_cds(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, person, _, target_g, _, actual_j in _iter_rows_cds_pab(ws_pab):
        if not person or is_excluded_person(person):
            continue
        out.setdefault(person, {"output": 0.0, "target": 0.0})
        if target_g is not None:
            out[person]["target"] += target_g
        if actual_j is not None:
            out[person]["output"] += actual_j
    return out
def compute_outputs_by_station_cds(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, _, cell_station, target_g, _, actual_j in _iter_rows_cds_pab(ws_pab):
        if not cell_station or _is_excluded_station_cds(cell_station):
            continue
        out.setdefault(cell_station, {"output": 0.0, "target": 0.0})
        if target_g is not None:
            out[cell_station]["target"] += target_g
        if actual_j is not None:
            out[cell_station]["output"] += actual_j
    return out
def compute_station_hours_cds(ws_pab: Worksheet) -> Tuple[Dict[str, float], Dict[str, Dict[str, float]]]:
    station_hours: Dict[str, float] = {}
    station_hours_by_person: Dict[str, Dict[str, float]] = {}
    for _, person, cell_station, _, hours_i, actual_j in _iter_rows_cds_pab(ws_pab):
        if not cell_station or _is_excluded_station_cds(cell_station):
            continue
        if hours_i is not None:
            station_hours[cell_station] = station_hours.get(cell_station, 0.0) + hours_i
        if not person or is_excluded_person(person):
            continue
        if actual_j is None:
            continue
        station_hours_by_person.setdefault(cell_station, {})
        station_hours_by_person[cell_station][person] = (
            station_hours_by_person[cell_station].get(person, 0.0) + actual_j
        )
    return station_hours, station_hours_by_person
def compute_output_by_station_by_person_cds(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, person, cell_station, _, hours_i, _ in _iter_rows_cds_pab(ws_pab):
        if not person or not cell_station or is_excluded_person(person) or _is_excluded_station_cds(cell_station):
            continue
        if hours_i is None:
            continue
        out.setdefault(cell_station, {})
        out[cell_station][person] = out[cell_station].get(person, 0.0) + hours_i
    return out
def _norm_path(p: str) -> str:
    return os.path.normpath(p)
def team_for_source(path: str) -> str:
    np = _norm_path(path)
    base = os.path.basename(np)
    if _is_lit_letters_path(np):
        return LIT_LETTERS_TEAM
    if np in TEAM_BY_SOURCE:
        return TEAM_BY_SOURCE[np]
    if np.startswith(DS_ROOT_HINT):
        return "DS"
    if np.startswith(CPT_ROOT_HINT):
        return "CPT"
    if np.startswith(CDS_ROOT_HINT):
        return "CDS"
    if np.startswith(MEIC_ROOT_HINT):
        return "NI & PM MEIC"
    if np.startswith(PM_CTS_ROOT_HINT):
        return "PM-CTS"
    if np.startswith(NI_ROOT_HINT):
        return "NI"
    if np.startswith(MCS_ROOT_HINT):
        return "MCS"
    return TEAM_BY_BASENAME.get(base, "")
NI_NEW_HOURS_START = _dt.date(2026, 4, 17)
def _ni_use_new_hours_layout(period: Optional[_dt.date]) -> bool:
    return isinstance(period, _dt.date) and period >= NI_NEW_HOURS_START
def _is_excluded_station_ni(v: Any) -> bool:
    s = str(v).strip().lower() if v is not None else ""
    return s in {"non-wip", "essential non-wip"}
def _iter_rows_ni_pab(
    ws_pab: Worksheet, start_row: int = 2
) -> Iterable[Tuple[int, str, str, Optional[float], Optional[float], Optional[float]]]:
    for r in range(start_row, ws_pab.max_row + 1):
        person = ws_pab[f"C{r}"].value
        cell_station = ws_pab[f"D{r}"].value
        target_g = _cell_number(ws_pab[f"G{r}"].value)
        hours_i = _cell_number(ws_pab[f"I{r}"].value)
        actual_j = _cell_number(ws_pab[f"J{r}"].value)
        p = str(person).strip() if person is not None else ""
        cs = str(cell_station).strip() if cell_station is not None else ""
        if p == "" and cs == "" and target_g is None and hours_i is None and actual_j is None:
            continue
        yield (r, p, cs, target_g, hours_i, actual_j)
def compute_period_date_ni(ws_metrics: Worksheet) -> Optional[_dt.date]:
    d = _cell_date(ws_metrics["B3"].value, default_year=2026)
    if d is None:
        return None
    return d - _dt.timedelta(days=4)
def compute_total_available_hours_ni(ws_wip_plan: Worksheet) -> Optional[float]:
    return _cell_number(ws_wip_plan["BS3"].value)
def compute_total_available_hours_pm_cts(ws_perf: Worksheet) -> Optional[float]:
    return _cell_number(ws_perf["T18"].value)
def compute_completed_hours_ni(
    ws_perf: Worksheet,
    period: Optional[_dt.date] = None,
) -> Tuple[Optional[float], Dict[str, float], List[str]]:
    if _use_af_actual_hours(period):
        total = _cell_number(ws_perf["AF12"].value)
        actual_col = "AF"
        row_stop = 12   # rows 5-11; AF12 should be the total if present
    elif _ni_use_new_hours_layout(period):
        total = _cell_number(ws_perf["C13"].value)
        actual_col = "W"
        row_stop = 12   # rows 5-11; W12 is the new total Completed Hours cell
    else:
        total = _cell_number(ws_perf["R13"].value)
        actual_col = "R"
        row_stop = 13   # old NI layout: rows 5-12
        if total is None or total == 0:
            total = _cell_number(ws_perf["AB13"].value)
            actual_col = "AB"
    actual_by_person: Dict[str, float] = {}
    people_in_wip: List[str] = []
    seen = set()
    summed_actual = 0.0
    for r in range(5, row_stop):
        person = ws_perf[f"A{r}"].value
        actual = _cell_number(ws_perf[f"{actual_col}{r}"].value)
        p = str(person).strip() if person is not None else ""
        if not p or is_excluded_person(p) or actual is None or actual == 0:
            continue
        actual_by_person[p] = actual_by_person.get(p, 0.0) + actual
        summed_actual += actual
        if p not in seen:
            seen.add(p)
            people_in_wip.append(p)
    if total is None and _use_af_actual_hours(period):
        total = summed_actual
    return total, actual_by_person, people_in_wip
def compute_completed_hours_pm_cts(
    ws_perf: Worksheet,
) -> Tuple[Optional[float], Dict[str, float], List[str]]:
    total = _cell_number(ws_perf["Q18"].value)
    actual_by_person: Dict[str, float] = {}
    people_in_wip: List[str] = []
    seen = set()
    for r in range(3, 17):
        person = ws_perf[f"A{r}"].value
        actual = _cell_number(ws_perf[f"Q{r}"].value)
        p = str(person).strip() if person is not None else ""
        if not p or is_excluded_person(p) or actual is None or actual == 0:
            continue
        actual_by_person[p] = actual_by_person.get(p, 0.0) + actual
        if p not in seen:
            seen.add(p)
            people_in_wip.append(p)
    return total, actual_by_person, people_in_wip
def compute_person_available_hours_ni(
    ws_perf: Worksheet,
    period: Optional[_dt.date] = None,
) -> Dict[str, float]:
    if _ni_use_new_hours_layout(period):
        available_col = "AA"
        row_stop = 12   # rows 5-11
    else:
        available_col = "AG"
        row_stop = 13   # old NI layout: rows 5-12
    out: Dict[str, float] = {}
    for r in range(5, row_stop):
        person = ws_perf[f"A{r}"].value
        available = _cell_number(ws_perf[f"{available_col}{r}"].value)
        p = str(person).strip() if person is not None else ""
        if not p or is_excluded_person(p) or available is None:
            continue
        out[p] = out.get(p, 0.0) + available
    return out
def compute_target_actual_output_ni(ws_pab: Worksheet) -> Tuple[float, float]:
    targ = 0.0
    act = 0.0
    for _, _, _, target_g, _, actual_j in _iter_rows_ni_pab(ws_pab):
        if target_g is not None:
            targ += target_g
        if actual_j is not None:
            act += actual_j
    return targ, act
def compute_outputs_by_person_ni(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, person, _, target_g, _, actual_j in _iter_rows_ni_pab(ws_pab):
        if not person or is_excluded_person(person):
            continue
        out.setdefault(person, {"output": 0.0, "target": 0.0})
        if target_g is not None:
            out[person]["target"] += target_g
        if actual_j is not None:
            out[person]["output"] += actual_j
    return out
def compute_outputs_by_station_ni(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, _, cell_station, target_g, _, actual_j in _iter_rows_ni_pab(ws_pab):
        if not cell_station or _is_excluded_station_ni(cell_station):
            continue
        out.setdefault(cell_station, {"output": 0.0, "target": 0.0})
        if target_g is not None:
            out[cell_station]["target"] += target_g
        if actual_j is not None:
            out[cell_station]["output"] += actual_j
    return out
def compute_station_hours_ni(ws_pab: Worksheet) -> Tuple[Dict[str, float], Dict[str, Dict[str, float]]]:
    station_hours: Dict[str, float] = {}
    station_hours_by_person: Dict[str, Dict[str, float]] = {}
    for _, person, cell_station, _, hours_i, actual_j in _iter_rows_ni_pab(ws_pab):
        if not cell_station or _is_excluded_station_ni(cell_station):
            continue
        if hours_i is not None:
            station_hours[cell_station] = station_hours.get(cell_station, 0.0) + hours_i
        if not person or is_excluded_person(person):
            continue
        if actual_j is None:
            continue
        station_hours_by_person.setdefault(cell_station, {})
        station_hours_by_person[cell_station][person] = (
            station_hours_by_person[cell_station].get(person, 0.0) + actual_j
        )
    return station_hours, station_hours_by_person
def compute_output_by_station_by_person_ni(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, person, cell_station, _, hours_i, _ in _iter_rows_ni_pab(ws_pab):
        if not person or not cell_station or is_excluded_person(person) or _is_excluded_station_ni(cell_station):
            continue
        if hours_i is None:
            continue
        out.setdefault(cell_station, {})
        out[cell_station][person] = out[cell_station].get(person, 0.0) + hours_i
    return out
def _meic_cell_station(category_d: Any, subarea_e: Any) -> str:
    category = str(category_d).strip() if category_d is not None else ""
    subarea = str(subarea_e).strip() if subarea_e is not None else ""
    if subarea and subarea.lower() != "no subareas":
        return subarea
    return category
def _is_excluded_station_meic(category_d: Any) -> bool:
    s = str(category_d).strip().lower() if category_d is not None else ""
    return s in {"non-wip", "essential non-wip"}
def _iter_rows_meic_pab(
    ws_pab: Worksheet, start_row: int = 2
) -> Iterable[Tuple[int, str, str, str, Optional[float], Optional[float], Optional[float]]]:
    for r in range(start_row, ws_pab.max_row + 1):
        person = ws_pab[f"C{r}"].value
        category_d = ws_pab[f"D{r}"].value
        subarea_e = ws_pab[f"E{r}"].value
        target_g = _cell_number(ws_pab[f"G{r}"].value)
        hours_i = _cell_number(ws_pab[f"I{r}"].value)
        actual_j = _cell_number(ws_pab[f"J{r}"].value)
        p = str(person).strip() if person is not None else ""
        category = str(category_d).strip() if category_d is not None else ""
        subarea = str(subarea_e).strip() if subarea_e is not None else ""
        if p == "" and category == "" and subarea == "" and target_g is None and hours_i is None and actual_j is None:
            continue
        yield (r, p, category, subarea, target_g, hours_i, actual_j)
def compute_total_available_hours_meic(ws_wip_plan: Worksheet) -> Optional[float]:
    return _cell_number(ws_wip_plan["DU3"].value)
def compute_completed_hours_meic(ws_perf: Worksheet) -> Tuple[Optional[float], Dict[str, float], List[str]]:
    total = _cell_number(ws_perf["R24"].value)
    actual_by_person: Dict[str, float] = {}
    people_in_wip: List[str] = []
    seen = set()
    for r in range(5, 24):
        person = ws_perf[f"A{r}"].value
        actual = _cell_number(ws_perf[f"R{r}"].value)
        p = str(person).strip() if person is not None else ""
        if not p or is_excluded_person(p) or actual is None or actual == 0:
            continue
        actual_by_person[p] = actual_by_person.get(p, 0.0) + actual
        if p not in seen:
            seen.add(p)
            people_in_wip.append(p)
    return total, actual_by_person, people_in_wip
def compute_person_available_hours_meic(ws_perf: Worksheet) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for r in range(5, 24):
        person = ws_perf[f"A{r}"].value
        available = _cell_number(ws_perf[f"Q{r}"].value)
        p = str(person).strip() if person is not None else ""
        if not p or is_excluded_person(p) or available is None:
            continue
        out[p] = out.get(p, 0.0) + available
    return out
def compute_person_available_hours_pm_cts(ws_perf: Worksheet) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for r in range(3, 17):
        person = ws_perf[f"A{r}"].value
        available = _cell_number(ws_perf[f"T{r}"].value)
        p = str(person).strip() if person is not None else ""
        if not p or is_excluded_person(p) or available is None:
            continue
        out[p] = out.get(p, 0.0) + available
    return out
def compute_target_actual_output_meic(ws_pab: Worksheet) -> Tuple[float, float]:
    targ = 0.0
    act = 0.0
    for _, _, _, _, target_g, _, actual_j in _iter_rows_meic_pab(ws_pab):
        if target_g is not None:
            targ += target_g
        if actual_j is not None:
            act += actual_j
    return targ, act
def compute_outputs_by_person_meic(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, person, _, _, target_g, _, actual_j in _iter_rows_meic_pab(ws_pab):
        if not person or is_excluded_person(person):
            continue
        out.setdefault(person, {"output": 0.0, "target": 0.0})
        if target_g is not None:
            out[person]["target"] += target_g
        if actual_j is not None:
            out[person]["output"] += actual_j
    return out
def compute_outputs_by_station_meic(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, _, category_d, subarea_e, target_g, _, actual_j in _iter_rows_meic_pab(ws_pab):
        if _is_excluded_station_meic(category_d):
            continue
        cell_station = _meic_cell_station(category_d, subarea_e)
        if not cell_station:
            continue
        out.setdefault(cell_station, {"output": 0.0, "target": 0.0})
        if target_g is not None:
            out[cell_station]["target"] += target_g
        if actual_j is not None:
            out[cell_station]["output"] += actual_j
    return out
def compute_station_hours_meic(ws_pab: Worksheet) -> Tuple[Dict[str, float], Dict[str, Dict[str, float]]]:
    station_hours: Dict[str, float] = {}
    station_hours_by_person: Dict[str, Dict[str, float]] = {}
    for _, person, category_d, subarea_e, _, hours_i, actual_j in _iter_rows_meic_pab(ws_pab):
        if _is_excluded_station_meic(category_d):
            continue
        cell_station = _meic_cell_station(category_d, subarea_e)
        if not cell_station:
            continue
        if hours_i is not None:
            station_hours[cell_station] = station_hours.get(cell_station, 0.0) + hours_i
        if not person or is_excluded_person(person):
            continue
        if actual_j is None:
            continue
        station_hours_by_person.setdefault(cell_station, {})
        station_hours_by_person[cell_station][person] = (
            station_hours_by_person[cell_station].get(person, 0.0) + actual_j
        )
    return station_hours, station_hours_by_person
def compute_output_by_station_by_person_meic(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, person, category_d, subarea_e, _, hours_i, _ in _iter_rows_meic_pab(ws_pab):
        if not person or is_excluded_person(person) or _is_excluded_station_meic(category_d):
            continue
        cell_station = _meic_cell_station(category_d, subarea_e)
        if not cell_station or hours_i is None:
            continue
        out.setdefault(cell_station, {})
        out[cell_station][person] = out[cell_station].get(person, 0.0) + hours_i
    return out
def scrape_one_workbook_meic(path: str) -> List[Dict[str, Any]]:
    team = team_for_source(path)
    wb = load_workbook(path, data_only=True)
    err_msgs: List[str] = []
    ws_wip_plan = wb[_sheet_ci(wb, "# 1 WIP plan")] if _sheet_ci(wb, "# 1 WIP plan") else None
    ws_pab = wb[_sheet_ci(wb, "#2 PAB")] if _sheet_ci(wb, "#2 PAB") else None
    ws_perf = wb[_sheet_ci(wb, "#5 Performance WIP Time")] if _sheet_ci(wb, "#5 Performance WIP Time") else None
    if ws_wip_plan is None:
        err_msgs.append("missing_#1_wip_plan_sheet")
    if ws_pab is None:
        err_msgs.append("missing_#2_pab_sheet")
    if ws_perf is None:
        err_msgs.append("missing_#5_performance_wip_time_sheet")
    period = parse_period_date_from_filename(path, default_year=2026)
    if period is None:
        err_msgs.append("missing_period_date_from_filename")
    total_available = None
    completed_hours = None
    actual_hours_by_person: Dict[str, float] = {}
    people: List[str] = []
    person_avail: Dict[str, float] = {}
    target_output = None
    actual_output = None
    outputs_by_person: Dict[str, Dict[str, float]] = {}
    outputs_by_station: Dict[str, Dict[str, float]] = {}
    station_hours: Dict[str, float] = {}
    station_hours_by_person: Dict[str, Dict[str, float]] = {}
    output_by_station_by_person: Dict[str, Dict[str, float]] = {}
    uplh_by_station_by_person: Dict[str, Dict[str, float]] = {}
    try:
        if ws_wip_plan is not None:
            total_available = compute_total_available_hours_meic(ws_wip_plan)
        if ws_perf is not None:
            completed_hours, actual_hours_by_person, people = compute_completed_hours_meic(ws_perf)
            person_avail = compute_person_available_hours_meic(ws_perf)
        if ws_pab is not None:
            target_output, actual_output = compute_target_actual_output_meic(ws_pab)
            outputs_by_person = compute_outputs_by_person_meic(ws_pab)
            outputs_by_station = compute_outputs_by_station_meic(ws_pab)
            station_hours, station_hours_by_person = compute_station_hours_meic(ws_pab)
            output_by_station_by_person = compute_output_by_station_by_person_meic(ws_pab)
            uplh_by_station_by_person = compute_uplh_by_station_by_person(output_by_station_by_person, station_hours_by_person)
    except Exception as e:
        err_msgs.append(f"meic_parse_error: {e!r}")
    target_uplh = safe_div(float(target_output or 0.0), float(completed_hours or 0.0))
    actual_uplh = safe_div(float(actual_output or 0.0), float(completed_hours or 0.0))
    hc_in_wip = len(people) if people else 0
    actual_hc_used = safe_div(float(completed_hours or 0.0), 32.5)
    return [{
        "team": team,
        "period_date": iso_date(period),
        "source_file": path,
        "Total Available Hours": float(total_available) if total_available is not None else "",
        "Completed Hours": float(completed_hours) if completed_hours is not None else "",
        "Target Output": float(target_output) if target_output is not None else "",
        "Actual Output": float(actual_output) if actual_output is not None else "",
        "Target UPLH": float(target_uplh) if target_uplh is not None else "",
        "Actual UPLH": float(actual_uplh) if actual_uplh is not None else "",
        "UPLH WP1": "",
        "UPLH WP2": "",
        "HC in WIP": hc_in_wip,
        "Actual HC Used": float(actual_hc_used) if actual_hc_used is not None else "",
        "People in WIP": dumps_json(people) if ws_perf is not None else "",
        "Person Hours": build_person_hours_json(person_avail, actual_hours_by_person) if ws_perf is not None else "",
        "Outputs by Person": dumps_json(outputs_by_person) if ws_pab is not None else "",
        "Outputs by Cell/Station": dumps_json(outputs_by_station) if ws_pab is not None else "",
        "Cell/Station Hours": dumps_json(station_hours) if ws_pab is not None else "",
        "Hours by Cell/Station - by person": dumps_json(station_hours_by_person) if ws_pab is not None else "",
        "Output by Cell/Station - by person": dumps_json(output_by_station_by_person) if ws_pab is not None else "",
        "UPLH by Cell/Station - by person": dumps_json(uplh_by_station_by_person) if ws_pab is not None else "",
        "error": "; ".join(err_msgs) if err_msgs else "",
        "Closures": "",
        "Opened": "",
    }]
def scrape_one_workbook_ni(path: str) -> List[Dict[str, Any]]:
    team = team_for_source(path)
    wb = load_workbook(path, data_only=True)
    err_msgs: List[str] = []
    ws_metrics = wb[_sheet_ci(wb, "#4 Performance Metrics")] if _sheet_ci(wb, "#4 Performance Metrics") else None
    ws_wip_plan = wb[_sheet_ci(wb, "# 1 WIP plan")] if _sheet_ci(wb, "# 1 WIP plan") else None
    ws_pab = wb[_sheet_ci(wb, "#2 PAB")] if _sheet_ci(wb, "#2 PAB") else None
    ws_perf = wb[_sheet_ci(wb, "#5 Performance WIP Time")] if _sheet_ci(wb, "#5 Performance WIP Time") else None
    if ws_metrics is None:
        err_msgs.append("missing_#4_performance_metrics_sheet")
    if ws_wip_plan is None:
        err_msgs.append("missing_#1_wip_plan_sheet")
    if ws_pab is None:
        err_msgs.append("missing_#2_pab_sheet")
    if ws_perf is None:
        err_msgs.append("missing_#5_performance_wip_time_sheet")
    period = None
    if ws_metrics is not None:
        period = compute_period_date_ni(ws_metrics)
    if period is None:
        period = parse_period_date_from_filename(path, default_year=2026)
    if period is None:
        err_msgs.append("missing_period_date_from_#4_performance_metrics_B3_and_filename")
    total_available = None
    completed_hours = None
    actual_hours_by_person: Dict[str, float] = {}
    people: List[str] = []
    person_avail: Dict[str, float] = {}
    target_output = None
    actual_output = None
    outputs_by_person: Dict[str, Dict[str, float]] = {}
    outputs_by_station: Dict[str, Dict[str, float]] = {}
    station_hours: Dict[str, float] = {}
    station_hours_by_person: Dict[str, Dict[str, float]] = {}
    output_by_station_by_person: Dict[str, Dict[str, float]] = {}
    uplh_by_station_by_person: Dict[str, Dict[str, float]] = {}
    try:
        if ws_wip_plan is not None:
            total_available = compute_total_available_hours_ni(ws_wip_plan)
        if ws_perf is not None:
            completed_hours, actual_hours_by_person, people = compute_completed_hours_ni(ws_perf, period)
            person_avail = compute_person_available_hours_ni(ws_perf, period)
        if ws_pab is not None:
            target_output, actual_output = compute_target_actual_output_ni(ws_pab)
            outputs_by_person = compute_outputs_by_person_ni(ws_pab)
            outputs_by_station = compute_outputs_by_station_ni(ws_pab)
            station_hours, station_hours_by_person = compute_station_hours_ni(ws_pab)
            output_by_station_by_person = compute_output_by_station_by_person_ni(ws_pab)
            uplh_by_station_by_person = compute_uplh_by_station_by_person(
                output_by_station_by_person, station_hours_by_person
            )
    except Exception as e:
        err_msgs.append(f"ni_parse_error: {e!r}")
    target_uplh = safe_div(float(target_output or 0.0), float(completed_hours or 0.0))
    actual_uplh = safe_div(float(actual_output or 0.0), float(completed_hours or 0.0))
    hc_in_wip = len(people) if people else 0
    actual_hc_used = safe_div(float(completed_hours or 0.0), 32.5)
    return [{
        "team": team,
        "period_date": iso_date(period),
        "source_file": path,
        "Total Available Hours": float(total_available) if total_available is not None else "",
        "Completed Hours": float(completed_hours) if completed_hours is not None else "",
        "Target Output": float(target_output) if target_output is not None else "",
        "Actual Output": float(actual_output) if actual_output is not None else "",
        "Target UPLH": float(target_uplh) if target_uplh is not None else "",
        "Actual UPLH": float(actual_uplh) if actual_uplh is not None else "",
        "UPLH WP1": "",
        "UPLH WP2": "",
        "HC in WIP": hc_in_wip,
        "Actual HC Used": float(actual_hc_used) if actual_hc_used is not None else "",
        "People in WIP": dumps_json(people) if ws_perf is not None else "",
        "Person Hours": build_person_hours_json(person_avail, actual_hours_by_person) if ws_perf is not None else "",
        "Outputs by Person": dumps_json(outputs_by_person) if ws_pab is not None else "",
        "Outputs by Cell/Station": dumps_json(outputs_by_station) if ws_pab is not None else "",
        "Cell/Station Hours": dumps_json(station_hours) if ws_pab is not None else "",
        "Hours by Cell/Station - by person": dumps_json(station_hours_by_person) if ws_pab is not None else "",
        "Output by Cell/Station - by person": dumps_json(output_by_station_by_person) if ws_pab is not None else "",
        "UPLH by Cell/Station - by person": dumps_json(uplh_by_station_by_person) if ws_pab is not None else "",
        "error": "; ".join(err_msgs) if err_msgs else "",
        "Closures": "",
        "Opened": "",
    }]
def scrape_one_workbook_pm_cts(path: str) -> List[Dict[str, Any]]:
    team = team_for_source(path)
    wb = load_workbook(path, data_only=True)
    err_msgs: List[str] = []
    ws_wip_plan = wb[_sheet_ci(wb, "# 1 WIP plan")] if _sheet_ci(wb, "# 1 WIP plan") else None
    ws_pab = wb[_sheet_ci(wb, "#2 PAB")] if _sheet_ci(wb, "#2 PAB") else None
    ws_perf = wb[_sheet_ci(wb, "#3 Performance WIP Time")] if _sheet_ci(wb, "#3 Performance WIP Time") else None
    if ws_wip_plan is None:
        err_msgs.append("missing_#1_wip_plan_sheet")
    if ws_pab is None:
        err_msgs.append("missing_#2_pab_sheet")
    if ws_perf is None:
        err_msgs.append("missing_#3_performance_wip_time_sheet")
    period = parse_period_date_from_filename(path, default_year=2026)
    if period is None:
        err_msgs.append("missing_period_date_from_filename")
    total_available = None
    completed_hours = None
    actual_hours_by_person: Dict[str, float] = {}
    people: List[str] = []
    person_avail: Dict[str, float] = {}
    target_output = None
    actual_output = None
    outputs_by_person: Dict[str, Dict[str, float]] = {}
    outputs_by_station: Dict[str, Dict[str, float]] = {}
    station_hours: Dict[str, float] = {}
    station_hours_by_person: Dict[str, Dict[str, float]] = {}
    output_by_station_by_person: Dict[str, Dict[str, float]] = {}
    uplh_by_station_by_person: Dict[str, Dict[str, float]] = {}
    try:
        if ws_perf is not None:
            total_available = compute_total_available_hours_pm_cts(ws_perf)
            completed_hours, actual_hours_by_person, people = compute_completed_hours_pm_cts(ws_perf)
            person_avail = compute_person_available_hours_pm_cts(ws_perf)
        if ws_pab is not None:
            target_output, actual_output = compute_target_actual_output_ni(ws_pab)
            outputs_by_person = compute_outputs_by_person_ni(ws_pab)
            outputs_by_station = compute_outputs_by_station_ni(ws_pab)
            station_hours, station_hours_by_person = compute_station_hours_ni(ws_pab)
            output_by_station_by_person = compute_output_by_station_by_person_ni(ws_pab)
            uplh_by_station_by_person = compute_uplh_by_station_by_person(
                output_by_station_by_person,
                station_hours_by_person,
            )
    except Exception as e:
        err_msgs.append(f"pm_cts_parse_error: {e!r}")
    target_uplh = safe_div(float(target_output or 0.0), float(completed_hours or 0.0))
    actual_uplh = safe_div(float(actual_output or 0.0), float(completed_hours or 0.0))
    hc_in_wip = len(people) if people else 0
    actual_hc_used = safe_div(float(completed_hours or 0.0), 32.5)
    return [{
        "team": team,
        "period_date": iso_date(period),
        "source_file": path,
        "Total Available Hours": float(total_available) if total_available is not None else "",
        "Completed Hours": float(completed_hours) if completed_hours is not None else "",
        "Target Output": float(target_output) if target_output is not None else "",
        "Actual Output": float(actual_output) if actual_output is not None else "",
        "Target UPLH": float(target_uplh) if target_uplh is not None else "",
        "Actual UPLH": float(actual_uplh) if actual_uplh is not None else "",
        "UPLH WP1": "",
        "UPLH WP2": "",
        "HC in WIP": hc_in_wip,
        "Actual HC Used": float(actual_hc_used) if actual_hc_used is not None else "",
        "People in WIP": dumps_json(people) if ws_perf is not None else "",
        "Person Hours": build_person_hours_json(person_avail, actual_hours_by_person) if ws_perf is not None else "",
        "Outputs by Person": dumps_json(outputs_by_person) if ws_pab is not None else "",
        "Outputs by Cell/Station": dumps_json(outputs_by_station) if ws_pab is not None else "",
        "Cell/Station Hours": dumps_json(station_hours) if ws_pab is not None else "",
        "Hours by Cell/Station - by person": dumps_json(station_hours_by_person) if ws_pab is not None else "",
        "Output by Cell/Station - by person": dumps_json(output_by_station_by_person) if ws_pab is not None else "",
        "UPLH by Cell/Station - by person": dumps_json(uplh_by_station_by_person) if ws_pab is not None else "",
        "error": "; ".join(err_msgs) if err_msgs else "",
        "Closures": "",
        "Opened": "",
    }]
def parse_period_date_from_text(text: str, *, default_year: Optional[int] = None) -> Optional[_dt.date]:
    if default_year is None:
        default_year = _dt.date.today().year
    s = (text or "").strip()
    if not s:
        return None
    for fmt in (
        "%m-%d-%Y", "%m-%d-%y",
        "%m/%d/%Y", "%m/%d/%y",
        "%Y-%m-%d",
    ):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    m = re.search(r"\b(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})\b", s)
    if m:
        month = int(m.group(1))
        day = int(m.group(2))
        year = int(m.group(3))
        if year < 100:
            year += 2000
        try:
            return _dt.date(year, month, day)
        except ValueError:
            return None
    m = re.search(r"\b(\d{4})-(\d{1,2})-(\d{1,2})\b", s)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        day = int(m.group(3))
        try:
            return _dt.date(year, month, day)
        except ValueError:
            return None
    patterns = [
        r"(\d{1,2})\s*[-/ ]\s*([A-Za-z]{3,9})(?:\s*[-/ ]\s*(\d{2,4}))?",
        r"\b(\d{1,2})([A-Za-z]{3,9})(\d{2,4})\b",
    ]
    for pat in patterns:
        m = re.search(pat, s)
        if not m:
            continue
        day = int(m.group(1))
        mon_raw = m.group(2).strip().lower()
        year_raw = m.group(3)
        if mon_raw not in _MONTH_MAP:
            continue
        month = _MONTH_MAP[mon_raw]
        year = default_year
        if year_raw:
            year = int(year_raw)
            if year < 100:
                year += 2000
        try:
            return _dt.date(year, month, day)
        except ValueError:
            return None

    return None
def parse_period_date_from_sheetname(sheet_name: str, *, default_year: Optional[int] = None) -> Optional[_dt.date]:
    return parse_period_date_from_text(sheet_name, default_year=default_year)
def parse_period_date_from_filename(path: str, *, default_year: Optional[int] = None) -> Optional[_dt.date]:
    name = os.path.splitext(os.path.basename(path))[0]
    return parse_period_date_from_text(name, default_year=default_year)
def iso_date(d: Optional[_dt.date]) -> str:
    return d.isoformat() if isinstance(d, _dt.date) else ""
def _cell_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        vs = v.strip().replace(",", "")
        if vs == "":
            return None
        try:
            return float(vs)
        except ValueError:
            return None
    return None
def is_excluded_person(person: str) -> bool:
    p = person.strip().lower()
    return p in {
        "do not use",
        "team tally",
        "tally",
        "total wip hours",
        "total non-wip hours",
    }
def sum_range(ws: Worksheet, cell1: str, cell2: str) -> float:
    total = 0.0
    for row in ws[cell1:cell2]:
        for c in row:
            n = _cell_number(c.value)
            if n is not None:
                total += n
    return total
def read_merged_value(ws: Worksheet, top_left_cell: str) -> str:
    v = ws[top_left_cell].value
    return str(v).strip() if v is not None else ""
def _norm_team(s: str) -> str:
    return (s or "").strip().upper()
def _norm_period_date(s: str) -> str:
    ss = (s or "").strip()
    if not ss:
        return ""
    try:
        return _dt.date.fromisoformat(ss).isoformat()
    except ValueError:
        return ss
def load_closures_lookup(path: str) -> Dict[Tuple[str, str], Tuple[str, str]]:
    lut: Dict[Tuple[str, str], Tuple[str, str]] = {}
    if not path or not os.path.exists(path):
        return lut
    with open(path, "r", newline="", encoding="utf-8-sig") as fp:
        r = csv.DictReader(fp)
        for row in r:
            team = _norm_team(row.get("team", ""))
            period = _norm_period_date(row.get("period_date", ""))
            closures = (row.get("Closures", "") or "").strip()
            opened = (row.get("Opened", "") or "").strip()
            if team and period:
                lut[(team, period)] = (closures, opened)
    return lut
def enrich_rows_with_metrics(
    rows: List[Dict[str, Any]],
    closures_lut: Dict[Tuple[str, str], Tuple[str, str]],
) -> None:
    for r in rows:
        key = (_norm_team(r.get("team", "")), _norm_period_date(r.get("period_date", "")))
        if key in closures_lut:
            c, o = closures_lut[key]
            r["Closures"] = c
            r["Opened"] = o
def find_sheets_by_period(wb, *, kind: str) -> Dict[_dt.date, str]:
    out: Dict[_dt.date, str] = {}
    for name in wb.sheetnames:
        if kind == "availability":
            if not _AVAIL_PAT.search(name):
                continue
        elif kind == "production":
            if not _PROD_PAT.search(name):
                continue
        else:
            raise ValueError("kind must be availability or production")
        d = parse_period_date_from_sheetname(name)
        if d:
            out[d] = name
    return out
def compute_total_available_hours(ws_av: Worksheet) -> float:
    blocks = [("B", "F"), ("I", "M"), ("P", "T")]
    rows = [5, 15, 25, 35, 45, 55]
    total = 0.0
    for r in rows:
        for c1, c2 in blocks:
            total += sum_range(ws_av, f"{c1}{r}", f"{c2}{r}")
    return total
def iter_prod_rows_mcs(ws_prod: Worksheet, start_row: int = 7) -> Iterable[Tuple[int, str, str, Optional[float], Optional[float]]]:
    maxr = min(ws_prod.max_row, 406)
    for r in range(start_row, maxr + 1):
        person = ws_prod[f"D{r}"].value
        cell_station = ws_prod[f"E{r}"].value
        target = _cell_number(ws_prod[f"F{r}"].value)
        output = _cell_number(ws_prod[f"G{r}"].value)
        p = str(person).strip() if person is not None else ""
        cs = str(cell_station).strip() if cell_station is not None else ""
        if p == "" and cs == "" and target is None and output is None:
            continue
        yield (r, p, cs, target, output)
def compute_completed_hours_mcs(ws_prod: Worksheet) -> Tuple[float, Dict[str, float]]:
    total = 0.0
    by_person: Dict[str, float] = {}
    for r, person, cell_station, target, output in iter_prod_rows_mcs(ws_prod, start_row=7):
        if output is None:
            continue
        h = 4.0 if (cell_station == "Promoted PE - Initial MDR") else 1.0
        total += h
        if person and not is_excluded_person(person):
            by_person[person] = by_person.get(person, 0.0) + h
    return total, by_person
def compute_target_actual_output_mcs(ws_prod: Worksheet) -> Tuple[float, float]:
    targ = 0.0
    act = 0.0
    for r, person, cell_station, target, output in iter_prod_rows_mcs(ws_prod, start_row=7):
        if output is None:
            continue
        act += output
        if target is not None:
            targ += target
    return targ, act
def unique_people_in_wip_mcs(ws_prod: Worksheet) -> List[str]:
    seen = set()
    for r, person, cell_station, target, output in iter_prod_rows_mcs(ws_prod, start_row=7):
        if output is None or not person or is_excluded_person(person):
            continue
        seen.add(person)
    return sorted(seen)
def compute_outputs_by_person_mcs(ws_prod: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for r, person, cell_station, target, output in iter_prod_rows_mcs(ws_prod, start_row=7):
        if output is None or not person or is_excluded_person(person):
            continue
        out.setdefault(person, {"output": 0.0, "target": 0.0})
        out[person]["output"] += output
        if target is not None:
            out[person]["target"] += target
    return out
def compute_outputs_by_station_mcs(ws_prod: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for r, person, cell_station, target, output in iter_prod_rows_mcs(ws_prod, start_row=7):
        if output is None or not cell_station:
            continue
        out.setdefault(cell_station, {"output": 0.0, "target": 0.0})
        out[cell_station]["output"] += output
        if target is not None:
            out[cell_station]["target"] += target
    return out
def compute_station_hours_mcs(ws_prod: Worksheet) -> Tuple[Dict[str, float], Dict[str, Dict[str, float]]]:
    station_hours: Dict[str, float] = {}
    station_hours_by_person: Dict[str, Dict[str, float]] = {}
    for r, person, cell_station, target, output in iter_prod_rows_mcs(ws_prod, start_row=7):
        if not person or not cell_station:
            continue
        if person.strip().lower() in {"do not use", "team member(s)"}:
            continue
        if cell_station.strip().lower() == "cell/station":
            continue
        h = 2.0 if cell_station == "Promoted PE - Initial MDR" else 1.0
        station_hours[cell_station] = station_hours.get(cell_station, 0.0) + h
        station_hours_by_person.setdefault(cell_station, {})
        station_hours_by_person[cell_station][person] = station_hours_by_person[cell_station].get(person, 0.0) + h
    return station_hours, station_hours_by_person
def compute_output_by_station_by_person_mcs(ws_prod: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for r, person, cell_station, target, output in iter_prod_rows_mcs(ws_prod, start_row=7):
        if output is None or not person or not cell_station:
            continue
        out.setdefault(cell_station, {})
        out[cell_station][person] = out[cell_station].get(person, 0.0) + output
    return out
def compute_uplh_by_station_by_person(
    output_by_station_by_person: Dict[str, Dict[str, float]],
    hours_by_station_by_person: Dict[str, Dict[str, float]],
) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for station, people_outputs in output_by_station_by_person.items():
        for person, out_val in people_outputs.items():
            hrs = hours_by_station_by_person.get(station, {}).get(person, 0.0)
            if hrs and hrs != 0.0:
                out.setdefault(station, {})
                out[station][person] = out_val / hrs
    return out
def compute_person_available_hours(ws_av: Worksheet) -> Dict[str, float]:
    pairs = [
        ("A13", ("B15", "F15")),
        ("A23", ("B25", "F25")),
        ("H3",  ("I5", "M5")),
        ("H23", ("I25", "M25")),
        ("H33", ("I35", "M35")),
        ("O3",  ("O5", "Q5")),
        ("O13", ("O15", "Q15")),
        ("O23", ("O25", "Q25")),
        ("O43", ("O45", "Q45")),
        ("O53", ("O55", "Q55")),
    ]
    out: Dict[str, float] = {}
    for name_cell, (c1, c2) in pairs:
        name = read_merged_value(ws_av, name_cell)
        if not name or is_excluded_person(name):
            continue
        avail = sum_range(ws_av, c1, c2)
        out[name] = out.get(name, 0.0) + avail
    return out
def _sheet_ci(wb, name: str) -> Optional[str]:
    want = name.strip().lower()
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() == want:
            return sheet_name
    return None
DS_NEW_HOURS_START = _dt.date(2026, 4, 24)
DS_COMPLETED_HOURS_UP_ONE_ROW_START = _dt.date(2026, 6, 1)
def _ds_use_new_hours_layout(period: Optional[_dt.date]) -> bool:
    return isinstance(period, _dt.date) and period >= DS_NEW_HOURS_START
def _ds_completed_hours_up_one_row(period: Optional[_dt.date]) -> bool:
    return isinstance(period, _dt.date) and period >= DS_COMPLETED_HOURS_UP_ONE_ROW_START
def _is_ds_excluded_category(v: Any) -> bool:
    s = str(v).strip().lower() if v is not None else ""
    return s in {"non-wip", "essential non-wip"}
def _iter_rows_ds_pab(ws_pab: Worksheet, start_row: int = 2) -> Iterable[Tuple[int, str, str, str, Optional[float], Optional[float], Optional[float]]]:
    for r in range(start_row, ws_pab.max_row + 1):
        person = ws_pab[f"C{r}"].value
        category = ws_pab[f"D{r}"].value
        cell_station = ws_pab[f"E{r}"].value
        target = _cell_number(ws_pab[f"G{r}"].value)
        hours_i = _cell_number(ws_pab[f"I{r}"].value)
        actual_j = _cell_number(ws_pab[f"J{r}"].value)
        p = str(person).strip() if person is not None else ""
        cat = str(category).strip() if category is not None else ""
        cs = str(cell_station).strip() if cell_station is not None else ""
        if p == "" and cat == "" and cs == "" and target is None and hours_i is None and actual_j is None:
            continue
        yield (r, p, cat, cs, target, hours_i, actual_j)
def compute_total_available_hours_ds(ws_wip_plan: Worksheet) -> Optional[float]:
    return _cell_number(ws_wip_plan["EG3"].value)
def _ds_use_r_layout(ws_perf: Worksheet) -> bool:
    for r in range(5, 46):
        if _cell_number(ws_perf[f"AA{r}"].value) is not None:
            return False
    return True
def _ds_column_sum(ws_perf: Worksheet, column: str) -> float:
    return sum(
        value
        for r in range(5, 46)
        if (value := _cell_number(ws_perf[f"{column}{r}"].value)) is not None
    )
def _ds_pick_actual_column(
    ws_perf: Worksheet, candidates: Iterable[str], total: Optional[float]
) -> str:
    column_sums = {column: _ds_column_sum(ws_perf, column) for column in candidates}
    if total is not None and total > 0:
        return min(column_sums, key=lambda column: abs(column_sums[column] - total))
    return max(column_sums, key=column_sums.get)
def compute_completed_hours_ds(
    ws_perf: Worksheet,
    period: Optional[_dt.date] = None,
) -> Tuple[Optional[float], Dict[str, float], List[str]]:
    if _ds_use_new_hours_layout(period):
        total_cell = "C46" if _ds_completed_hours_up_one_row(period) else "C47"
        total = _cell_number(ws_perf[total_cell].value)
        actual_col = _ds_pick_actual_column(ws_perf, ("W", "AF"), total)
    else:
        use_r_layout = _ds_use_r_layout(ws_perf)
        total_col = "R" if use_r_layout else "AB"
        actual_col = "R" if use_r_layout else "AB"
        total = _cell_number(ws_perf[f"{total_col}46"].value)
    actual_by_person: Dict[str, float] = {}
    people_in_wip: List[str] = []
    seen = set()
    for r in range(5, 46):
        person = ws_perf[f"A{r}"].value
        actual = _cell_number(ws_perf[f"{actual_col}{r}"].value)
        p = str(person).strip() if person is not None else ""
        if not p or is_excluded_person(p) or actual is None or actual == 0:
            continue
        actual_by_person[p] = actual_by_person.get(p, 0.0) + actual
        if p not in seen:
            seen.add(p)
            people_in_wip.append(p)
    return total, actual_by_person, people_in_wip
def compute_person_available_hours_ds(
    ws_perf: Worksheet,
    period: Optional[_dt.date] = None,
) -> Dict[str, float]:
    if _ds_use_new_hours_layout(period):
        available_col = max(("AA", "AB"), key=lambda column: _ds_column_sum(ws_perf, column))
    else:
        use_r_layout = _ds_use_r_layout(ws_perf)
        available_col = "R" if use_r_layout else "AA"
    out: Dict[str, float] = {}
    for r in range(5, 46):
        person = ws_perf[f"A{r}"].value
        available = _cell_number(ws_perf[f"{available_col}{r}"].value)
        p = str(person).strip() if person is not None else ""
        if not p or is_excluded_person(p) or available is None:
            continue
        out[p] = out.get(p, 0.0) + available
    return out
def compute_target_actual_output_ds(ws_pab: Worksheet) -> Tuple[float, float]:
    targ = 0.0
    act = 0.0
    for _, _, _, _, target, _, actual in _iter_rows_ds_pab(ws_pab):
        if target is not None:
            targ += target
        if actual is not None:
            act += actual
    return targ, act
def compute_outputs_by_person_ds(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, person, _, _, target, _, actual in _iter_rows_ds_pab(ws_pab):
        if not person or is_excluded_person(person):
            continue
        out.setdefault(person, {"output": 0.0, "target": 0.0})
        if target is not None:
            out[person]["target"] += target
        if actual is not None:
            out[person]["output"] += actual
    return out
def compute_outputs_by_station_ds(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, _, category, cell_station, target, _, actual in _iter_rows_ds_pab(ws_pab):
        if not cell_station or _is_ds_excluded_category(category):
            continue
        out.setdefault(cell_station, {"output": 0.0, "target": 0.0})
        if target is not None:
            out[cell_station]["target"] += target
        if actual is not None:
            out[cell_station]["output"] += actual
    return out
def compute_station_hours_ds(ws_pab: Worksheet) -> Tuple[Dict[str, float], Dict[str, Dict[str, float]]]:
    station_hours: Dict[str, float] = {}
    station_hours_by_person: Dict[str, Dict[str, float]] = {}
    for _, person, category, cell_station, _, hours_i, actual_j in _iter_rows_ds_pab(ws_pab):
        if not cell_station or _is_ds_excluded_category(category):
            continue
        if hours_i is not None:
            station_hours[cell_station] = station_hours.get(cell_station, 0.0) + hours_i
        if not person or is_excluded_person(person):
            continue
        hours_by_person_val = actual_j
        if hours_by_person_val is None:
            continue
        station_hours_by_person.setdefault(cell_station, {})
        station_hours_by_person[cell_station][person] = station_hours_by_person[cell_station].get(person, 0.0) + hours_by_person_val
    return station_hours, station_hours_by_person
def compute_output_by_station_by_person_ds(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, person, category, cell_station, _, hours_i, _ in _iter_rows_ds_pab(ws_pab):
        if not person or not cell_station or is_excluded_person(person) or _is_ds_excluded_category(category):
            continue
        if hours_i is None:
            continue
        out.setdefault(cell_station, {})
        out[cell_station][person] = out[cell_station].get(person, 0.0) + hours_i
    return out
def _iter_rows_cpt_pab(ws_pab: Worksheet, start_row: int = 2) -> Iterable[Tuple[int, str, str, str, Optional[float], Optional[float], Optional[float]]]:
    for r in range(start_row, ws_pab.max_row + 1):
        person = ws_pab[f"C{r}"].value
        category = ws_pab[f"D{r}"].value
        cell_station = ws_pab[f"E{r}"].value
        target = _cell_number(ws_pab[f"G{r}"].value)
        hours_i = _cell_number(ws_pab[f"I{r}"].value)
        actual_j = _cell_number(ws_pab[f"J{r}"].value)
        p = str(person).strip() if person is not None else ""
        cat = str(category).strip() if category is not None else ""
        cs = str(cell_station).strip() if cell_station is not None else ""
        if p == "" and cat == "" and cs == "" and target is None and hours_i is None and actual_j is None:
            continue
        yield (r, p, cat, cs, target, hours_i, actual_j)
CPT_NEW_HOURS_START = _dt.date(2026, 5, 4)
def _cpt_use_new_hours_layout(period: Optional[_dt.date]) -> bool:
    return isinstance(period, _dt.date) and period >= CPT_NEW_HOURS_START
def _first_number_from_cells(ws: Worksheet, *cells: str) -> Optional[float]:
    for cell in cells:
        n = _cell_number(ws[cell].value)
        if n is not None:
            return n
    return None
def compute_total_available_hours_cpt(
    ws_wip_plan: Optional[Worksheet],
    ws_perf: Optional[Worksheet] = None,
    period: Optional[_dt.date] = None,
) -> Optional[float]:
    if _cpt_use_new_hours_layout(period) and ws_perf is not None:
        total = _first_number_from_cells(ws_perf, "B49", "C49")
        if total is not None:
            return total
    if ws_wip_plan is None:
        return None
    return _cell_number(ws_wip_plan["DU3"].value)
def compute_completed_hours_cpt(
    ws_perf: Worksheet,
    period: Optional[_dt.date] = None,
) -> Tuple[Optional[float], Dict[str, float], List[str]]:
    if _cpt_use_new_hours_layout(period):
        actual_col = "AF"
        total = _first_number_from_cells(ws_perf, "B49", "C49", "AF48")
    else:
        b48 = _cell_number(ws_perf["B48"].value)
        c48 = _cell_number(ws_perf["C48"].value)
        if b48 is not None:
            actual_col = "AB"
            total = b48
        elif c48 is not None:
            actual_col = "AF"
            total = c48
        else:
            actual_col = "AB"
            total = None
    actual_by_person: Dict[str, float] = {}
    people_in_wip: List[str] = []
    seen = set()
    summed_actual = 0.0
    for r in range(5, 49):
        person = ws_perf[f"A{r}"].value
        actual = _cell_number(ws_perf[f"{actual_col}{r}"].value)
        p = str(person).strip() if person is not None else ""
        if not p or is_excluded_person(p) or actual is None or actual == 0:
            continue
        actual_by_person[p] = actual_by_person.get(p, 0.0) + actual
        summed_actual += actual
        if p not in seen:
            seen.add(p)
            people_in_wip.append(p)
    if total is None:
        total = summed_actual
    return total, actual_by_person, people_in_wip
def compute_person_available_hours_cpt(
    ws_perf: Worksheet,
    period: Optional[_dt.date] = None,
) -> Dict[str, float]:
    if _cpt_use_new_hours_layout(period):
        available_col = "AA"
    else:
        b48 = _cell_number(ws_perf["B48"].value)
        c48 = _cell_number(ws_perf["C48"].value)
        if c48 is not None:
            available_col = "AA"
        elif b48 is not None:
            available_col = "V"
        else:
            available_col = "AA"
    out: Dict[str, float] = {}
    for r in range(5, 48):
        person = ws_perf[f"A{r}"].value
        available = _cell_number(ws_perf[f"{available_col}{r}"].value)
        p = str(person).strip() if person is not None else ""
        if not p or is_excluded_person(p) or available is None:
            continue
        out[p] = out.get(p, 0.0) + available
    return out
def compute_target_actual_output_cpt(ws_pab: Worksheet) -> Tuple[float, float]:
    targ = 0.0
    act = 0.0
    for _, _, _, _, target, _, actual in _iter_rows_cpt_pab(ws_pab):
        if target is not None:
            targ += target
        if actual is not None:
            act += actual
    return targ, act
def compute_outputs_by_person_cpt(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, person, _, _, target, _, actual in _iter_rows_cpt_pab(ws_pab):
        if not person or is_excluded_person(person):
            continue
        out.setdefault(person, {"output": 0.0, "target": 0.0})
        if target is not None:
            out[person]["target"] += target
        if actual is not None:
            out[person]["output"] += actual
    return out
def compute_outputs_by_station_cpt(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, _, category, cell_station, target, _, actual in _iter_rows_cpt_pab(ws_pab):
        if not cell_station or _is_ds_excluded_category(category):
            continue
        out.setdefault(cell_station, {"output": 0.0, "target": 0.0})
        if target is not None:
            out[cell_station]["target"] += target
        if actual is not None:
            out[cell_station]["output"] += actual
    return out
def compute_station_hours_cpt(ws_pab: Worksheet) -> Tuple[Dict[str, float], Dict[str, Dict[str, float]]]:
    station_hours: Dict[str, float] = {}
    station_hours_by_person: Dict[str, Dict[str, float]] = {}
    for _, person, category, cell_station, _, hours_i, actual_j in _iter_rows_cpt_pab(ws_pab):
        if not cell_station or _is_ds_excluded_category(category):
            continue
        if hours_i is not None:
            station_hours[cell_station] = station_hours.get(cell_station, 0.0) + hours_i
        if not person or is_excluded_person(person):
            continue
        hours_by_person_val = actual_j
        if hours_by_person_val is None:
            continue
        station_hours_by_person.setdefault(cell_station, {})
        station_hours_by_person[cell_station][person] = station_hours_by_person[cell_station].get(person, 0.0) + hours_by_person_val
    return station_hours, station_hours_by_person
def compute_output_by_station_by_person_cpt(ws_pab: Worksheet) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for _, person, category, cell_station, _, hours_i, _ in _iter_rows_cpt_pab(ws_pab):
        if not person or not cell_station or is_excluded_person(person) or _is_ds_excluded_category(category):
            continue
        if hours_i is None:
            continue
        out.setdefault(cell_station, {})
        out[cell_station][person] = out[cell_station].get(person, 0.0) + hours_i
    return out
def scrape_one_workbook_cpt(path: str) -> List[Dict[str, Any]]:
    team = team_for_source(path)
    wb = load_workbook(path, data_only=True)
    err_msgs: List[str] = []
    ws_wip_plan = wb[_sheet_ci(wb, "# 1 WIP plan")] if _sheet_ci(wb, "# 1 WIP plan") else None
    ws_pab = wb[_sheet_ci(wb, "#3 PAB")] if _sheet_ci(wb, "#3 PAB") else None
    ws_perf = wb[_sheet_ci(wb, "#6 Performance WIP Time")] if _sheet_ci(wb, "#6 Performance WIP Time") else None
    if ws_wip_plan is None:
        err_msgs.append("missing_#1_wip_plan_sheet")
    if ws_pab is None:
        err_msgs.append("missing_#3_pab_sheet")
    if ws_perf is None:
        err_msgs.append("missing_#6_performance_wip_time_sheet")
    period = parse_period_date_from_filename(path, default_year=2026)
    if period is None:
        return []
    if period < _dt.date(2026, 4, 1):
        return []
    total_available = None
    completed_hours = None
    actual_hours_by_person: Dict[str, float] = {}
    people: List[str] = []
    person_avail: Dict[str, float] = {}
    target_output = None
    actual_output = None
    outputs_by_person: Dict[str, Dict[str, float]] = {}
    outputs_by_station: Dict[str, Dict[str, float]] = {}
    station_hours: Dict[str, float] = {}
    station_hours_by_person: Dict[str, Dict[str, float]] = {}
    output_by_station_by_person: Dict[str, Dict[str, float]] = {}
    uplh_by_station_by_person: Dict[str, Dict[str, float]] = {}
    try:
        total_available = compute_total_available_hours_cpt(ws_wip_plan, ws_perf, period)
        if ws_perf is not None:
            completed_hours, actual_hours_by_person, people = compute_completed_hours_cpt(ws_perf, period)
            person_avail = compute_person_available_hours_cpt(ws_perf, period)
        if ws_pab is not None:
            target_output, actual_output = compute_target_actual_output_cpt(ws_pab)
            outputs_by_person = compute_outputs_by_person_cpt(ws_pab)
            outputs_by_station = compute_outputs_by_station_cpt(ws_pab)
            station_hours, station_hours_by_person = compute_station_hours_cpt(ws_pab)
            output_by_station_by_person = compute_output_by_station_by_person_cpt(ws_pab)
            uplh_by_station_by_person = compute_uplh_by_station_by_person(output_by_station_by_person, station_hours_by_person)
    except Exception as e:
        err_msgs.append(f"cpt_parse_error: {e!r}")
    target_uplh = safe_div(float(target_output or 0.0), float(completed_hours or 0.0))
    actual_uplh = safe_div(float(actual_output or 0.0), float(completed_hours or 0.0))
    hc_in_wip = len(people) if ws_perf is not None else ""
    actual_hc_used = safe_div(float(completed_hours or 0.0), 32.5)
    return [{
        "team": team,
        "period_date": iso_date(period),
        "source_file": path,
        "Total Available Hours": float(total_available) if total_available is not None else "",
        "Completed Hours": float(completed_hours) if completed_hours is not None else "",
        "Target Output": float(target_output) if target_output is not None else "",
        "Actual Output": float(actual_output) if actual_output is not None else "",
        "Target UPLH": float(target_uplh) if target_uplh is not None else "",
        "Actual UPLH": float(actual_uplh) if actual_uplh is not None else "",
        "UPLH WP1": "",
        "UPLH WP2": "",
        "HC in WIP": hc_in_wip,
        "Actual HC Used": float(actual_hc_used) if actual_hc_used is not None else "",
        "People in WIP": dumps_json(people) if ws_perf is not None else "",
        "Person Hours": build_person_hours_json(person_avail, actual_hours_by_person) if ws_perf is not None else "",
        "Outputs by Person": dumps_json(outputs_by_person) if ws_pab is not None else "",
        "Outputs by Cell/Station": dumps_json(outputs_by_station) if ws_pab is not None else "",
        "Cell/Station Hours": dumps_json(station_hours) if ws_pab is not None else "",
        "Hours by Cell/Station - by person": dumps_json(station_hours_by_person) if ws_pab is not None else "",
        "Output by Cell/Station - by person": dumps_json(output_by_station_by_person) if ws_pab is not None else "",
        "UPLH by Cell/Station - by person": dumps_json(uplh_by_station_by_person) if ws_pab is not None else "",
        "error": "; ".join(err_msgs) if err_msgs else "",
        "Closures": "",
        "Opened": "",
    }]
def scrape_one_workbook_ds(path: str) -> List[Dict[str, Any]]:
    team = team_for_source(path)
    wb = load_workbook(path, data_only=True)
    err_msgs: List[str] = []
    ws_wip_plan = wb[_sheet_ci(wb, "# 1 WIP plan")] if _sheet_ci(wb, "# 1 WIP plan") else None
    ws_pab = wb[_sheet_ci(wb, "#2 PAB")] if _sheet_ci(wb, "#2 PAB") else None
    ws_perf = wb[_sheet_ci(wb, "#5 Performance WIP Time")] if _sheet_ci(wb, "#5 Performance WIP Time") else None
    if ws_wip_plan is None:
        err_msgs.append("missing_#1_wip_plan_sheet")
    if ws_pab is None:
        err_msgs.append("missing_#2_pab_sheet")
    if ws_perf is None:
        err_msgs.append("missing_#5_performance_wip_time_sheet")
    period = parse_period_date_from_filename(path, default_year=2026)
    if period is None:
        return []
    total_available = None
    completed_hours = None
    actual_hours_by_person: Dict[str, float] = {}
    people: List[str] = []
    person_avail: Dict[str, float] = {}
    target_output = None
    actual_output = None
    outputs_by_person: Dict[str, Dict[str, float]] = {}
    outputs_by_station: Dict[str, Dict[str, float]] = {}
    station_hours: Dict[str, float] = {}
    station_hours_by_person: Dict[str, Dict[str, float]] = {}
    output_by_station_by_person: Dict[str, Dict[str, float]] = {}
    uplh_by_station_by_person: Dict[str, Dict[str, float]] = {}
    try:
        if ws_wip_plan is not None:
            total_available = compute_total_available_hours_ds(ws_wip_plan)
        if ws_perf is not None:
            completed_hours, actual_hours_by_person, people = compute_completed_hours_ds(ws_perf, period)
            person_avail = compute_person_available_hours_ds(ws_perf, period)
        if ws_pab is not None:
            target_output, actual_output = compute_target_actual_output_ds(ws_pab)
            outputs_by_person = compute_outputs_by_person_ds(ws_pab)
            outputs_by_station = compute_outputs_by_station_ds(ws_pab)
            station_hours, station_hours_by_person = compute_station_hours_ds(ws_pab)
            output_by_station_by_person = compute_output_by_station_by_person_ds(ws_pab)
            uplh_by_station_by_person = compute_uplh_by_station_by_person(output_by_station_by_person, station_hours_by_person)
    except Exception as e:
        err_msgs.append(f"ds_parse_error: {e!r}")
    target_uplh = safe_div(float(target_output or 0.0), float(completed_hours or 0.0))
    actual_uplh = safe_div(float(actual_output or 0.0), float(completed_hours or 0.0))
    hc_in_wip = len(people) if ws_perf is not None else ""
    actual_hc_used = safe_div(float(completed_hours or 0.0), 32.5)
    return [{
        "team": team,
        "period_date": iso_date(period),
        "source_file": path,
        "Total Available Hours": float(total_available) if total_available is not None else "",
        "Completed Hours": float(completed_hours) if completed_hours is not None else "",
        "Target Output": float(target_output) if target_output is not None else "",
        "Actual Output": float(actual_output) if actual_output is not None else "",
        "Target UPLH": float(target_uplh) if target_uplh is not None else "",
        "Actual UPLH": float(actual_uplh) if actual_uplh is not None else "",
        "UPLH WP1": "",
        "UPLH WP2": "",
        "HC in WIP": hc_in_wip,
        "Actual HC Used": float(actual_hc_used) if actual_hc_used is not None else "",
        "People in WIP": dumps_json(people) if ws_perf is not None else "",
        "Person Hours": build_person_hours_json(person_avail, actual_hours_by_person) if (ws_perf is not None) else "",
        "Outputs by Person": dumps_json(outputs_by_person) if ws_pab is not None else "",
        "Outputs by Cell/Station": dumps_json(outputs_by_station) if ws_pab is not None else "",
        "Cell/Station Hours": dumps_json(station_hours) if ws_pab is not None else "",
        "Hours by Cell/Station - by person": dumps_json(station_hours_by_person) if ws_pab is not None else "",
        "Output by Cell/Station - by person": dumps_json(output_by_station_by_person) if ws_pab is not None else "",
        "UPLH by Cell/Station - by person": dumps_json(uplh_by_station_by_person) if ws_pab is not None else "",
        "error": "; ".join(err_msgs) if err_msgs else "",
        "Closures": "",
        "Opened": "",
    }]
def build_person_hours_json(available_by_person: Dict[str, float], actual_by_person: Dict[str, float]) -> str:
    all_people = sorted(set(available_by_person.keys()) | set(actual_by_person.keys()))
    payload = {}
    for p in all_people:
        if is_excluded_person(p):
            continue
        payload[p] = {
            "actual": float(actual_by_person.get(p, 0.0)),
            "available": float(available_by_person.get(p, 0.0)),
        }
    return json.dumps(payload, ensure_ascii=False)
def dumps_json(obj: Any) -> str:
    return json.dumps(obj, ensure_ascii=False)
def safe_div(n: float, d: float) -> Optional[float]:
    if d is None or d == 0:
        return None
    return n / d
def scrape_one_workbook_mcs(path: str) -> List[Dict[str, Any]]:
    team = team_for_source(path)
    wb = load_workbook(path, data_only=True)
    avail_sheets = find_sheets_by_period(wb, kind="availability")
    prod_sheets = find_sheets_by_period(wb, kind="production")
    periods = sorted(set(avail_sheets.keys()) | set(prod_sheets.keys()))
    rows: List[Dict[str, Any]] = []
    for period in periods:
        err_msgs: List[str] = []
        ws_av = wb[avail_sheets[period]] if period in avail_sheets else None
        ws_prod = wb[prod_sheets[period]] if period in prod_sheets else None
        total_available = None
        person_avail = {}
        if ws_av is not None:
            try:
                total_available = compute_total_available_hours(ws_av)
                person_avail = compute_person_available_hours(ws_av)
            except Exception as e:
                err_msgs.append(f"availability_parse_error: {e!r}")
        else:
            err_msgs.append("missing_availability_sheet")
        completed_hours = None
        actual_hours_by_person = {}
        target_output = None
        actual_output = None
        people = []
        outputs_by_person = {}
        outputs_by_station = {}
        station_hours = {}
        station_hours_by_person = {}
        output_by_station_by_person = {}
        uplh_by_station_by_person = {}
        if ws_prod is not None:
            try:
                completed_hours, actual_hours_by_person = compute_completed_hours_mcs(ws_prod)
                target_output, actual_output = compute_target_actual_output_mcs(ws_prod)
                people = unique_people_in_wip_mcs(ws_prod)
                outputs_by_person = compute_outputs_by_person_mcs(ws_prod)
                outputs_by_station = compute_outputs_by_station_mcs(ws_prod)
                station_hours, station_hours_by_person = compute_station_hours_mcs(ws_prod)
                output_by_station_by_person = compute_output_by_station_by_person_mcs(ws_prod)
                uplh_by_station_by_person = compute_uplh_by_station_by_person(output_by_station_by_person, station_hours_by_person)
            except Exception as e:
                err_msgs.append(f"production_parse_error: {e!r}")
        else:
            err_msgs.append("missing_production_analysis_sheet")
        target_uplh = safe_div(float(target_output or 0.0), float(completed_hours or 0.0))
        actual_uplh = safe_div(float(actual_output or 0.0), float(completed_hours or 0.0))
        hc_in_wip = len(people) if people else 0
        actual_hc_used = safe_div(float(completed_hours or 0.0), 32.0)
        rows.append({
            "team": team,
            "period_date": iso_date(period),
            "source_file": path,
            "Total Available Hours": float(total_available) if total_available is not None else "",
            "Completed Hours": float(completed_hours) if completed_hours is not None else "",
            "Target Output": float(target_output) if target_output is not None else "",
            "Actual Output": float(actual_output) if actual_output is not None else "",
            "Target UPLH": float(target_uplh) if target_uplh is not None else "",
            "Actual UPLH": float(actual_uplh) if actual_uplh is not None else "",
            "UPLH WP1": "",
            "UPLH WP2": "",
            "HC in WIP": hc_in_wip if completed_hours is not None else "",
            "Actual HC Used": float(actual_hc_used) if actual_hc_used is not None else "",
            "People in WIP": dumps_json(people) if people else dumps_json([]) if ws_prod is not None else "",
            "Person Hours": build_person_hours_json(person_avail, actual_hours_by_person) if (ws_av is not None or ws_prod is not None) else "",
            "Outputs by Person": dumps_json(outputs_by_person) if ws_prod is not None else "",
            "Outputs by Cell/Station": dumps_json(outputs_by_station) if ws_prod is not None else "",
            "Cell/Station Hours": dumps_json(station_hours) if ws_prod is not None else "",
            "Hours by Cell/Station - by person": dumps_json(station_hours_by_person) if ws_prod is not None else "",
            "Output by Cell/Station - by person": dumps_json(output_by_station_by_person) if ws_prod is not None else "",
            "UPLH by Cell/Station - by person": dumps_json(uplh_by_station_by_person) if ws_prod is not None else "",
            "error": "; ".join(err_msgs) if err_msgs else "",
            "Closures": "",
            "Opened": "",
        })
    return rows
def scrape_one_workbook_cds(path: str) -> List[Dict[str, Any]]:
    team = team_for_source(path)
    wb = load_workbook(path, data_only=True)
    err_msgs: List[str] = []
    ws_metrics = wb[_sheet_ci(wb, "#4 Performance Metrics")] if _sheet_ci(wb, "#4 Performance Metrics") else None
    ws_wip_plan = wb[_sheet_ci(wb, "# 1 WIP plan")] if _sheet_ci(wb, "# 1 WIP plan") else None
    ws_pab = wb[_sheet_ci(wb, "#2 PAB")] if _sheet_ci(wb, "#2 PAB") else None
    ws_perf = wb[_sheet_ci(wb, "#5 Performance WIP Time")] if _sheet_ci(wb, "#5 Performance WIP Time") else None
    if ws_metrics is None:
        err_msgs.append("missing_#4_performance_metrics_sheet")
    if ws_wip_plan is None:
        err_msgs.append("missing_#1_wip_plan_sheet")
    if ws_pab is None:
        err_msgs.append("missing_#2_pab_sheet")
    if ws_perf is None:
        err_msgs.append("missing_#5_performance_wip_time_sheet")
    period = None
    if ws_metrics is not None:
        period = compute_period_date_cds(ws_metrics)
    if period is None:
        period = parse_period_date_from_filename(path, default_year=2026)
    if period is None:
        err_msgs.append("missing_period_date_from_#4_performance_metrics_B3_and_filename")
    total_available = None
    completed_hours = None
    actual_hours_by_person: Dict[str, float] = {}
    people: List[str] = []
    person_avail: Dict[str, float] = {}
    target_output = None
    actual_output = None
    outputs_by_person: Dict[str, Dict[str, float]] = {}
    outputs_by_station: Dict[str, Dict[str, float]] = {}
    station_hours: Dict[str, float] = {}
    station_hours_by_person: Dict[str, Dict[str, float]] = {}
    output_by_station_by_person: Dict[str, Dict[str, float]] = {}
    uplh_by_station_by_person: Dict[str, Dict[str, float]] = {}
    try:
        if ws_wip_plan is not None:
            total_available = compute_total_available_hours_cds(ws_wip_plan)
        if ws_perf is not None:
            completed_hours, actual_hours_by_person, people = compute_completed_hours_cds(ws_perf, period)
            person_avail = compute_person_available_hours_cds(ws_perf, period)
        if ws_pab is not None:
            target_output, actual_output = compute_target_actual_output_cds(ws_pab)
            outputs_by_person = compute_outputs_by_person_cds(ws_pab)
            outputs_by_station = compute_outputs_by_station_cds(ws_pab)
            station_hours, station_hours_by_person = compute_station_hours_cds(ws_pab)
            output_by_station_by_person = compute_output_by_station_by_person_cds(ws_pab)
            uplh_by_station_by_person = compute_uplh_by_station_by_person(
                output_by_station_by_person, station_hours_by_person
            )
    except Exception as e:
        err_msgs.append(f"cds_parse_error: {e!r}")
    target_uplh = safe_div(float(target_output or 0.0), float(completed_hours or 0.0))
    actual_uplh = safe_div(float(actual_output or 0.0), float(completed_hours or 0.0))
    hc_in_wip = len(people) if people else 0
    actual_hc_used = safe_div(float(completed_hours or 0.0), 32.5)
    return [{
        "team": team,
        "period_date": iso_date(period),
        "source_file": path,
        "Total Available Hours": float(total_available) if total_available is not None else "",
        "Completed Hours": float(completed_hours) if completed_hours is not None else "",
        "Target Output": float(target_output) if target_output is not None else "",
        "Actual Output": float(actual_output) if actual_output is not None else "",
        "Target UPLH": float(target_uplh) if target_uplh is not None else "",
        "Actual UPLH": float(actual_uplh) if actual_uplh is not None else "",
        "UPLH WP1": "",
        "UPLH WP2": "",
        "HC in WIP": hc_in_wip,
        "Actual HC Used": float(actual_hc_used) if actual_hc_used is not None else "",
        "People in WIP": dumps_json(people) if ws_perf is not None else "",
        "Person Hours": build_person_hours_json(person_avail, actual_hours_by_person) if ws_perf is not None else "",
        "Outputs by Person": dumps_json(outputs_by_person) if ws_pab is not None else "",
        "Outputs by Cell/Station": dumps_json(outputs_by_station) if ws_pab is not None else "",
        "Cell/Station Hours": dumps_json(station_hours) if ws_pab is not None else "",
        "Hours by Cell/Station - by person": dumps_json(station_hours_by_person) if ws_pab is not None else "",
        "Output by Cell/Station - by person": dumps_json(output_by_station_by_person) if ws_pab is not None else "",
        "UPLH by Cell/Station - by person": dumps_json(uplh_by_station_by_person) if ws_pab is not None else "",
        "error": "; ".join(err_msgs) if err_msgs else "",
        "Closures": "",
        "Opened": "",
    }]
def scrape_one_workbook(path: str) -> List[Dict[str, Any]]:
    team = team_for_source(path)
    try:
        if team == "MCS":
            return scrape_one_workbook_mcs(path)
        if team == "CDS":
            return scrape_one_workbook_cds(path)
        if team == LIT_LETTERS_TEAM:
            return scrape_one_workbook_lit_letters(path)
        if team == "DS":
            return scrape_one_workbook_ds(path)
        if team == "CPT":
            return scrape_one_workbook_cpt(path)
        if team == "NI":
            return scrape_one_workbook_ni(path)
        if team == "NI & PM MEIC":
            return scrape_one_workbook_meic(path)
        if team == "PM-CTS":
            return scrape_one_workbook_pm_cts(path)
        return [dict(blank_row_for_missing_file(path), error=f"unknown_team_for_source: {path}")]
    except FileNotFoundError:
        return [blank_row_for_missing_file(path)]
    except PermissionError as e:
        return [dict(blank_row_for_missing_file(path), error=f"permission_error: {e}")]
    except BadZipFile:
        return [dict(blank_row_for_missing_file(path), error="bad_zip_file_or_invalid_excel")]
    except Exception as e:
        return [dict(blank_row_for_missing_file(path), error=f"unhandled_error: {e!r}")]
CSV_COLUMNS = [
    "team", "period_date", "source_file", "Total Available Hours", "Completed Hours",
    "Target Output", "Actual Output", "Target UPLH", "Actual UPLH", "UPLH WP1", "UPLH WP2",
    "HC in WIP", "Actual HC Used", "People in WIP", "Person Hours", "Outputs by Person",
    "Outputs by Cell/Station", "Cell/Station Hours", "Hours by Cell/Station - by person",
    "Output by Cell/Station - by person", "UPLH by Cell/Station - by person",
    "error", "Closures", "Opened",
]
def iter_input_files(paths: List[str]) -> Iterable[str]:
    for p in paths:
        if os.path.isdir(p):
            np = _norm_path(p)
            for name in sorted(os.listdir(p)):
                lower = name.lower()
                if name.startswith("~$"):
                    continue
                if not lower.endswith((".xlsx", ".xlsm", ".xls")):
                    continue
                full_path = _norm_path(os.path.join(p, name))
                if full_path in EXCLUDED_FILES:
                    continue
                if np.startswith(CDS_ROOT_HINT) and "pab" not in lower:
                    continue
                if np.startswith(NI_ROOT_HINT) and not np.startswith(PM_CTS_ROOT_HINT) and "pab" not in lower:
                    continue
                if np.startswith(MEIC_ROOT_HINT) and "heijunka tracker" not in lower:
                    continue
                yield full_path
        else:
            full_path = _norm_path(p)
            if full_path in EXCLUDED_FILES:
                continue
            yield full_path
def blank_row_for_missing_file(f: str) -> Dict[str, Any]:
    return {
        "team": team_for_source(f),
        "period_date": "",
        "source_file": f,
        "Total Available Hours": "",
        "Completed Hours": "",
        "Target Output": "",
        "Actual Output": "",
        "Target UPLH": "",
        "Actual UPLH": "",
        "UPLH WP1": "",
        "UPLH WP2": "",
        "HC in WIP": "",
        "Actual HC Used": "",
        "People in WIP": "",
        "Person Hours": "",
        "Outputs by Person": "",
        "Outputs by Cell/Station": "",
        "Cell/Station Hours": "",
        "Hours by Cell/Station - by person": "",
        "Output by Cell/Station - by person": "",
        "UPLH by Cell/Station - by person": "",
        "error": f"file_not_found: {f}",
        "Closures": "",
        "Opened": "",
    }
def main() -> int:
    default_paths = [
        r"C:\Users\wadec8\Medtronic PLC\MCS COS Transformation - VMB Scheduling\Heijunka Current.xlsm",
        r"C:\Users\wadec8\Medtronic PLC\Diagnostics MDR - Heijunka and Production Analysis",
        r"C:\Users\wadec8\Medtronic PLC\Diagnostics MDR - Heijunka and Production Analysis\Archived PAB",
        r"C:\Users\wadec8\Medtronic PLC\Defibrillation Solutions - Schedule and PAB",
        r"C:\Users\wadec8\Medtronic PLC\Defibrillation Solutions - Schedule and PAB\Archive",
        r"C:\Users\wadec8\Medtronic PLC\Cardiac Pacing Therapies CQXM - Heijunka & PAB",
        r"C:\Users\wadec8\Medtronic PLC\Cardiac Pacing Therapies CQXM - Heijunka & PAB\Archive\2026\4. April 2026",
        r"C:\Users\wadec8\Medtronic PLC\Cardiac Pacing Therapies CQXM - Heijunka & PAB\Archive\2026\5. May 2026",
        r"C:\Users\wadec8\Medtronic PLC\Cardiac Pacing Therapies CQXM - Heijunka & PAB\Archive\2026\6. Jun 2026",
        r"C:\Users\wadec8\Medtronic PLC\Tier1 PXM - Non Implantables - Heijunka",
        r"C:\Users\wadec8\Medtronic PLC\Tier1 PXM - Non Implantables - Heijunka\Archived PAB\April 2026 - PAB",
        r"C:\Users\wadec8\Medtronic PLC\Tier1 PXM - Non Implantables - Heijunka\Archived PAB",
        r"C:\Users\wadec8\Medtronic PLC\CRM CQXM Reports - 1.9 Heijunka Tracker",
        r"C:\Users\wadec8\Medtronic PLC\CRM CQXM Reports - 1.9 Heijunka Tracker\Archive",
        r"C:\Users\wadec8\Medtronic PLC\Tier1 PXM - Non Implantables - Heijunka\PM-CTS PAB",
    ]
    ap = argparse.ArgumentParser()
    ap.add_argument("files", nargs="*", help="Excel workbook(s) and/or folders to scrape (.xlsx/.xlsm).")
    ap.add_argument("--out", default="CRM_DATA\\CRM_WIP.csv", help="Output CSV path (default: CRM_WIP.csv).")
    args = ap.parse_args()
    inputs = args.files or default_paths
    all_rows: List[Dict[str, Any]] = []
    for f in iter_input_files(inputs):
        if not os.path.exists(f):
            all_rows.append(blank_row_for_missing_file(f))
            continue
        all_rows.extend(scrape_one_workbook(f))
    all_rows.sort(key=lambda r: ((r.get("team") or ""), (r.get("period_date") or "")))
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", newline="", encoding="utf-8") as fp:
        w = csv.DictWriter(fp, fieldnames=CSV_COLUMNS)
        w.writeheader()
        for r in all_rows:
            w.writerow({k: r.get(k, "") for k in CSV_COLUMNS})
    print(f"Wrote {len(all_rows)} row(s) to {args.out}")
    return 0
if __name__ == "__main__":
    raise SystemExit(main())