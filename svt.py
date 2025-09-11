import re
import sys
import argparse
from pathlib import Path
from datetime import datetime as _dt, date as _date, timedelta
from dateutil import parser as dateparser
import pandas as pd
import numpy as np
import csv
from openpyxl import load_workbook

# -------------------- Defaults & Config --------------------
# You can override these via CLI flags if needed
SVT_GLOB_DEFAULT = r"C:\Users\wadec8\Medtronic PLC\SVT PXM Team - Heijunka_Schedule_Finding Work\SVT Heijunka*.xls*"
OUT_CSV_DEFAULT = Path.cwd() / "svt.csv"
TIMELINESS_DEFAULT = Path(r"C:\heijunka-dev\timeliness.csv")

EXCLUDED_SOURCE_FILES = {
    r"C:\Users\wadec8\Medtronic PLC\SVT PXM Team - Archived Heijunka\SVT Future Heijunka.xlsm"
}
EXCLUDED_DIRS = {
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\Weekly Heijunka Archived",
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\Clinical",
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\Commercial",
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\Remediation",
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\WIP Blitz Power Hour",
    r"C:\Users\wadec8\Medtronic PLC\SVT PXM Team - Archived Heijunka",
}
EXCLUDED_DIRS = {s.lower().rstrip("\\").replace("/", "\\") for s in EXCLUDED_DIRS}
EXCLUDED_SOURCE_FILES = {s.lower().replace("/", "\\") for s in EXCLUDED_SOURCE_FILES}

SKIP_PATTERNS = [r"~\\$", r"\.tmp$"]
DATE_REGEXES = [
    r"\b\d{4}[-_]\d{2}[-_]\d{2}\b",
    r"\b\d{8}\b",
    r"\b\d{1,2}[-_]\d{1,2}[-_]\d{2,4}\b",
    r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*[-_ ]\d{1,2}[-_, ]\d{2,4}\b",
]

# -------------------- Utilities --------------------

def looks_like_temp(name: str) -> bool:
    return any(re.search(pat, name, flags=re.IGNORECASE) for pat in SKIP_PATTERNS)

def _is_excluded_path(p: Path) -> bool:
    try:
        sp = str(p).lower().replace("/", "\\")
        if sp in EXCLUDED_SOURCE_FILES:
            return True
        for d in EXCLUDED_DIRS:
            if sp == d or sp.startswith(d + "\\"):
                return True
        if p.name.lower() == "svt future heijunka.xlsm":
            return True
    except Exception:
        pass
    return False

def parse_date_from_text(text: str):
    for rx in DATE_REGEXES:
        m = re.search(rx, text, flags=re.IGNORECASE)
        if m:
            try:
                return dateparser.parse(m.group(0), dayfirst=False, yearfirst=False).date()
            except Exception:
                pass
    return None

def _excel_serial_to_date(n) -> _date | None:
    try:
        return (_dt(1899, 12, 30) + timedelta(days=float(n))).date()
    except Exception:
        return None

YEAR_RX = re.compile(r"\b(?:19|20)\d{2}\b")  # 1900–2099

def _coerce_to_date_for_filter(v, require_explicit_year: bool = False) -> _date | None:
    if isinstance(v, _dt):
        return v.date()
    if isinstance(v, _date):
        return v
    if isinstance(v, (int, float)):
        d = _excel_serial_to_date(v)
        if d and _dt(1900, 1, 1).date() <= d <= _dt(2100, 1, 1).date():
            return d
        return None
    s = str(v)
    if require_explicit_year and not YEAR_RX.search(s):
        return None
    try:
        return dateparser.parse(s, dayfirst=False, yearfirst=False).date()
    except Exception:
        return None

def infer_period_date(path: Path):
    d = parse_date_from_text(str(path))
    if d:
        return d
    try:
        return _dt.fromtimestamp(path.stat().st_mtime).date()
    except Exception:
        return None

def col_letter_to_index(letter: str) -> int:
    letter = letter.strip().upper()
    num = 0
    for ch in letter:
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f"Invalid column letter: {letter}")
        num = num * 26 + (ord(ch) - ord('A') + 1)
    return num

def a1_to_rowcol(a1: str) -> tuple[int, int]:
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", a1.strip())
    if not m:
        raise ValueError(f"Bad cell address: {a1}")
    letters, row = m.groups()
    col = 0
    for ch in letters.upper():
        col = col * 26 + (ord(ch) - ord('A') + 1)
    return int(row), col

def read_one_cell_openpyxl(ws, a1: str):
    r, c = a1_to_rowcol(a1)
    vals = list(ws.iter_rows(min_row=r, max_row=r, min_col=c, max_col=c, values_only=True))
    return vals[0][0] if vals else None

def sum_cells_openpyxl(ws, addrs: list[str]):
    total, any_vals = 0.0, False
    for a in addrs:
        v = read_one_cell_openpyxl(ws, a)
        if v is None:
            continue
        try:
            v = float(str(v).replace(",", "").strip())
            total += v; any_vals = True
        except Exception:
            pass
    return total if any_vals else None

def sum_column(ws, col_letter):
    col = 0
    for ch in col_letter.strip().upper():
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f"Invalid column letter: {col_letter}")
        col = col * 26 + (ord(ch) - ord('A') + 1)
    total = 0.0
    any_vals = False
    for (val,) in ws.iter_rows(min_col=col, max_col=col, values_only=True):
        try:
            if val is None:
                continue
            if isinstance(val, (int, float)):
                total += float(val); any_vals = True
            else:
                v = float(str(val).replace(",", "").strip())
                total += v; any_vals = True
        except Exception:
            continue
    return total if any_vals else None

# Filtering + conditional summation used for Completed Hours (skip hidden rows)

def _passes_filters(row_map: dict, include_contains: dict | None, exclude_regex: dict | None) -> bool:
    if include_contains:
        for col, needle in include_contains.items():
            v = row_map.get(col)
            if not isinstance(v, str) or needle.lower() not in v.lower():
                return False
    if exclude_regex:
        for col, rx in exclude_regex.items():
            v = row_map.get(col)
            if isinstance(v, str) and re.search(rx, v, flags=re.IGNORECASE):
                return False
    return True

def sum_column_openpyxl_filtered(ws, target_col: str,
                                 include_contains: dict | None = None,
                                 exclude_regex: dict | None = None,
                                 row_start: int | None = None,
                                 row_end: int | None = None,
                                 skip_hidden: bool = False) -> float | None:
    t_idx = col_letter_to_index(target_col)
    need_cols = {t_idx}
    if include_contains:
        need_cols |= {col_letter_to_index(c) for c in include_contains.keys()}
    if exclude_regex:
        need_cols |= {col_letter_to_index(c) for c in exclude_regex.keys()}
    min_c, max_c = min(need_cols), max(need_cols)
    row_start = row_start or 1
    row_end = row_end or ws.max_row
    total, any_vals = 0.0, False
    for r_idx, row_vals in enumerate(ws.iter_rows(min_row=row_start, max_row=row_end,
                                                  min_col=min_c, max_col=max_c,
                                                  values_only=True), start=row_start):
        if skip_hidden and hasattr(ws, "row_dimensions"):
            rd = ws.row_dimensions.get(r_idx) if hasattr(ws.row_dimensions, "get") else None
            if rd is not None and getattr(rd, "hidden", False):
                continue
        row_map = {}
        for c_idx_off, val in enumerate(row_vals, start=min_c):
            row_map[chr(ord('A') + c_idx_off - 1)] = val
        if not _passes_filters(row_map, include_contains, exclude_regex):
            continue
        val = row_map.get(chr(ord('A') + t_idx - 1))
        try:
            if val is None:
                continue
            if isinstance(val, (int, float)):
                total += float(val); any_vals = True
            else:
                total += float(str(val).replace(",", "").strip()); any_vals = True
        except Exception:
            continue
    return total if any_vals else None

# SVT-specific: count unique IDs in col C where effort (col G) > 0

def _svt_hc_in_wip_openpyxl(ws,
                            key_col_letter: str = "C",
                            cond_col_letter: str = "G",
                            row_start: int = 7,
                            row_end: int = 200) -> int:
    kc = col_letter_to_index(key_col_letter)
    cc = col_letter_to_index(cond_col_letter)
    unique_keys = set()
    for r, row in enumerate(ws.iter_rows(min_row=row_start, max_row=row_end,
                                         min_col=min(kc, cc), max_col=max(kc, cc),
                                         values_only=True), start=row_start):
        key_val = row[kc - min(kc, cc)]
        cond_val = row[cc - min(kc, cc)]
        try:
            if cond_val is None:
                continue
            cond_num = float(str(cond_val).replace(",", "").strip())
            if cond_num <= 0:
                continue
        except Exception:
            continue
        if key_val is None:
            continue
        key_str = str(key_val).strip()
        if key_str:
            unique_keys.add(key_str)
    return len(unique_keys)

# -------------------- Collection --------------------

def collect_svt_rows(file_glob: str) -> list[dict]:
    rows: list[dict] = []
    for p in Path().glob("**/*"):
        # no-op; just to calm linters in some IDEs about Path import usage
        break
    paths = [Path(s) for s in Path().glob("**/*")]  # placeholder to satisfy type checkers
    # Use glob from pathlib for Windows wildcards reliably
    import glob as _glob
    candidates = [Path(s) for s in _glob.glob(file_glob)]
    if not candidates:
        try:
            parent = Path(file_glob).parent
            name_pat = Path(file_glob).name
            if parent.exists():
                candidates = list(parent.rglob(name_pat))
        except Exception:
            pass
    if not candidates:
        print(f"[svt] No files matched glob: {file_glob}")
    for path in candidates:
        if path.is_dir() or looks_like_temp(path.name) or _is_excluded_path(path):
            continue
        ext = path.suffix.lower()
        if ext not in (".xlsx", ".xlsm"):
            rows.append({
            "team": "SVT",
            "period_date": infer_period_date(path),
            "source_file": str(path),
            "error": str(e),
            "fallback_used": None,
        })
    return rows

# -------------------- Transform & Output --------------------

def _filter_future_periods(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "period_date" not in df.columns:
        return df
    today = pd.Timestamp.today().normalize()
    d = pd.to_datetime(df["period_date"], errors="coerce").dt.normalize()
    keep = d.isna() | (d <= today)
    return df.loc[keep].copy()

def add_open_complaint_timeliness(df: pd.DataFrame, timeliness_csv: Path) -> pd.DataFrame:
    """Join timeliness by exact period_date; if missing, fall back to Monday-of-week match.
    Also normalizes team names (maps anything containing 'SVT' -> 'SVT').
    """
    if not timeliness_csv or not Path(timeliness_csv).exists():
        print(f"[timeliness] {timeliness_csv} not found; skipping join.")
        # Ensure column exists even if we can't join
        if "Open Complaint Timeliness" not in df.columns:
            df = df.copy(); df["Open Complaint Timeliness"] = np.nan
        return df
    try:
        t = pd.read_csv(timeliness_csv, dtype=str, keep_default_na=False)
    except Exception as e:
        print(f"[timeliness] Failed to read {timeliness_csv}: {e}")
        if "Open Complaint Timeliness" not in df.columns:
            df = df.copy(); df["Open Complaint Timeliness"] = np.nan
        return df
    if t.shape[1] < 3:
        print("[timeliness] Expected at least 3 columns (A, B, value). Skipping join.")
        if "Open Complaint Timeliness" not in df.columns:
            df = df.copy(); df["Open Complaint Timeliness"] = np.nan
        return df

    lower_cols = [str(c).strip().lower() for c in t.columns]
    def _first_match(names):
        for want in names:
            if want in lower_cols:
                return t.columns[lower_cols.index(want)]
        return None
    team_col = _first_match(["team"]) or t.columns[0]
    date_col = _first_match(["period_date", "period", "date"]) or t.columns[1]
    val_col  = _first_match(["open complaint timeliness", "timeliness", "value", "metric"]) or t.columns[2]

    t = t.rename(columns={team_col: "team", date_col: "period_date", val_col: "Open Complaint Timeliness"})

    # Normalize team and dates
    t["team"] = t["team"].astype(str).str.strip()
    # Map any team values that contain 'svt' -> 'SVT'
    t["team_norm"] = t["team"].str.upper()
    t.loc[t["team_norm"].str.contains("SVT", na=False), "team_norm"] = "SVT"

    # Coerce dates robustly
    t["period_date"] = pd.to_datetime(t["period_date"], errors="coerce").dt.normalize()
    t = t.dropna(subset=["team_norm", "period_date"]).drop_duplicates(subset=["team_norm", "period_date"], keep="last")

    out = df.copy()
    if "team" not in out.columns or "period_date" not in out.columns:
        if "Open Complaint Timeliness" not in out.columns:
            out["Open Complaint Timeliness"] = np.nan
        return out

    out["team_norm"] = out["team"].astype(str).str.strip().str.upper()
    out.loc[out["team_norm"].str.contains("SVT", na=False), "team_norm"] = "SVT"
    out["period_date"] = pd.to_datetime(out["period_date"], errors="coerce").dt.normalize()

    base = out.reset_index(drop=True)

    # First: exact date join
    m1 = base.merge(
        t[["team_norm", "period_date", "Open Complaint Timeliness"]],
        on=["team_norm", "period_date"], how="left"
    )

    # Second: week-start join (Monday) as a fallback
    base_ws = base.copy()
    base_ws["week_start"] = base_ws["period_date"] - pd.to_timedelta(base_ws["period_date"].dt.weekday, unit="D")
    t_ws = t.copy()
    t_ws["week_start"] = t_ws["period_date"] - pd.to_timedelta(t_ws["period_date"].dt.weekday, unit="D")

    m2 = base_ws[["team_norm", "week_start"]].merge(
        t_ws[["team_norm", "week_start", "Open Complaint Timeliness"]],
        on=["team_norm", "week_start"], how="left"
    )

    # Prefer exact match, then fill from week match
    m1["Open Complaint Timeliness"] = m1["Open Complaint Timeliness"].combine_first(m2["Open Complaint Timeliness"])

    # Clean up and return in original column order + the OCT column
    out_cols = list(out.columns)
    result = m1[out_cols + ["Open Complaint Timeliness"]].copy()
    return result

def compute_calculations(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    # Normalize period date
    if "period_date" in df.columns:
        df["period_date"] = pd.to_datetime(df["period_date"], errors="coerce").dt.normalize()

    # Calculations
    def _safe_div(n, d):
        try:
            n = float(str(n).replace(",", "").strip())
            d = float(str(d).replace(",", "").strip())
            return None if d == 0 else n / d
        except Exception:
            return None

    df["Target UPLH"] = df.apply(lambda r: _safe_div(r.get("Target Output"), r.get("Total Available Hours")), axis=1)
    df["Actual UPLH"] = df.apply(lambda r: _safe_div(r.get("Actual Output"), r.get("Completed Hours")), axis=1)
    if "Target UPLH" in df.columns:
        df["Target UPLH"] = pd.to_numeric(df["Target UPLH"], errors="coerce").round(2)
    if "Actual UPLH" in df.columns:
        df["Actual UPLH"] = pd.to_numeric(df["Actual UPLH"], errors="coerce").round(2)
    if "Completed Hours" in df.columns:
        df["Actual HC Used"] = pd.to_numeric(df["Completed Hours"], errors="coerce") / 32.5
        df["Actual HC Used"] = pd.to_numeric(df["Actual HC Used"], errors="coerce").round(2)

    return df

def save_csv(df: pd.DataFrame, out_csv: Path):
    if df.empty:
        print("No SVT rows collected. Check paths/sheets/cells.")
        return
    # Ensure expected columns exist even if empty
    for c in ["fallback_used", "error", "Open Complaint Timeliness"]:
        if c not in df.columns:
            df[c] = ""
    preferred_cols = [
        "team", "period_date", "source_file",
        "Total Available Hours", "Completed Hours",
        "Target Output", "Actual Output",
        "Target UPLH", "Actual UPLH",
        "HC in WIP", "Actual HC Used",
        "Open Complaint Timeliness", "fallback_used", "error",
    ]
    cols = [c for c in preferred_cols if c in df.columns]
    out = df.loc[:, cols].copy()
    if "period_date" in out.columns:
        out["period_date"] = pd.to_datetime(out["period_date"], errors="coerce").dt.strftime("%Y-%m-%d")
    numeric_cols = {"Total Available Hours", "Completed Hours", "Target Output", "Actual Output",
                    "Target UPLH", "Actual UPLH", "Actual HC Used", "Open Complaint Timeliness"} & set(out.columns)
    for c in numeric_cols:
        out[c] = pd.to_numeric(out[c], errors="coerce")
    out = out.replace({np.nan: ""})
    out.to_csv(
        out_csv,
        index=False,
        sep=",",
        encoding="utf-8",
        lineterminator="",
        quoting=csv.QUOTE_MINIMAL,
        date_format="%Y-%m-%d",
    )
    print(f"Saved CSV: {out_csv.resolve()}")

# -------------------- Main --------------------

def main():
    parser = argparse.ArgumentParser(description="Extract SVT metrics from Heijunka Excel files into svt.csv")
    parser.add_argument("--glob", default=SVT_GLOB_DEFAULT, help="Glob pattern to locate SVT Heijunka files")
    parser.add_argument("--out", default=str(OUT_CSV_DEFAULT), help="Output CSV path (default: ./svt.csv)")
    parser.add_argument("--timeliness", default=str(TIMELINESS_DEFAULT), help="Path to timeliness CSV for joining")
    args = parser.parse_args()

    rows = collect_svt_rows(args.glob)
    df = pd.DataFrame(rows)

    # Keep only SVT, filter future weeks
    if not df.empty:
        df = df.copy()
        df["team"] = "SVT"
        df = _filter_future_periods(df)

    df = compute_calculations(df)
    df = add_open_complaint_timeliness(df, Path(args.timeliness) if args.timeliness else None)

    save_csv(df, Path(args.out))

    if not df.empty:
        with pd.option_context("display.max_columns", None, "display.width", 180):
            print("\nPreview:")
            try:
                print(df.head(12).to_string(index=False))
            except Exception:
                print(df.head(12))

if __name__ == "__main__":
    main()
