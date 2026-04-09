from __future__ import annotations
import csv
import json
import re
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
TEAM = "ECT"
FOLDERS = [
    Path(r"C:\Users\wadec8\Medtronic PLC\Doran, Elaine - Heijunka Production Analysis\Archived Heijunka"),
    Path(r"C:\Users\wadec8\Medtronic PLC\Doran, Elaine - Heijunka Production Analysis"),
]
START_DATE = date(2026, 1, 4)
SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_CSV = SCRIPT_DIR / f"{TEAM.lower()}_heijunka_extract.csv"
TIMELINESS_CSV = SCRIPT_DIR / "timeliness.csv"
CLOSURES_CSV = SCRIPT_DIR / "closures.csv"
METRICS_AGGREGATE_CSV = SCRIPT_DIR / "IV_DATA\\metrics_aggregate_dev.csv"
AVAILABLE_SHEET = "Available WIP Hours"
PRODUCTION_SHEET = "#12 Production Analysis"
AVAILABLE_ROWS = [7, 9, 11, 13, 15, 17, 19, 21, 23]
AVAILABLE_COLS = ["C", "D", "E", "F", "G"]
PROD_START_ROW = 7
PROD_END_ROW = 200
OUTPUT_COLUMNS = [
    "team",
    "period_date",
    "source_file",
    "Total Available Hours",
    "Completed Hours",
    "Target Output",
    "Actual Output",
    "Target UPLH",
    "Actual UPLH",
    "UPLH WP1",
    "UPLH WP2",
    "HC in WIP",
    "Actual HC Used",
    "People in WIP",
    "Person Hours",
    "Outputs by Person",
    "Outputs by Cell/Station",
    "Cell/Station Hours",
    "Hours by Cell/Station - by person",
    "Output by Cell/Station - by person",
    "UPLH by Cell/Station - by person",
    "Open Complaint Timeliness",
    "error",
    "Closures",
    "Opened",
]
def safe_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        if cleaned == "":
            return 0.0
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    return 0.0
def normalize_name(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
def normalize_team(value: Any) -> str:
    return normalize_name(value).upper()
def first_name_only(name: str) -> str:
    parts = normalize_name(name).split()
    return parts[0] if parts else ""
def parse_iso_date(value: Any) -> date | None:
    text = normalize_name(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None
def excel_date_to_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in (
            "%d-%b-%Y",
            "%d-%B-%Y",
            "%m-%d-%Y",
            "%m/%d/%Y",
            "%Y-%m-%d",
            "%d %B %Y",
            "%d %b %Y",
        ):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None
def parse_date_from_filename(path: Path) -> date | None:
    stem = path.stem
    prefix = f"{TEAM} Future Heijunka "
    if stem.startswith(prefix):
        tail = stem[len(prefix):].strip()
    else:
        tail = stem
    for fmt in ("%d %B %Y", "%d %b %Y", "%m-%d-%Y", "%m-%d-%y", "%m %d %Y"):
        try:
            return datetime.strptime(tail, fmt).date()
        except ValueError:
            continue
    m = re.search(r"(\d{1,2}[ -][A-Za-z]{3,9}[ -]\d{2,4}|\d{1,2}-\d{1,2}-\d{2,4})", tail)
    if m:
        token = m.group(1)
        for fmt in ("%d %B %Y", "%d %b %Y", "%d-%B-%Y", "%d-%b-%Y", "%m-%d-%Y", "%m-%d-%y"):
            try:
                return datetime.strptime(token, fmt).date()
            except ValueError:
                continue
    return None
def round_dict(obj: Any, digits: int = 2) -> Any:
    if isinstance(obj, dict):
        return {k: round_dict(v, digits) for k, v in obj.items()}
    if isinstance(obj, float):
        return round(obj, digits)
    return obj
def json_string(obj: Any) -> str:
    return json.dumps(round_dict(obj), ensure_ascii=False, sort_keys=True)
def divide(numerator: float, denominator: float) -> float | None:
    if denominator == 0:
        return None
    return numerator / denominator
def get_available_name_for_row(ws, row_num: int) -> str:
    candidates = [
        ws[f"A{row_num-1}"].value if row_num > 1 else None,
        ws[f"A{row_num}"].value,
        ws[f"A{row_num-2}"].value if row_num > 2 else None,
    ]
    for candidate in candidates:
        name = first_name_only(candidate)
        if name:
            return name
    return ""
def get_available_person_hours(ws) -> dict[str, dict[str, float]]:
    person_hours: dict[str, dict[str, float]] = {}
    for row_num in AVAILABLE_ROWS:
        name = get_available_name_for_row(ws, row_num)
        if not name:
            continue
        available = sum(safe_float(ws[f"{col}{row_num}"].value) for col in AVAILABLE_COLS)
        if name not in person_hours:
            person_hours[name] = {"actual": 0.0, "available": 0.0}
        person_hours[name]["available"] += available
    return person_hours
def load_timeliness_lookup(path: Path) -> dict[tuple[str, str], float]:
    lookup: dict[tuple[str, str], float] = {}
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for record in reader:
            team = normalize_team(record.get("team"))
            period = parse_iso_date(record.get("period_date"))
            if not team or not period:
                continue
            lookup[(team, period.isoformat())] = safe_float(record.get("Open Complaint Timeliness"))
    return lookup
def load_closures_lookup(path: Path) -> dict[tuple[str, str], dict[str, float]]:
    lookup: dict[tuple[str, str], dict[str, float]] = {}
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for record in reader:
            team = normalize_team(record.get("team"))
            period = parse_iso_date(record.get("period_date"))
            if not team or not period:
                continue
            lookup[(team, period.isoformat())] = {
                "Closures": safe_float(record.get("Closures")),
                "Opened": safe_float(record.get("Opened")),
            }
    return lookup
def process_workbook(
    path: Path,
    timeliness_lookup: dict[tuple[str, str], float],
    closures_lookup: dict[tuple[str, str], dict[str, float]],
) -> dict[str, Any]:
    row = {col: "" for col in OUTPUT_COLUMNS}
    row["team"] = TEAM
    row["source_file"] = str(path)
    parsed_file_date = parse_date_from_filename(path)
    if parsed_file_date:
        row["period_date"] = parsed_file_date.isoformat()
    wb = load_workbook(path, data_only=True, read_only=True)
    if AVAILABLE_SHEET not in wb.sheetnames:
        raise ValueError(f"Missing sheet: {AVAILABLE_SHEET}")
    if PRODUCTION_SHEET not in wb.sheetnames:
        raise ValueError(f"Missing sheet: {PRODUCTION_SHEET}")
    ws_avail = wb[AVAILABLE_SHEET]
    ws_prod = wb[PRODUCTION_SHEET]
    workbook_date = excel_date_to_date(ws_avail["B3"].value)
    if workbook_date:
        row["period_date"] = workbook_date.isoformat()
    person_hours = get_available_person_hours(ws_avail)
    total_available_hours = sum(info["available"] for info in person_hours.values())
    completed_minutes = 0.0
    target_output = 0.0
    actual_output = 0.0
    unique_people: set[str] = set()
    outputs_by_person: dict[str, dict[str, float]] = defaultdict(lambda: {"output": 0.0, "target": 0.0})
    outputs_by_station: dict[str, dict[str, float]] = defaultdict(lambda: {"output": 0.0, "target": 0.0})
    station_hours: dict[str, float] = defaultdict(float)
    hours_by_station_person: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    output_by_station_person: dict[str, dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: {"output": 0.0, "target": 0.0})
    )
    actual_hours_by_person: dict[str, float] = defaultdict(float)
    for excel_row in range(PROD_START_ROW, PROD_END_ROW + 1):
        person = first_name_only(ws_prod[f"C{excel_row}"].value)
        station = normalize_name(ws_prod[f"D{excel_row}"].value)
        target_minutes = safe_float(ws_prod[f"F{excel_row}"].value)
        completed_row_minutes = safe_float(ws_prod[f"G{excel_row}"].value)
        actual_output_value = safe_float(ws_prod[f"I{excel_row}"].value)
        completed_minutes += completed_row_minutes
        target_output += target_minutes
        actual_output += actual_output_value
        if person:
            unique_people.add(person)
            actual_hours_by_person[person] += completed_row_minutes / 60.0
            outputs_by_person[person]["target"] += target_minutes
            outputs_by_person[person]["output"] += actual_output_value
            if person not in person_hours:
                person_hours[person] = {"actual": 0.0, "available": 0.0}
        if station:
            outputs_by_station[station]["target"] += target_minutes
            outputs_by_station[station]["output"] += actual_output_value
            station_hours[station] += completed_row_minutes / 60.0
        if station and person:
            hours_by_station_person[station][person] += completed_row_minutes / 60.0
            output_by_station_person[station][person]["target"] += target_minutes
            output_by_station_person[station][person]["output"] += actual_output_value
    for person, actual in actual_hours_by_person.items():
        person_hours.setdefault(person, {"actual": 0.0, "available": 0.0})
        person_hours[person]["actual"] = actual
    completed_hours = completed_minutes / 60.0
    uplh_by_station_person: dict[str, dict[str, dict[str, float | None]]] = {}
    for station, person_map in output_by_station_person.items():
        uplh_by_station_person[station] = {}
        for person, values in person_map.items():
            hours = hours_by_station_person.get(station, {}).get(person, 0.0)
            uplh_by_station_person[station][person] = {
                "actual": divide(values["output"], hours),
                "target": divide(values["target"], hours),
            }
    row["Total Available Hours"] = round(total_available_hours, 2)
    row["Completed Hours"] = round(completed_hours, 2)
    row["Target Output"] = round(target_output, 2)
    row["Actual Output"] = round(actual_output, 2)
    row["Target UPLH"] = round(divide(target_output, completed_hours) or 0.0, 4)
    row["Actual UPLH"] = round(divide(actual_output, completed_hours) or 0.0, 4)
    row["UPLH WP1"] = ""
    row["UPLH WP2"] = ""
    row["HC in WIP"] = len(unique_people)
    row["Actual HC Used"] = round((completed_hours / 32.5) if completed_hours else 0.0, 2)
    row["People in WIP"] = ""
    row["Person Hours"] = json_string(person_hours)
    row["Outputs by Person"] = json_string(outputs_by_person)
    row["Outputs by Cell/Station"] = json_string(outputs_by_station)
    row["Cell/Station Hours"] = json_string(station_hours)
    row["Hours by Cell/Station - by person"] = json_string(hours_by_station_person)
    row["Output by Cell/Station - by person"] = json_string(output_by_station_person)
    row["UPLH by Cell/Station - by person"] = json_string(uplh_by_station_person)
    key = (normalize_team(row["team"]), row["period_date"])
    timeliness_value = timeliness_lookup.get(key)
    if timeliness_value is not None:
        row["Open Complaint Timeliness"] = round(timeliness_value, 1)
    closures_info = closures_lookup.get(key)
    if closures_info:
        row["Closures"] = round(closures_info["Closures"], 1)
        row["Opened"] = round(closures_info["Opened"], 1)
    row["error"] = ""
    wb.close()
    return row
def find_candidate_files(folders: list[Path]) -> list[Path]:
    seen: set[Path] = set()
    files: list[Path] = []
    for folder in folders:
        if not folder.exists():
            continue
        for path in sorted(folder.rglob(f"{TEAM} Future Heijunka *.xls*")):
            resolved = path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            parsed = parse_date_from_filename(path)
            if parsed is None or parsed >= START_DATE:
                files.append(path)
    return sorted(files)
def row_period_date_for_sort(row: dict[str, Any]) -> date:
    parsed = parse_iso_date(row.get("period_date"))
    return parsed if parsed else date.max
def sort_rows_by_date(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(rows, key=row_period_date_for_sort)
def aggregate_sort_key(row: dict[str, Any]) -> tuple[str, date]:
    return (
        normalize_name(row.get("team")).casefold(),
        row_period_date_for_sort(row),
    )
def read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    if not path.exists():
        return OUTPUT_COLUMNS.copy(), []
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames[:] if reader.fieldnames else OUTPUT_COLUMNS.copy()
        rows = list(reader)
    return fieldnames, rows
def merge_rows_by_team_period(
    existing_rows: list[dict[str, Any]],
    new_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    merged: dict[tuple[str, str], dict[str, Any]] = {}
    for row in existing_rows:
        key = (
            normalize_team(row.get("team")),
            normalize_name(row.get("period_date")),
        )
        merged[key] = row
    for row in new_rows:
        key = (
            normalize_team(row.get("team")),
            normalize_name(row.get("period_date")),
        )
        merged[key] = row
    return sorted(merged.values(), key=aggregate_sort_key)
def ensure_fieldnames(existing_fieldnames: list[str], required_fieldnames: list[str]) -> list[str]:
    final = existing_fieldnames[:] if existing_fieldnames else []
    for col in required_fieldnames:
        if col not in final:
            final.append(col)
    return final
def write_csv_rows(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            normalized_row = {col: row.get(col, "") for col in fieldnames}
            writer.writerow(normalized_row)
def main() -> None:
    files = find_candidate_files(FOLDERS)
    if not files:
        searched = ", ".join(str(f) for f in FOLDERS)
        raise FileNotFoundError(f"No matching files found in: {searched}")
    timeliness_lookup = load_timeliness_lookup(TIMELINESS_CSV)
    closures_lookup = load_closures_lookup(CLOSURES_CSV)
    rows: list[dict[str, Any]] = []
    for path in files:
        try:
            row = process_workbook(path, timeliness_lookup, closures_lookup)
            period = row.get("period_date")
            if period:
                period_date = datetime.strptime(period, "%Y-%m-%d").date()
                if period_date < START_DATE:
                    continue
            rows.append(row)
        except Exception as exc:
            parsed = parse_date_from_filename(path)
            rows.append(
                {
                    **{col: "" for col in OUTPUT_COLUMNS},
                    "team": TEAM,
                    "source_file": str(path),
                    "period_date": parsed.isoformat() if parsed else "",
                    "error": str(exc),
                }
            )
    rows = sort_rows_by_date(rows)
    write_csv_rows(OUTPUT_CSV, OUTPUT_COLUMNS, rows)
    existing_fieldnames, existing_rows = read_csv_rows(METRICS_AGGREGATE_CSV)
    final_fieldnames = ensure_fieldnames(existing_fieldnames, OUTPUT_COLUMNS)
    merged_rows = merge_rows_by_team_period(existing_rows, rows)
    write_csv_rows(METRICS_AGGREGATE_CSV, final_fieldnames, merged_rows)
    print(f"Wrote {len(rows)} rows to {OUTPUT_CSV}")
    print(f"Merged into {METRICS_AGGREGATE_CSV} with {len(merged_rows)} total rows")
if __name__ == "__main__":
    main()