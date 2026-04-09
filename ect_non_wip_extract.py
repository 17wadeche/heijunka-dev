from __future__ import annotations
import csv
import json
import re
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
TEAM = "ECT"
FOLDERS = [
    Path(r"C:\Users\wadec8\Medtronic PLC\Doran, Elaine - Heijunka Production Analysis\Archived Heijunka"),
    Path(r"C:\Users\wadec8\Medtronic PLC\Doran, Elaine - Heijunka Production Analysis"),
]
START_DATE = date(2026, 1, 4)
SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_CSV = SCRIPT_DIR / f"{TEAM.lower()}_non_wip_extract.csv"
NON_WIP_CSV = SCRIPT_DIR / "IV_DATA\\non_wip.csv"
NON_WIP_ACTIVITIES_CSV = SCRIPT_DIR / "non_wip_activities.csv"
AVAILABLE_SHEET = "Available WIP Hours"
PRODUCTION_SHEET = "#12 Production Analysis"
PEOPLE_NAME_START_ROW = 6
PEOPLE_NAME_END_ROW = 30
PROD_START_ROW = 7
PROD_END_ROW = 200
OUTPUT_COLUMNS = [
    "Team",
    "Week",
    "People Count",
    "Total Non-WIP Hours",
    "OOO Hours",
    "% in WIP",
    "Non-WIP by Person",
    "Non-WIP Activities",
    "WIP Workers",
    "WIP Workers Count",
    "WIP Workers OOO Hours",
]
NON_WIP_ACTIVITIES_COLUMNS = [
    "team",
    "period_date",
    "source_file",
    "people_count",
    "total_non_wip_hours",
    "% in WIP",
    "non_wip_by_person",
    "non_wip_activities",
    "OOO Hours",
]
def to_non_wip_activities_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "team": row.get("Team", ""),
        "period_date": row.get("Week", ""),
        "source_file": row.get("source_file", ""),
        "people_count": row.get("People Count", ""),
        "total_non_wip_hours": row.get("Total Non-WIP Hours", ""),
        "% in WIP": row.get("% in WIP", ""),
        "non_wip_by_person": row.get("Non-WIP by Person", ""),
        "non_wip_activities": row.get("Non-WIP Activities", ""),
        "OOO Hours": row.get("OOO Hours", ""),
    }
def normalize_name(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
def normalize_team(value: Any) -> str:
    return normalize_name(value).upper()
def first_name_only(value: Any) -> str:
    parts = normalize_name(value).split()
    return parts[0] if parts else ""
def safe_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        if not cleaned:
            return 0.0
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    return 0.0
def parse_iso_date(value: Any) -> date | None:
    text = normalize_name(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None
def excel_date_to_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in (
            "%d-%b-%Y",
            "%d-%B-%Y",
            "%m-%d-%Y",
            "%m/%d/%Y",
            "%Y-%m-%d",
            "%d %B %Y",
            "%d %b %Y",
        ):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None
def parse_date_from_filename(path: Path) -> date | None:
    stem = path.stem
    prefix = f"{TEAM} Future Heijunka "
    if stem.startswith(prefix):
        tail = stem[len(prefix):].strip()
    else:
        tail = stem
    for fmt in (
        "%d %B %Y",
        "%d %b %Y",
        "%m-%d-%Y",
        "%m-%d-%y",
        "%m %d %Y",
    ):
        try:
            return datetime.strptime(tail, fmt).date()
        except ValueError:
            continue
    m = re.search(
        r"(\d{1,2}[ -][A-Za-z]{3,9}[ -]\d{2,4}|\d{1,2}[ -]\d{1,2}[ -]\d{2,4})",
        tail,
    )
    if m:
        token = m.group(1).strip()
        for fmt in (
            "%d %B %Y",
            "%d %b %Y",
            "%d-%B-%Y",
            "%d-%b-%Y",
            "%d %B %Y",
            "%d %b %Y",
            "%m-%d-%Y",
            "%m-%d-%y",
            "%m %d %Y",
            "%d-%m-%Y",
            "%d-%m-%y",
        ):
            try:
                return datetime.strptime(token, fmt).date()
            except ValueError:
                continue
    return None
def json_string(obj: Any) -> str:
    return json.dumps(obj, ensure_ascii=False, sort_keys=False)
def round_hours(value: float) -> float:
    return round(value, 2)
def get_people_count(ws_avail) -> int:
    unique_names: set[str] = set()
    excluded = {"0", "total available hours"}
    for row_num in range(PEOPLE_NAME_START_ROW, PEOPLE_NAME_END_ROW + 1):
        raw_value = normalize_name(ws_avail[f"A{row_num}"].value)
        if not raw_value:
            continue
        if raw_value.casefold() in excluded:
            continue
        name = first_name_only(raw_value)
        if not name:
            continue
        if name.casefold() in excluded:
            continue
        unique_names.add(name)
    return len(unique_names)
def find_candidate_files(folders: list[Path]) -> list[Path]:
    seen: set[Path] = set()
    files: list[Path] = []
    for folder in folders:
        if not folder.exists():
            continue
        for path in sorted(folder.rglob("*.xls*")):
            if f"{TEAM} Future Heijunka" not in path.stem:
                continue
            resolved = path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            parsed = parse_date_from_filename(path)
            if parsed is None or parsed >= START_DATE:
                files.append(path)
    return sorted(files)
def row_has_real_non_wip_data(row: dict[str, Any]) -> bool:
    return any([
        safe_float(row.get("Total Non-WIP Hours")) > 0,
        safe_float(row.get("OOO Hours")) > 0,
        normalize_name(row.get("Non-WIP by Person")) not in {"", "{}", "[]"},
        normalize_name(row.get("Non-WIP Activities")) not in {"", "[]", "{}"},
        normalize_name(row.get("WIP Workers")) not in {"", "[]", "{}"},
        int(safe_float(row.get("WIP Workers Count"))) > 0,
        safe_float(row.get("WIP Workers OOO Hours")) > 0,
    ])
def row_quality_score(row: dict[str, Any]) -> tuple[int, float, float, int]:
    return (
        1 if row_has_real_non_wip_data(row) else 0,
        safe_float(row.get("Total Non-WIP Hours")),
        safe_float(row.get("OOO Hours")),
        int(safe_float(row.get("WIP Workers Count"))),
    )
def choose_better_row(existing: dict[str, Any], candidate: dict[str, Any]) -> dict[str, Any]:
    return candidate if row_quality_score(candidate) > row_quality_score(existing) else existing
def process_workbook(path: Path) -> dict[str, Any]:
    row = {col: "" for col in OUTPUT_COLUMNS}
    row["Team"] = TEAM
    parsed_file_date = parse_date_from_filename(path)
    if parsed_file_date:
        row["Week"] = parsed_file_date.isoformat()
    wb = load_workbook(path, data_only=True, read_only=True)
    if AVAILABLE_SHEET not in wb.sheetnames:
        raise ValueError(f"Missing sheet: {AVAILABLE_SHEET}")
    if PRODUCTION_SHEET not in wb.sheetnames:
        raise ValueError(f"Missing sheet: {PRODUCTION_SHEET}")
    ws_avail = wb[AVAILABLE_SHEET]
    ws_prod = wb[PRODUCTION_SHEET]
    workbook_date = excel_date_to_date(ws_avail["B3"].value)
    if workbook_date:
        row["Week"] = workbook_date.isoformat()
    people_count = get_people_count(ws_avail)
    total_non_wip_minutes = 0.0
    ooo_minutes = 0.0
    non_wip_by_person: dict[str, float] = defaultdict(float)
    non_wip_activities: list[dict[str, Any]] = []
    person_activity_types: dict[str, set[str]] = defaultdict(set)
    ooo_minutes_by_person: dict[str, float] = defaultdict(float)
    for excel_row in range(PROD_START_ROW, PROD_END_ROW + 1):
        person = first_name_only(ws_prod[f"C{excel_row}"].value)
        activity_type = normalize_name(ws_prod[f"D{excel_row}"].value)
        minutes = safe_float(ws_prod[f"G{excel_row}"].value)
        activity_detail = normalize_name(ws_prod[f"K{excel_row}"].value)
        if not person or minutes == 0:
            continue
        normalized_activity = activity_type.casefold()
        person_activity_types[person].add(normalized_activity)
        if normalized_activity == "non-wip":
            total_non_wip_minutes += minutes
            non_wip_by_person[person] += minutes / 60.0
            non_wip_activities.append(
                {
                    "name": person,
                    "activity": activity_detail if activity_detail else "Non-WIP",
                    "hours": round_hours(minutes / 60.0),
                }
            )
        elif normalized_activity == "other team wip":
            non_wip_activities.append(
                {
                    "name": person,
                    "activity": activity_detail if activity_detail else "Other Team WIP",
                    "hours": round_hours(minutes / 60.0),
                }
            )
        if normalized_activity == "ooo":
            ooo_minutes += minutes
            ooo_minutes_by_person[person] += minutes
            non_wip_activities.append(
                {
                    "name": person,
                    "activity": activity_detail if activity_detail else "OOO",
                    "hours": round_hours(minutes / 60.0),
                }
            )
    wip_workers: list[str] = []
    for person, activity_types in person_activity_types.items():
        if activity_types and activity_types.issubset({"non-wip", "ooo"}):
            continue
        wip_workers.append(person)
    wip_workers = sorted(set(wip_workers))
    wip_workers_ooo_hours = round_hours(
        sum(ooo_minutes_by_person.get(person, 0.0) for person in wip_workers) / 60.0
    )
    non_wip_by_person_rounded = {
        person: round_hours(hours)
        for person, hours in sorted(non_wip_by_person.items())
    }
    row["People Count"] = people_count
    row["Total Non-WIP Hours"] = round_hours(total_non_wip_minutes / 60.0)
    row["OOO Hours"] = round_hours(ooo_minutes / 60.0)
    row["% in WIP"] = ""
    row["Non-WIP by Person"] = json_string(non_wip_by_person_rounded)
    row["Non-WIP Activities"] = json_string(non_wip_activities)
    row["WIP Workers"] = json_string(wip_workers)
    row["WIP Workers Count"] = len(wip_workers)
    row["WIP Workers OOO Hours"] = wip_workers_ooo_hours
    wb.close()
    return row
def row_week_for_sort(row: dict[str, Any]) -> date:
    parsed = parse_iso_date(row.get("Week"))
    return parsed if parsed else date.max
def non_wip_sort_key(row: dict[str, Any]) -> tuple[str, date]:
    return (
        normalize_name(row.get("Team")).casefold(),
        row_week_for_sort(row),
    )
def sort_rows_by_week(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(rows, key=row_week_for_sort)
def read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    if not path.exists():
        return OUTPUT_COLUMNS.copy(), []
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames[:] if reader.fieldnames else OUTPUT_COLUMNS.copy()
        rows = list(reader)
    return fieldnames, rows
def ensure_fieldnames(existing_fieldnames: list[str], required_fieldnames: list[str]) -> list[str]:
    final = existing_fieldnames[:] if existing_fieldnames else []
    for col in required_fieldnames:
        if col not in final:
            final.append(col)
    return final
def merge_rows_by_team_week(
    existing_rows: list[dict[str, Any]],
    new_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    merged: dict[tuple[str, str], dict[str, Any]] = {}
    for row in existing_rows:
        key = (
            normalize_team(row.get("Team")),
            normalize_name(row.get("Week")),
        )
        merged[key] = row
    for row in new_rows:
        key = (
            normalize_team(row.get("Team")),
            normalize_name(row.get("Week")),
        )
        if key in merged:
            merged[key] = choose_better_row(merged[key], row)
        else:
            merged[key] = row
    return sorted(merged.values(), key=non_wip_sort_key)
def merge_rows_by_team_period(
    existing_rows: list[dict[str, Any]],
    new_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    merged: dict[tuple[str, str], dict[str, Any]] = {}
    for row in existing_rows:
        key = (
            normalize_team(row.get("team")),
            normalize_name(row.get("period_date")),
        )
        merged[key] = row
    for row in new_rows:
        key = (
            normalize_team(row.get("team")),
            normalize_name(row.get("period_date")),
        )
        if key in merged:
            existing = merged[key]
            candidate_score = (
                safe_float(row.get("total_non_wip_hours")),
                safe_float(row.get("OOO Hours")),
                1 if normalize_name(row.get("non_wip_by_person")) not in {"", "{}", "[]"} else 0,
                1 if normalize_name(row.get("non_wip_activities")) not in {"", "[]", "{}"} else 0,
            )
            existing_score = (
                safe_float(existing.get("total_non_wip_hours")),
                safe_float(existing.get("OOO Hours")),
                1 if normalize_name(existing.get("non_wip_by_person")) not in {"", "{}", "[]"} else 0,
                1 if normalize_name(existing.get("non_wip_activities")) not in {"", "[]", "{}"} else 0,
            )
            merged[key] = row if candidate_score > existing_score else existing
        else:
            merged[key] = row
    return sorted(
        merged.values(),
        key=lambda r: (
            normalize_name(r.get("team")).casefold(),
            row_week_for_sort({"Week": r.get("period_date", "")}),
        ),
    )
def dedupe_rows_by_team_week(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    deduped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (
            normalize_team(row.get("Team")),
            normalize_name(row.get("Week")),
        )
        if key in deduped:
            deduped[key] = choose_better_row(deduped[key], row)
        else:
            deduped[key] = row
    return sorted(deduped.values(), key=non_wip_sort_key)
def write_csv_rows(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            normalized_row = {col: row.get(col, "") for col in fieldnames}
            writer.writerow(normalized_row)
def main() -> None:
    files = find_candidate_files(FOLDERS)
    if not files:
        searched = ", ".join(str(f) for f in FOLDERS)
        raise FileNotFoundError(f"No matching files found in: {searched}")
    rows: list[dict[str, Any]] = []
    for path in files:
        try:
            row = process_workbook(path)
            week = row.get("Week")
            if week:
                week_date = datetime.strptime(week, "%Y-%m-%d").date()
                if week_date < START_DATE:
                    continue
            rows.append(row)
        except Exception as exc:
            parsed = parse_date_from_filename(path)
            rows.append(
                {
                    **{col: "" for col in OUTPUT_COLUMNS},
                    "Team": TEAM,
                    "Week": parsed.isoformat() if parsed else "",
                }
            )
            print(f"Error processing {path}: {exc}")
    rows = dedupe_rows_by_team_week(rows)
    write_csv_rows(OUTPUT_CSV, OUTPUT_COLUMNS, rows)
    existing_fieldnames_non_wip, existing_rows_non_wip = read_csv_rows(NON_WIP_CSV)
    final_fieldnames_non_wip = ensure_fieldnames(existing_fieldnames_non_wip, OUTPUT_COLUMNS)
    merged_rows_non_wip = merge_rows_by_team_week(existing_rows_non_wip, rows)
    write_csv_rows(NON_WIP_CSV, final_fieldnames_non_wip, merged_rows_non_wip)
    activity_rows = [to_non_wip_activities_row(row) for row in rows]
    existing_fieldnames_activities, existing_rows_activities = read_csv_rows(NON_WIP_ACTIVITIES_CSV)
    final_fieldnames_activities = ensure_fieldnames(
        existing_fieldnames_activities,
        NON_WIP_ACTIVITIES_COLUMNS,
    )
    merged_rows_activities = merge_rows_by_team_period(existing_rows_activities, activity_rows)
    write_csv_rows(
        NON_WIP_ACTIVITIES_CSV,
        final_fieldnames_activities,
        merged_rows_activities,
    )
    print(f"Wrote {len(rows)} rows to {OUTPUT_CSV}")
    print(f"Merged into {NON_WIP_CSV} with {len(merged_rows_non_wip)} total rows")
    print(f"Merged into {NON_WIP_ACTIVITIES_CSV} with {len(merged_rows_activities)} total rows")
if __name__ == "__main__":
    main()