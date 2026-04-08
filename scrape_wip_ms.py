from __future__ import annotations
import argparse
import csv
import datetime as _dt
import json
import os
from collections import defaultdict
from typing import Any, Dict, Iterable, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from collections import defaultdict
import json
from typing import Any, Dict, List, Optional, Tuple
TEAM_BY_SOURCE: Dict[str, str] = {
    r"C:\Users\wadec8\Medtronic PLC\CQXM RI-Heijunka live spreadsheet shared - Documents\WIP+Non-WIP Heijunka Template CQXM  VSS 2026 03.xlsm": "VSS",
    r"C:\Users\wadec8\Medtronic PLC\Robotics Complaint Intake - Heijunka\RST(US)-Heijunka Surgical.xlsm":"Surgical Robotics",
    r"C:\Users\wadec8\Medtronic PLC\Beeman, Amy - Heijunka Dashboard Endoscopy\WIP+Non-WIP Heijunka Template.xlsm":"Endoscopy",
    r"C:\Users\wadec8\Medtronic PLC\Surgical CQXM Team - !Heijunka\AST-GST(US) - Heijunka Surgical.xlsm":"Surgical AST-GST",
    r"C:\Users\wadec8\Medtronic PLC\SI INV Backlog - Heijunka\Surgical Inv (MEIC) - Heijunka.xlsm":"Surgical AST-GST MEIC",
    r"C:\Users\wadec8\Medtronic PLC\SI INV Backlog - Heijunka\Surgical Inv (US)-Heijunka.xlsm":"Surgical AST-GST",
    r"C:\Users\wadec8\Medtronic PLC\ACM Documents - General\ACM INV (US)-Heijunka v1.0 (002).xlsm":"ACM",
    r"C:\Users\wadec8\Medtronic PLC\Surgical CQXM Team - Legal Mesh -- Shared Folder\WIP+Non-WIP Heijunka Surgical - Legal Team.xlsm":"Surgical Legal",
    r"C:\Users\wadec8\Medtronic PLC\Robotics Complaint Intake - Heijunka\RST(MEIC)-Heijunka Surgical.xlsm":"Surgical Robotics MEIC",
    r"C:\Users\wadec8\Medtronic PLC\Heijunka -VSS,ENDO,ACM - Heijunka -VSS,ENDO,ACM\Heijunka v1.0- ACM - In Use.xlsm":"ACM MEIC",
    r"C:\Users\wadec8\Medtronic PLC\Heijunka -VSS,ENDO,ACM - Heijunka -VSS,ENDO,ACM\Heijunka v1.0- VSS.xlsm":"VSS MEIC",
    r"C:\Users\wadec8\Medtronic PLC\Heijunka -VSS,ENDO,ACM - Heijunka -VSS,ENDO,ACM\Heijunka v1.0-ENDO.xlsm":"Endo MEIC",
    
}
TEAM_BY_BASENAME: Dict[str, str] = {
    "WIP+Non-WIP Heijunka Template CQXM  VSS 2026 03.xlsm": "VSS",
    "RST(US)-Heijunka Surgical.xlsm":"Surgical Robotics",
    "WIP+Non-WIP Heijunka Template.xlsm":"Endoscopy",
    "AST-GST(US) - Heijunka Surgical.xlsm":"Surgical AST-GST",
    "ACM INV (US)-Heijunka v1.0 (002).xlsm":"ACM",
    "WIP+Non-WIP Heijunka Surgical - Legal Team.xlsm":"Surgical Legal",
    "Surgical Inv (MEIC) - Heijunka.xlsm":"Surgical AST-GST MEIC",
    "Surgical Inv (US)-Heijunka.xlsm":"Surgical AST-GST",
    "Heijunka v1.0- ACM - In Use.xlsm":"ACM MEIC",
    "Heijunka v1.0- VSS.xlsm":"VSS MEIC",
    "Heijunka v1.0-ENDO.xlsm":"Endo MEIC",
    "RST(MEIC)-Heijunka Surgical.xlsm":"Surgical Robotics MEIC",
}
DEFAULT_FILES: List[str] = [
    r"C:\Users\wadec8\Medtronic PLC\CQXM RI-Heijunka live spreadsheet shared - Documents\WIP+Non-WIP Heijunka Template CQXM  VSS 2026 03.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Robotics Complaint Intake - Heijunka\RST(US)-Heijunka Surgical.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Beeman, Amy - Heijunka Dashboard Endoscopy\WIP+Non-WIP Heijunka Template.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Surgical CQXM Team - !Heijunka\AST-GST(US) - Heijunka Surgical.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\ACM Documents - General\ACM INV (US)-Heijunka v1.0 (002).xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Surgical CQXM Team - Legal Mesh -- Shared Folder\WIP+Non-WIP Heijunka Surgical - Legal Team.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\SI INV Backlog - Heijunka\Surgical Inv (MEIC) - Heijunka.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\SI INV Backlog - Heijunka\Surgical Inv (US)-Heijunka.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Robotics Complaint Intake - Heijunka\RST(MEIC)-Heijunka Surgical.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Heijunka -VSS,ENDO,ACM - Heijunka -VSS,ENDO,ACM\Heijunka v1.0- ACM - In Use.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Heijunka -VSS,ENDO,ACM - Heijunka -VSS,ENDO,ACM\Heijunka v1.0- VSS.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Heijunka -VSS,ENDO,ACM - Heijunka -VSS,ENDO,ACM\Heijunka v1.0-ENDO.xlsm",
]
CSV_COLUMNS = [
    "team",
    "period_date",
    "source_file",
    "Total Available Hours",
    "Completed Hours",
    "Target Output",
    "Actual Output",
    "Target UPLH",
    "Actual UPLH",
    "UPLH WP1",
    "UPLH WP2",
    "HC in WIP",
    "Actual HC Used",
    "People in WIP",
    "Person Hours",
    "Outputs by Person",
    "Outputs by Cell/Station",
    "Cell/Station Hours",
    "Hours by Cell/Station - by person",
    "Output by Cell/Station - by person",
    "UPLH by Cell/Station - by person",
    "Open Complaint Timeliness",
    "error",
    "Closures",
    "Opened",
]
EXCLUDED_STATIONS = {"ooo", "non wip", ""}
AVAILABILITY_SHEET = "Available WIP+Non-WIP Hours"
PRODUCTION_SHEET = "Production Analysis"
EXCLUDED_NAMES = {
    "x", "0", "",
    "user1", "user2", "user3", "user4",
    "user5", "user6", "user7", "user8",
    "user9", "user10", "user11"
}
def json_load_safe(v: Any) -> Any:
    if v is None:
        return {}
    if isinstance(v, (dict, list)):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return {}
        try:
            return json.loads(s)
        except Exception:
            return {}
    return {}
def safe_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s:
            return 0.0
        try:
            return float(s)
        except ValueError:
            return 0.0
    return 0.0
def _sum_nested_person_map(dst: dict, src: dict, keys=("actual", "available")) -> None:
    for name, rec in (src or {}).items():
        if not isinstance(rec, dict):
            continue
        drec = dst.setdefault(name, {k: 0.0 for k in keys})
        for k in keys:
            drec[k] = safe_float(drec.get(k)) + safe_float(rec.get(k))
def _sum_nested_output_target_map(dst: dict, src: dict) -> None:
    for name, rec in (src or {}).items():
        if not isinstance(rec, dict):
            continue
        drec = dst.setdefault(name, {"output": 0.0, "target": 0.0})
        drec["output"] = safe_float(drec.get("output")) + safe_float(rec.get("output"))
        drec["target"] = safe_float(drec.get("target")) + safe_float(rec.get("target"))
def _sum_simple_map(dst: dict, src: dict) -> None:
    for k, v in (src or {}).items():
        dst[k] = safe_float(dst.get(k)) + safe_float(v)

def _sum_cell_person_map(dst: dict, src: dict) -> None:
    for cell, people in (src or {}).items():
        if not isinstance(people, dict):
            continue
        dcell = dst.setdefault(cell, {})
        for person, val in people.items():
            dcell[person] = safe_float(dcell.get(person)) + safe_float(val)

def _recalc_uplh_by_station_by_person(
    output_by_station_by_person: Dict[str, Dict[str, float]],
    hours_by_station_by_person: Dict[str, Dict[str, float]],
) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for station in sorted(set(output_by_station_by_person) | set(hours_by_station_by_person)):
        omap = output_by_station_by_person.get(station, {}) or {}
        hmap = hours_by_station_by_person.get(station, {}) or {}
        for person in sorted(set(omap) | set(hmap)):
            hrs = safe_float(hmap.get(person))
            outv = safe_float(omap.get(person))
            if hrs:
                out.setdefault(station, {})
                out[station][person] = outv / hrs
    return out
TEAM_ROLLUP_MAP: Dict[str, str] = {
    "VSS": "VSS",
    "VSS MEIC": "VSS",
    "ACM": "ACM",
    "ACM MEIC": "ACM",
    "Endoscopy": "Endoscopy",
    "Endo MEIC": "Endoscopy",
    "Surgical Robotics": "Surgical Robotics",
    "Surgical Robotics MEIC": "Surgical Robotics",
    "Surgical AST-GST": "Surgical AST-GST",
    "Surgical AST-GST MEIC": "Surgical AST-GST",
    "Surgical Legal": "Surgical Legal",
}
def rollup_team_name(team: str) -> str:
    t = (team or "").strip()
    return TEAM_ROLLUP_MAP.get(t, t)
def rollup_rows_by_team_period(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    buckets: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        team = rollup_team_name((row.get("team", "") or "").strip())
        period_date = (row.get("period_date", "") or "").strip()
        buckets[(team, period_date)].append(row)
    out_rows: List[Dict[str, Any]] = []
    for (team, period_date), group in buckets.items():
        if len(group) == 1 or not team or not period_date:
            only = dict(group[0])
            only["team"] = team
            out_rows.append(only)
            continue
        total_available_hours = 0.0
        completed_hours = 0.0
        target_output = 0.0
        actual_output = 0.0
        people_in_wip_set = set()
        person_hours: Dict[str, Dict[str, float]] = {}
        outputs_by_person: Dict[str, Dict[str, float]] = {}
        outputs_by_station: Dict[str, Dict[str, float]] = {}
        station_hours: Dict[str, float] = {}
        hours_by_station_by_person: Dict[str, Dict[str, float]] = {}
        output_by_station_by_person: Dict[str, Dict[str, float]] = {}
        open_complaint_timeliness = ""
        closures = ""
        opened = ""
        source_files: List[str] = []
        errors: List[str] = []
        for row in group:
            total_available_hours += safe_float(row.get("Total Available Hours"))
            completed_hours += safe_float(row.get("Completed Hours"))
            target_output += safe_float(row.get("Target Output"))
            actual_output += safe_float(row.get("Actual Output"))
            people_in_wip = json_load_safe(row.get("People in WIP"))
            if isinstance(people_in_wip, list):
                people_in_wip_set.update(str(x).strip() for x in people_in_wip if str(x).strip())
            _sum_nested_person_map(
                person_hours,
                json_load_safe(row.get("Person Hours")),
                keys=("actual", "available"),
            )
            _sum_nested_output_target_map(
                outputs_by_person,
                json_load_safe(row.get("Outputs by Person")),
            )
            _sum_nested_output_target_map(
                outputs_by_station,
                json_load_safe(row.get("Outputs by Cell/Station")),
            )
            _sum_simple_map(
                station_hours,
                json_load_safe(row.get("Cell/Station Hours")),
            )
            _sum_cell_person_map(
                hours_by_station_by_person,
                json_load_safe(row.get("Hours by Cell/Station - by person")),
            )
            _sum_cell_person_map(
                output_by_station_by_person,
                json_load_safe(row.get("Output by Cell/Station - by person")),
            )
            sf = (row.get("source_file", "") or "").strip()
            if sf:
                source_files.append(sf)
            er = (row.get("error", "") or "").strip()
            if er:
                errors.append(er)
            if not open_complaint_timeliness:
                open_complaint_timeliness = (row.get("Open Complaint Timeliness", "") or "").strip()
            if not closures:
                closures = (row.get("Closures", "") or "").strip()
            if not opened:
                opened = (row.get("Opened", "") or "").strip()
        target_uplh = safe_div(target_output, completed_hours)
        actual_uplh = safe_div(actual_output, completed_hours)
        actual_hc_used = safe_div(completed_hours, 32.5)
        uplh_by_station_by_person = _recalc_uplh_by_station_by_person(
            output_by_station_by_person,
            hours_by_station_by_person,
        )
        wp1 = outputs_by_station.get("WP1", {})
        wp2 = outputs_by_station.get("WP2", {})
        wp1_hours = safe_float(station_hours.get("WP1"))
        wp2_hours = safe_float(station_hours.get("WP2"))
        uplh_wp1 = safe_div(safe_float(wp1.get("output")), wp1_hours) if wp1_hours else ""
        uplh_wp2 = safe_div(safe_float(wp2.get("output")), wp2_hours) if wp2_hours else ""
        out_rows.append({
            "team": team,
            "period_date": period_date,
            "source_file": " | ".join(sorted(set(source_files))),
            "Total Available Hours": total_available_hours,
            "Completed Hours": completed_hours,
            "Target Output": target_output,
            "Actual Output": actual_output,
            "Target UPLH": float(target_uplh) if target_uplh is not None else "",
            "Actual UPLH": float(actual_uplh) if actual_uplh is not None else "",
            "UPLH WP1": float(uplh_wp1) if uplh_wp1 not in (None, "") else "",
            "UPLH WP2": float(uplh_wp2) if uplh_wp2 not in (None, "") else "",
            "HC in WIP": len(people_in_wip_set),
            "Actual HC Used": float(actual_hc_used) if actual_hc_used is not None else "",
            "People in WIP": dumps_json(sorted(people_in_wip_set)),
            "Person Hours": dumps_json(person_hours),
            "Outputs by Person": dumps_json(outputs_by_person),
            "Outputs by Cell/Station": dumps_json(outputs_by_station),
            "Cell/Station Hours": dumps_json(station_hours),
            "Hours by Cell/Station - by person": dumps_json(hours_by_station_by_person),
            "Output by Cell/Station - by person": dumps_json(output_by_station_by_person),
            "UPLH by Cell/Station - by person": dumps_json(uplh_by_station_by_person),
            "Open Complaint Timeliness": open_complaint_timeliness,
            "error": "; ".join(sorted(set(errors))) if errors else "",
            "Closures": closures,
            "Opened": opened,
        })
    out_rows.sort(key=lambda r: ((r.get("team", "") or "").lower(), r.get("period_date", "") or "9999-12-31"))
    return out_rows
def is_valid_name(name: str) -> bool:
    return name.strip().lower() not in EXCLUDED_NAMES
def _norm_path(p: str) -> str:
    return os.path.normpath(p)
def team_for_source(path: str) -> str:
    np = _norm_path(path)
    if np in TEAM_BY_SOURCE:
        return TEAM_BY_SOURCE[np]
    return TEAM_BY_BASENAME.get(os.path.basename(np), "")
def _cell_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None
def _norm_team(s: str) -> str:
    return (s or "").strip().upper()
def _norm_period_date(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    try:
        return _dt.date.fromisoformat(s).isoformat()
    except ValueError:
        return s
def script_dir() -> str:
    return os.path.dirname(os.path.abspath(__file__))
def load_timeliness_lookup(path: str) -> Dict[Tuple[str, str], str]:
    lut: Dict[Tuple[str, str], str] = {}
    if not os.path.exists(path):
        return lut
    with open(path, "r", newline="", encoding="utf-8-sig") as fp:
        reader = csv.DictReader(fp)
        for row in reader:
            team = _norm_team(row.get("team", ""))
            period = _norm_period_date(row.get("period_date", ""))
            val = (row.get("Open Complaint Timeliness", "") or "").strip()
            if team and period:
                lut[(team, period)] = val
    return lut
def load_closures_lookup(path: str) -> Dict[Tuple[str, str], Tuple[str, str]]:
    lut: Dict[Tuple[str, str], Tuple[str, str]] = {}
    if not os.path.exists(path):
        return lut
    with open(path, "r", newline="", encoding="utf-8-sig") as fp:
        reader = csv.DictReader(fp)
        for row in reader:
            team = _norm_team(row.get("team", ""))
            period = _norm_period_date(row.get("period_date", ""))
            closures = (row.get("Closures", "") or "").strip()
            opened = (row.get("Opened", "") or "").strip()
            if team and period:
                lut[(team, period)] = (closures, opened)
    return lut
def enrich_rows_with_metrics(
    rows: List[Dict[str, Any]],
    timeliness_lut: Dict[Tuple[str, str], str],
    closures_lut: Dict[Tuple[str, str], Tuple[str, str]],
) -> None:
    for row in rows:
        key = (_norm_team(row.get("team", "")), _norm_period_date(row.get("period_date", "")))
        if key in timeliness_lut:
            row["Open Complaint Timeliness"] = timeliness_lut[key]
        if key in closures_lut:
            closures, opened = closures_lut[key]
            row["Closures"] = closures
            row["Opened"] = opened
def _as_text(v: Any) -> str:
    return str(v).strip() if v is not None else ""
def _as_date(v: Any) -> Optional[_dt.date]:
    if v is None:
        return None
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
            try:
                return _dt.datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None
def monday_of_week(d: _dt.date) -> _dt.date:
    return d - _dt.timedelta(days=d.weekday())
def iso_date(d: Optional[_dt.date]) -> str:
    return d.isoformat() if isinstance(d, _dt.date) else ""
def dumps_json(obj: Any) -> str:
    return json.dumps(obj, ensure_ascii=False)
def safe_div(n: float, d: float) -> Optional[float]:
    if d in (0, 0.0) or d is None:
        return None
    return n / d
def is_wip_station(station: str) -> bool:
    return station.strip().lower() not in EXCLUDED_STATIONS
def iter_available_blocks(ws: Worksheet) -> Iterable[Tuple[int, _dt.date]]:
    for row in range(1, ws.max_row + 1):
        label = _as_text(ws.cell(row=row, column=1).value)
        if label.lower() == "week starting:":
            week_val = ws.cell(row=row, column=2).value
            week_date = _as_date(week_val)
            if week_date is not None:
                yield row, week_date
def find_next_week_row(starts: List[Tuple[int, _dt.date]], idx: int, max_row: int) -> int:
    if idx + 1 < len(starts):
        return starts[idx + 1][0] - 1
    return max_row
def parse_available_sheet(ws: Worksheet) -> Dict[_dt.date, Dict[str, float]]:
    result: Dict[_dt.date, Dict[str, float]] = {}
    starts = list(iter_available_blocks(ws))
    for idx, (start_row, week_date) in enumerate(starts):
        end_row = find_next_week_row(starts, idx, ws.max_row)
        people_avail: Dict[str, float] = {}
        current_person = ""
        row = start_row + 1
        while row <= end_row:
            person_cell = _as_text(ws.cell(row=row, column=3).value)
            category = _as_text(ws.cell(row=row, column=4).value)
            if person_cell:
                current_person = person_cell
            if is_valid_name(current_person) and category.lower() == "available wip":
                total = 0.0
                for col in range(5, 10):  # E:I
                    total += _cell_number(ws.cell(row=row, column=col).value) or 0.0
                people_avail[current_person] = total
            row += 1
        result[week_date] = people_avail
    return result
def iter_production_rows(ws: Worksheet, start_row: int = 3) -> Iterable[Tuple[_dt.date, str, str, float, float, float]]:
    for row in range(start_row, ws.max_row + 1):
        date_val = _as_date(ws.cell(row=row, column=1).value)
        if date_val is None:
            continue
        person = _as_text(ws.cell(row=row, column=2).value)
        station = _as_text(ws.cell(row=row, column=3).value)
        mins = _cell_number(ws.cell(row=row, column=5).value) or 0.0
        actual_output = _cell_number(ws.cell(row=row, column=6).value) or 0.0
        target_output = _cell_number(ws.cell(row=row, column=8).value) or 0.0
        yield (date_val, person, station, mins, actual_output, target_output)
def nested_float_dict() -> defaultdict:
    return defaultdict(float)
def parse_production_sheet(ws: Worksheet) -> Dict[_dt.date, Dict[str, Any]]:
    weekly: Dict[_dt.date, Dict[str, Any]] = {}
    for date_val, person, station, mins, actual_output, target_output in iter_production_rows(ws):
        period = monday_of_week(date_val)
        if period not in weekly:
            weekly[period] = {
                "completed_minutes": 0.0,
                "target_output": 0.0,
                "actual_output": 0.0,
                "wip_people": set(),
                "actual_hours_by_person": defaultdict(float),
                "outputs_by_person": defaultdict(lambda: {"output": 0.0, "target": 0.0}),
                "outputs_by_station": defaultdict(lambda: {"output": 0.0, "target": 0.0}),
                "station_hours": defaultdict(float),
                "hours_by_station_by_person": defaultdict(nested_float_dict),
                "output_by_station_by_person": defaultdict(nested_float_dict),
            }
        bucket = weekly[period]
        bucket["target_output"] += target_output
        bucket["actual_output"] += actual_output
        station_is_wip = is_wip_station(station)
        if station_is_wip:
            bucket["completed_minutes"] += mins
            if person and is_valid_name(person):
                bucket["wip_people"].add(person)
                bucket["actual_hours_by_person"][person] += mins / 60.0
                bucket["outputs_by_person"][person]["output"] += actual_output
                bucket["outputs_by_person"][person]["target"] += target_output
                bucket["hours_by_station_by_person"][station][person] += mins / 60.0
                bucket["output_by_station_by_person"][station][person] += actual_output
            bucket["outputs_by_station"][station]["output"] += actual_output
            bucket["outputs_by_station"][station]["target"] += target_output
            bucket["station_hours"][station] += mins / 60.0
    return weekly
def finalize_nested_defaultdict(obj: Any) -> Any:
    if isinstance(obj, defaultdict):
        return {k: finalize_nested_defaultdict(v) for k, v in obj.items()}
    if isinstance(obj, dict):
        return {k: finalize_nested_defaultdict(v) for k, v in obj.items()}
    if isinstance(obj, set):
        return sorted(obj)
    return obj
def compute_uplh_by_station_by_person(
    output_by_station_by_person: Dict[str, Dict[str, float]],
    hours_by_station_by_person: Dict[str, Dict[str, float]],
) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for station, person_outputs in output_by_station_by_person.items():
        for person, output_val in person_outputs.items():
            hrs = hours_by_station_by_person.get(station, {}).get(person, 0.0)
            if hrs:
                out.setdefault(station, {})
                out[station][person] = output_val / hrs
    return out
def build_person_hours_json(available_by_person: Dict[str, float], actual_by_person: Dict[str, float]) -> str:
    all_people = sorted(set(available_by_person.keys()) | set(actual_by_person.keys()))
    payload: Dict[str, Dict[str, float]] = {}
    for person in all_people:
        payload[person] = {
            "actual": float(actual_by_person.get(person, 0.0)),
            "available": float(available_by_person.get(person, 0.0)),
        }
    return dumps_json(payload)
def blank_row_for_missing_file(path: str) -> Dict[str, Any]:
    return {
        "team": team_for_source(path),
        "period_date": "",
        "source_file": path,
        "Total Available Hours": "",
        "Completed Hours": "",
        "Target Output": "",
        "Actual Output": "",
        "Target UPLH": "",
        "Actual UPLH": "",
        "UPLH WP1": "",
        "UPLH WP2": "",
        "HC in WIP": "",
        "Actual HC Used": "",
        "People in WIP": "",
        "Person Hours": "",
        "Outputs by Person": "",
        "Outputs by Cell/Station": "",
        "Cell/Station Hours": "",
        "Hours by Cell/Station - by person": "",
        "Output by Cell/Station - by person": "",
        "UPLH by Cell/Station - by person": "",
        "Open Complaint Timeliness": "",
        "error": f"file_not_found: {path}",
        "Closures": "",
        "Opened": "",
    }
def scrape_one_workbook(path: str) -> List[Dict[str, Any]]:
    team = team_for_source(path)
    rows: List[Dict[str, Any]] = []
    err_msgs: List[str] = []
    wb = load_workbook(path, data_only=True)
    available_by_week: Dict[_dt.date, Dict[str, float]] = {}
    production_by_week: Dict[_dt.date, Dict[str, Any]] = {}
    if AVAILABILITY_SHEET in wb.sheetnames:
        try:
            available_by_week = parse_available_sheet(wb[AVAILABILITY_SHEET])
        except Exception as e:
            err_msgs.append(f"availability_parse_error: {e!r}")
    else:
        err_msgs.append(f"missing_sheet: {AVAILABILITY_SHEET}")
    if PRODUCTION_SHEET in wb.sheetnames:
        try:
            production_by_week = parse_production_sheet(wb[PRODUCTION_SHEET])
        except Exception as e:
            err_msgs.append(f"production_parse_error: {e!r}")
    else:
        err_msgs.append(f"missing_sheet: {PRODUCTION_SHEET}")
    periods = sorted(set(available_by_week.keys()) | set(production_by_week.keys()))
    for period in periods:
        week_errors = list(err_msgs)
        person_avail = available_by_week.get(period, {})
        prod = production_by_week.get(period, {})
        completed_hours = float((prod.get("completed_minutes", 0.0) or 0.0) / 60.0)
        target_output = float(prod.get("target_output", 0.0) or 0.0)
        actual_output = float(prod.get("actual_output", 0.0) or 0.0)
        target_uplh = safe_div(target_output, completed_hours)
        actual_uplh = safe_div(actual_output, completed_hours)
        people_in_wip = sorted(prod.get("wip_people", set()))
        actual_by_person = finalize_nested_defaultdict(prod.get("actual_hours_by_person", {}))
        outputs_by_person = finalize_nested_defaultdict(prod.get("outputs_by_person", {}))
        outputs_by_station = finalize_nested_defaultdict(prod.get("outputs_by_station", {}))
        station_hours = finalize_nested_defaultdict(prod.get("station_hours", {}))
        hours_by_station_by_person = finalize_nested_defaultdict(prod.get("hours_by_station_by_person", {}))
        output_by_station_by_person = finalize_nested_defaultdict(prod.get("output_by_station_by_person", {}))
        uplh_by_station_by_person = compute_uplh_by_station_by_person(
            output_by_station_by_person, hours_by_station_by_person
        )
        total_available_hours = float(sum(person_avail.values()))
        actual_hc_used = safe_div(completed_hours, 32.5)
        row: Dict[str, Any] = {
            "team": team,
            "period_date": iso_date(period),
            "source_file": path,
            "Total Available Hours": total_available_hours,
            "Completed Hours": completed_hours,
            "Target Output": target_output,
            "Actual Output": actual_output,
            "Target UPLH": float(target_uplh) if target_uplh is not None else "",
            "Actual UPLH": float(actual_uplh) if actual_uplh is not None else "",
            "UPLH WP1": "",
            "UPLH WP2": "",
            "HC in WIP": len(people_in_wip),
            "Actual HC Used": float(actual_hc_used) if actual_hc_used is not None else "",
            "People in WIP": dumps_json(people_in_wip),
            "Person Hours": build_person_hours_json(person_avail, actual_by_person),
            "Outputs by Person": dumps_json(outputs_by_person),
            "Outputs by Cell/Station": dumps_json(outputs_by_station),
            "Cell/Station Hours": dumps_json(station_hours),
            "Hours by Cell/Station - by person": dumps_json(hours_by_station_by_person),
            "Output by Cell/Station - by person": dumps_json(output_by_station_by_person),
            "UPLH by Cell/Station - by person": dumps_json(uplh_by_station_by_person),
            "Open Complaint Timeliness": "",
            "error": "; ".join(week_errors) if week_errors else "",
            "Closures": "",
            "Opened": "",
        }
        rows.append(row)
    if not rows:
        rows.append({
            "team": team,
            "period_date": "",
            "source_file": path,
            "Total Available Hours": "",
            "Completed Hours": "",
            "Target Output": "",
            "Actual Output": "",
            "Target UPLH": "",
            "Actual UPLH": "",
            "UPLH WP1": "",
            "UPLH WP2": "",
            "HC in WIP": "",
            "Actual HC Used": "",
            "People in WIP": "",
            "Person Hours": "",
            "Outputs by Person": "",
            "Outputs by Cell/Station": "",
            "Cell/Station Hours": "",
            "Hours by Cell/Station - by person": "",
            "Output by Cell/Station - by person": "",
            "UPLH by Cell/Station - by person": "",
            "Open Complaint Timeliness": "",
            "error": "; ".join(err_msgs) if err_msgs else "no_periods_found",
            "Closures": "",
            "Opened": "",
        })
    return rows
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("files", nargs="*", help="Optional workbook(s) to scrape. If omitted, uses DEFAULT_FILES in the script.")
    ap.add_argument("--out", default="MS_WIP.csv", help="Output CSV path.")
    args = ap.parse_args()
    files = args.files or DEFAULT_FILES
    all_rows: List[Dict[str, Any]] = []
    for path in files:
        if not os.path.exists(path):
            all_rows.append(blank_row_for_missing_file(path))
            continue
        all_rows.extend(scrape_one_workbook(path))
    base_dir = script_dir()
    timeliness_csv = os.path.join(base_dir, "timeliness.csv")
    closures_csv = os.path.join(base_dir, "closures.csv")
    timeliness_lut = load_timeliness_lookup(timeliness_csv)
    closures_lut = load_closures_lookup(closures_csv)
    enrich_rows_with_metrics(all_rows, timeliness_lut, closures_lut)
    final_rows = rollup_rows_by_team_period(all_rows)
    with open(args.out, "w", newline="", encoding="utf-8") as fp:
        writer = csv.DictWriter(fp, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for row in final_rows:
            writer.writerow({k: row.get(k, "") for k in CSV_COLUMNS})
    print(f"Wrote {len(final_rows)} row(s) to {args.out}")
    return 0
if __name__ == "__main__":
    raise SystemExit(main())