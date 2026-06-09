import csv
import json
import os
import re
import shutil
import tempfile
import uuid
from datetime import datetime, date, timedelta
from typing import Any, Dict, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import win32com.client
import time
import pythoncom
import pywintypes
import argparse
import logging
import sys
import traceback
from contextlib import contextmanager
from threading import Thread, Event
import math
WIP_HEADERS = [
    "team",
    "period_date",
    "Total Available Hours",
    "Completed Hours",
    "Target Output",
    "Actual Output",
    "Target UPLH",
    "Actual UPLH",
    "UPLH WP1",
    "UPLH WP2",
    "HC in WIP",
    "Actual HC Used",
    "People in WIP",
    "Person Hours",
    "Outputs by Person",
    "Outputs by Cell/Station",
    "Cell/Station Hours",
    "Hours by Cell/Station - by person",
    "Output by Cell/Station - by person",
    "UPLH by Cell/Station - by person",
    "error",
]
def json_load_safe(s: Any) -> dict:
    if s is None:
        return {}
    if isinstance(s, dict):
        return s
    txt = safe_str(s)
    if not txt:
        return {}
    try:
        return json.loads(txt)
    except Exception:
        return {}
def parse_sheet_date_requires_year(sheet_name: str) -> str:
    raw = (sheet_name or "").strip()
    if not re.search(r"\b\d{4}\b", raw):
        return ""
    raw = raw.replace("\u00a0", " ")
    raw = re.sub(r"(\d{1,2})(st|nd|rd|th)\b", r"\1", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\s+", " ", raw).strip()
    fmts = [
        "%b %d, %Y",
        "%B %d, %Y",
        "%b %d,%Y",
        "%B %d,%Y",
        "%b %d %Y",
        "%B %d %Y",
        "%Y-%m-%d",
        "%m/%d/%Y",
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return ""
def _sum_nested_person_map(dst: dict, src: dict, keys=("actual", "available")) -> None:
    for name, rec in (src or {}).items():
        if not isinstance(rec, dict):
            continue
        drec = dst.setdefault(name, {k: 0.0 for k in keys})
        for k in keys:
            drec[k] = safe_float(drec.get(k)) + safe_float(rec.get(k))
def _sum_nested_output_target_map(dst: dict, src: dict) -> None:
    for name, rec in (src or {}).items():
        if not isinstance(rec, dict):
            continue
        drec = dst.setdefault(name, {"output": 0.0, "target": 0.0})
        drec["output"] = safe_float(drec.get("output")) + safe_float(rec.get("output"))
        drec["target"] = safe_float(drec.get("target")) + safe_float(rec.get("target"))
def _sum_cell_map(dst: dict, src: dict) -> None:
    for cell, rec in (src or {}).items():
        if not isinstance(rec, dict):
            continue
        drec = dst.setdefault(cell, {"output": 0.0, "target": 0.0})
        drec["output"] = safe_float(drec.get("output")) + safe_float(rec.get("output"))
        drec["target"] = safe_float(drec.get("target")) + safe_float(rec.get("target"))
def _sum_simple_map(dst: dict, src: dict) -> None:
    for k, v in (src or {}).items():
        dst[k] = safe_float(dst.get(k)) + safe_float(v)
def _sum_cell_person_map(dst: dict, src: dict) -> None:
    for cell, people in (src or {}).items():
        if not isinstance(people, dict):
            continue
        dcell = dst.setdefault(cell, {})
        for person, val in people.items():
            dcell[person] = safe_float(dcell.get(person)) + safe_float(val)
def _recalc_uplh_by_cell_person(hours_by_cell_person: dict, out_by_cell_person: dict) -> dict:
    uplh = {}
    for cell in ("WP1", "WP2"):
        uplh[cell] = {}
        hmap = (hours_by_cell_person or {}).get(cell, {}) or {}
        omap = (out_by_cell_person or {}).get(cell, {}) or {}
        for person in set(list(hmap.keys()) + list(omap.keys())):
            hrs = safe_float(hmap.get(person))
            outv = safe_float(omap.get(person))
            uplh[cell][person] = safe_div(outv, hrs)
    return uplh
def build_ns_wip_rows(all_rows: list[dict]) -> list[dict]:
    et_combine_teams = {"O-Arm MEIC", "Nav", "Mazor", "AE MEIC", "CSF"}
    et_label = "Enabling Technologies"
    rollups = [
        ({"MEIC PH", "DBS MEIC", "Other MEIC"}, "PH-NM MEIC"),
        ({"DBS C13", "DBS C14"}, "DBS"),
        ({"PH", "PH Cell 17"}, "PH"),
        ({"SCS Cell 1", "SCS Super Cell"}, "SCS"),
        ({"PSS US", "PSS MEIC", "PSS MEIC Intern"}, "PSS"),
        (et_combine_teams, et_label),
    ]
    rename_map = {"TDD COS 1": "TDD"}
    buckets_by_label: Dict[str, Dict[str, list[dict]]] = {label: {} for _, label in rollups}
    passthrough: list[dict] = []
    for r in all_rows:
        team = safe_str(r.get("team"))
        if team in rename_map:
            r = dict(r)
            r["team"] = rename_map[team]
            team = r["team"]
        placed = False
        pd = safe_str(r.get("period_date"))
        for team_set, label in rollups:
            if team in team_set and pd:
                buckets_by_label[label].setdefault(pd, []).append(r)
                placed = True
                break
        if not placed:
            passthrough.append(r)
    out_rows: list[dict] = []
    out_rows.extend(passthrough)
    def _emit_rollup(label: str, period_date: str, rows: list[dict]) -> None:
        taa = 0.0
        ch = 0.0
        tgt_out = 0.0
        act_out = 0.0
        hc_wip = 0.0
        wp1_out_total = 0.0
        wp2_out_total = 0.0
        wp1_uplh_weighted_sum = 0.0
        wp2_uplh_weighted_sum = 0.0
        person_hours = {}
        outputs_by_person = {}
        outputs_by_cell = {}
        cell_station_hours = {}
        hours_by_cell_person = {}
        out_by_cell_person = {}
        errs = []
        for r in rows:
            taa += safe_float(r.get("Total Available Hours"))
            ch += safe_float(r.get("Completed Hours"))
            tgt_out += safe_float(r.get("Target Output"))
            act_out += safe_float(r.get("Actual Output"))
            hc_wip += int(safe_float(r.get("HC in WIP")))
            cell_json = json_load_safe(r.get("Outputs by Cell/Station"))
            wp1_out = safe_float(((cell_json.get("WP1") or {}).get("output")))
            wp2_out = safe_float(((cell_json.get("WP2") or {}).get("output")))
            wp1_u = r.get("UPLH WP1")
            wp2_u = r.get("UPLH WP2")
            if wp1_out > 0:
                wp1_out_total += wp1_out
                wp1_uplh_weighted_sum += safe_float(wp1_u) * wp1_out
            if wp2_out > 0:
                wp2_out_total += wp2_out
                wp2_uplh_weighted_sum += safe_float(wp2_u) * wp2_out
            _sum_nested_person_map(person_hours, json_load_safe(r.get("Person Hours")), keys=("actual", "available"))
            _sum_nested_output_target_map(outputs_by_person, json_load_safe(r.get("Outputs by Person")))
            _sum_cell_map(outputs_by_cell, cell_json)
            _sum_simple_map(cell_station_hours, json_load_safe(r.get("Cell/Station Hours")))
            _sum_cell_person_map(hours_by_cell_person, json_load_safe(r.get("Hours by Cell/Station - by person")))
            _sum_cell_person_map(out_by_cell_person, json_load_safe(r.get("Output by Cell/Station - by person")))
            er = safe_str(r.get("error"))
            if er:
                errs.append(er)
        target_uplh = safe_div(tgt_out, ch)
        actual_uplh = safe_div(act_out, ch)
        uplh_wp1 = safe_div(wp1_uplh_weighted_sum, wp1_out_total) 
        uplh_wp2 = safe_div(wp2_uplh_weighted_sum, wp2_out_total) 
        actual_hc_used = safe_div(ch, 32.5)
        uplh_by_cell_person = _recalc_uplh_by_cell_person(hours_by_cell_person, out_by_cell_person)
        out_rows.append({
            "team": label,
            "period_date": period_date,
            "Total Available Hours": taa,
            "Completed Hours": ch,
            "Target Output": tgt_out,
            "Actual Output": act_out,
            "Target UPLH": target_uplh,
            "Actual UPLH": actual_uplh,
            "UPLH WP1": uplh_wp1,
            "UPLH WP2": uplh_wp2,
            "HC in WIP": hc_wip,
            "Actual HC Used": actual_hc_used,
            "People in WIP": "",
            "Person Hours": json.dumps(person_hours, ensure_ascii=False),
            "Outputs by Person": json.dumps(outputs_by_person, ensure_ascii=False),
            "Outputs by Cell/Station": json.dumps(outputs_by_cell, ensure_ascii=False),
            "Cell/Station Hours": json.dumps(cell_station_hours, ensure_ascii=False),
            "Hours by Cell/Station - by person": json.dumps(hours_by_cell_person, ensure_ascii=False),
            "Output by Cell/Station - by person": json.dumps(out_by_cell_person, ensure_ascii=False),
            "UPLH by Cell/Station - by person": json.dumps(uplh_by_cell_person, ensure_ascii=False),
            "error": " | ".join(errs) if errs else "",
        })
    for _, label in rollups:
        for period_date in sorted(buckets_by_label[label].keys()):
            _emit_rollup(label, period_date, buckets_by_label[label][period_date])
    def sort_key_wip(r: dict) -> tuple:
        team = safe_str(r.get("team")).lower()
        d = safe_str(r.get("period_date"))
        date_key = d if (len(d) == 10 and d[4] == "-" and d[7] == "-") else "9999-12-31"
        return (team, date_key)
    out_rows.sort(key=sort_key_wip)
    return out_rows
def write_csv_wip(rows: list[dict], out_path: str) -> None:
    out_dir = os.path.dirname(os.path.abspath(out_path)) or "."
    tmp_path = os.path.join(out_dir, f".{os.path.basename(out_path)}.tmp")
    try:
        with open(tmp_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=WIP_HEADERS)
            w.writeheader()
            for r in rows:
                w.writerow({h: r.get(h, "") for h in WIP_HEADERS})
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp_path, out_path)
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass
def setup_logging() -> logging.Logger:
    logger = logging.getLogger("ns_metrics")
    logger.setLevel(logging.INFO)
    logger.propagate = False
    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    for handler in list(logger.handlers):
        logger.removeHandler(handler)
        try:
            handler.close()
        except Exception:
            pass
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    ch.setLevel(logging.INFO)
    logger.addHandler(ch)
    return logger
@contextmanager
def heartbeat(logger: logging.Logger, label: str, every_seconds: int = 120):
    stop = Event()
    def _run():
        while not stop.wait(every_seconds):
            logger.info(f"[{label}] still running...")
    t = Thread(target=_run, daemon=True)
    t.start()
    try:
        yield
    finally:
        stop.set()
        t.join(timeout=1)
def load_pss_us_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    pss_us_source_file: str,
    pss_us_cfg: Dict[str, Any],
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    weeks_to_refresh = set(iso_monday_weeks_back(date.today(), weeks_back=2))
    existing_pss_us_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "PSS US"
    ]
    min_keep_date = safe_str(pss_us_cfg.get("min_period_date")) or "2025-06-02"
    frozen_pss_us_rows = [
        r for r in existing_pss_us_rows
        if safe_str(r.get("period_date")) not in weeks_to_refresh
        and safe_str(r.get("period_date")) >= min_keep_date
    ]
    cfg = dict(pss_us_cfg)
    cfg["min_period_date"] = min(weeks_to_refresh)
    cfg["max_period_date"] = max(weeks_to_refresh)
    refreshed_pss_us_rows = scrape_previous_weeks_xlsm_with_filters(
        pss_us_source_file,
        "PSS US",
        cfg,
        dropdown_override=sorted(weeks_to_refresh),
    )
    refreshed_pss_us_rows = [
        r for r in refreshed_pss_us_rows
        if safe_str(r.get("period_date")) in weeks_to_refresh
    ]
    if not refreshed_pss_us_rows:
        if logger:
            logger.warning(
                f"[PSS US] refresh returned 0 rows; keeping existing cached PSS US rows only | "
                f"existing={len(existing_pss_us_rows)} | "
                f"weeks_to_refresh={sorted(weeks_to_refresh)}"
            )
        return existing_pss_us_rows
    merged_pss_us_rows = merge_rows_by_team_period(
        frozen_pss_us_rows + refreshed_pss_us_rows
    )
    if logger:
        logger.info(
            f"[PSS US] loaded cached rows from NS_metrics: {len(existing_pss_us_rows)} | "
            f"kept={len(frozen_pss_us_rows)} | "
            f"refreshed={len(refreshed_pss_us_rows)} | "
            f"final={len(merged_pss_us_rows)} | "
            f"weeks_to_refresh={sorted(weeks_to_refresh)}"
        )
    return merged_pss_us_rows
def load_scs_super_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    scs_super_source_file: str,
    scs_super_old_cfg: Dict[str, Any],
    scs_super_new_cfg: Dict[str, Any],
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    weeks_to_refresh = sorted(set(iso_monday_weeks_back(date.today(), weeks_back=2)))
    weeks_set = set(weeks_to_refresh)
    existing_super_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "SCS Super Cell"
    ]
    frozen_super_rows = [
        r for r in existing_super_rows
        if safe_str(r.get("period_date")) not in weeks_set
        and safe_str(r.get("period_date")) >= "2025-06-30"
    ]
    refreshed_rows: list[dict] = []
    switch_date = safe_str(scs_super_new_cfg.get("min_period_date")) or "9999-12-31"
    if existing_super_rows:
        old_weeks = [w for w in weeks_to_refresh if w < switch_date]
        new_weeks = [w for w in weeks_to_refresh if w >= switch_date]
        if old_weeks:
            old_cfg = dict(scs_super_old_cfg)
            old_cfg["min_period_date"] = min(old_weeks)
            old_cfg["max_period_date"] = max(old_weeks)
            refreshed_rows.extend(scrape_workbook_with_config(scs_super_source_file, old_cfg))
        if new_weeks:
            new_cfg = dict(scs_super_new_cfg)
            new_cfg["min_period_date"] = min(new_weeks)
            new_cfg["max_period_date"] = max(new_weeks)
            refreshed_rows.extend(scrape_workbook_with_config(scs_super_source_file, new_cfg))
        refreshed_rows = [
            r for r in refreshed_rows
            if safe_str(r.get("period_date")) in weeks_set
        ]
        merged_super_rows = merge_rows_by_team_period(frozen_super_rows + refreshed_rows)
        if logger:
            logger.info(
                f"[SCS Super Cell] loaded cached rows from NS_metrics: {len(existing_super_rows)} | "
                f"kept={len(frozen_super_rows)} | refreshed={len(refreshed_rows)} | "
                f"final={len(merged_super_rows)} | "
                f"old_weeks={old_weeks} | new_weeks={new_weeks}"
            )
        return merged_super_rows
    old_cfg = dict(scs_super_old_cfg)
    old_cfg["min_period_date"] = "2025-06-30"
    old_cfg["max_period_date"] = safe_str(scs_super_old_cfg.get("max_period_date")) or "2026-02-23"
    new_cfg = dict(scs_super_new_cfg)
    new_cfg["min_period_date"] = safe_str(scs_super_new_cfg.get("min_period_date")) or "2026-03-01"
    new_cfg["max_period_date"] = max(weeks_to_refresh)
    refreshed_rows.extend(scrape_workbook_with_config(scs_super_source_file, old_cfg))
    refreshed_rows.extend(scrape_workbook_with_config(scs_super_source_file, new_cfg))
    merged_super_rows = merge_rows_by_team_period(refreshed_rows)
    if logger:
        logger.info(
            f"[SCS Super Cell] no cached rows found; performed backfill | "
            f"refreshed={len(refreshed_rows)} | final={len(merged_super_rows)}"
        )
    return merged_super_rows
def load_tdd_cos1_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    cos_source_file: str,
    cos_cfg: Dict[str, Any],
    logger: Optional[logging.Logger] = None,
    cos_new_cfg: Optional[Dict[str, Any]] = None,
    cos_newer_cfg: Optional[Dict[str, Any]] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    weeks_to_refresh = sorted(set(iso_monday_weeks_back(date.today(), weeks_back=2)))
    weeks_set = set(weeks_to_refresh)
    existing_cos_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "TDD COS 1"
    ]
    frozen_cos_rows = [
        r for r in existing_cos_rows
        if safe_str(r.get("period_date")) not in weeks_set
        and safe_str(r.get("period_date")) >= "2025-06-02"
    ]

    refreshed_cos_rows: list[dict] = []
    layout_cfgs = [cos_cfg]
    if cos_new_cfg:
        layout_cfgs.append(cos_new_cfg)
    if cos_newer_cfg:
        layout_cfgs.append(cos_newer_cfg)

    layout_cfgs.sort(key=lambda cfg: safe_str(cfg.get("min_period_date")) or "0000-01-01")

    for idx, layout_cfg in enumerate(layout_cfgs):
        layout_start = safe_str(layout_cfg.get("min_period_date")) or "0000-01-01"
        next_start = (
            safe_str(layout_cfgs[idx + 1].get("min_period_date"))
            if idx + 1 < len(layout_cfgs)
            else ""
        )
        layout_weeks = [
            w for w in weeks_to_refresh
            if w >= layout_start and (not next_start or w < next_start)
        ]
        if not layout_weeks:
            continue

        cfg = dict(layout_cfg)
        cfg["min_period_date"] = min(layout_weeks)
        cfg["max_period_date"] = max(layout_weeks)
        refreshed_cos_rows.extend(scrape_workbook_with_config(cos_source_file, cfg))

    refreshed_cos_rows = [
        r for r in refreshed_cos_rows
        if safe_str(r.get("period_date")) in weeks_set
    ]
    merged_cos_rows = merge_rows_by_team_period(frozen_cos_rows + refreshed_cos_rows)
    if logger:
        logger.info(
            f"[TDD COS 1] loaded cached rows from NS_metrics: {len(existing_cos_rows)} | "
            f"kept={len(frozen_cos_rows)} | refreshed={len(refreshed_cos_rows)} | "
            f"final={len(merged_cos_rows)}"
        )
    return merged_cos_rows
def load_mazor_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    mazor_source_file: str,
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    weeks_to_refresh = set(iso_monday_weeks_back(date.today(), weeks_back=2))
    existing_mazor_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "Mazor"
    ]
    frozen_mazor_rows = [
        r for r in existing_mazor_rows
        if safe_str(r.get("period_date")) not in weeks_to_refresh
        and safe_str(r.get("period_date")) >= "2025-06-02"
    ]
    refreshed_mazor_rows = scrape_meic_ae_oarm_previous_weeks_xlsm(
        mazor_source_file,
        "Mazor",
        dropdown_override=sorted(weeks_to_refresh),
    )
    refreshed_mazor_rows = [
        r for r in refreshed_mazor_rows
        if safe_str(r.get("period_date")) in weeks_to_refresh
    ]
    merged_mazor_rows = merge_rows_by_team_period(frozen_mazor_rows + refreshed_mazor_rows)
    if logger:
        logger.info(
            f"[Mazor] loaded cached rows from NS_metrics: {len(existing_mazor_rows)} | "
            f"kept={len(frozen_mazor_rows)} | refreshed={len(refreshed_mazor_rows)} | "
            f"final={len(merged_mazor_rows)}"
        )
    return merged_mazor_rows
def load_csf_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    csf_source_file: str,
    CSF_CFG: Dict[str, Any],
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    weeks_to_refresh = set(iso_monday_weeks_back(date.today(), weeks_back=2))
    existing_csf_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "CSF"
    ]
    cfg = dict(CSF_CFG)
    cfg["min_period_date"] = min(weeks_to_refresh)
    cfg["max_period_date"] = max(weeks_to_refresh)
    refreshed_csf_rows = scrape_csf_previous_weeks_with_config(
        csf_source_file,
        cfg,
        dropdown_override=sorted(weeks_to_refresh),
    )
    refreshed_csf_rows = [
        r for r in refreshed_csf_rows
        if safe_str(r.get("period_date")) in weeks_to_refresh
    ]
    if not refreshed_csf_rows:
        if logger:
            logger.warning(
                f"[CSF] refresh returned 0 rows; keeping existing cached CSF rows only | "
                f"existing={len(existing_csf_rows)} | weeks_to_refresh={sorted(weeks_to_refresh)}"
            )
        return existing_csf_rows
    refreshed_by_period = {
        safe_str(r.get("period_date")): r
        for r in refreshed_csf_rows
    }
    merged_csf_rows = []
    for old_row in existing_csf_rows:
        pd = safe_str(old_row.get("period_date"))
        if pd in refreshed_by_period:
            merged_csf_rows.append(refreshed_by_period.pop(pd))
        else:
            merged_csf_rows.append(old_row)
    merged_csf_rows.extend(refreshed_by_period.values())
    merged_csf_rows = merge_rows_by_team_period(merged_csf_rows)
    if logger:
        logger.info(
            f"[CSF] loaded cached rows from NS_metrics: {len(existing_csf_rows)} | "
            f"refreshed={len(refreshed_csf_rows)} | "
            f"final={len(merged_csf_rows)} | "
            f"weeks_to_refresh={sorted(weeks_to_refresh)}"
        )
    return merged_csf_rows
def load_oarm_meic_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    oarm_meic_source_file: str,
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    weeks_to_refresh = set(iso_monday_weeks_back(date.today(), weeks_back=2))
    existing_oarm_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "O-Arm MEIC"
    ]
    frozen_oarm_rows = [
        r for r in existing_oarm_rows
        if safe_str(r.get("period_date")) not in weeks_to_refresh
        and safe_str(r.get("period_date")) >= "2025-06-02"
    ]
    refreshed_oarm_rows = scrape_meic_ae_oarm_previous_weeks_xlsm(
        oarm_meic_source_file,
        "O-Arm MEIC",
        dropdown_override=sorted(weeks_to_refresh),
    )
    refreshed_oarm_rows = [
        r for r in refreshed_oarm_rows
        if safe_str(r.get("period_date")) in weeks_to_refresh
    ]
    merged_oarm_rows = merge_rows_by_team_period(frozen_oarm_rows + refreshed_oarm_rows)
    if logger:
        logger.info(
            f"[O-Arm MEIC] loaded cached rows from NS_metrics: {len(existing_oarm_rows)} | "
            f"kept={len(frozen_oarm_rows)} | refreshed={len(refreshed_oarm_rows)} | "
            f"final={len(merged_oarm_rows)}"
        )
    return merged_oarm_rows
def load_ae_meic_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    ae_meic_source_file: str,
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    weeks_to_refresh = set(iso_monday_weeks_back(date.today(), weeks_back=2))
    existing_ae_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "AE MEIC"
    ]
    frozen_ae_rows = [
        r for r in existing_ae_rows
        if safe_str(r.get("period_date")) not in weeks_to_refresh
        and safe_str(r.get("period_date")) >= "2025-06-02"
    ]
    refreshed_ae_rows = scrape_meic_ae_oarm_previous_weeks_xlsm(
        ae_meic_source_file,
        "AE MEIC",
        dropdown_override=sorted(weeks_to_refresh),
    )
    refreshed_ae_rows = [
        r for r in refreshed_ae_rows
        if safe_str(r.get("period_date")) in weeks_to_refresh
    ]
    merged_ae_rows = merge_rows_by_team_period(frozen_ae_rows + refreshed_ae_rows)
    if logger:
        logger.info(
            f"[AE MEIC] loaded cached rows from NS_metrics: {len(existing_ae_rows)} | "
            f"kept={len(frozen_ae_rows)} | refreshed={len(refreshed_ae_rows)} | "
            f"final={len(merged_ae_rows)}"
        )
    return merged_ae_rows
def load_nav_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    nav_source_file: str,
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    weeks_to_refresh = set(iso_monday_weeks_back(date.today(), weeks_back=2))
    existing_nav_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "Nav"
    ]
    frozen_nav_rows = [
        r for r in existing_nav_rows
        if safe_str(r.get("period_date")) not in weeks_to_refresh
        and safe_str(r.get("period_date")) >= "2025-06-02"
    ]
    refreshed_nav_rows = scrape_nav_previous_weeks_xlsm(
        nav_source_file,
        "Nav",
        dropdown_override=sorted(weeks_to_refresh),
    )
    refreshed_nav_rows = [
        r for r in refreshed_nav_rows
        if safe_str(r.get("period_date")) in weeks_to_refresh
    ]
    merged_nav_rows = merge_rows_by_team_period(frozen_nav_rows + refreshed_nav_rows)
    if logger:
        logger.info(
            f"[Nav] loaded cached rows from NS_metrics: {len(existing_nav_rows)} | "
            f"kept={len(frozen_nav_rows)} | refreshed={len(refreshed_nav_rows)} | "
            f"final={len(merged_nav_rows)}"
        )
    return merged_nav_rows
def read_existing_metrics_rows(path: str) -> list[dict]:
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", newline="", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))
    except Exception:
        return []
def load_nv_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    nv_source_file: str,
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    weeks_to_refresh = set(iso_monday_weeks_back(date.today(), weeks_back=2))
    existing_nv_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "NV"
    ]
    frozen_nv_rows = [
        r for r in existing_nv_rows
        if safe_str(r.get("period_date")) not in weeks_to_refresh
        and safe_str(r.get("period_date")) >= "2025-07-07"
    ]
    refreshed_nv_rows = scrape_dbs_previous_weeks_xlsm(
        nv_source_file,
        "NV",
        dropdown_override=sorted(weeks_to_refresh),
    )
    refreshed_nv_rows = [
        r for r in refreshed_nv_rows
        if safe_str(r.get("period_date")) in weeks_to_refresh
    ]
    merged_nv_rows = merge_rows_by_team_period(frozen_nv_rows + refreshed_nv_rows)
    if logger:
        logger.info(
            f"[NV] loaded cached rows from NS_metrics: {len(existing_nv_rows)} | "
            f"kept={len(frozen_nv_rows)} | refreshed={len(refreshed_nv_rows)} | "
            f"final={len(merged_nv_rows)}"
        )
    return merged_nv_rows
def load_meic_ph_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    meic_source_file: str,
    meic_cfg: Dict[str, Any],
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    weeks_to_refresh = set(iso_monday_weeks_back(date.today(), weeks_back=2))
    existing_meic_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "MEIC PH"
    ]
    frozen_meic_rows = [
        r for r in existing_meic_rows
        if safe_str(r.get("period_date")) not in weeks_to_refresh
        and safe_str(r.get("period_date")) >= "2025-09-01"
    ]
    cfg = dict(meic_cfg)
    cfg["min_period_date"] = min(weeks_to_refresh)
    cfg["max_period_date"] = max(weeks_to_refresh)
    refreshed_meic_rows = scrape_workbook_with_config(meic_source_file, cfg)
    refreshed_meic_rows = [
        r for r in refreshed_meic_rows
        if safe_str(r.get("period_date")) in weeks_to_refresh
    ]
    merged_meic_rows = merge_rows_by_team_period(frozen_meic_rows + refreshed_meic_rows)
    if logger:
        logger.info(
            f"[MEIC PH] loaded cached rows from NS_metrics: {len(existing_meic_rows)} | "
            f"kept={len(frozen_meic_rows)} | refreshed={len(refreshed_meic_rows)} | "
            f"final={len(merged_meic_rows)}"
        )
    return merged_meic_rows
def load_scs_cell1_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    scs_source_file: str,
    scs_old_cfg: Dict[str, Any],
    scs_new_cfg: Dict[str, Any],
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    scs_weeks_to_refresh = sorted(set(iso_monday_weeks_back(date.today(), weeks_back=2)))
    existing_scs_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "SCS Cell 1"
    ]
    frozen_scs_rows = [
        r for r in existing_scs_rows
        if safe_str(r.get("period_date")) not in scs_weeks_to_refresh
        and safe_str(r.get("period_date")) >= "2025-06-30"
    ]
    switch_date = safe_str(scs_new_cfg.get("min_period_date"))
    old_weeks = [w for w in scs_weeks_to_refresh if w < switch_date]
    new_weeks = [w for w in scs_weeks_to_refresh if w >= switch_date]
    refreshed_rows: list[dict] = []
    if old_weeks:
        old_cfg = dict(scs_old_cfg)
        old_cfg["min_period_date"] = min(old_weeks)
        old_cfg["max_period_date"] = max(old_weeks)
        refreshed_rows.extend(scrape_workbook_with_config(scs_source_file, old_cfg))
    if new_weeks:
        new_cfg = dict(scs_new_cfg)
        new_cfg["min_period_date"] = min(new_weeks)
        new_cfg["max_period_date"] = max(new_weeks)
        refreshed_rows.extend(scrape_workbook_with_config(scs_source_file, new_cfg))
    refreshed_rows = [
        r for r in refreshed_rows
        if safe_str(r.get("period_date")) in scs_weeks_to_refresh
    ]
    merged_scs_rows = merge_rows_by_team_period(frozen_scs_rows + refreshed_rows)
    if logger:
        logger.info(
            f"[SCS Cell 1] loaded cached rows from NS_metrics: {len(existing_scs_rows)} | "
            f"kept={len(frozen_scs_rows)} | refreshed={len(refreshed_rows)} | "
            f"final={len(merged_scs_rows)} | "
            f"old_weeks={old_weeks} | new_weeks={new_weeks}"
        )
    return merged_scs_rows
def recalc_workbook_to_temp_copy(source_file: str) -> str:
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    excel.EnableEvents = False
    excel.AutomationSecurity = 3
    src = os.path.abspath(os.path.expandvars(source_file))
    base = os.path.splitext(os.path.basename(src))[0]
    ext = os.path.splitext(src)[1]
    tmp_path = os.path.join(tempfile.gettempdir(), f"{base}__calc__{uuid.uuid4().hex}{ext}")
    shutil.copy2(src, tmp_path)
    wb = None
    try:
        wb = excel.Workbooks.Open(
            tmp_path,
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            Notify=False,
            AddToMru=False,
        )
        try:
            wb.RefreshAll()
        except Exception:
            pass
        try:
            excel.CalculateFullRebuild()
        except Exception:
            excel.Calculate()
        wb.Save()
        return tmp_path
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=True)
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
def load_pss_meic_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    pss_meic_source_file: str,
    pss_meic_cfg: Dict[str, Any],
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    weeks_to_refresh = set(iso_monday_weeks_back(date.today(), weeks_back=2))
    existing_pss_meic_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "PSS MEIC"
    ]
    min_keep_date = safe_str(pss_meic_cfg.get("min_period_date")) or "2025-06-02"
    frozen_pss_meic_rows = [
        r for r in existing_pss_meic_rows
        if safe_str(r.get("period_date")) not in weeks_to_refresh
        and safe_str(r.get("period_date")) >= min_keep_date
    ]
    cfg = dict(pss_meic_cfg)
    cfg["min_period_date"] = min(weeks_to_refresh)
    cfg["max_period_date"] = max(weeks_to_refresh)
    refreshed_pss_meic_rows = scrape_previous_weeks_xlsm_with_filters(
        pss_meic_source_file,
        "PSS MEIC",
        cfg,
        dropdown_override=sorted(weeks_to_refresh),
    )
    refreshed_pss_meic_rows = [
        r for r in refreshed_pss_meic_rows
        if safe_str(r.get("period_date")) in weeks_to_refresh
    ]
    if not refreshed_pss_meic_rows:
        if logger:
            logger.warning(
                f"[PSS MEIC] refresh returned 0 rows; keeping existing cached PSS MEIC rows only | "
                f"existing={len(existing_pss_meic_rows)} | "
                f"weeks_to_refresh={sorted(weeks_to_refresh)}"
            )
        return existing_pss_meic_rows
    merged_pss_meic_rows = merge_rows_by_team_period(
        frozen_pss_meic_rows + refreshed_pss_meic_rows
    )
    if logger:
        logger.info(
            f"[PSS MEIC] loaded cached rows from NS_metrics: {len(existing_pss_meic_rows)} | "
            f"kept={len(frozen_pss_meic_rows)} | "
            f"refreshed={len(refreshed_pss_meic_rows)} | "
            f"final={len(merged_pss_meic_rows)} | "
            f"weeks_to_refresh={sorted(weeks_to_refresh)}"
        )
    return merged_pss_meic_rows
def load_pss_meic_intern_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    pss_intern_source_file: str,
    pss_meic_intern_cfg: Dict[str, Any],
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    weeks_to_refresh = set(iso_monday_weeks_back(date.today(), weeks_back=2))
    existing_pss_intern_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "PSS MEIC Intern"
    ]
    min_keep_date = safe_str(pss_meic_intern_cfg.get("min_period_date")) or "2025-06-02"
    frozen_pss_intern_rows = [
        r for r in existing_pss_intern_rows
        if safe_str(r.get("period_date")) not in weeks_to_refresh
        and safe_str(r.get("period_date")) >= min_keep_date
    ]
    cfg = dict(pss_meic_intern_cfg)
    cfg["min_period_date"] = min(weeks_to_refresh)
    cfg["max_period_date"] = max(weeks_to_refresh)
    refreshed_pss_intern_rows = scrape_previous_weeks_xlsm_with_filters(
        pss_intern_source_file,
        "PSS MEIC Intern",
        cfg,
        dropdown_override=sorted(weeks_to_refresh),
    )
    refreshed_pss_intern_rows = [
        r for r in refreshed_pss_intern_rows
        if safe_str(r.get("period_date")) in weeks_to_refresh
    ]
    if not refreshed_pss_intern_rows:
        if logger:
            logger.warning(
                f"[PSS MEIC Intern] refresh returned 0 rows; keeping existing cached PSS MEIC Intern rows only | "
                f"existing={len(existing_pss_intern_rows)} | "
                f"weeks_to_refresh={sorted(weeks_to_refresh)}"
            )
        return existing_pss_intern_rows
    merged_pss_intern_rows = merge_rows_by_team_period(
        frozen_pss_intern_rows + refreshed_pss_intern_rows
    )
    if logger:
        logger.info(
            f"[PSS MEIC Intern] loaded cached rows from NS_metrics: {len(existing_pss_intern_rows)} | "
            f"kept={len(frozen_pss_intern_rows)} | "
            f"refreshed={len(refreshed_pss_intern_rows)} | "
            f"final={len(merged_pss_intern_rows)} | "
            f"weeks_to_refresh={sorted(weeks_to_refresh)}"
        )
    return merged_pss_intern_rows
def load_pss_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    pss_source_file: str,
    pss_cfg: Dict[str, Any],
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    cutoff = safe_str(pss_cfg.get("min_period_date")) or "2026-05-11"
    legacy_teams = {"PSS US", "PSS MEIC", "PSS MEIC Intern"}
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    weeks_to_refresh = sorted(set(iso_monday_weeks_back(date.today(), weeks_back=2)))
    weeks_set = set(weeks_to_refresh)
    new_weeks = [w for w in weeks_to_refresh if w >= cutoff]
    frozen_legacy_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) in legacy_teams
        and safe_str(r.get("period_date")) < cutoff
        and safe_str(r.get("period_date")) >= "2025-06-02"
    ]
    existing_pss_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "PSS"
        and safe_str(r.get("period_date")) >= cutoff
    ]
    frozen_pss_rows = [
        r for r in existing_pss_rows
        if safe_str(r.get("period_date")) not in weeks_set
    ]
    refreshed_pss_rows: list[dict] = []
    if new_weeks:
        cfg = dict(pss_cfg)
        cfg["min_period_date"] = min(new_weeks)
        cfg["max_period_date"] = max(new_weeks)
        refreshed_pss_rows = scrape_previous_weeks_xlsm_with_filters(
            pss_source_file,
            "PSS",
            cfg,
            dropdown_override=new_weeks,
        )
        refreshed_pss_rows = [
            r for r in refreshed_pss_rows
            if safe_str(r.get("period_date")) in set(new_weeks)
        ]
    if new_weeks and not refreshed_pss_rows:
        if logger:
            logger.warning(
                f"[PSS] refresh returned 0 rows from combined PSS workbook; "
                f"keeping cached PSS rows only | existing={len(existing_pss_rows)} | "
                f"weeks_to_refresh={new_weeks} | source={pss_source_file}"
            )
        merged_pss_rows = merge_rows_by_team_period(existing_pss_rows)
    else:
        merged_pss_rows = merge_rows_by_team_period(frozen_pss_rows + refreshed_pss_rows)
    out = frozen_legacy_rows + merged_pss_rows
    out = merge_rows_by_team_period(out)
    if logger:
        logger.info(
            f"[PSS] combined workbook active from {cutoff} | "
            f"legacy_kept_before_cutoff={len(frozen_legacy_rows)} | "
            f"cached_combined={len(existing_pss_rows)} | "
            f"refreshed_combined={len(refreshed_pss_rows)} | "
            f"final={len(out)} | weeks_to_refresh={weeks_to_refresh} | new_weeks={new_weeks}"
        )
    return out
def load_spine_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    spine_source_file: str,
    spine_new_source_file: str,
    spine_old_cfg: Dict[str, Any],
    spine_new_cfg: Dict[str, Any],
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    weeks_to_refresh = sorted(set(iso_monday_weeks_back(date.today(), weeks_back=2)))
    weeks_set = set(weeks_to_refresh)
    existing_spine_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "Spine"
    ]
    spine_min_date = safe_str(spine_old_cfg.get("min_period_date")) or "2025-06-30"
    frozen_spine_rows = [
        r for r in existing_spine_rows
        if safe_str(r.get("period_date")) not in weeks_set
        and safe_str(r.get("period_date")) >= spine_min_date
    ]
    switch_date = safe_str(spine_new_cfg.get("min_period_date")) or "9999-12-31"
    old_weeks = [w for w in weeks_to_refresh if w < switch_date]
    new_weeks = [w for w in weeks_to_refresh if w >= switch_date]
    refreshed_rows: list[dict] = []
    if existing_spine_rows:
        if old_weeks:
            old_cfg = dict(spine_old_cfg)
            old_cfg["min_period_date"] = min(old_weeks)
            old_cfg["max_period_date"] = max(old_weeks)
            refreshed_rows.extend(scrape_workbook_with_config(spine_source_file, old_cfg))
        if new_weeks:
            new_cfg = dict(spine_new_cfg)
            new_cfg["min_period_date"] = min(new_weeks)
            new_cfg["max_period_date"] = max(new_weeks)
            refreshed_rows.extend(
                scrape_spine_previous_weeks_xlsm(
                    spine_new_source_file,
                    new_cfg,
                    team="Spine",
                    dropdown_override=sorted(new_weeks),
                )
            )
        refreshed_rows = [
            r for r in refreshed_rows
            if safe_str(r.get("period_date")) in weeks_set
        ]
        if not refreshed_rows:
            if logger:
                logger.warning(
                    f"[Spine] refresh returned 0 rows; keeping existing cached Spine rows only | "
                    f"existing={len(existing_spine_rows)} | weeks_to_refresh={sorted(weeks_set)}"
                )
            return existing_spine_rows
        merged_spine_rows = merge_rows_by_team_period(frozen_spine_rows + refreshed_rows)

        if logger:
            logger.info(
                f"[Spine] loaded cached rows from NS_metrics: {len(existing_spine_rows)} | "
                f"kept={len(frozen_spine_rows)} | refreshed={len(refreshed_rows)} | "
                f"final={len(merged_spine_rows)} | "
                f"old_weeks={old_weeks} | new_weeks={new_weeks}"
            )
        return merged_spine_rows
    old_cfg = dict(spine_old_cfg)
    old_cfg["min_period_date"] = safe_str(spine_old_cfg.get("min_period_date")) or "2025-06-30"
    old_cfg["max_period_date"] = safe_str(spine_old_cfg.get("max_period_date")) or "2026-02-23"
    refreshed_rows.extend(scrape_workbook_with_config(spine_source_file, old_cfg))
    new_cfg = dict(spine_new_cfg)
    new_cfg["min_period_date"] = safe_str(spine_new_cfg.get("min_period_date")) or "2026-03-02"
    new_cfg["max_period_date"] = max(weeks_to_refresh)
    refreshed_rows.extend(
        scrape_spine_previous_weeks_xlsm(
            spine_new_source_file,
            new_cfg,
            team="Spine",
        )
    )
    merged_spine_rows = merge_rows_by_team_period(refreshed_rows)
    if logger:
        logger.info(
            f"[Spine] no cached rows found; performed backfill | "
            f"refreshed={len(refreshed_rows)} | final={len(merged_spine_rows)}"
        )
    return merged_spine_rows
def load_ph_from_existing_metrics_and_refresh_3_weeks(
    ns_metrics_path: str,
    ph_source_file: str,
    ph_new_cfg: Dict[str, Any],
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    existing_rows = read_existing_metrics_rows(ns_metrics_path)
    ph_weeks_to_refresh = set(iso_monday_weeks_back(date.today(), weeks_back=2))
    existing_ph_rows = [
        r for r in existing_rows
        if safe_str(r.get("team")) == "PH"
    ]
    frozen_ph_rows = [
        r for r in existing_ph_rows
        if safe_str(r.get("period_date")) not in ph_weeks_to_refresh
        and safe_str(r.get("period_date")) >= "2025-09-01"
    ]
    cfg = dict(ph_new_cfg)
    cfg["min_period_date"] = min(ph_weeks_to_refresh)
    cfg["max_period_date"] = max(ph_weeks_to_refresh)
    tmp_calc_path = None
    try:
        tmp_calc_path = recalc_workbook_to_temp_copy(ph_source_file)
        refreshed_ph_rows = scrape_workbook_with_config(tmp_calc_path, cfg)
    finally:
        if tmp_calc_path and os.path.exists(tmp_calc_path):
            try:
                os.remove(tmp_calc_path)
            except OSError:
                pass
    refreshed_ph_rows = [
        r for r in refreshed_ph_rows
        if safe_str(r.get("period_date")) in ph_weeks_to_refresh
    ]
    merged_ph_rows = merge_rows_by_team_period(frozen_ph_rows + refreshed_ph_rows)
    if logger:
        logger.info(
            f"[PH] loaded cached PH rows from NS_metrics: {len(existing_ph_rows)} | "
            f"kept={len(frozen_ph_rows)} | refreshed={len(refreshed_ph_rows)} | "
            f"final={len(merged_ph_rows)}"
        )
    return merged_ph_rows
def run_team(logger: logging.Logger, team_name: str, fn):
    start = datetime.now()
    logger.info(f"[{team_name}] START")
    try:
        with heartbeat(logger, team_name, every_seconds=180): 
            rows = fn()
        elapsed = datetime.now() - start
        logger.info(f"[{team_name}] DONE | rows={len(rows)} | elapsed={elapsed}")
        return rows
    except Exception as e:
        elapsed = datetime.now() - start
        logger.error(f"[{team_name}] FAIL | elapsed={elapsed} | error={e}")
        logger.error(traceback.format_exc())
        return []
HEADERS = [
    "team",
    "period_date",
    "source_file",
    "Total Available Hours",
    "Completed Hours",
    "Target Output",
    "Actual Output",
    "Target UPLH",
    "Actual UPLH",
    "UPLH WP1",
    "UPLH WP2",
    "HC in WIP",
    "Actual HC Used",
    "People in WIP",
    "Person Hours",
    "Outputs by Person",
    "Outputs by Cell/Station",
    "Cell/Station Hours",
    "Hours by Cell/Station - by person",
    "Output by Cell/Station - by person",
    "UPLH by Cell/Station - by person",
    "error",
]
def safe_float(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return 0.0
        try:
            return float(s)
        except ValueError:
            return 0.0
    return 0.0
def safe_str(v: Any) -> str:
    return "" if v is None else str(v).strip()
ENT_NAME_REPLACEMENTS = {
    "AG, Girish": "Girish AG",
    "Sharma, Anurag": "Anurag Sharma",
    "Badugu, Aravind Kumar": "Aravind Kumar Badugu",
    "Boya, Kranthi Kumar": "Kranthi Kumar Boya",
    "Kumari, Taruna": "Taruna Kumari",
    "Pavani Uppari":"Uppari Pavani",
    "Raju, Surekha": "Surekha Raju Anantarapu",
    "S, Selvarasu": "Selvarasu Sampathu",
    "Uppari, Pavani": "Uppari Pavani",
    "Megan R":"Megan",
}
def normalize_ent_name(name: Any) -> str:
    s = safe_str(name)
    return ENT_NAME_REPLACEMENTS.get(s, s)
def apply_ent_name_replacements_to_sheet(
    xlsx_path: str,
    sheet_name: str = "Next Week Forecast",
    start_row: int = 2,
    end_row: int = 30,
    name_col_letter: str = "A",
) -> str:
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        ws = wb.active
    else:
        ws = wb[sheet_name]
    name_col = column_index_from_string(name_col_letter)
    changes = 0
    for r in range(start_row, end_row + 1):
        cell = ws.cell(row=r, column=name_col)
        old_val = safe_str(cell.value)
        new_val = normalize_ent_name(old_val)
        if old_val and new_val != old_val:
            cell.value = new_val
            changes += 1
    wb.save(xlsx_path)
    return f"ENT name replacements applied in {os.path.basename(xlsx_path)} ({changes} changes)"
def safe_div(n: float, d: float) -> Optional[float]:
    return None if d == 0 else (n / d)
def parse_sheet_date(sheet_name: str) -> str:
    name = sheet_name.strip()
    fmts = [
        "%b %d %Y",
        "%b %d, %Y",
        "%B %d %Y",
        "%B %d, %Y",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
    ]
    for fmt in fmts:
        try:
            dt = datetime.strptime(name, fmt).date()
            return dt.isoformat()
        except ValueError:
            pass
    try:
        from dateutil import parser  # type: ignore
        dt = parser.parse(name, fuzzy=True).date()
        return dt.isoformat()
    except Exception:
        return name
import re
from datetime import datetime, date
def parse_sheet_date_scs_missing_year(sheet_name: str) -> str:
    raw = (sheet_name or "").strip()
    low = raw.lower()
    if any(k in low for k in ["template", "agenda", "work instruction", "instructions"]):
        return ""
    m = re.search(r"\b([A-Za-z]{3,9})\s+(\d{1,2})\b", raw)
    if not m:
        return ""
    mon_txt = m.group(1)
    day_txt = m.group(2)
    mm = dd = None
    for fmt in ("%b %d", "%B %d"):
        try:
            dt = datetime.strptime(f"{mon_txt} {day_txt}", fmt)
            mm, dd = dt.month, dt.day
            break
        except ValueError:
            pass
    if mm is None or dd is None:
        return ""
    for y in range(date.today().year, 1999, -1):
        try:
            d = date(y, mm, dd)
        except ValueError:
            continue
        if d.weekday() == 0:
            return d.isoformat()
    return ""
def col_range(start_col_letter: str, end_col_letter: str) -> range:
    start = column_index_from_string(start_col_letter.upper())
    end = column_index_from_string(end_col_letter.upper())
    return range(start, end + 1)
def sum_rows(ws, rows: list[int], col: int) -> float:
    return sum(safe_float(ws.cell(row=r, column=col).value) for r in rows)
def read_lookup_csv(path: str) -> Tuple[Dict[Tuple[str, str], Dict[str, Any]], str]:
    lookup: Dict[Tuple[str, str], Dict[str, Any]] = {}
    try:
        with open(path, "r", newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                team = safe_str(row.get("team"))
                period_date = safe_str(row.get("period_date"))
                if team and period_date:
                    lookup[(team, period_date)] = row
        return lookup, ""
    except Exception as e:
        return lookup, f"Failed reading {os.path.basename(path)}: {e}"
def scrape_csf_previous_weeks_with_config(
    source_file: str,
    cfg: Dict[str, Any],
    dropdown_override: Optional[list[Any]] = None,
) -> list[dict]:
    import shutil
    import tempfile
    import uuid
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    excel.EnableEvents = False
    excel.AutomationSecurity = 3
    wb = None
    tmp_path = None
    rows_out: list[dict] = []
    def _open_via_temp_copy(src_path: str):
        nonlocal tmp_path
        src = os.path.abspath(os.path.expandvars(src_path))
        if not os.path.exists(src):
            raise FileNotFoundError(f"File not found on disk: {src}")
        base = os.path.splitext(os.path.basename(src))[0]
        ext = os.path.splitext(src)[1]
        tmp_path = os.path.join(tempfile.gettempdir(), f"{base}__{uuid.uuid4().hex}{ext}")
        shutil.copy2(src, tmp_path)
        return _com_call(lambda: excel.Workbooks.Open(
            tmp_path,
            UpdateLinks=0,
            ReadOnly=True,
            IgnoreReadOnlyRecommended=True,
            Notify=False,
            AddToMru=False,
            CorruptLoad=0,
        ))
    def _range_val(ws, ref: str):
        return _com_call(lambda: ws.Range(ref).Value)
    def _cell_val(ws, row: int, col: int):
        return _com_call(lambda: ws.Cells(row, col).Value)
    try:
        wb = _open_via_temp_copy(source_file)
        ws = _com_call(lambda: wb.Worksheets("Previous Weeks"))
        dd = _com_call(lambda: ws.Range("A2"))
        dropdown_values = dropdown_override if dropdown_override is not None else _get_dropdown_values_from_validation(dd)
        seen = set()
        dropdown_values = [v for v in dropdown_values if not (safe_str(v) in seen or seen.add(safe_str(v)))]
        cols = col_range(cfg["person_cols"][0], cfg["person_cols"][1])
        for choice in dropdown_values:
            _com_call(lambda: setattr(dd, "Value", choice))
            try:
                _com_call(lambda: wb.RefreshAll())
            except Exception:
                pass
            try:
                _com_call(lambda: excel.CalculateFullRebuild())
            except Exception:
                _com_call(lambda: excel.Calculate())
            time.sleep(1)
            period_date = _as_iso_date(_com_call(lambda: dd.Value))
            if not period_date:
                continue
            min_pd = safe_str(cfg.get("min_period_date"))
            if min_pd and period_date < min_pd:
                continue
            max_pd = safe_str(cfg.get("max_period_date"))
            if max_pd and period_date > max_pd:
                continue
            taa_spec = cfg["cells"]["total_available_hours"]
            if isinstance(taa_spec, str):
                total_available_hours = safe_float(_range_val(ws, taa_spec))
            else:
                if taa_spec.get("type") == "sum_range":
                    start_col, start_row, end_col, end_row = re.match(
                        r"([A-Z]+)(\d+):([A-Z]+)(\d+)", taa_spec["range"]
                    ).groups()
                    total_available_hours = 0.0
                    for c in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1):
                        for r in range(int(start_row), int(end_row) + 1):
                            total_available_hours += safe_float(_cell_val(ws, r, c))
                elif taa_spec.get("type") == "sum_cells":
                    total_available_hours = sum(safe_float(_range_val(ws, cell_ref)) for cell_ref in taa_spec["cells"])
                else:
                    total_available_hours = 0.0
            completed_spec = cfg["cells"]["completed_hours"]
            if isinstance(completed_spec, str):
                completed_hours = safe_float(_range_val(ws, completed_spec))
            else:
                if completed_spec.get("type") == "sum_range":
                    start_col, start_row, end_col, end_row = re.match(
                        r"([A-Z]+)(\d+):([A-Z]+)(\d+)", completed_spec["range"]
                    ).groups()
                    completed_hours = 0.0
                    for c in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1):
                        for r in range(int(start_row), int(end_row) + 1):
                            completed_hours += safe_float(_cell_val(ws, r, c))
                elif completed_spec.get("type") == "sum_cells":
                    completed_hours = sum(safe_float(_range_val(ws, cell_ref)) for cell_ref in completed_spec["cells"])
                else:
                    completed_hours = 0.0
            wp1_out = safe_float(_range_val(ws, cfg["cells"]["wp1_output"]))
            wp2_out = safe_float(_range_val(ws, cfg["cells"]["wp2_output"]))
            wp1_tgt = safe_float(_range_val(ws, cfg["cells"]["wp1_target"]))
            wp2_tgt = safe_float(_range_val(ws, cfg["cells"]["wp2_target"]))
            target_output = wp1_tgt + wp2_tgt
            actual_output = wp1_out + wp2_out
            target_uplh = safe_div(target_output, completed_hours)
            actual_uplh = safe_div(actual_output, completed_hours)
            uplh_wp1 = safe_float(_range_val(ws, cfg["cells"]["uplh_wp1"]))
            uplh_wp2 = safe_float(_range_val(ws, cfg["cells"]["uplh_wp2"]))
            rows_cfg = cfg.get("rows", {})
            team_name = cfg.get("team", "CSF")
            hc_row = rows_cfg.get("hc_row", rows_cfg.get("person_actual_row_for_person_hours"))
            hc_in_wip = 0
            if hc_row:
                for c in cols:
                    if safe_float(_cell_val(ws, hc_row, c)) != 0.0:
                        hc_in_wip += 1
            actual_hc_used = safe_div(completed_hours, 32.5)
            person_hours: Dict[str, Dict[str, float]] = {}
            name_row_ph = cfg["rows"]["person_name_row_for_person_hours"]
            actual_row_ph = cfg["rows"]["person_actual_row_for_person_hours"]
            avail_row_ph = cfg["rows"]["person_available_row_for_person_hours"]
            for c in cols:
                name = safe_str(_cell_val(ws, name_row_ph, c))
                if not name:
                    continue
                actual = safe_float(_cell_val(ws, actual_row_ph, c))
                available = safe_float(_cell_val(ws, avail_row_ph, c))
                person_hours[name] = {"actual": actual, "available": available}
            outputs_by_person: Dict[str, Dict[str, float]] = {}
            name_row_op = rows_cfg.get("person_name_row_for_outputs_by_person")
            target_row_op = rows_cfg.get("person_target_row_for_outputs_by_person")
            output_spec = cfg.get("outputs_by_person_output")
            if name_row_op and target_row_op and output_spec:
                for c in cols:
                    name = safe_str(_cell_val(ws, name_row_op, c))
                    if not name:
                        continue
                    if output_spec.get("type") == "row":
                        output_val = safe_float(_cell_val(ws, output_spec["row"], c))
                    elif output_spec.get("type") == "sum_rows":
                        output_val = sum(
                            safe_float(_cell_val(ws, r, c))
                            for r in output_spec.get("rows", [])
                        )
                    else:
                        output_val = 0.0
                    target_val = safe_float(_cell_val(ws, target_row_op, c))
                    if output_val != 0.0 or target_val != 0.0:
                        outputs_by_person[name] = {"output": output_val, "target": target_val}
            outputs_by_cell = {
                "WP1": {"output": wp1_out, "target": wp1_tgt},
                "WP2": {"output": wp2_out, "target": wp2_tgt},
            }
            cell_station_hours = {
                "WP1": safe_float(_range_val(ws, cfg["cells"]["wp1_hours"])),
                "WP2": safe_float(_range_val(ws, cfg["cells"]["wp2_hours"])),
            }
            wp3_total_cells = cfg.get("cells", {}).get("wp3_hours_sum_cells")
            if wp3_total_cells:
                cell_station_hours["WP3"] = sum(safe_float(_range_val(ws, cell)) for cell in wp3_total_cells)
            hours_by_cell_by_person = {"WP1": {}, "WP2": {}}
            wp3_hour_rows = cfg.get("rows", {}).get("wp3_hour_rows")
            if wp3_hour_rows:
                hours_by_cell_by_person["WP3"] = {}
            name_row_hc = cfg["rows"]["person_name_row_for_hours_by_cell_by_person"]
            wp1_hour_rows = cfg["rows"]["wp1_hour_rows"]
            wp2_hour_rows = cfg["rows"]["wp2_hour_rows"]
            for c in cols:
                name = safe_str(_cell_val(ws, name_row_hc, c))
                if not name:
                    continue
                wp1_hrs = sum(safe_float(_cell_val(ws, r, c)) for r in wp1_hour_rows)
                wp2_hrs = sum(safe_float(_cell_val(ws, r, c)) for r in wp2_hour_rows)
                if wp1_hrs != 0.0:
                    hours_by_cell_by_person["WP1"][name] = wp1_hrs
                if wp2_hrs != 0.0:
                    hours_by_cell_by_person["WP2"][name] = wp2_hrs
                if wp3_hour_rows:
                    wp3_hrs = sum(safe_float(_cell_val(ws, r, c)) for r in wp3_hour_rows)
                    if wp3_hrs != 0.0:
                        hours_by_cell_by_person["WP3"][name] = wp3_hrs
            output_by_cell_by_person = {"WP1": {}, "WP2": {}}
            name_row_oc = cfg["rows"]["person_name_row_for_output_by_cell_by_person"]
            wp1_out_rows = cfg["rows"]["wp1_output_rows_by_person"]
            wp2_out_rows = cfg["rows"]["wp2_output_rows_by_person"]
            for c in cols:
                name = safe_str(_cell_val(ws, name_row_oc, c))
                if not name:
                    continue
                wp1_o = sum(safe_float(_cell_val(ws, r, c)) for r in wp1_out_rows)
                wp2_o = sum(safe_float(_cell_val(ws, r, c)) for r in wp2_out_rows)
                if wp1_o != 0.0:
                    output_by_cell_by_person["WP1"][name] = wp1_o
                if wp2_o != 0.0:
                    output_by_cell_by_person["WP2"][name] = wp2_o
            uplh_by_cell_by_person: Dict[str, Dict[str, Optional[float]]] = {"WP1": {}, "WP2": {}}
            for wp in ("WP1", "WP2"):
                for person, out_val in output_by_cell_by_person[wp].items():
                    hrs = safe_float(hours_by_cell_by_person[wp].get(person, 0.0))
                    uplh_by_cell_by_person[wp][person] = safe_div(out_val, hrs)
            rows_out.append({
                "team": team_name,
                "period_date": period_date,
                "source_file": os.path.abspath(os.path.expandvars(source_file)),
                "Total Available Hours": total_available_hours,
                "Completed Hours": completed_hours,
                "Target Output": target_output,
                "Actual Output": actual_output,
                "Target UPLH": target_uplh,
                "Actual UPLH": actual_uplh,
                "UPLH WP1": uplh_wp1,
                "UPLH WP2": uplh_wp2,
                "HC in WIP": hc_in_wip,
                "Actual HC Used": actual_hc_used,
                "People in WIP": "",
                "Person Hours": json.dumps(person_hours, ensure_ascii=False),
                "Outputs by Person": json.dumps(outputs_by_person, ensure_ascii=False),
                "Outputs by Cell/Station": json.dumps(outputs_by_cell, ensure_ascii=False),
                "Cell/Station Hours": json.dumps(cell_station_hours, ensure_ascii=False),
                "Hours by Cell/Station - by person": json.dumps(hours_by_cell_by_person, ensure_ascii=False),
                "Output by Cell/Station - by person": json.dumps(output_by_cell_by_person, ensure_ascii=False),
                "UPLH by Cell/Station - by person": json.dumps(uplh_by_cell_by_person, ensure_ascii=False),
                "error": "",
            })
        return rows_out
    finally:
        try:
            if wb is not None:
                _com_call(lambda: wb.Close(SaveChanges=False), tries=10, sleep_s=0.3)
        except Exception:
            pass
        try:
            _com_call(lambda: excel.Quit(), tries=10, sleep_s=0.3)
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        try:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
def scrape_dbs_dated_tabs_xlsx(
    source_file: str,
    team: str,
    min_period_date: str = "2025-06-02",
    max_period_date: Optional[str] = None,
) -> list[dict]:
    wb = load_workbook(source_file, data_only=True, read_only=True, keep_links=False)
    rows_out: list[dict] = []
    cols = _excel_col_range("B", "R")
    today_iso = date.today().isoformat()
    upper_bound = max_period_date or today_iso
    for ws in wb.worksheets:
        period_date = parse_sheet_date_requires_year(ws.title)
        if not period_date:
            continue
        if period_date < min_period_date:
            continue
        if period_date > upper_bound:
            continue
        total_available_hours = safe_float(ws["T61"].value)
        completed_hours = safe_float(ws["T50"].value)
        wp1_tgt = safe_float(ws["Z7"].value)
        wp2_tgt = safe_float(ws["AB7"].value)
        wp1_out = safe_float(ws["Z2"].value)
        wp2_out = safe_float(ws["AB2"].value)
        target_output = wp1_tgt + wp2_tgt
        actual_output = wp1_out + wp2_out
        if target_output < 0:
            continue
        target_uplh = safe_div(target_output, completed_hours)
        actual_uplh = safe_div(actual_output, completed_hours)
        uplh_wp1 = safe_float(ws["Z5"].value)
        uplh_wp2 = safe_float(ws["AB5"].value)
        hc_in_wip = 0
        for c in cols:
            if safe_float(ws.cell(row=50, column=c).value) != 0.0:
                hc_in_wip += 1
        actual_hc_used = safe_div(completed_hours, 32.5)
        person_hours: Dict[str, Dict[str, float]] = {}
        for c in cols:
            name = safe_str(ws.cell(row=30, column=c).value)
            if not name:
                continue
            actual = safe_float(ws.cell(row=50, column=c).value)
            available = safe_float(ws.cell(row=61, column=c).value)
            person_hours[name] = {"actual": actual, "available": available}
        outputs_by_person: Dict[str, Dict[str, float]] = {}
        for c in cols:
            name = safe_str(ws.cell(row=10, column=c).value)
            if not name:
                continue
            out_val = sum(
                safe_float(ws.cell(row=r, column=c).value)
                for r in range(11, 24)
            )
            tgt_val = safe_float(ws.cell(row=25, column=c).value)
            if out_val != 0.0 or tgt_val != 0.0:
                outputs_by_person[name] = {"output": out_val, "target": tgt_val}
        outputs_by_cell = {
            "WP1": {"output": wp1_out, "target": wp1_tgt},
            "WP2": {"output": wp2_out, "target": wp2_tgt},
        }
        cell_station_hours = {
            "WP1": safe_float(ws["Z4"].value),
            "WP2": safe_float(ws["AB4"].value),
            "WP3": sum(
                safe_float(ws[cell].value)
                for cell in ["T33", "T37", "T41", "T45", "T49"]
            ),
        }
        hours_by_cell_by_person = {"WP1": {}, "WP2": {}, "WP3": {}}
        wp1_hour_rows = [31, 35, 39, 43, 47]
        wp2_hour_rows = [32, 36, 40, 44, 48]
        wp3_hour_rows = [33, 37, 41, 45, 49]
        for c in cols:
            name = safe_str(ws.cell(row=30, column=c).value)
            if not name:
                continue
            wp1_hrs = sum(safe_float(ws.cell(row=r, column=c).value) for r in wp1_hour_rows)
            wp2_hrs = sum(safe_float(ws.cell(row=r, column=c).value) for r in wp2_hour_rows)
            wp3_hrs = sum(safe_float(ws.cell(row=r, column=c).value) for r in wp3_hour_rows)
            if wp1_hrs != 0.0:
                hours_by_cell_by_person["WP1"][name] = wp1_hrs
            if wp2_hrs != 0.0:
                hours_by_cell_by_person["WP2"][name] = wp2_hrs
            if wp3_hrs != 0.0:
                hours_by_cell_by_person["WP3"][name] = wp3_hrs
        output_by_cell_by_person = {"WP1": {}, "WP2": {}}
        wp1_out_rows = [11, 14, 17, 20, 23]
        wp2_out_rows = [12, 15, 18, 21, 24]
        for c in cols:
            name = safe_str(ws.cell(row=13, column=c).value)
            if not name:
                continue
            wp1_o = sum(safe_float(ws.cell(row=r, column=c).value) for r in wp1_out_rows)
            wp2_o = sum(safe_float(ws.cell(row=r, column=c).value) for r in wp2_out_rows)
            if wp1_o != 0.0:
                output_by_cell_by_person["WP1"][name] = wp1_o
            if wp2_o != 0.0:
                output_by_cell_by_person["WP2"][name] = wp2_o
        uplh_by_cell_by_person: Dict[str, Dict[str, Optional[float]]] = {"WP1": {}, "WP2": {}}
        for wp in ("WP1", "WP2"):
            for person, out_val in output_by_cell_by_person[wp].items():
                hrs = safe_float(hours_by_cell_by_person[wp].get(person, 0.0))
                uplh_by_cell_by_person[wp][person] = safe_div(out_val, hrs)
        errs = []
        rows_out.append({
            "team": team,
            "period_date": period_date,
            "source_file": os.path.abspath(os.path.expandvars(source_file)),
            "Total Available Hours": total_available_hours,
            "Completed Hours": completed_hours,
            "Target Output": target_output,
            "Actual Output": actual_output,
            "Target UPLH": target_uplh,
            "Actual UPLH": actual_uplh,
            "UPLH WP1": uplh_wp1,
            "UPLH WP2": uplh_wp2,
            "HC in WIP": hc_in_wip,
            "Actual HC Used": actual_hc_used,
            "People in WIP": "",
            "Person Hours": json.dumps(person_hours, ensure_ascii=False),
            "Outputs by Person": json.dumps(outputs_by_person, ensure_ascii=False),
            "Outputs by Cell/Station": json.dumps(outputs_by_cell, ensure_ascii=False),
            "Cell/Station Hours": json.dumps(cell_station_hours, ensure_ascii=False),
            "Hours by Cell/Station - by person": json.dumps(hours_by_cell_by_person, ensure_ascii=False),
            "Output by Cell/Station - by person": json.dumps(output_by_cell_by_person, ensure_ascii=False),
            "UPLH by Cell/Station - by person": json.dumps(uplh_by_cell_by_person, ensure_ascii=False),
            "error": " | ".join(errs) if errs else "",
        })
    return rows_out
def scrape_workbook_with_config(source_file: str, cfg: Dict[str, Any]) -> list[dict]:
    wb = load_workbook(source_file, data_only=True, read_only=True, keep_links=False)
    rows_out: list[dict] = []
    cols = col_range(cfg["person_cols"][0], cfg["person_cols"][1])
    for ws in wb.worksheets:
        date_parser = cfg.get("date_parser", parse_sheet_date)
        period_date = date_parser(ws.title)
        if not period_date:
            continue
        min_pd = safe_str(cfg.get("min_period_date"))
        if min_pd and period_date < min_pd:
            continue
        max_pd = safe_str(cfg.get("max_period_date"))
        if max_pd and period_date > max_pd:
            continue
        taa_spec = cfg["cells"]["total_available_hours"]
        if isinstance(taa_spec, str):
            total_available_hours = safe_float(ws[taa_spec].value)
        else:
            if taa_spec.get("type") == "sum_range":
                rng = taa_spec["range"]
                total_available_hours = sum(safe_float(cell.value) for row in ws[rng] for cell in row)
            elif taa_spec.get("type") == "sum_cells":
                total_available_hours = sum(safe_float(ws[cell_ref].value) for cell_ref in taa_spec["cells"])
            else:
                total_available_hours = 0.0
        completed_spec = cfg["cells"]["completed_hours"]
        if isinstance(completed_spec, str):
            completed_hours = safe_float(ws[completed_spec].value)
        else:
            if completed_spec.get("type") == "sum_range":
                rng = completed_spec["range"]
                completed_hours = sum(safe_float(cell.value) for row in ws[rng] for cell in row)
            elif completed_spec.get("type") == "sum_cells":
                completed_hours = sum(safe_float(ws[cell_ref].value) for cell_ref in completed_spec["cells"])
            else:
                completed_hours = 0.0
        wp1_out = safe_float(ws[cfg["cells"]["wp1_output"]].value)
        wp2_out = safe_float(ws[cfg["cells"]["wp2_output"]].value)
        wp1_tgt = safe_float(ws[cfg["cells"]["wp1_target"]].value)
        wp2_tgt = safe_float(ws[cfg["cells"]["wp2_target"]].value)
        target_output = wp1_tgt + wp2_tgt
        actual_output = wp1_out + wp2_out
        target_uplh = safe_div(target_output, completed_hours)
        actual_uplh = safe_div(actual_output, completed_hours)
        uplh_wp1 = safe_float(ws[cfg["cells"]["uplh_wp1"]].value)
        uplh_wp2 = safe_float(ws[cfg["cells"]["uplh_wp2"]].value)
        hc_row = cfg["rows"]["hc_row"]
        hc_in_wip = 0
        for c in cols:
            if safe_float(ws.cell(row=hc_row, column=c).value) != 0.0:
                hc_in_wip += 1
        actual_hc_used = safe_div(completed_hours, 32.5)
        person_hours: Dict[str, Dict[str, float]] = {}
        name_row_ph = cfg["rows"]["person_name_row_for_person_hours"]
        actual_row_ph = cfg["rows"]["person_actual_row_for_person_hours"]
        avail_row_ph = cfg["rows"]["person_available_row_for_person_hours"]
        for c in cols:
            name = safe_str(ws.cell(row=name_row_ph, column=c).value)
            if not name:
                continue
            actual = safe_float(ws.cell(row=actual_row_ph, column=c).value)
            available = safe_float(ws.cell(row=avail_row_ph, column=c).value)
            person_hours[name] = {"actual": actual, "available": available}
        outputs_by_person: Dict[str, Dict[str, float]] = {}
        name_row_op = cfg["rows"]["person_name_row_for_outputs_by_person"]
        target_row_op = cfg["rows"]["person_target_row_for_outputs_by_person"]
        output_spec = cfg["outputs_by_person_output"]  
        for c in cols:
            name = safe_str(ws.cell(row=name_row_op, column=c).value)
            if not name:
                continue
            if output_spec["type"] == "row":
                output_val = safe_float(ws.cell(row=output_spec["row"], column=c).value)
            elif output_spec["type"] == "sum_rows":
                output_val = sum_rows(ws, output_spec["rows"], c)
            else:
                output_val = 0.0
            target_val = safe_float(ws.cell(row=target_row_op, column=c).value)
            if output_val != 0.0 or target_val != 0.0:
                outputs_by_person[name] = {"output": output_val, "target": target_val}
        outputs_by_cell = {
            "WP1": {"output": wp1_out, "target": wp1_tgt},
            "WP2": {"output": wp2_out, "target": wp2_tgt},
        }
        cell_station_hours = {
            "WP1": safe_float(ws[cfg["cells"]["wp1_hours"]].value),
            "WP2": safe_float(ws[cfg["cells"]["wp2_hours"]].value),
        }
        wp3_total_cells = cfg.get("cells", {}).get("wp3_hours_sum_cells")
        if wp3_total_cells:
            cell_station_hours["WP3"] = sum(
                safe_float(ws[cell].value)
                for cell in wp3_total_cells
            )
        hours_by_cell_by_person = {"WP1": {}, "WP2": {}}
        wp3_hour_rows = cfg.get("rows", {}).get("wp3_hour_rows")
        if wp3_hour_rows:
            hours_by_cell_by_person["WP3"] = {}
        name_row_hc = cfg["rows"]["person_name_row_for_hours_by_cell_by_person"]
        wp1_hour_rows = cfg["rows"]["wp1_hour_rows"]
        wp2_hour_rows = cfg["rows"]["wp2_hour_rows"]
        for c in cols:
            name = safe_str(ws.cell(row=name_row_hc, column=c).value)
            if not name:
                continue
            wp1_hrs = sum_rows(ws, wp1_hour_rows, c)
            wp2_hrs = sum_rows(ws, wp2_hour_rows, c)
            if wp1_hrs != 0.0:
                hours_by_cell_by_person["WP1"][name] = wp1_hrs
            if wp2_hrs != 0.0:
                hours_by_cell_by_person["WP2"][name] = wp2_hrs
            if wp3_hour_rows:
                wp3_hrs = sum_rows(ws, wp3_hour_rows, c)
                if wp3_hrs != 0.0:
                    hours_by_cell_by_person["WP3"][name] = wp3_hrs
        output_by_cell_by_person = {"WP1": {}, "WP2": {}}
        name_row_oc = cfg["rows"]["person_name_row_for_output_by_cell_by_person"]
        wp1_out_rows = cfg["rows"]["wp1_output_rows_by_person"]
        wp2_out_rows = cfg["rows"]["wp2_output_rows_by_person"]
        for c in cols:
            name = safe_str(ws.cell(row=name_row_oc, column=c).value)
            if not name:
                continue
            wp1_o = sum_rows(ws, wp1_out_rows, c)
            wp2_o = sum_rows(ws, wp2_out_rows, c)
            if wp1_o != 0.0:
                output_by_cell_by_person["WP1"][name] = wp1_o
            if wp2_o != 0.0:
                output_by_cell_by_person["WP2"][name] = wp2_o
        uplh_by_cell_by_person: Dict[str, Dict[str, Optional[float]]] = {"WP1": {}, "WP2": {}}
        for wp in ("WP1", "WP2"):
            for person, out_val in output_by_cell_by_person[wp].items():
                hrs = safe_float(hours_by_cell_by_person[wp].get(person, 0.0))
                uplh_by_cell_by_person[wp][person] = safe_div(out_val, hrs)
        team = cfg["team"]
        rows_out.append(
            {
                "team": team,
                "period_date": period_date,
                "source_file": source_file,
                "Total Available Hours": total_available_hours,
                "Completed Hours": completed_hours,
                "Target Output": target_output,
                "Actual Output": actual_output,
                "Target UPLH": target_uplh,
                "Actual UPLH": actual_uplh,
                "UPLH WP1": uplh_wp1,
                "UPLH WP2": uplh_wp2,
                "HC in WIP": hc_in_wip,
                "Actual HC Used": actual_hc_used,
                "People in WIP": "",
                "Person Hours": json.dumps(person_hours, ensure_ascii=False),
                "Outputs by Person": json.dumps(outputs_by_person, ensure_ascii=False),
                "Outputs by Cell/Station": json.dumps(outputs_by_cell, ensure_ascii=False),
                "Cell/Station Hours": json.dumps(cell_station_hours, ensure_ascii=False),
                "Hours by Cell/Station - by person": json.dumps(hours_by_cell_by_person, ensure_ascii=False),
                "Output by Cell/Station - by person": json.dumps(output_by_cell_by_person, ensure_ascii=False),
                "UPLH by Cell/Station - by person": json.dumps(uplh_by_cell_by_person, ensure_ascii=False),
            }
        )
    return rows_out
def write_csv(rows: list, out_path: str) -> None:
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=HEADERS)
        w.writeheader()
        for r in rows:
            w.writerow({h: r.get(h, "") for h in HEADERS})
def _ph_history_csv_path(out_file: str) -> str:
    out_dir = os.path.dirname(os.path.abspath(out_file)) or "."
    return os.path.join(out_dir, "ph_historical_rows.csv")
def read_rows_csv(path: str) -> list[dict]:
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", newline="", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))
    except Exception:
        return []
def write_rows_csv(rows: list[dict], out_path: str, fieldnames: list[str]) -> None:
    out_dir = os.path.dirname(os.path.abspath(out_path)) or "."
    tmp_path = os.path.join(out_dir, f".{os.path.basename(out_path)}.tmp")
    try:
        with open(tmp_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            for r in rows:
                w.writerow({h: r.get(h, "") for h in fieldnames})
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp_path, out_path)
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass
def merge_rows_by_team_period(rows: list[dict]) -> list[dict]:
    merged: dict[tuple[str, str], dict] = {}
    for r in rows:
        key = (safe_str(r.get("team")), safe_str(r.get("period_date")))
        merged[key] = r
    out = list(merged.values())
    out.sort(key=lambda r: (safe_str(r.get("team")).lower(), safe_str(r.get("period_date"))))
    return out
def monday_of_week(d: date) -> date:
    return d - timedelta(days=d.weekday())
def iso_monday_weeks_back(today: Optional[date] = None, weeks_back: int = 2) -> list[str]:
    if today is None:
        today = date.today()
    start = monday_of_week(today)
    return [(start - timedelta(days=7 * i)).isoformat() for i in range(weeks_back + 1)]
def refresh_ph_recent_3_weeks(
    ph_source_file: str,
    out_file: str,
    ph_new_cfg: Dict[str, Any],
    logger: Optional[logging.Logger] = None,
) -> list[dict]:
    ph_recent_csv = os.path.join(
        os.path.dirname(os.path.abspath(out_file)) or ".",
        "ph_recent_rows.csv",
    )
    existing_recent = read_rows_csv(ph_recent_csv)
    keep_weeks = set(iso_monday_weeks_back(date.today(), weeks_back=2))
    frozen_recent = [
        r for r in existing_recent
        if safe_str(r.get("period_date")) not in keep_weeks
    ]
    cfg = dict(ph_new_cfg)
    cfg["min_period_date"] = min(keep_weeks)
    cfg["max_period_date"] = max(keep_weeks)
    rescanned_recent = scrape_workbook_with_config(ph_source_file, cfg)
    rescanned_recent = [
        r for r in rescanned_recent
        if safe_str(r.get("period_date")) in keep_weeks
    ]
    merged_recent = merge_rows_by_team_period(frozen_recent + rescanned_recent)
    write_rows_csv(merged_recent, ph_recent_csv, HEADERS)
    if logger:
        logger.info(
            f"[PH] refreshed only these weeks: {sorted(keep_weeks)} | "
            f"kept_cached={len(frozen_recent)} | rescanned={len(rescanned_recent)} | "
            f"recent_cache_total={len(merged_recent)}"
        )
    return merged_recent
PH_HISTORY_MIN_DATE = "2025-09-01"
PH_HISTORY_OLD_END = "2026-03-16"
def freeze_ph_history_once(
    ph_source_file: str,
    out_file: str,
    ph_old_cfg: Dict[str, Any],
    logger: Optional[logging.Logger] = None,
) -> None:
    ph_hist_csv = _ph_history_csv_path(out_file)
    if os.path.exists(ph_hist_csv):
        if logger:
            logger.info(f"[PH] history csv already exists: {ph_hist_csv}")
        return
    cfg = dict(ph_old_cfg)
    cfg["min_period_date"] = PH_HISTORY_MIN_DATE
    cfg["max_period_date"] = PH_HISTORY_OLD_END
    old_rows = scrape_workbook_with_config(ph_source_file, cfg)
    old_rows = merge_rows_by_team_period(old_rows)
    write_rows_csv(old_rows, ph_hist_csv, HEADERS)
    if logger:
        logger.info(
            f"[PH] froze {len(old_rows)} historical rows "
            f"from {PH_HISTORY_MIN_DATE} to {PH_HISTORY_OLD_END} "
            f"to {ph_hist_csv}"
        )
def _excel_col_range(start_col: str, end_col: str) -> list[int]:
    s = column_index_from_string(start_col)
    e = column_index_from_string(end_col)
    return list(range(s, e + 1))
def _as_iso_date(v: Any) -> str:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = safe_str(v)
    if not s:
        return ""
    return parse_sheet_date(s)
def _get_dropdown_values_from_validation(cell) -> list[Any]:
    try:
        formula1 = cell.Validation.Formula1
    except Exception:
        return []
    if not formula1:
        return []
    f = str(formula1).strip()
    if not f.startswith("="):
        vals = [x.strip() for x in f.split(",")]
        return [v for v in vals if v]
    try:
        rng = cell.Parent.Range(f.lstrip("="))
        vals = []
        v = rng.Value
        if isinstance(v, tuple):
            for row in v:
                if isinstance(row, tuple):
                    for item in row:
                        if safe_str(item):
                            vals.append(item)
                else:
                    if safe_str(row):
                        vals.append(row)
        else:
            if safe_str(v):
                vals.append(v)
        return vals
    except Exception:
        try:
            app = cell.Application
            rng = app.Range(f.lstrip("="))
            v = rng.Value
            vals = []
            if isinstance(v, tuple):
                for row in v:
                    if isinstance(row, tuple):
                        for item in row:
                            if safe_str(item):
                                vals.append(item)
                    else:
                        if safe_str(row):
                            vals.append(row)
            else:
                if safe_str(v):
                    vals.append(v)
            return vals
        except Exception:
            return []
def _com_call(fn, tries: int = 30, sleep_s: float = 0.25):
    for i in range(tries):
        try:
            return fn()
        except pywintypes.com_error as e:
            if e.args and e.args[0] == -2147418111:
                time.sleep(sleep_s)
                continue
            raise
    return fn()
def scrape_dbs_previous_weeks_xlsm(source_file: str, team: str, dropdown_override: Optional[list[Any]] = None) -> list[dict]:
    import shutil
    import tempfile
    import uuid
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")  # new isolated Excel instance
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    excel.EnableEvents = False
    excel.AutomationSecurity = 3  
    wb = None
    tmp_path = None
    rows_out: list[dict] = []
    def _open_via_temp_copy(src_path: str):
        nonlocal tmp_path
        src = os.path.abspath(os.path.expandvars(src_path))
        if not os.path.exists(src):
            raise FileNotFoundError(f"File not found on disk: {src}")
        base = os.path.splitext(os.path.basename(src))[0]
        ext = os.path.splitext(src)[1]
        tmp_path = os.path.join(tempfile.gettempdir(), f"{base}__{uuid.uuid4().hex}{ext}")
        shutil.copy2(src, tmp_path)
        return _com_call(lambda: excel.Workbooks.Open(
            tmp_path,
            UpdateLinks=0,
            ReadOnly=True,
            IgnoreReadOnlyRecommended=True,
            Notify=False,
            AddToMru=False,
            CorruptLoad=0,  # xlNormalLoad
        ))
    try:
        wb = _open_via_temp_copy(source_file)
        ws = _com_call(lambda: wb.Worksheets("Previous Weeks"))
        dd = _com_call(lambda: ws.Range("A2"))
        dropdown_values = dropdown_override if dropdown_override is not None else _get_dropdown_values_from_validation(dd)
        seen = set()
        dropdown_values = [v for v in dropdown_values if not (safe_str(v) in seen or seen.add(safe_str(v)))]
        cols = _excel_col_range("B", "M")
        for choice in dropdown_values:
            _com_call(lambda: setattr(dd, "Value", choice))
            _com_call(lambda: excel.Calculate())
            period_date = _as_iso_date(_com_call(lambda: dd.Value))
            if not period_date:
                continue
            total_available_hours = safe_float(_com_call(lambda: ws.Range("O69").Value))
            completed_hours = safe_float(_com_call(lambda: ws.Range("O59").Value))
            wp1_tgt = safe_float(_com_call(lambda: ws.Range("T10").Value))
            wp2_tgt = safe_float(_com_call(lambda: ws.Range("V10").Value))
            wp1_out = safe_float(_com_call(lambda: ws.Range("T5").Value))
            wp2_out = safe_float(_com_call(lambda: ws.Range("V5").Value))
            target_output = wp1_tgt + wp2_tgt
            actual_output = wp1_out + wp2_out
            target_uplh = safe_div(target_output, completed_hours)
            actual_uplh = safe_div(actual_output, completed_hours)
            uplh_wp1 = safe_float(_com_call(lambda: ws.Range("T8").Value))
            uplh_wp2 = safe_float(_com_call(lambda: ws.Range("V8").Value))
            hc_in_wip = 0
            for c in cols:
                if safe_float(_com_call(lambda c=c: ws.Cells(59, c).Value)) != 0.0:
                    hc_in_wip += 1
            actual_hc_used = safe_div(completed_hours, 32.5)
            person_hours: Dict[str, Dict[str, float]] = {}
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(33, c).Value))
                if not name:
                    continue
                actual = safe_float(_com_call(lambda c=c: ws.Cells(59, c).Value))
                available = safe_float(_com_call(lambda c=c: ws.Cells(69, c).Value))
                person_hours[name] = {"actual": actual, "available": available}
            outputs_by_person: Dict[str, Dict[str, float]] = {}
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(13, c).Value))
                if not name:
                    continue
                out_val = sum(
                    safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value))
                    for r in range(14, 28)
                )
                tgt_val = safe_float(_com_call(lambda c=c: ws.Cells(28, c).Value))
                if out_val != 0.0 or tgt_val != 0.0:
                    outputs_by_person[name] = {"output": out_val, "target": tgt_val}
            outputs_by_cell = {
                "WP1": {"output": wp1_out, "target": wp1_tgt},
                "WP2": {"output": wp2_out, "target": wp2_tgt},
            }
            cell_station_hours = {
                "WP1": safe_float(_com_call(lambda: ws.Range("T7").Value)),
                "WP2": safe_float(_com_call(lambda: ws.Range("V7").Value)),
            }
            hours_by_cell_by_person = {"WP1": {}, "WP2": {}}
            wp1_hour_rows = [34, 39, 44, 49, 54]
            wp2_hour_rows = [35, 40, 45, 50, 55]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(33, c).Value))
                if not name:
                    continue
                wp1_hrs = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp1_hour_rows)
                wp2_hrs = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp2_hour_rows)
                if wp1_hrs != 0.0:
                    hours_by_cell_by_person["WP1"][name] = wp1_hrs
                if wp2_hrs != 0.0:
                    hours_by_cell_by_person["WP2"][name] = wp2_hrs
            output_by_cell_by_person = {"WP1": {}, "WP2": {}}
            wp1_out_rows = [14, 17, 20, 23, 26]
            wp2_out_rows = [15, 18, 21, 24, 27]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(13, c).Value))
                if not name:
                    continue
                wp1_o = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp1_out_rows)
                wp2_o = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp2_out_rows)
                if wp1_o != 0.0:
                    output_by_cell_by_person["WP1"][name] = wp1_o
                if wp2_o != 0.0:
                    output_by_cell_by_person["WP2"][name] = wp2_o
            uplh_by_cell_by_person: Dict[str, Dict[str, Optional[float]]] = {"WP1": {}, "WP2": {}}
            for wp in ("WP1", "WP2"):
                for person, out_val in output_by_cell_by_person[wp].items():
                    hrs = safe_float(hours_by_cell_by_person[wp].get(person, 0.0))
                    uplh_by_cell_by_person[wp][person] = safe_div(out_val, hrs)
            rows_out.append({
                "team": team,
                "period_date": period_date,
                "source_file": os.path.abspath(os.path.expandvars(source_file)),
                "Total Available Hours": total_available_hours,
                "Completed Hours": completed_hours,
                "Target Output": target_output,
                "Actual Output": actual_output,
                "Target UPLH": target_uplh,
                "Actual UPLH": actual_uplh,
                "UPLH WP1": uplh_wp1,
                "UPLH WP2": uplh_wp2,
                "HC in WIP": hc_in_wip,
                "Actual HC Used": actual_hc_used,
                "People in WIP": "",
                "Person Hours": json.dumps(person_hours, ensure_ascii=False),
                "Outputs by Person": json.dumps(outputs_by_person, ensure_ascii=False),
                "Outputs by Cell/Station": json.dumps(outputs_by_cell, ensure_ascii=False),
                "Cell/Station Hours": json.dumps(cell_station_hours, ensure_ascii=False),
                "Hours by Cell/Station - by person": json.dumps(hours_by_cell_by_person, ensure_ascii=False),
                "Output by Cell/Station - by person": json.dumps(output_by_cell_by_person, ensure_ascii=False),
                "UPLH by Cell/Station - by person": json.dumps(uplh_by_cell_by_person, ensure_ascii=False),
            })
        return rows_out
    finally:
        try:
            if wb is not None:
                _com_call(lambda: wb.Close(SaveChanges=False), tries=10, sleep_s=0.3)
        except Exception:
            pass
        try:
            _com_call(lambda: excel.Quit(), tries=10, sleep_s=0.3)
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        try:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
def scrape_nav_previous_weeks_xlsm(source_file: str, team: str = "Nav", dropdown_override: Optional[list[Any]] = None) -> list[dict]:
    import shutil
    import tempfile
    import uuid
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")  # isolated Excel instance
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    excel.EnableEvents = False
    excel.AutomationSecurity = 3  # force-disable macros
    wb = None
    tmp_path = None
    rows_out: list[dict] = []
    def _open_via_temp_copy(src_path: str):
        nonlocal tmp_path
        src = os.path.abspath(os.path.expandvars(src_path))
        if not os.path.exists(src):
            raise FileNotFoundError(f"File not found on disk: {src}")
        base = os.path.splitext(os.path.basename(src))[0]
        ext = os.path.splitext(src)[1]
        tmp_path = os.path.join(tempfile.gettempdir(), f"{base}__{uuid.uuid4().hex}{ext}")
        shutil.copy2(src, tmp_path)
        return _com_call(
            lambda: excel.Workbooks.Open(
                tmp_path,
                UpdateLinks=0,
                ReadOnly=True,
                IgnoreReadOnlyRecommended=True,
                Notify=False,
                AddToMru=False,
                CorruptLoad=0,
            )
        )
    try:
        wb = _open_via_temp_copy(source_file)
        ws = _com_call(lambda: wb.Worksheets("Previous Weeks"))
        dd = _com_call(lambda: ws.Range("A2"))
        dropdown_values = dropdown_override if dropdown_override is not None else _get_dropdown_values_from_validation(dd)
        seen = set()
        dropdown_values = [v for v in dropdown_values if not (safe_str(v) in seen or seen.add(safe_str(v)))]
        cols = _excel_col_range("B", "V")
        today_iso = date.today().isoformat() 
        for choice in dropdown_values:
            _com_call(lambda: setattr(dd, "Value", choice))
            _com_call(lambda: excel.Calculate())
            period_date = _as_iso_date(_com_call(lambda: dd.Value))
            if not period_date:
                continue
            if period_date < "2025-06-02":
                continue
            if period_date > today_iso:
                continue
            total_available_hours = safe_float(_com_call(lambda: ws.Range("X64").Value))
            completed_hours = safe_float(_com_call(lambda: ws.Range("X54").Value))
            wp1_tgt = safe_float(_com_call(lambda: ws.Range("AD10").Value))
            wp2_tgt = safe_float(_com_call(lambda: ws.Range("AF10").Value))
            wp1_out = safe_float(_com_call(lambda: ws.Range("AD5").Value))
            wp2_out = safe_float(_com_call(lambda: ws.Range("AF5").Value))
            target_output = wp1_tgt + wp2_tgt
            actual_output = wp1_out + wp2_out
            if target_output < 0:
                continue
            target_uplh = safe_div(target_output, completed_hours)
            actual_uplh = safe_div(actual_output, completed_hours)
            uplh_wp1 = safe_float(_com_call(lambda: ws.Range("AD8").Value))
            uplh_wp2 = safe_float(_com_call(lambda: ws.Range("AF8").Value))
            hc_in_wip = 0
            for c in cols:
                if safe_float(_com_call(lambda c=c: ws.Cells(28, c).Value)) != 0.0:
                    hc_in_wip += 1
            actual_hc_used = safe_div(completed_hours, 32.5)
            person_hours: Dict[str, Dict[str, float]] = {}
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(33, c).Value))
                if not name:
                    continue
                actual = safe_float(_com_call(lambda c=c: ws.Cells(54, c).Value))
                available = safe_float(_com_call(lambda c=c: ws.Cells(64, c).Value))
                person_hours[name] = {"actual": actual, "available": available}
            outputs_by_person: Dict[str, Dict[str, float]] = {}
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(13, c).Value))
                if not name:
                    continue
                out_val = sum(
                    safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value))
                    for r in range(14, 28)  # 14..27 inclusive
                )
                tgt_val = safe_float(_com_call(lambda c=c: ws.Cells(28, c).Value))
                if out_val != 0.0 or tgt_val != 0.0:
                    outputs_by_person[name] = {"output": out_val, "target": tgt_val}
            outputs_by_cell = {
                "WP1": {"output": wp1_out, "target": wp1_tgt},
                "WP2": {"output": wp2_out, "target": wp2_tgt},
            }
            cell_station_hours = {
                "WP1": safe_float(_com_call(lambda: ws.Range("AD7").Value)),
                "WP2": safe_float(_com_call(lambda: ws.Range("AF7").Value)),
            }
            hours_by_cell_by_person = {"WP1": {}, "WP2": {}}
            wp1_hour_rows = [34, 38, 42, 46, 50]
            wp2_hour_rows = [35, 39, 43, 47, 51]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(33, c).Value))
                if not name:
                    continue
                wp1_hrs = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp1_hour_rows)
                wp2_hrs = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp2_hour_rows)
                if wp1_hrs != 0.0:
                    hours_by_cell_by_person["WP1"][name] = wp1_hrs
                if wp2_hrs != 0.0:
                    hours_by_cell_by_person["WP2"][name] = wp2_hrs
            output_by_cell_by_person = {"WP1": {}, "WP2": {}}
            wp1_out_rows = [14, 17, 20, 23, 26]
            wp2_out_rows = [15, 18, 21, 24, 27]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(13, c).Value))
                if not name:
                    continue
                wp1_o = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp1_out_rows)
                wp2_o = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp2_out_rows)
                if wp1_o != 0.0:
                    output_by_cell_by_person["WP1"][name] = wp1_o
                if wp2_o != 0.0:
                    output_by_cell_by_person["WP2"][name] = wp2_o
            uplh_by_cell_by_person: Dict[str, Dict[str, Optional[float]]] = {"WP1": {}, "WP2": {}}
            for wp in ("WP1", "WP2"):
                for person, out_val in output_by_cell_by_person[wp].items():
                    hrs = safe_float(hours_by_cell_by_person[wp].get(person, 0.0))
                    uplh_by_cell_by_person[wp][person] = safe_div(out_val, hrs)
            key = (team, period_date)
            rows_out.append(
                {
                    "team": team,
                    "period_date": period_date,
                    "source_file": os.path.abspath(os.path.expandvars(source_file)),
                    "Total Available Hours": total_available_hours,
                    "Completed Hours": completed_hours,
                    "Target Output": target_output,
                    "Actual Output": actual_output,
                    "Target UPLH": target_uplh,
                    "Actual UPLH": actual_uplh,
                    "UPLH WP1": uplh_wp1,
                    "UPLH WP2": uplh_wp2,
                    "HC in WIP": hc_in_wip,
                    "Actual HC Used": actual_hc_used,
                    "People in WIP": "",
                    "Person Hours": json.dumps(person_hours, ensure_ascii=False),
                    "Outputs by Person": json.dumps(outputs_by_person, ensure_ascii=False),
                    "Outputs by Cell/Station": json.dumps(outputs_by_cell, ensure_ascii=False),
                    "Cell/Station Hours": json.dumps(cell_station_hours, ensure_ascii=False),
                    "Hours by Cell/Station - by person": json.dumps(hours_by_cell_by_person, ensure_ascii=False),
                    "Output by Cell/Station - by person": json.dumps(output_by_cell_by_person, ensure_ascii=False),
                    "UPLH by Cell/Station - by person": json.dumps(uplh_by_cell_by_person, ensure_ascii=False),
                }
            )
        return rows_out
    finally:
        try:
            if wb is not None:
                _com_call(lambda: wb.Close(SaveChanges=False), tries=10, sleep_s=0.3)
        except Exception:
            pass
        try:
            _com_call(lambda: excel.Quit(), tries=10, sleep_s=0.3)
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        try:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
def scrape_meic_ae_oarm_previous_weeks_xlsm(source_file: str, team: str, dropdown_override: Optional[list[Any]] = None) -> list[dict]:
    import shutil
    import tempfile
    import uuid
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    excel.EnableEvents = False
    excel.AutomationSecurity = 3  # disable macros
    wb = None
    tmp_path = None
    rows_out: list[dict] = []
    def _open_via_temp_copy(src_path: str):
        nonlocal tmp_path
        src = os.path.abspath(os.path.expandvars(src_path))
        if not os.path.exists(src):
            raise FileNotFoundError(f"File not found on disk: {src}")
        base = os.path.splitext(os.path.basename(src))[0]
        ext = os.path.splitext(src)[1]
        tmp_path = os.path.join(tempfile.gettempdir(), f"{base}__{uuid.uuid4().hex}{ext}")
        shutil.copy2(src, tmp_path)
        return _com_call(
            lambda: excel.Workbooks.Open(
                tmp_path,
                UpdateLinks=0,
                ReadOnly=True,
                IgnoreReadOnlyRecommended=True,
                Notify=False,
                AddToMru=False,
                CorruptLoad=0,
            )
        )
    try:
        wb = _open_via_temp_copy(source_file)
        ws = _com_call(lambda: wb.Worksheets("Previous Weeks"))
        dd = _com_call(lambda: ws.Range("A2"))
        dropdown_values = dropdown_override if dropdown_override is not None else _get_dropdown_values_from_validation(dd)
        seen = set()
        dropdown_values = [v for v in dropdown_values if not (safe_str(v) in seen or seen.add(safe_str(v)))]
        cols = _excel_col_range("B", "P") 
        today_iso = date.today().isoformat()
        for choice in dropdown_values:
            _com_call(lambda: setattr(dd, "Value", choice))
            try:
                _com_call(lambda: wb.RefreshAll())
            except Exception:
                pass
            try:
                _com_call(lambda: excel.CalculateFullRebuild())
            except Exception:
                _com_call(lambda: excel.Calculate())
            time.sleep(1)
            period_date = _as_iso_date(_com_call(lambda: dd.Value))
            if not period_date:
                continue
            if period_date < "2025-06-02":
                continue
            if period_date > today_iso:
                continue
            total_available_hours = safe_float(_com_call(lambda: ws.Range("R64").Value))
            completed_hours = safe_float(_com_call(lambda: ws.Range("R54").Value))
            wp1_tgt = safe_float(_com_call(lambda: ws.Range("X10").Value))
            wp2_tgt = safe_float(_com_call(lambda: ws.Range("Z10").Value))
            wp1_out = safe_float(_com_call(lambda: ws.Range("X5").Value))
            wp2_out = safe_float(_com_call(lambda: ws.Range("Z5").Value))
            target_output = wp1_tgt + wp2_tgt
            actual_output = wp1_out + wp2_out
            if target_output < 0:
                continue
            target_uplh = safe_div(target_output, completed_hours)
            actual_uplh = safe_div(actual_output, completed_hours)
            uplh_wp1 = safe_float(_com_call(lambda: ws.Range("X8").Value))
            uplh_wp2 = safe_float(_com_call(lambda: ws.Range("Z8").Value))
            hc_in_wip = 0
            for c in cols:
                if safe_float(_com_call(lambda c=c: ws.Cells(28, c).Value)) != 0.0:
                    hc_in_wip += 1
            actual_hc_used = safe_div(completed_hours, 32.5)
            person_hours: Dict[str, Dict[str, float]] = {}
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(33, c).Value))
                if not name:
                    continue
                actual = safe_float(_com_call(lambda c=c: ws.Cells(54, c).Value))
                available = safe_float(_com_call(lambda c=c: ws.Cells(64, c).Value))
                person_hours[name] = {"actual": actual, "available": available}
            outputs_by_person: Dict[str, Dict[str, float]] = {}
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(13, c).Value))
                if not name:
                    continue
                out_val = sum(
                    safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value))
                    for r in range(14, 28)  # 14..27 inclusive
                )
                tgt_val = safe_float(_com_call(lambda c=c: ws.Cells(28, c).Value))
                if out_val != 0.0 or tgt_val != 0.0:
                    outputs_by_person[name] = {"output": out_val, "target": tgt_val}
            outputs_by_cell = {
                "WP1": {"output": wp1_out, "target": wp1_tgt},
                "WP2": {"output": wp2_out, "target": wp2_tgt},
            }
            cell_station_hours = {
                "WP1": safe_float(_com_call(lambda: ws.Range("X7").Value)),
                "WP2": safe_float(_com_call(lambda: ws.Range("Z7").Value)),
            }
            hours_by_cell_by_person = {"WP1": {}, "WP2": {}}
            wp1_hour_rows = [34, 38, 42, 46, 50]
            wp2_hour_rows = [35, 39, 43, 47, 51]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(33, c).Value))
                if not name:
                    continue
                wp1_hrs = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp1_hour_rows)
                wp2_hrs = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp2_hour_rows)
                if wp1_hrs != 0.0:
                    hours_by_cell_by_person["WP1"][name] = wp1_hrs
                if wp2_hrs != 0.0:
                    hours_by_cell_by_person["WP2"][name] = wp2_hrs
            output_by_cell_by_person = {"WP1": {}, "WP2": {}}
            wp1_out_rows = [14, 17, 20, 23, 26]
            wp2_out_rows = [15, 18, 21, 24, 27]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(13, c).Value))
                if not name:
                    continue
                wp1_o = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp1_out_rows)
                wp2_o = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp2_out_rows)
                if wp1_o != 0.0:
                    output_by_cell_by_person["WP1"][name] = wp1_o
                if wp2_o != 0.0:
                    output_by_cell_by_person["WP2"][name] = wp2_o
            uplh_by_cell_by_person: Dict[str, Dict[str, Optional[float]]] = {"WP1": {}, "WP2": {}}
            for wp in ("WP1", "WP2"):
                for person, out_val in output_by_cell_by_person[wp].items():
                    hrs = safe_float(hours_by_cell_by_person[wp].get(person, 0.0))
                    uplh_by_cell_by_person[wp][person] = safe_div(out_val, hrs)
            rows_out.append(
                {
                    "team": team,
                    "period_date": period_date,
                    "source_file": os.path.abspath(os.path.expandvars(source_file)),
                    "Total Available Hours": total_available_hours,
                    "Completed Hours": completed_hours,
                    "Target Output": target_output,
                    "Actual Output": actual_output,
                    "Target UPLH": target_uplh,
                    "Actual UPLH": actual_uplh,
                    "UPLH WP1": uplh_wp1,
                    "UPLH WP2": uplh_wp2,
                    "HC in WIP": hc_in_wip,
                    "Actual HC Used": actual_hc_used,
                    "People in WIP": "",
                    "Person Hours": json.dumps(person_hours, ensure_ascii=False),
                    "Outputs by Person": json.dumps(outputs_by_person, ensure_ascii=False),
                    "Outputs by Cell/Station": json.dumps(outputs_by_cell, ensure_ascii=False),
                    "Cell/Station Hours": json.dumps(cell_station_hours, ensure_ascii=False),
                    "Hours by Cell/Station - by person": json.dumps(hours_by_cell_by_person, ensure_ascii=False),
                    "Output by Cell/Station - by person": json.dumps(output_by_cell_by_person, ensure_ascii=False),
                    "UPLH by Cell/Station - by person": json.dumps(uplh_by_cell_by_person, ensure_ascii=False),
                }
            )
        return rows_out
    finally:
        try:
            if wb is not None:
                _com_call(lambda: wb.Close(SaveChanges=False), tries=10, sleep_s=0.3)
        except Exception:
            pass
        try:
            _com_call(lambda: excel.Quit(), tries=10, sleep_s=0.3)
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        try:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
def scrape_previous_weeks_xlsm_with_filters(source_file: str, team: str, cfg: Dict[str, Any], dropdown_override: Optional[list[Any]] = None) -> list[dict]:
    import shutil
    import tempfile
    import uuid
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    excel.EnableEvents = False
    excel.AutomationSecurity = 3  # disable macros
    wb = None
    tmp_path = None
    rows_out: list[dict] = []
    def _open_via_temp_copy(src_path: str):
        nonlocal tmp_path
        src = os.path.abspath(os.path.expandvars(src_path))
        if not os.path.exists(src):
            raise FileNotFoundError(f"File not found on disk: {src}")
        base = os.path.splitext(os.path.basename(src))[0]
        ext = os.path.splitext(src)[1]
        tmp_path = os.path.join(tempfile.gettempdir(), f"{base}__{uuid.uuid4().hex}{ext}")
        shutil.copy2(src, tmp_path)
        return _com_call(
            lambda: excel.Workbooks.Open(
                tmp_path,
                UpdateLinks=0,
                ReadOnly=True,
                IgnoreReadOnlyRecommended=True,
                Notify=False,
                AddToMru=False,
                CorruptLoad=0,
            )
        )
    try:
        wb = _open_via_temp_copy(source_file)
        ws = _com_call(lambda: wb.Worksheets("Previous Weeks"))
        dd = _com_call(lambda: ws.Range("A2"))
        dropdown_values = dropdown_override if dropdown_override is not None else _get_dropdown_values_from_validation(dd)
        seen = set()
        dropdown_values = [v for v in dropdown_values if not (safe_str(v) in seen or seen.add(safe_str(v)))]
        cols = _excel_col_range(cfg["person_cols"][0], cfg["person_cols"][1])
        today_iso = date.today().isoformat()
        for choice in dropdown_values:
            _com_call(lambda: setattr(dd, "Value", choice))
            _com_call(lambda: excel.Calculate())
            period_date = _as_iso_date(_com_call(lambda: dd.Value))
            if not period_date:
                continue
            if period_date < "2025-06-02":
                continue
            if period_date > today_iso:
                continue
            total_available_hours = safe_float(_com_call(lambda: ws.Range(cfg["cells"]["total_available_hours"]).Value))
            completed_hours = safe_float(_com_call(lambda: ws.Range(cfg["cells"]["completed_hours"]).Value))
            wp1_tgt = safe_float(_com_call(lambda: ws.Range(cfg["cells"]["wp1_target"]).Value))
            wp2_tgt = safe_float(_com_call(lambda: ws.Range(cfg["cells"]["wp2_target"]).Value))
            wp1_out = safe_float(_com_call(lambda: ws.Range(cfg["cells"]["wp1_output"]).Value))
            wp2_out = safe_float(_com_call(lambda: ws.Range(cfg["cells"]["wp2_output"]).Value))
            target_output = wp1_tgt + wp2_tgt
            actual_output = wp1_out + wp2_out
            if target_output < 0:
                continue
            target_uplh = safe_div(target_output, completed_hours)
            actual_uplh = safe_div(actual_output, completed_hours)
            uplh_wp1 = safe_float(_com_call(lambda: ws.Range(cfg["cells"]["uplh_wp1"]).Value))
            uplh_wp2 = safe_float(_com_call(lambda: ws.Range(cfg["cells"]["uplh_wp2"]).Value))
            hc_in_wip = 0
            for c in cols:
                if safe_float(_com_call(lambda c=c: ws.Cells(28, c).Value)) != 0.0:
                    hc_in_wip += 1
            actual_hc_used = safe_div(completed_hours, 32.5)
            person_hours: Dict[str, Dict[str, float]] = {}
            name_row_ph = cfg["rows"]["person_name_row_for_person_hours"]
            actual_row_ph = cfg["rows"]["person_actual_row_for_person_hours"]
            avail_row_ph = cfg["rows"]["person_available_row_for_person_hours"]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(name_row_ph, c).Value))
                if not name:
                    continue
                actual = safe_float(_com_call(lambda c=c: ws.Cells(actual_row_ph, c).Value))
                available = safe_float(_com_call(lambda c=c: ws.Cells(avail_row_ph, c).Value))
                person_hours[name] = {"actual": actual, "available": available}
            outputs_by_person: Dict[str, Dict[str, float]] = {}
            name_row_op = cfg["rows"]["person_name_row_for_outputs_by_person"]
            target_row_op = cfg["rows"]["person_target_row_for_outputs_by_person"]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(name_row_op, c).Value))
                if not name:
                    continue
                out_val = sum(
                    safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value))
                    for r in range(14, 28)  # 14..27
                )
                tgt_val = safe_float(_com_call(lambda c=c: ws.Cells(target_row_op, c).Value))
                if out_val != 0.0 or tgt_val != 0.0:
                    outputs_by_person[name] = {"output": out_val, "target": tgt_val}
            outputs_by_cell = {
                "WP1": {"output": wp1_out, "target": wp1_tgt},
                "WP2": {"output": wp2_out, "target": wp2_tgt},
            }
            cell_station_hours = {
                "WP1": safe_float(_com_call(lambda: ws.Range(cfg["cells"]["wp1_hours"]).Value)),
                "WP2": safe_float(_com_call(lambda: ws.Range(cfg["cells"]["wp2_hours"]).Value)),
            }
            hours_by_cell_by_person = {"WP1": {}, "WP2": {}}
            name_row_hc = cfg["rows"]["person_name_row_for_hours_by_cell_by_person"]
            wp1_hour_rows = cfg["rows"]["wp1_hour_rows"]
            wp2_hour_rows = cfg["rows"]["wp2_hour_rows"]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(name_row_hc, c).Value))
                if not name:
                    continue
                wp1_hrs = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp1_hour_rows)
                wp2_hrs = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp2_hour_rows)
                if wp1_hrs != 0.0:
                    hours_by_cell_by_person["WP1"][name] = wp1_hrs
                if wp2_hrs != 0.0:
                    hours_by_cell_by_person["WP2"][name] = wp2_hrs
            output_by_cell_by_person = {"WP1": {}, "WP2": {}}
            name_row_oc = cfg["rows"]["person_name_row_for_output_by_cell_by_person"]
            wp1_out_rows = cfg["rows"]["wp1_output_rows_by_person"]
            wp2_out_rows = cfg["rows"]["wp2_output_rows_by_person"]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(name_row_oc, c).Value))
                if not name:
                    continue
                wp1_o = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp1_out_rows)
                wp2_o = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp2_out_rows)
                if wp1_o != 0.0:
                    output_by_cell_by_person["WP1"][name] = wp1_o
                if wp2_o != 0.0:
                    output_by_cell_by_person["WP2"][name] = wp2_o
            uplh_by_cell_by_person: Dict[str, Dict[str, Optional[float]]] = {"WP1": {}, "WP2": {}}
            for wp in ("WP1", "WP2"):
                for person, out_val in output_by_cell_by_person[wp].items():
                    hrs = safe_float(hours_by_cell_by_person[wp].get(person, 0.0))
                    uplh_by_cell_by_person[wp][person] = safe_div(out_val, hrs)
            rows_out.append(
                {
                    "team": team,
                    "period_date": period_date,
                    "source_file": os.path.abspath(os.path.expandvars(source_file)),
                    "Total Available Hours": total_available_hours,
                    "Completed Hours": completed_hours,
                    "Target Output": target_output,
                    "Actual Output": actual_output,
                    "Target UPLH": target_uplh,
                    "Actual UPLH": actual_uplh,
                    "UPLH WP1": uplh_wp1,
                    "UPLH WP2": uplh_wp2,
                    "HC in WIP": hc_in_wip,
                    "Actual HC Used": actual_hc_used,
                    "People in WIP": "",
                    "Person Hours": json.dumps(person_hours, ensure_ascii=False),
                    "Outputs by Person": json.dumps(outputs_by_person, ensure_ascii=False),
                    "Outputs by Cell/Station": json.dumps(outputs_by_cell, ensure_ascii=False),
                    "Cell/Station Hours": json.dumps(cell_station_hours, ensure_ascii=False),
                    "Hours by Cell/Station - by person": json.dumps(hours_by_cell_by_person, ensure_ascii=False),
                    "Output by Cell/Station - by person": json.dumps(output_by_cell_by_person, ensure_ascii=False),
                    "UPLH by Cell/Station - by person": json.dumps(uplh_by_cell_by_person, ensure_ascii=False),
                }
            )
        return rows_out
    finally:
        try:
            if wb is not None:
                _com_call(lambda: wb.Close(SaveChanges=False), tries=10, sleep_s=0.3)
        except Exception:
            pass
        try:
            _com_call(lambda: excel.Quit(), tries=10, sleep_s=0.3)
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        try:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
def parse_sheet_date_day_first_requires_year(sheet_name: str) -> str:
    raw = (sheet_name or "").strip()
    if not re.search(r"\b\d{4}\b", raw):
        return ""
    raw = raw.replace("\u00a0", " ")
    raw = re.sub(r"\s+", " ", raw).strip()
    fmts = [
        "%d %b %Y",     # 23 Feb 2026
        "%d %B %Y",     # 23 February 2026
        "%d-%b-%Y",
        "%d-%B-%Y",
        "%d/%m/%Y",
        "%Y-%m-%d",
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return ""
def scrape_previous_weeks_xlsm_with_filters(source_file: str, team: str, cfg: Dict[str, Any], dropdown_override: Optional[list[Any]] = None) -> list[dict]:
    import shutil
    import tempfile
    import uuid
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    excel.EnableEvents = False
    excel.AutomationSecurity = 3  # disable macros
    wb = None
    tmp_path = None
    rows_out: list[dict] = []
    def _open_via_temp_copy(src_path: str):
        nonlocal tmp_path
        src = os.path.abspath(os.path.expandvars(src_path))
        if not os.path.exists(src):
            raise FileNotFoundError(f"File not found on disk: {src}")
        base = os.path.splitext(os.path.basename(src))[0]
        ext = os.path.splitext(src)[1]
        tmp_path = os.path.join(tempfile.gettempdir(), f"{base}__{uuid.uuid4().hex}{ext}")
        shutil.copy2(src, tmp_path)
        return _com_call(
            lambda: excel.Workbooks.Open(
                tmp_path,
                UpdateLinks=0,
                ReadOnly=True,
                IgnoreReadOnlyRecommended=True,
                Notify=False,
                AddToMru=False,
                CorruptLoad=0,
            )
        )
    def _cfg_float(ws, key: str) -> float:
        return safe_float(_com_call(lambda: ws.Range(cfg["cells"][key]).Value))
    try:
        wb = _open_via_temp_copy(source_file)
        ws = _com_call(lambda: wb.Worksheets("Previous Weeks"))
        dd = _com_call(lambda: ws.Range("A2"))
        dropdown_values = dropdown_override if dropdown_override is not None else _get_dropdown_values_from_validation(dd)
        seen = set()
        dropdown_values = [v for v in dropdown_values if not (safe_str(v) in seen or seen.add(safe_str(v)))]
        cols = _excel_col_range(cfg["person_cols"][0], cfg["person_cols"][1])
        today_iso = date.today().isoformat()
        min_pd = safe_str(cfg.get("min_period_date")) or "2025-06-02"
        max_pd = safe_str(cfg.get("max_period_date")) or today_iso
        for choice in dropdown_values:
            _com_call(lambda: setattr(dd, "Value", choice))
            try:
                _com_call(lambda: wb.RefreshAll())
            except Exception:
                pass
            try:
                _com_call(lambda: excel.CalculateFullRebuild())
            except Exception:
                _com_call(lambda: excel.Calculate())
            time.sleep(0.5)
            period_date = _as_iso_date(_com_call(lambda: dd.Value))
            if not period_date:
                continue
            if period_date < min_pd:
                continue
            if period_date > max_pd:
                continue
            total_available_hours = _cfg_float(ws, "total_available_hours")
            completed_hours = _cfg_float(ws, "completed_hours")
            wp1_tgt = _cfg_float(ws, "wp1_target")
            wp2_tgt = _cfg_float(ws, "wp2_target")
            wp1_out = _cfg_float(ws, "wp1_output")
            wp2_out = _cfg_float(ws, "wp2_output")
            target_output = wp1_tgt + wp2_tgt
            actual_output = wp1_out + wp2_out
            if target_output < 0:
                continue
            target_uplh = safe_div(target_output, completed_hours)
            actual_uplh = safe_div(actual_output, completed_hours)
            uplh_wp1 = _cfg_float(ws, "uplh_wp1")
            uplh_wp2 = _cfg_float(ws, "uplh_wp2")
            rows_cfg = cfg["rows"]
            hc_row = rows_cfg.get("hc_row", rows_cfg.get("person_target_row_for_outputs_by_person", 28))
            hc_in_wip = 0
            for c in cols:
                if safe_float(_com_call(lambda c=c: ws.Cells(hc_row, c).Value)) != 0.0:
                    hc_in_wip += 1
            actual_hc_used = safe_div(completed_hours, 32.5)
            person_hours: Dict[str, Dict[str, float]] = {}
            name_row_ph = rows_cfg["person_name_row_for_person_hours"]
            actual_row_ph = rows_cfg["person_actual_row_for_person_hours"]
            avail_row_ph = rows_cfg["person_available_row_for_person_hours"]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(name_row_ph, c).Value))
                if not name:
                    continue
                actual = safe_float(_com_call(lambda c=c: ws.Cells(actual_row_ph, c).Value))
                available = safe_float(_com_call(lambda c=c: ws.Cells(avail_row_ph, c).Value))
                person_hours[name] = {"actual": actual, "available": available}
            outputs_by_person: Dict[str, Dict[str, float]] = {}
            name_row_op = rows_cfg["person_name_row_for_outputs_by_person"]
            target_row_op = rows_cfg["person_target_row_for_outputs_by_person"]
            output_spec = cfg.get("outputs_by_person_output")
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(name_row_op, c).Value))
                if not name:
                    continue
                if isinstance(output_spec, dict) and output_spec.get("type") == "sum_rows":
                    out_val = sum(
                        safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value))
                        for r in output_spec.get("rows", [])
                    )
                elif isinstance(output_spec, dict) and output_spec.get("type") == "row":
                    out_val = safe_float(_com_call(lambda c=c: ws.Cells(output_spec["row"], c).Value))
                else:
                    out_val = sum(
                        safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value))
                        for r in range(14, 28)  # legacy PSS layout
                    )
                tgt_val = safe_float(_com_call(lambda c=c: ws.Cells(target_row_op, c).Value))
                if out_val != 0.0 or tgt_val != 0.0:
                    outputs_by_person[name] = {"output": out_val, "target": tgt_val}
            outputs_by_cell = {
                "WP1": {"output": wp1_out, "target": wp1_tgt},
                "WP2": {"output": wp2_out, "target": wp2_tgt},
            }
            cell_station_hours = {
                "WP1": _cfg_float(ws, "wp1_hours"),
                "WP2": _cfg_float(ws, "wp2_hours"),
            }
            hours_by_cell_by_person = {"WP1": {}, "WP2": {}}
            name_row_hc = rows_cfg["person_name_row_for_hours_by_cell_by_person"]
            wp1_hour_rows = rows_cfg["wp1_hour_rows"]
            wp2_hour_rows = rows_cfg["wp2_hour_rows"]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(name_row_hc, c).Value))
                if not name:
                    continue
                wp1_hrs = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp1_hour_rows)
                wp2_hrs = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp2_hour_rows)
                if wp1_hrs != 0.0:
                    hours_by_cell_by_person["WP1"][name] = wp1_hrs
                if wp2_hrs != 0.0:
                    hours_by_cell_by_person["WP2"][name] = wp2_hrs
            output_by_cell_by_person = {"WP1": {}, "WP2": {}}
            name_row_oc = rows_cfg["person_name_row_for_output_by_cell_by_person"]
            wp1_out_rows = rows_cfg["wp1_output_rows_by_person"]
            wp2_out_rows = rows_cfg["wp2_output_rows_by_person"]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(name_row_oc, c).Value))
                if not name:
                    continue
                wp1_o = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp1_out_rows)
                wp2_o = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp2_out_rows)
                if wp1_o != 0.0:
                    output_by_cell_by_person["WP1"][name] = wp1_o
                if wp2_o != 0.0:
                    output_by_cell_by_person["WP2"][name] = wp2_o
            uplh_by_cell_by_person: Dict[str, Dict[str, Optional[float]]] = {"WP1": {}, "WP2": {}}
            for wp in ("WP1", "WP2"):
                for person, out_val in output_by_cell_by_person[wp].items():
                    hrs = safe_float(hours_by_cell_by_person[wp].get(person, 0.0))
                    uplh_by_cell_by_person[wp][person] = safe_div(out_val, hrs)
            rows_out.append(
                {
                    "team": team,
                    "period_date": period_date,
                    "source_file": os.path.abspath(os.path.expandvars(source_file)),
                    "Total Available Hours": total_available_hours,
                    "Completed Hours": completed_hours,
                    "Target Output": target_output,
                    "Actual Output": actual_output,
                    "Target UPLH": target_uplh,
                    "Actual UPLH": actual_uplh,
                    "UPLH WP1": uplh_wp1,
                    "UPLH WP2": uplh_wp2,
                    "HC in WIP": hc_in_wip,
                    "Actual HC Used": actual_hc_used,
                    "People in WIP": "",
                    "Person Hours": json.dumps(person_hours, ensure_ascii=False),
                    "Outputs by Person": json.dumps(outputs_by_person, ensure_ascii=False),
                    "Outputs by Cell/Station": json.dumps(outputs_by_cell, ensure_ascii=False),
                    "Cell/Station Hours": json.dumps(cell_station_hours, ensure_ascii=False),
                    "Hours by Cell/Station - by person": json.dumps(hours_by_cell_by_person, ensure_ascii=False),
                    "Output by Cell/Station - by person": json.dumps(output_by_cell_by_person, ensure_ascii=False),
                    "UPLH by Cell/Station - by person": json.dumps(uplh_by_cell_by_person, ensure_ascii=False),
                }
            )
        return rows_out
    finally:
        try:
            if wb is not None:
                _com_call(lambda: wb.Close(SaveChanges=False), tries=10, sleep_s=0.3)
        except Exception:
            pass
        try:
            _com_call(lambda: excel.Quit(), tries=10, sleep_s=0.3)
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        try:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
def week_monday_iso(d: date) -> str:
    monday = d - timedelta(days=d.weekday())
    return monday.isoformat()
def _parse_any_date_to_date(v: Any) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = safe_str(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    try:
        iso = parse_sheet_date(s)
        return date.fromisoformat(iso)
    except Exception:
        return None
def read_ent_team_tenure_mapping(
    xlsx_path: str,
    sheet_name: str = "Next Week Forecast",
    start_row: int = 2,
    end_row: int = 30,
    start_col_letter: str = "D",
    end_col_letter: str = "J",
    name_col_letter: str = "A",
) -> Tuple[float, Dict[str, float]]:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        ws = wb.active
    else:
        ws = wb[sheet_name]
    sc = column_index_from_string(start_col_letter)
    ec = column_index_from_string(end_col_letter)
    name_c = column_index_from_string(name_col_letter)
    total = 0.0
    per_person: Dict[str, float] = {}
    for r in range(start_row, end_row + 1):
        row_sum = 0.0
        for c in range(sc, ec + 1):
            row_sum += safe_float(ws.cell(row=r, column=c).value)
        total += row_sum
        nm = normalize_ent_name(ws.cell(row=r, column=name_c).value)
        if nm:
            per_person[nm] = row_sum
    return total, per_person
def _ent_cache_path(base_dir: str) -> str:
    return os.path.join(base_dir, "ent_total_available_cache.json")
def get_ent_total_available_for_week(
    mapping_xlsx: str,
    week_monday: str,
    today: Optional[date] = None,
) -> Tuple[float, Dict[str, float], str]:
    if today is None:
        today = date.today()
    base_dir = os.path.dirname(os.path.abspath(os.path.expandvars(mapping_xlsx)))
    cache_file = _ent_cache_path(base_dir)
    cache: Dict[str, Any] = {}
    if os.path.exists(cache_file):
        try:
            with open(cache_file, "r", encoding="utf-8") as f:
                cache = json.load(f) or {}
        except Exception:
            cache = {}
    is_monday = (today.weekday() == 0)
    if is_monday:
        total, per_person = read_ent_team_tenure_mapping(mapping_xlsx)
        cache[week_monday] = {"total": total, "per_person": per_person, "refreshed_on": today.isoformat()}
        try:
            with open(cache_file, "w", encoding="utf-8") as f:
                json.dump(cache, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
        return total, per_person, f"ENT mapping refreshed (Monday) and cached in {os.path.basename(cache_file)}"
    wk = cache.get(week_monday)
    if isinstance(wk, dict) and ("total" in wk) and ("per_person" in wk):
        return safe_float(wk.get("total")), (wk.get("per_person") or {}), f"ENT mapping loaded from cache {os.path.basename(cache_file)}"
    total, per_person = read_ent_team_tenure_mapping(mapping_xlsx)
    cache[week_monday] = {"total": total, "per_person": per_person, "refreshed_on": today.isoformat(), "note": "cache-miss fallback"}
    try:
        with open(cache_file, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    return total, per_person, f"ENT mapping cache miss; read mapping and cached in {os.path.basename(cache_file)}"
def scrape_csv_team_fixed_availability(
    csv_path: str,
    team: str,
    hours_per_person: float = 20.0,
) -> list[dict]:
    if not os.path.exists(csv_path):
        return [{
            "team": team,
            "period_date": "",
            "source_file": csv_path,
            "Total Available Hours": "",
            "Completed Hours": "",
            "Target Output": "",
            "Actual Output": "",
            "Target UPLH": "",
            "Actual UPLH": "",
            "UPLH WP1": "",
            "UPLH WP2": "",
            "HC in WIP": "",
            "Actual HC Used": "",
            "People in WIP": "",
            "Person Hours": "",
            "Outputs by Person": "",
            "Outputs by Cell/Station": "",
            "Cell/Station Hours": "",
            "Hours by Cell/Station - by person": "",
            "Output by Cell/Station - by person": "",
            "UPLH by Cell/Station - by person": "",
            "error": f"Missing file: {os.path.basename(csv_path)}",
        }]
    weekly: Dict[str, Dict[str, Any]] = {}
    with open(csv_path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        _ = next(reader, None)  # header
        for row in reader:
            if not row or len(row) < 8:
                continue
            name = safe_str(row[0])
            d_parsed = _parse_any_date_to_date(row[1])
            if not d_parsed:
                continue
            period_date = week_monday_iso(d_parsed)
            wp1_out = safe_float(row[2])
            wp2_out = safe_float(row[3])
            wp3_out = safe_float(row[4])
            wp1_hrs = safe_float(row[5])
            wp2_hrs = safe_float(row[6])
            wp3_hrs = safe_float(row[7])
            rec = weekly.setdefault(period_date, {
                "wp1_out": 0.0,
                "wp2_out": 0.0,
                "wp3_out": 0.0,
                "wp1_hrs": 0.0,
                "wp2_hrs": 0.0,
                "wp3_hrs": 0.0,
                "by_person": {},
            })
            rec["wp1_out"] += wp1_out
            rec["wp2_out"] += wp2_out
            rec["wp3_out"] += wp3_out
            rec["wp1_hrs"] += wp1_hrs
            rec["wp2_hrs"] += wp2_hrs
            rec["wp3_hrs"] += wp3_hrs
            if name:
                p = rec["by_person"].setdefault(
                    name, {"wp1_out": 0.0, "wp2_out": 0.0, "wp3_out": 0.0, "wp1_hrs": 0.0, "wp2_hrs": 0.0,"wp3_hrs": 0.0}
                )
                p["wp1_out"] += wp1_out
                p["wp2_out"] += wp2_out
                p["wp3_out"] += wp3_out
                p["wp1_hrs"] += wp1_hrs
                p["wp2_hrs"] += wp2_hrs
                p["wp3_hrs"] += wp3_hrs
    rows_out: list[dict] = []
    for period_date in sorted(weekly.keys()):
        agg = weekly[period_date]
        completed_hours = agg["wp1_hrs"] + agg["wp2_hrs"] + agg["wp3_hrs"]
        actual_output = agg["wp1_out"] + agg["wp2_out"]+ agg["wp3_out"]
        active_people = []
        for nm, pdata in (agg["by_person"] or {}).items():
            if (pdata["wp1_out"] + pdata["wp2_out"] + pdata["wp3_out"]) > 0:
                active_people.append(nm)
        hc_in_wip = len(set(active_people))
        total_available_hours = hc_in_wip * float(hours_per_person)
        actual_uplh = safe_div(actual_output, completed_hours)
        uplh_wp1 = safe_div(agg["wp1_out"], agg["wp1_hrs"])
        uplh_wp2 = safe_div(agg["wp2_out"], agg["wp2_hrs"])
        uplh_wp3 = safe_div(agg["wp3_out"], agg["wp3_hrs"])
        actual_hc_used = safe_div(completed_hours, 32.5)
        person_hours: Dict[str, Dict[str, float]] = {}
        for nm in set(active_people):
            pdata = agg["by_person"][nm]
            actual_person = pdata["wp1_hrs"] + pdata["wp2_hrs"] + pdata["wp3_hrs"]
            person_hours[nm] = {"actual": actual_person, "available": float(hours_per_person)}
        outputs_by_person: Dict[str, Dict[str, float]] = {}
        for nm in set(active_people):
            pdata = agg["by_person"][nm]
            out_person = pdata["wp1_out"] + pdata["wp2_out"] + pdata["wp3_out"]
            outputs_by_person[nm] = {"output": out_person, "target": 0.0}
        outputs_by_cell = {
            "WP1": {"output": agg["wp1_out"], "target": 0.0},
            "WP2": {"output": agg["wp2_out"], "target": 0.0},
            "WP3": {"output": agg["wp3_out"], "target": 0.0},
        }
        cell_station_hours = {"WP1": agg["wp1_hrs"], "WP2": agg["wp2_hrs"],  "WP3": agg["wp3_hrs"]}
        hours_by_cell_by_person = {"WP1": {}, "WP2": {}, "WP3": {}}
        for nm in set(active_people):
            pdata = agg["by_person"][nm]
            hours_by_cell_by_person["WP1"][nm] = pdata["wp1_hrs"]
            hours_by_cell_by_person["WP2"][nm] = pdata["wp2_hrs"]
            hours_by_cell_by_person["WP3"][nm] = pdata["wp3_hrs"]
        output_by_cell_by_person = {"WP1": {}, "WP2": {}, "WP3": {}}
        for nm in set(active_people):
            pdata = agg["by_person"][nm]
            output_by_cell_by_person["WP1"][nm] = pdata["wp1_out"]
            output_by_cell_by_person["WP2"][nm] = pdata["wp2_out"]
            output_by_cell_by_person["WP3"][nm] = pdata["wp3_out"]
        uplh_by_cell_by_person: Dict[str, Dict[str, Optional[float]]] = {"WP1": {}, "WP2": {}, "WP3":{}}
        for nm in set(active_people):
            uplh_by_cell_by_person["WP1"][nm] = safe_div(output_by_cell_by_person["WP1"][nm], hours_by_cell_by_person["WP1"][nm])
            uplh_by_cell_by_person["WP2"][nm] = safe_div(output_by_cell_by_person["WP2"][nm], hours_by_cell_by_person["WP2"][nm])
            uplh_by_cell_by_person["WP3"][nm] = safe_div(output_by_cell_by_person["WP3"][nm], hours_by_cell_by_person["WP3"][nm])
        rows_out.append({
            "team": team,
            "period_date": period_date,
            "source_file": os.path.abspath(os.path.expandvars(csv_path)),
            "Total Available Hours": total_available_hours,
            "Completed Hours": completed_hours,
            "Target Output": "",
            "Actual Output": actual_output,
            "Target UPLH": "",
            "Actual UPLH": actual_uplh,
            "UPLH WP1": uplh_wp1,
            "UPLH WP2": uplh_wp2,
            "HC in WIP": hc_in_wip,
            "Actual HC Used": actual_hc_used,
            "People in WIP": "",
            "Person Hours": json.dumps(person_hours, ensure_ascii=False),
            "Outputs by Person": json.dumps(outputs_by_person, ensure_ascii=False),
            "Outputs by Cell/Station": json.dumps(outputs_by_cell, ensure_ascii=False),
            "Cell/Station Hours": json.dumps(cell_station_hours, ensure_ascii=False),
            "Hours by Cell/Station - by person": json.dumps(hours_by_cell_by_person, ensure_ascii=False),
            "Output by Cell/Station - by person": json.dumps(output_by_cell_by_person, ensure_ascii=False),
            "UPLH by Cell/Station - by person": json.dumps(uplh_by_cell_by_person, ensure_ascii=False),
        })
    return rows_out
def scrape_ent_from_csv(
    ent_csv_path: str,
    mapping_xlsx_path: str,
    team: str = "ENT",
) -> list[dict]:
    if not os.path.exists(ent_csv_path):
        return [{
            "team": team,
            "period_date": "",
            "source_file": ent_csv_path,
        }]
    weekly: Dict[str, Dict[str, Any]] = {}
    with open(ent_csv_path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        headers = next(reader, None)  # skip header row
        for row in reader:
            if not row or len(row) < 6:
                continue
            name = normalize_ent_name(row[0])
            d_raw = row[1]
            d_parsed = _parse_any_date_to_date(d_raw)
            if not d_parsed:
                continue
            wk = week_monday_iso(d_parsed)
            wp1_out = safe_float(row[2])
            wp2_out = safe_float(row[3])
            wp1_hrs = safe_float(row[4])
            wp2_hrs = safe_float(row[5])
            rec = weekly.setdefault(wk, {
                "wp1_out": 0.0, "wp2_out": 0.0,
                "wp1_hrs": 0.0, "wp2_hrs": 0.0,
                "by_person": {},
            })
            rec["wp1_out"] += wp1_out
            rec["wp2_out"] += wp2_out
            rec["wp1_hrs"] += wp1_hrs
            rec["wp2_hrs"] += wp2_hrs
            if name:
                p = rec["by_person"].setdefault(
                    name,
                    {"wp1_out": 0.0, "wp2_out": 0.0, "wp1_hrs": 0.0, "wp2_hrs": 0.0}
                )
                p["wp1_out"] += wp1_out
                p["wp2_out"] += wp2_out
                p["wp1_hrs"] += wp1_hrs
                p["wp2_hrs"] += wp2_hrs
    rows_out: list[dict] = []
    for period_date in sorted(weekly.keys()):
        agg = weekly[period_date]
        completed_hours = agg["wp1_hrs"] + agg["wp2_hrs"]
        actual_output = agg["wp1_out"] + agg["wp2_out"]
        taa, per_person_avail, taa_note = get_ent_total_available_for_week(mapping_xlsx_path, period_date)
        actual_uplh = safe_div(actual_output, completed_hours)
        uplh_wp1 = safe_div(agg["wp1_out"], agg["wp1_hrs"])
        uplh_wp2 = safe_div(agg["wp2_out"], agg["wp2_hrs"])
        hc_in_wip = 0
        for nm, pdata in (agg["by_person"] or {}).items():
            if (pdata["wp1_out"] + pdata["wp2_out"]) > 0:
                hc_in_wip += 1
        actual_hc_used = safe_div(completed_hours, 32.5)
        person_hours: Dict[str, Dict[str, float]] = {}
        for nm, pdata in (agg["by_person"] or {}).items():
            actual_person = pdata["wp1_hrs"] + pdata["wp2_hrs"]
            available_person = safe_float(per_person_avail.get(nm, 0.0))
            person_hours[nm] = {"actual": actual_person, "available": available_person}
        outputs_by_person: Dict[str, Dict[str, float]] = {}
        for nm, pdata in (agg["by_person"] or {}).items():
            out_person = pdata["wp1_out"] + pdata["wp2_out"]
            outputs_by_person[nm] = {"output": out_person, "target": 0.0}
        outputs_by_cell = {
            "WP1": {"output": agg["wp1_out"], "target": 0.0},
            "WP2": {"output": agg["wp2_out"], "target": 0.0},
        }
        cell_station_hours = {
            "WP1": agg["wp1_hrs"],
            "WP2": agg["wp2_hrs"],
        }
        hours_by_cell_by_person = {"WP1": {}, "WP2": {}}
        for nm, pdata in (agg["by_person"] or {}).items():
            hours_by_cell_by_person["WP1"][nm] = pdata["wp1_hrs"]
            hours_by_cell_by_person["WP2"][nm] = pdata["wp2_hrs"]
        output_by_cell_by_person = {"WP1": {}, "WP2": {}}
        for nm, pdata in (agg["by_person"] or {}).items():
            output_by_cell_by_person["WP1"][nm] = pdata["wp1_out"]
            output_by_cell_by_person["WP2"][nm] = pdata["wp2_out"]
        uplh_by_cell_by_person: Dict[str, Dict[str, Optional[float]]] = {"WP1": {}, "WP2": {}}
        for nm in (agg["by_person"] or {}).keys():
            uplh_by_cell_by_person["WP1"][nm] = safe_div(output_by_cell_by_person["WP1"][nm], hours_by_cell_by_person["WP1"][nm])
            uplh_by_cell_by_person["WP2"][nm] = safe_div(output_by_cell_by_person["WP2"][nm], hours_by_cell_by_person["WP2"][nm])
        errs = []
        if taa_note:
            errs.append(taa_note)
        rows_out.append({
            "team": team,
            "period_date": period_date,  # ALWAYS Monday
            "source_file": f"{os.path.abspath(os.path.expandvars(ent_csv_path))} | {os.path.abspath(os.path.expandvars(mapping_xlsx_path))}",
            "Total Available Hours": taa,
            "Completed Hours": completed_hours,
            "Target Output": "",  # blank per spec
            "Actual Output": actual_output,
            "Target UPLH": "",    # blank per spec
            "Actual UPLH": actual_uplh,
            "UPLH WP1": uplh_wp1,
            "UPLH WP2": uplh_wp2,
            "HC in WIP": hc_in_wip,
            "Actual HC Used": actual_hc_used,
            "People in WIP": "",  # blank per spec
            "Person Hours": json.dumps(person_hours, ensure_ascii=False),
            "Outputs by Person": json.dumps(outputs_by_person, ensure_ascii=False),
            "Outputs by Cell/Station": json.dumps(outputs_by_cell, ensure_ascii=False),
            "Cell/Station Hours": json.dumps(cell_station_hours, ensure_ascii=False),
            "Hours by Cell/Station - by person": json.dumps(hours_by_cell_by_person, ensure_ascii=False),
            "Output by Cell/Station - by person": json.dumps(output_by_cell_by_person, ensure_ascii=False),
            "UPLH by Cell/Station - by person": json.dumps(uplh_by_cell_by_person, ensure_ascii=False),
        })
    return rows_out
def scrape_spine_previous_weeks_xlsm(
    source_file: str,
    cfg: Dict[str, Any],
    team: str = "Spine",
    dropdown_override: Optional[list[Any]] = None,
) -> list[dict]:
    import shutil
    import tempfile
    import uuid
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    excel.EnableEvents = False
    excel.AutomationSecurity = 3
    wb = None
    tmp_path = None
    rows_out: list[dict] = []
    def _open_via_temp_copy(src_path: str):
        nonlocal tmp_path
        src = os.path.abspath(os.path.expandvars(src_path))
        if not os.path.exists(src):
            raise FileNotFoundError(f"File not found on disk: {src}")
        base = os.path.splitext(os.path.basename(src))[0]
        ext = os.path.splitext(src)[1]
        tmp_path = os.path.join(tempfile.gettempdir(), f"{base}__{uuid.uuid4().hex}{ext}")
        shutil.copy2(src, tmp_path)
        return _com_call(
            lambda: excel.Workbooks.Open(
                tmp_path,
                UpdateLinks=0,
                ReadOnly=True,
                IgnoreReadOnlyRecommended=True,
                Notify=False,
                AddToMru=False,
                CorruptLoad=0,
            )
        )
    try:
        wb = _open_via_temp_copy(source_file)
        ws = _com_call(lambda: wb.Worksheets("Previous Weeks"))
        try:
            _com_call(lambda: wb.RefreshAll())
        except Exception:
            pass
        try:
            _com_call(lambda: excel.CalculateFullRebuild())
        except Exception:
            _com_call(lambda: excel.Calculate())
        time.sleep(5)
        dd = _com_call(lambda: ws.Range("A2"))
        dropdown_values = dropdown_override if dropdown_override is not None else _get_dropdown_values_from_validation(dd)
        seen = set()
        dropdown_values = [v for v in dropdown_values if not (safe_str(v) in seen or seen.add(safe_str(v)))]
        cols = _excel_col_range("B", "T")
        min_pd = safe_str(cfg.get("min_period_date")) or "2026-03-02"
        max_pd = safe_str(cfg.get("max_period_date")) or date.today().isoformat()
        for choice in dropdown_values:
            _com_call(lambda: setattr(dd, "Value", choice))
            _com_call(lambda: excel.Calculate())
            period_date = _as_iso_date(_com_call(lambda: dd.Value))
            if not period_date:
                continue
            if period_date < min_pd:
                continue
            if period_date > max_pd:
                continue
            total_available_hours = safe_float(
                _com_call(lambda: ws.Range(cfg["cells"]["total_available_hours"]).Value)
            )
            completed_hours = safe_float(
                _com_call(lambda: ws.Range(cfg["cells"]["completed_hours"]).Value)
            )
            wp1_tgt = safe_float(
                _com_call(lambda: ws.Range(cfg["cells"]["wp1_target"]).Value)
            )
            wp2_tgt = safe_float(
                _com_call(lambda: ws.Range(cfg["cells"]["wp2_target"]).Value)
            )
            wp1_out = safe_float(
                _com_call(lambda: ws.Range(cfg["cells"]["wp1_output"]).Value)
            )
            wp2_out = safe_float(
                _com_call(lambda: ws.Range(cfg["cells"]["wp2_output"]).Value)
            )
            target_output = wp1_tgt + wp2_tgt
            actual_output = wp1_out + wp2_out
            if target_output < 0:
                continue
            target_uplh = safe_div(target_output, completed_hours)
            actual_uplh = safe_div(actual_output, completed_hours)
            uplh_wp1 = safe_float(
                _com_call(lambda: ws.Range(cfg["cells"]["uplh_wp1"]).Value)
            )
            uplh_wp2 = safe_float(
                _com_call(lambda: ws.Range(cfg["cells"]["uplh_wp2"]).Value)
            )
            hc_in_wip = 0
            for c in cols:
                if safe_float(_com_call(lambda c=c: ws.Cells(28, c).Value)) != 0.0:
                    hc_in_wip += 1
            actual_hc_used = safe_div(completed_hours, 32.5)
            person_hours: Dict[str, Dict[str, float]] = {}
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(33, c).Value))
                if not name:
                    continue
                actual = safe_float(_com_call(lambda c=c: ws.Cells(54, c).Value))
                available = safe_float(_com_call(lambda c=c: ws.Cells(64, c).Value))
                person_hours[name] = {"actual": actual, "available": available}
            outputs_by_person: Dict[str, Dict[str, float]] = {}
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(13, c).Value))
                if not name:
                    continue
                out_val = sum(
                    safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value))
                    for r in range(14, 28)
                )
                tgt_val = safe_float(_com_call(lambda c=c: ws.Cells(28, c).Value))
                if out_val != 0.0 or tgt_val != 0.0:
                    outputs_by_person[name] = {"output": out_val, "target": tgt_val}
            outputs_by_cell = {
                "WP1": {"output": wp1_out, "target": wp1_tgt},
                "WP2": {"output": wp2_out, "target": wp2_tgt},
            }
            cell_station_hours = {
                "WP1": safe_float(_com_call(lambda: ws.Range("AD7").Value)),
                "WP2": safe_float(_com_call(lambda: ws.Range("AF7").Value)),
            }
            hours_by_cell_by_person = {"WP1": {}, "WP2": {}}
            wp1_hour_rows = [34, 38, 42, 46, 50]
            wp2_hour_rows = [35, 39, 43, 47, 51]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(33, c).Value))
                if not name:
                    continue
                wp1_hrs = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp1_hour_rows)
                wp2_hrs = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp2_hour_rows)
                if wp1_hrs != 0.0:
                    hours_by_cell_by_person["WP1"][name] = wp1_hrs
                if wp2_hrs != 0.0:
                    hours_by_cell_by_person["WP2"][name] = wp2_hrs
            output_by_cell_by_person = {"WP1": {}, "WP2": {}}
            wp1_out_rows = [14, 17, 20, 23, 26]
            wp2_out_rows = [15, 18, 21, 24, 27]
            for c in cols:
                name = safe_str(_com_call(lambda c=c: ws.Cells(13, c).Value))
                if not name:
                    continue
                wp1_o = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp1_out_rows)
                wp2_o = sum(safe_float(_com_call(lambda r=r, c=c: ws.Cells(r, c).Value)) for r in wp2_out_rows)
                if wp1_o != 0.0:
                    output_by_cell_by_person["WP1"][name] = wp1_o
                if wp2_o != 0.0:
                    output_by_cell_by_person["WP2"][name] = wp2_o
            uplh_by_cell_by_person: Dict[str, Dict[str, Optional[float]]] = {"WP1": {}, "WP2": {}}
            for wp in ("WP1", "WP2"):
                for person, out_val in output_by_cell_by_person[wp].items():
                    hrs = safe_float(hours_by_cell_by_person[wp].get(person, 0.0))
                    uplh_by_cell_by_person[wp][person] = safe_div(out_val, hrs)
            rows_out.append({
                "team": team,
                "period_date": period_date,
                "source_file": os.path.abspath(os.path.expandvars(source_file)),
                "Total Available Hours": total_available_hours,
                "Completed Hours": completed_hours,
                "Target Output": target_output,
                "Actual Output": actual_output,
                "Target UPLH": target_uplh,
                "Actual UPLH": actual_uplh,
                "UPLH WP1": uplh_wp1,
                "UPLH WP2": uplh_wp2,
                "HC in WIP": hc_in_wip,
                "Actual HC Used": actual_hc_used,
                "People in WIP": "",
                "Person Hours": json.dumps(person_hours, ensure_ascii=False),
                "Outputs by Person": json.dumps(outputs_by_person, ensure_ascii=False),
                "Outputs by Cell/Station": json.dumps(outputs_by_cell, ensure_ascii=False),
                "Cell/Station Hours": json.dumps(cell_station_hours, ensure_ascii=False),
                "Hours by Cell/Station - by person": json.dumps(hours_by_cell_by_person, ensure_ascii=False),
                "Output by Cell/Station - by person": json.dumps(output_by_cell_by_person, ensure_ascii=False),
                "UPLH by Cell/Station - by person": json.dumps(uplh_by_cell_by_person, ensure_ascii=False),
            })
        return rows_out
    finally:
        try:
            if wb is not None:
                _com_call(lambda: wb.Close(SaveChanges=False), tries=10, sleep_s=0.3)
        except Exception:
            pass
        try:
            _com_call(lambda: excel.Quit(), tries=10, sleep_s=0.3)
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        try:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
def filter_rows_before(rows: list[dict], cutoff_iso: str) -> list[dict]:
    return [r for r in rows if safe_str(r.get("period_date")) < cutoff_iso]
def dedupe_rows_by_team_period(rows: list[dict]) -> list[dict]:
    by_key: dict[tuple[str, str], dict] = {}
    for r in rows:
        key = (safe_str(r.get("team")), safe_str(r.get("period_date")))
        if key[0] and key[1]:
            by_key[key] = r
    out = list(by_key.values())
    def _sort_key(r: dict) -> tuple[str, str]:
        return (
            safe_str(r.get("team")).lower(),
            safe_str(r.get("period_date")),
        )
    out.sort(key=_sort_key)
    return out
def filter_rows_on_or_after(rows: list[dict], cutoff_iso: str) -> list[dict]:
    return [r for r in rows if safe_str(r.get("period_date")) >= cutoff_iso]
def _looks_like_iso_date(s: str) -> bool:
    s = safe_str(s)
    return (len(s) == 10 and s[4] == "-" and s[7] == "-")
def _is_monday_iso(s: str) -> bool:
    s = safe_str(s)
    if not _looks_like_iso_date(s):
        return False
    try:
        d = date.fromisoformat(s)
        return d.weekday() == 0
    except Exception:
        return False
def read_csv_rows(path: str) -> list[dict]:
    if not os.path.exists(path):
        return []
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))
def append_missing_placeholders_from_wip(
    wip_csv_path: str,
    logger: Optional[logging.Logger] = None,
) -> None:
    wip_rows = read_csv_rows(wip_csv_path)
    wip_keys: set[tuple[str, str]] = set()
    for r in wip_rows:
        team = safe_str(r.get("team"))
        pd = safe_str(r.get("period_date"))
        if not team or not pd:
            continue
        if _is_monday_iso(pd):
            wip_keys.add((team, pd))
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--team", default="all", help="Team to run (or 'all'). Example: --team PH")
    args = parser.parse_args()
    logger = setup_logging()
    logger.info("=== NS Metrics Run START ===")
    logger.info(f"Selected team: {args.team}")
    ph_source_file = r"C:\Users\wadec8\Medtronic PLC\Customer Quality Pelvic Health - Daily Tracker\PH Cell Heijunka.xlsx"
    meic_source_file = r"C:\Users\wadec8\Medtronic PLC\Customer Quality Pelvic Health - Daily Tracker\MEIC\New MEIC PH Heijunka.xlsx"
    scs_source_file = r"C:\Users\wadec8\Medtronic PLC\Customer Quality SCS - Cell 17\Cell 1 - Heijunka.xlsx"
    scs_super_source_file = r"C:\Users\wadec8\Medtronic PLC\Customer Quality SCS - SCS Super Cell\Super Cell Heijunka.xlsx"
    cos_source_file = r"C:\Users\wadec8\Medtronic PLC\COS Cell - Documents\Heijunka v2 TDD.xlsx"
    spine_source_file = r"C:\Users\wadec8\Medtronic PLC\MEIC - RTG - Documents\Heijunka - RTG Spine CQ_V2.0_Apr2023.xlsx"
    spine_new_source_file = r"C:\Users\wadec8\Medtronic PLC\MEIC - RTG - Documents\Spine_Heijunka.xlsm"
    nv_source_file = r"C:\Users\wadec8\Medtronic PLC\RTG Customer Quality Neurovascular - Documents\Cell\NV_Heijunka.xlsm"
    dbs_c13_source_file = r"C:\Users\wadec8\Medtronic PLC\DBS CQ Team - Documents\Cell 13 Heijunka V2.xlsx"
    dbs_c14_source_file = r"C:\Users\wadec8\Medtronic PLC\DBS CQ Team - Documents\Cell 14 Heijunka V2.xlsx"
    nav_source_file = r"C:\Users\wadec8\Medtronic PLC\MNAV Sharepoint - Navigation Work Reports\Heijunka_MNAV_Ranges_May2025.xlsm"
    ae_meic_source_file = r"C:\Users\wadec8\Medtronic PLC\MNAV Sharepoint - MEIC AE + OARM\AE_MEIC_Heijunka.xlsm"
    oarm_meic_source_file = r"C:\Users\wadec8\Medtronic PLC\MNAV Sharepoint - MEIC AE + OARM\OARM_MEIC_Heijunka.xlsm"
    mazor_source_file = r"C:\Users\wadec8\Medtronic PLC\MNAV Sharepoint - Caesarea Team\CAE - Heijunka_v2.xlsm"
    csf_source_file   = r"C:\Users\wadec8\Medtronic PLC\CQ CSF Management - Documents\CSF_Heijunka.xlsm"
    pss_source_file   = r"C:\Users\wadec8\Medtronic PLC\PSS Sharepoint - Documents\PSS Team Heijunka Tool.xlsm"
    ent_mapping_xlsx = r"C:\Users\wadec8\Medtronic PLC\ENT GEMBA Board - Heijunka 2.0 Files\Team & Tenure Mapping.xlsx"
    ent_data_csv     = r"C:\Users\wadec8\OneDrive - Medtronic PLC\ENT\ENT_Data.csv"
    dbs_meic_csv = r"C:\Users\wadec8\OneDrive - Medtronic PLC\DBS\DBS_Data.csv"
    other_meic_csv = r"C:\Users\wadec8\OneDrive - Medtronic PLC\Other MEIC\Other_MEIC_Data.csv"
    ph_cell17_source_file = r"C:\Users\wadec8\Medtronic PLC\Customer Quality Pelvic Health - Cell 17\Cell 17 New Heijunka.xlsx"
    out_file = "NS_DATA\\NS_metrics.csv"
    if not os.path.exists(ph_source_file):
        raise FileNotFoundError(f"Input file not found: {ph_source_file}")
    if not os.path.exists(meic_source_file):
        raise FileNotFoundError(f"Input file not found: {meic_source_file}")
    if not os.path.exists(scs_source_file):
        raise FileNotFoundError(f"Input file not found: {scs_source_file}")
    SPINE_CFG = {
        "team": "Spine",
        "person_cols": ("B", "O"),
        "date_parser": parse_sheet_date_day_first_requires_year,
        "min_period_date": "2025-06-30",
        "max_period_date": "2026-02-23",
        "cells": {
            "total_available_hours": "P54",
            "completed_hours": {"type": "sum_cells", "cells": ["Q4", "S4"]},
            "wp1_output": "Q2",
            "wp1_target": "Q7",
            "wp2_output": "S2",
            "wp2_target": "S7",
            "uplh_wp1": "Q5",
            "uplh_wp2": "S5",
            "wp1_hours": "Q4",
            "wp2_hours": "S4",
        },
        "rows": {
            "hc_row": 25,
            "person_name_row_for_person_hours": 30,
            "person_actual_row_for_person_hours": 45,
            "person_available_row_for_person_hours": 54,
            "person_name_row_for_outputs_by_person": 10,
            "person_target_row_for_outputs_by_person": 25,
            "person_name_row_for_hours_by_cell_by_person": 30,
            "wp1_hour_rows": [31, 34, 37, 40, 43],
            "wp2_hour_rows": [32, 35, 38, 41, 44],
            "person_name_row_for_output_by_cell_by_person": 10,
            "wp1_output_rows_by_person": [11, 14, 17, 20, 23],
            "wp2_output_rows_by_person": [12, 15, 18, 21, 24],
        },
        "outputs_by_person_output": {"type": "sum_rows", "rows": list(range(11, 25))},
    }
    SPINE_NEW_CFG = {
        "team": "Spine",
        "person_cols": ("B", "Q"),
        "date_parser": parse_sheet_date_day_first_requires_year,
        "min_period_date": "2026-02-24",
        "cells": {
            "total_available_hours": "W64",
            "completed_hours": "W54",
            "wp1_output": "AD5",
            "wp1_target": "AD9",
            "wp2_output": "AF5",
            "wp2_target": "AF9",
            "uplh_wp1": "AD8",
            "uplh_wp2": "AF8",
            "wp1_hours": "AD7",
            "wp2_hours": "AF7",
        },
        "rows": {
            "hc_row": 28,
            "person_name_row_for_person_hours": 33,
            "person_actual_row_for_person_hours": 54,
            "person_available_row_for_person_hours": 64,
            "person_name_row_for_outputs_by_person": 13,
            "person_target_row_for_outputs_by_person": 28,
            "person_name_row_for_hours_by_cell_by_person": 33,
            "wp1_hour_rows": [34, 38, 42, 46, 50],
            "wp2_hour_rows": [35, 39, 43, 47, 51],
            "person_name_row_for_output_by_cell_by_person": 13,
            "wp1_output_rows_by_person": [14, 17, 20, 23, 26],
            "wp2_output_rows_by_person": [15, 18, 21, 24, 27],
        },
        "outputs_by_person_output": {"type": "sum_rows", "rows": list(range(14, 27))},
    }
    MAZOR_CFG = {
        "person_cols": ("B", "J"),
        "cells": {
            "total_available_hours": "R64",
            "completed_hours": "R54",
            "wp1_target": "X10",
            "wp2_target": "Z10",
            "wp1_output": "X5",
            "wp2_output": "Z5",
            "uplh_wp1": "X8",
            "uplh_wp2": "Z8",
            "wp1_hours": "X7",
            "wp2_hours": "Z7",
        },
        "rows": {
            "person_name_row_for_person_hours": 33,
            "person_actual_row_for_person_hours": 54,
            "person_available_row_for_person_hours": 64,
            "person_name_row_for_outputs_by_person": 13,
            "person_target_row_for_outputs_by_person": 28,
            "person_name_row_for_hours_by_cell_by_person": 33,
            "wp1_hour_rows": [34, 38, 42, 46, 50],
            "wp2_hour_rows": [35, 39, 43, 47, 51],
            "person_name_row_for_output_by_cell_by_person": 13,
            "wp1_output_rows_by_person": [14, 17, 20, 23, 26],
            "wp2_output_rows_by_person": [15, 18, 21, 24, 27],
        },
    }
    CSF_CFG = {
        "person_cols": ("B", "G"),
        "cells": {
            "total_available_hours": "I69",
            "completed_hours": "I59",  # per your spec
            "wp1_target": "N10",
            "wp2_target": "P10",
            "wp1_output": "N5",
            "wp2_output": "P5",
            "uplh_wp1": "N8",
            "uplh_wp2": "P8",
            "wp1_hours": "N7",
            "wp2_hours": "P7",
        },
        "rows": {
            "person_name_row_for_person_hours": 33,
            "person_actual_row_for_person_hours": 59,
            "person_available_row_for_person_hours": 69,
            "person_name_row_for_outputs_by_person": 13,
            "person_target_row_for_outputs_by_person": 28,
            "person_name_row_for_hours_by_cell_by_person": 33,
            "wp1_hour_rows": [34, 39, 44, 49, 54],
            "wp2_hour_rows": [35, 40, 45, 50, 55],
            "person_name_row_for_output_by_cell_by_person": 13,
            "wp1_output_rows_by_person": [14, 17, 20, 23, 26],
            "wp2_output_rows_by_person": [15, 18, 21, 24, 27],
        },
    }
    PSS_CFG = {
        "team": "PSS",
        "min_period_date": "2026-05-11",
        "person_cols": ("B", "BF"),
        "cells": {
            "total_available_hours": "BG64",
            "completed_hours": "BG54",
            "wp1_target": "BN10",
            "wp2_target": "BP10",
            "wp1_output": "BN5",
            "wp2_output": "BP5",
            "uplh_wp1": "BN8",
            "uplh_wp2": "BP8",
            "wp1_hours": "BN7",
            "wp2_hours": "BP7",
        },
        "rows": {
            "hc_row": 28,
            "person_name_row_for_person_hours": 33,
            "person_actual_row_for_person_hours": 54,
            "person_available_row_for_person_hours": 64,
            "person_name_row_for_outputs_by_person": 66,
            "person_target_row_for_outputs_by_person": 28,
            "person_name_row_for_hours_by_cell_by_person": 33,
            "wp1_hour_rows": [34, 38, 42, 46, 50],
            "wp2_hour_rows": [35, 39, 43, 47, 51],
            "person_name_row_for_output_by_cell_by_person": 66,
            "wp1_output_rows_by_person": [67, 69, 71, 73, 75],
            "wp2_output_rows_by_person": [68, 70, 72, 74, 76],
        },
        "outputs_by_person_output": {"type": "sum_rows", "rows": list(range(67, 77))},
    }
    PH_NEW_CFG = {
        "team": "PH",
        "person_cols": ("B", "R"),
        "min_period_date": "2026-03-23",
        "cells": {
            "total_available_hours": "T61",
            "completed_hours": "T50",
            "wp1_output": "Z2",
            "wp1_target": "Z7",
            "wp2_output": "AB2",
            "wp2_target": "AB7",
            "uplh_wp1": "Z5",
            "uplh_wp2": "AB5",
            "wp1_hours": "Z4",
            "wp2_hours": "AB4",
            "wp3_hours_sum_cells": ["T33", "T37", "T41", "T45", "T49"],
        },
        "rows": {
            "hc_row": 50,
            "person_name_row_for_person_hours": 55,
            "person_actual_row_for_person_hours": 50,
            "person_available_row_for_person_hours": 61,
            "person_name_row_for_outputs_by_person": 55,
            "person_target_row_for_outputs_by_person": 25,
            "person_name_row_for_hours_by_cell_by_person": 30,
            "wp1_hour_rows": [31, 35, 39, 43, 47],
            "wp2_hour_rows": [32, 36, 40, 44, 48],
            "wp3_hour_rows": [33, 37, 41, 45, 49],
            "person_name_row_for_output_by_cell_by_person": 10,
            "wp1_output_rows_by_person": [11, 14, 17, 20, 23],
            "wp2_output_rows_by_person": [12, 15, 18, 21, 24],
        },
        "outputs_by_person_output": {"type": "sum_rows", "rows": list(range(11, 25))},
    }
    PH_CELL17_CFG = {
        "team": "PH Cell 17",
        "person_cols": ("B", "L"),
        "date_parser": parse_sheet_date_requires_year,
        "min_period_date": "2025-09-01",
        "cells": {
            "total_available_hours": "N113",
            "completed_hours": "N50",
            "wp1_output": "T2",
            "wp1_target": "T7",
            "wp2_output": "U2",
            "wp2_target": "U7",
            "uplh_wp1": "T5",
            "uplh_wp2": "U5",
            "wp1_hours": "T4",
            "wp2_hours": "U4",
            "wp3_hours_sum_cells": ["M33", "M37", "M41", "M45", "M49"],
        },
        "rows": {
            "hc_row": 76,
            "person_name_row_for_person_hours": 30,
            "person_actual_row_for_person_hours": 50,
            "person_available_row_for_person_hours": 113,
            "person_name_row_for_outputs_by_person": 10,
            "person_target_row_for_outputs_by_person": 25,
            "person_name_row_for_hours_by_cell_by_person": 30,
            "wp1_hour_rows": [31, 35, 39, 43, 47],
            "wp2_hour_rows": [32, 36, 40, 44, 48],
            "wp3_hour_rows": [33, 34, 41, 45, 49],
            "person_name_row_for_output_by_cell_by_person": 10,
            "wp1_output_rows_by_person": [56, 60, 64, 68, 72],
            "wp2_output_rows_by_person": [57, 61, 65, 69, 73],
        },
        "outputs_by_person_output": {"type": "sum_rows", "rows": list(range(56, 74))},
    }
    MEIC_PH_CFG = {
        "team": "MEIC PH",
        "person_cols": ("B", "Q"),
        "cells": {
            "total_available_hours": "S111",
            "completed_hours": "S50",
            "wp1_output": "Y2",
            "wp1_target": "Y7",
            "wp2_output": "AA2",
            "wp2_target": "AA7",
            "uplh_wp1": "Y5",
            "uplh_wp2": "AA5",
            "wp1_hours": "Y4",
            "wp2_hours": "AA4",
        },
        "rows": {
            "hc_row": 50,
            "person_name_row_for_person_hours": 30,
            "person_actual_row_for_person_hours": 50,
            "person_available_row_for_person_hours": 111,
            "person_name_row_for_outputs_by_person": 30,
            "person_target_row_for_outputs_by_person": 25,
            "person_name_row_for_hours_by_cell_by_person": 30,
            "wp1_hour_rows": [31, 35, 39, 43, 47],
            "wp2_hour_rows": [32, 36, 40, 44, 48],
            "person_name_row_for_output_by_cell_by_person": 53,
            "wp1_output_rows_by_person": [54, 58, 62, 66, 70],
            "wp2_output_rows_by_person": [55, 57, 63, 67, 71],
        },
        "outputs_by_person_output": {"type": "row", "row": 73},
    }
    SCS_CELL1_OLD_CFG = {
        "team": "SCS Cell 1",
        "person_cols": ("B", "R"),
        "date_parser": parse_sheet_date_scs_missing_year,
        "max_period_date": "2026-03-23",
        "cells": {
            "total_available_hours": "S111",
            "completed_hours": {"type": "sum_cells", "cells": ["T4", "V4"]},
            "wp1_output": "T2",
            "wp1_target": "T7",
            "wp2_output": "V2",
            "wp2_target": "V7",
            "uplh_wp1": "T5",
            "uplh_wp2": "V5",
            "wp1_hours": "T4",
            "wp2_hours": "V4",
        },
        "rows": {
            "hc_row": 25,
            "person_name_row_for_person_hours": 30,
            "person_actual_row_for_person_hours": 50,
            "person_available_row_for_person_hours": 111,
            "person_name_row_for_outputs_by_person": 53,
            "person_target_row_for_outputs_by_person": 25,
            "person_name_row_for_hours_by_cell_by_person": 30,
            "wp1_hour_rows": [31, 35, 39, 43, 47],
            "wp2_hour_rows": [32, 36, 40, 44, 48],
            "person_name_row_for_output_by_cell_by_person": 10,
            "wp1_output_rows_by_person": [54, 58, 62, 66, 70],
            "wp2_output_rows_by_person": [55, 59, 63, 67, 71],
        },
        "outputs_by_person_output": {"type": "row", "row": 73},
    }
    SCS_CELL1_NEW_CFG = {
        "team": "SCS Cell 1",
        "person_cols": ("B", "L"),
        "date_parser": parse_sheet_date_scs_missing_year,
        "min_period_date": "2026-03-30",
        "cells": {
            "total_available_hours": "T61",
            "completed_hours": "T50",
            "wp1_output": "Z2",
            "wp1_target": "Z7",
            "wp2_output": "AB2",
            "wp2_target": "AB7",
            "uplh_wp1": "Z5",
            "uplh_wp2": "AB5",
            "wp1_hours": "Z4",
            "wp2_hours": "AB4",
            "wp3_hours_sum_cells": ["T33", "T37", "T41", "T45", "T49"],
        },
        "rows": {
            "hc_row": 25,
            "person_name_row_for_person_hours": 30,
            "person_actual_row_for_person_hours": 50,
            "person_available_row_for_person_hours": 61,
            "person_name_row_for_outputs_by_person": 10,
            "person_target_row_for_outputs_by_person": 25,
            "person_name_row_for_hours_by_cell_by_person": 30,
            "wp1_hour_rows": [31, 35, 39, 43, 47],
            "wp2_hour_rows": [32, 36, 40, 44, 48],
            "wp3_hour_rows": [33, 37, 41, 45, 49],
            "person_name_row_for_output_by_cell_by_person": 10,
            "wp1_output_rows_by_person": [11, 14, 17, 20, 23],
            "wp2_output_rows_by_person": [12, 15, 18, 21, 24],
        },
        "outputs_by_person_output": {"type": "row", "row": 25},
    }
    SCS_SUPER_OLD_CFG = {
        "team": "SCS Super Cell",
        "person_cols": ("B", "V"),
        "date_parser": parse_sheet_date_scs_missing_year, 
        "max_period_date": "2026-02-23",
        "cells": {
            "total_available_hours": {"type": "sum_range", "range": "B60:V60"},
            "completed_hours": {"type": "sum_range", "range": "B60:V60"},
            "wp1_output": "AE2",
            "wp1_target": "AE7",
            "wp2_output": "AG2",
            "wp2_target": "AG7",
            "uplh_wp1": "AE5",
            "uplh_wp2": "AG5",
            "wp1_hours": "AE4",
            "wp2_hours": "AG4",
        },
        "rows": {
            "hc_row": 25,
            "person_name_row_for_person_hours": 30,
            "person_actual_row_for_person_hours": 50,
            "person_available_row_for_person_hours": 60,  # available from row 60
            "person_name_row_for_outputs_by_person": 10,
            "person_target_row_for_outputs_by_person": 25,
            "person_name_row_for_hours_by_cell_by_person": 30,
            "wp1_hour_rows": [31, 35, 39, 43, 47],
            "wp2_hour_rows": [32, 36, 40, 44, 48],
            "person_name_row_for_output_by_cell_by_person": 10,
            "wp1_output_rows_by_person": [11, 14, 17, 20, 23],
            "wp2_output_rows_by_person": [12, 15, 18, 21, 24],
        },
        "outputs_by_person_output": {"type": "sum_rows", "rows": list(range(11, 25))},
    }
    SCS_SUPER_NEW_CFG = {
        "team": "SCS Super Cell",
        "person_cols": ("B", "Y"),
        "date_parser": parse_sheet_date_scs_missing_year, 
        "min_period_date": "2026-03-01",
        "cells": {
            "total_available_hours": {"type": "sum_range", "range": "B61:Y61"},
            "completed_hours": {"type": "sum_range", "range": "B50:Y50"},
            "wp1_output": "AH2",
            "wp1_target": "AH7",
            "wp2_output": "AJ2",
            "wp2_target": "AJ7",
            "uplh_wp1": "AH5",
            "uplh_wp2": "AJ5",
            "wp1_hours": "AH4",
            "wp2_hours": "AJ4",
            "wp3_hours_sum_cells": ["Z33", "Z37", "Z41", "Z45", "Z49"],
        },
        "rows": {
            "hc_row": 25,
            "person_name_row_for_person_hours": 30,
            "person_actual_row_for_person_hours": 50,
            "person_available_row_for_person_hours": 61,
            "person_name_row_for_outputs_by_person": 10,
            "person_target_row_for_outputs_by_person": 25,
            "person_name_row_for_hours_by_cell_by_person": 30,
            "wp1_hour_rows": [31, 35, 39, 43, 47],
            "wp2_hour_rows": [32, 36, 40, 44, 48],
            "wp3_hour_rows": [33, 37, 41, 45, 49],
            "person_name_row_for_output_by_cell_by_person": 10,
            "wp1_output_rows_by_person": [11, 14, 17, 20, 23],
            "wp2_output_rows_by_person": [12, 15, 18, 21, 24],
        },
        "outputs_by_person_output": {"type": "sum_rows", "rows": list(range(11, 25))},
    }
    TDD_COS1_OLD_CFG = {
        "team": "TDD COS 1",
        "person_cols": ("B", "T"),
        "date_parser": parse_sheet_date_scs_missing_year,
        "cells": {
            "total_available_hours": "V65",
            "completed_hours": "U55",
            "wp1_output": "AB2",
            "wp1_target": "AB7",
            "wp2_output": "AD2",
            "wp2_target": "AD7",
            "uplh_wp1": "AB5",
            "uplh_wp2": "AD5",
            "wp1_hours": "AB4",
            "wp2_hours": "AD4",
            "wp3_hours_sum_cells": ["U33", "U38", "U43", "U48", "U53"],
        },
        "rows": {
            "hc_row": 25,
            "person_name_row_for_person_hours": 30,
            "person_actual_row_for_person_hours": 55,
            "person_available_row_for_person_hours": 64,
            "person_name_row_for_outputs_by_person": 10,
            "person_target_row_for_outputs_by_person": 25,
            "person_name_row_for_hours_by_cell_by_person": 30,
            "wp1_hour_rows": [31, 36, 41, 46, 51],
            "wp2_hour_rows": [32, 37, 42, 47, 52],
            "wp3_hour_rows": [33, 38, 43, 48, 53],
            "person_name_row_for_output_by_cell_by_person": 10,
            "wp1_output_rows_by_person": [11, 14, 17, 20, 23],
            "wp2_output_rows_by_person": [12, 15, 18, 21, 24],
        },
        "outputs_by_person_output": {"type": "sum_rows", "rows": list(range(11, 25))},
    }
    TDD_COS1_NEW_CFG = {
        **TDD_COS1_OLD_CFG,
        "min_period_date": "2026-05-11",
        "person_cols": ("B", "U"),
        "cells": {
            **TDD_COS1_OLD_CFG["cells"],
            "total_available_hours": "U65",
            "completed_hours": "T55",
            "wp1_hours": "AA4",
            "wp2_hours": "AC4",
            "wp3_hours_sum_cells": ["T33", "T38", "T43", "T48", "T53"],
        },
    }
    TDD_COS1_2026_06_CFG = {
        **TDD_COS1_NEW_CFG,
        "min_period_date": "2026-06-01",
        "person_cols": ("B", "R"),
        "cells": {
            **TDD_COS1_NEW_CFG["cells"],
            "total_available_hours": "T65",
            "completed_hours": "S55",
            "wp3_hours_sum_cells": ["S33", "S38", "S43", "S48", "S53"],
        },
    }
    rows: list[dict] = []
    selected_team = safe_str(args.team).lower()
    def should_run(team_name: str) -> bool:
        return selected_team in ("all", "", team_name.lower())
    def should_run_pss() -> bool:
        return selected_team in ("all", "", "pss", "pss us", "pss meic", "pss meic intern")
    def extend_team(team_name: str, fn):
        out = run_team(logger, team_name, fn)   # logs START/DONE/FAIL + rows + elapsed
        rows.extend(out)
        return out
    def mondays_since(start_iso: str, end_d: date) -> list[str]:
        start = date.fromisoformat(start_iso)
        start = start - timedelta(days=start.weekday()) 
        out: list[str] = []
        d = start
        while d <= end_d:
            out.append(d.isoformat())
            d += timedelta(days=7)
        return out
    ALL_MONDAYS_SINCE_2025_06_02 = mondays_since("2025-06-02", date.today())
    if should_run("TDD COS 1"):
        cos_rows = run_team(
            logger,
            "TDD COS 1",
            lambda: load_tdd_cos1_from_existing_metrics_and_refresh_3_weeks(
                out_file,
                cos_source_file,
                TDD_COS1_OLD_CFG,
                logger=logger,
                cos_new_cfg=TDD_COS1_NEW_CFG,
                cos_newer_cfg=TDD_COS1_2026_06_CFG,
            )
        )
        rows.extend(cos_rows)
    if should_run("Spine"):
        spine_rows = run_team(
            logger,
            "Spine",
            lambda: load_spine_from_existing_metrics_and_refresh_3_weeks(
                out_file,
                spine_source_file,
                spine_new_source_file,
                SPINE_CFG,
                SPINE_NEW_CFG,
                logger=logger,
            )
        )
        rows.extend(spine_rows)
    if should_run("PH"):
        ph_rows = run_team(
            logger,
            "PH",
            lambda: load_ph_from_existing_metrics_and_refresh_3_weeks(
                out_file,
                ph_source_file,
                PH_NEW_CFG,
                logger=logger,
            )
        )
        rows.extend(ph_rows)
    if should_run("PH Cell 17"):
        extend_team("PH Cell 17", lambda: scrape_workbook_with_config(ph_cell17_source_file, PH_CELL17_CFG))
    if should_run("SCS Cell 1"):
        scs_cell1_rows = run_team(
            logger,
            "SCS Cell 1",
            lambda: load_scs_cell1_from_existing_metrics_and_refresh_3_weeks(
                out_file,
                scs_source_file,
                SCS_CELL1_OLD_CFG,
                SCS_CELL1_NEW_CFG,
                logger=logger,
            )
        )
        rows.extend(scs_cell1_rows)
    if should_run("MEIC PH"):
        meic_rows = run_team(
            logger,
            "MEIC PH",
            lambda: load_meic_ph_from_existing_metrics_and_refresh_3_weeks(
                out_file,
                meic_source_file,
                MEIC_PH_CFG,
                logger=logger,
            )
        )
        rows.extend(meic_rows)
    cutoff_dbs = "2025-07-07"
    dbs_c13_rows = run_team(
        logger,
        "DBS C13",
        lambda: scrape_dbs_dated_tabs_xlsx(
            dbs_c13_source_file,
            "DBS C13",
            min_period_date="2025-06-02",
        ),
    )
    before = len(dbs_c13_rows)
    dbs_c13_rows = filter_rows_on_or_after(dbs_c13_rows, cutoff_dbs)
    logger.info(f"[DBS C13] filter >= {cutoff_dbs}: {before} -> {len(dbs_c13_rows)}")
    rows.extend(dbs_c13_rows)
    dbs_c14_rows = run_team(
        logger,
        "DBS C14",
        lambda: scrape_dbs_dated_tabs_xlsx(
            dbs_c14_source_file,
            "DBS C14",
            min_period_date="2025-06-02",
        ),
    )
    before = len(dbs_c14_rows)
    dbs_c14_rows = filter_rows_on_or_after(dbs_c14_rows, cutoff_dbs)
    logger.info(f"[DBS C14] filter >= {cutoff_dbs}: {before} -> {len(dbs_c14_rows)}")
    rows.extend(dbs_c14_rows)
    if should_run("NV"):
        nv_rows = run_team(
            logger,
            "NV",
            lambda: load_nv_from_existing_metrics_and_refresh_3_weeks(
                out_file,
                nv_source_file,
                logger=logger,
            )
        )
        rows.extend(nv_rows)
    if should_run("Nav"):
        nav_rows = run_team(
            logger,
            "Nav",
            lambda: load_nav_from_existing_metrics_and_refresh_3_weeks(
                out_file,
                nav_source_file,
                logger=logger,
            )
        )
        rows.extend(nav_rows)
    if should_run("AE MEIC"):
        ae_meic_rows = run_team(
            logger,
            "AE MEIC",
            lambda: load_ae_meic_from_existing_metrics_and_refresh_3_weeks(
                out_file,
                ae_meic_source_file,
                logger=logger,
            )
        )
        rows.extend(ae_meic_rows)
    if should_run("O-Arm MEIC"):
        oarm_meic_rows = run_team(
            logger,
            "O-Arm MEIC",
            lambda: load_oarm_meic_from_existing_metrics_and_refresh_3_weeks(
                out_file,
                oarm_meic_source_file,
                logger=logger,
            )
        )
        rows.extend(oarm_meic_rows)
    if should_run("Mazor"):
        mazor_rows = run_team(
            logger,
            "Mazor",
            lambda: load_mazor_from_existing_metrics_and_refresh_3_weeks(
                out_file,
                mazor_source_file,
                logger=logger,
            )
        )
        rows.extend(mazor_rows)
    if should_run("CSF"):
        csf_rows = run_team(
            logger,
            "CSF",
            lambda: load_csf_from_existing_metrics_and_refresh_3_weeks(
                out_file,
                csf_source_file,
                CSF_CFG,
                logger=logger,
            )
        )
        rows.extend(csf_rows)
    if should_run_pss():
        pss_rows = run_team(
            logger,
            "PSS",
            lambda: load_pss_from_existing_metrics_and_refresh_3_weeks(
                out_file,
                pss_source_file,
                PSS_CFG,
                logger=logger,
            )
        )
        rows.extend(pss_rows)
    if should_run("ENT"):
        try:
            ent_name_note = apply_ent_name_replacements_to_sheet(
                ent_mapping_xlsx,
                sheet_name="Next Week Forecast",
                start_row=2,
                end_row=30,
                name_col_letter="A",
            )
            logger.info(f"[ENT] {ent_name_note}")
        except Exception as e:
            logger.error(f"[ENT] Failed applying name replacements to mapping sheet: {e}")
        extend_team("ENT", lambda: scrape_ent_from_csv(ent_data_csv, ent_mapping_xlsx, team="ENT"))
    if should_run("DBS MEIC"):
        extend_team("DBS MEIC", lambda: scrape_csv_team_fixed_availability(dbs_meic_csv, team="DBS MEIC", hours_per_person=20.0))
    if should_run("Other MEIC"):
        extend_team("Other MEIC", lambda: scrape_csv_team_fixed_availability(other_meic_csv, team="Other MEIC", hours_per_person=20.0))
    if should_run("SCS Super Cell"):
        scs_super_rows = run_team(
            logger,
            "SCS Super Cell",
            lambda: load_scs_super_from_existing_metrics_and_refresh_3_weeks(
                out_file,
                scs_super_source_file,
                SCS_SUPER_OLD_CFG,
                SCS_SUPER_NEW_CFG,
                logger=logger,
            )
        )
        rows.extend(scs_super_rows)
    ph17_before = sum(1 for r in rows if r.get("team") == "PH Cell 17")
    logger.info(f"[PH Cell 17] rows before TAA filter = {ph17_before}")
    before = len(rows)
    rows = [
        r for r in rows
        if (r.get("team") in ("SCS Super Cell", "PH Cell 17", "Spine", "TDD COS 1"))
        or (safe_float(r.get("Total Available Hours")) != 0.0)
    ]
    logger.info(f"[ALL] filter TAA!=0 (except SCS Super Cell, PH Cell 17): {before} -> {len(rows)}")
    ph17_after = sum(1 for r in rows if r.get("team") == "PH Cell 17")
    logger.info(f"[PH Cell 17] rows after TAA filter = {ph17_after}")
    logger.info(f"[ALL] filter TAA!=0 (except SCS Super Cell): {before} -> {len(rows)}")
    for bad in ("2023-11-06", "2026-09-07"):
        before = len(rows)
        rows = [r for r in rows if safe_str(r.get("period_date")) != bad]
        logger.info(f"[ALL] drop period_date == {bad}: {before} -> {len(rows)}")
    def sort_key(r: dict) -> tuple:
        team = safe_str(r.get("team")).lower()
        d = safe_str(r.get("period_date"))
        date_key = d if (len(d) == 10 and d[4] == "-" and d[7] == "-") else "9999-12-31"
        return (team, date_key)
    rows.sort(key=sort_key)
    write_csv(rows, out_file)
    logger.info(f"Wrote {len(rows)} rows to {out_file}")
    wip_rows = build_ns_wip_rows(rows)
    wip_out_file = "NS_DATA\\NS_WIP.csv"
    write_csv_wip(wip_rows, wip_out_file)
if __name__ == "__main__":
    main()