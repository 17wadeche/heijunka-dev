# collect_non_wip.py
import csv
import json
from pathlib import Path
from datetime import datetime as _dt, date as _date, timedelta
import re
import pandas as pd
from openpyxl import load_workbook
from dateutil import parser as dateparser
_DAY_RANGES = {
    "Monday":    (7, 40),
    "Tuesday":   (42, 77),
    "Wednesday": (79, 118),
    "Thursday":  (120, 161),
    "Friday":    (163, 200),
}
NWW_CATEGORIES = {"ooo", "workshop", "audit", "yellow belt", "nextgen"}
import regex as _re
_HALF_DAY_RE = _re.compile(
    r"""(?ix)
    (?:\bAM\b|\bPM\b|half\s*day|1/2(?:\s*day)?|~\s*12|
       @\s*(?:noon|12(?:\s*:?\s*30)?|PM|1)|
       out\s+at\s+(?:12:00\s*pm|1:00\s*pm)|
       afternoon|
       -\s*out\s+in\s*(?:am|pm)
    )
    """
)
_OOO_TOKEN_RE = _re.compile(r"""(?ix)\bO\s*O\s*O\b""")
def _has_inline_ooo(s: str) -> bool:
    return bool(_OOO_TOKEN_RE.search(str(s or "")))
def _strip_inline_ooo(s: str) -> str:
    return _OOO_TOKEN_RE.sub("", str(s or "")).strip()
def _is_half_day(text: str) -> bool:
    return bool(_HALF_DAY_RE.search(str(text or "")))
def _split_people(cell_c: str) -> list[str]:
    parts = _re.split(r"[,&]+", str(cell_c or ""))
    return [p.strip() for p in parts if p and p.strip()]
def _to_float(v) -> float | None:
    try:
        s = str(v).strip()
        if s == "" or s.lower() in {"nan", "none", "-"}:
            return None
        return float(s)
    except Exception:
        return None
def _hours_for_activity(note_text: str) -> float:
    return 4.0 if _is_half_day(note_text) else 8.0
def _strip_name_annotations(s: str) -> str:
    txt = str(s or "")
    txt = _re.sub(r"\s*@\s*\d{1,2}(?::[0-5]\d)?\s*(?:am|pm)?\b.*$", "", txt, flags=_re.I)
    txt = _re.sub(r"\s*-\s*out.*$", "", txt, flags=_re.I)
    txt = _re.sub(r"\s*\([^)]*\)\s*$", "", txt)
    return txt.strip()
_CAS_WEEK_ROWS = [
    {"row":  7786, "date": "1/1/2024"},
    {"row":  7918, "date": "1/8/2024"},
    {"row":  8054, "date": "1/15/2024"},
    {"row":  8190, "date": "1/22/2024"},
    {"row":  8331, "date": "1/29/2024"},
    {"row":  8473, "date": "2/5/2024"},
    {"row":  8616, "date": "2/12/2024"},
    {"row":  8760, "date": "2/19/2024"},
    {"row":  8901, "date": "2/26/2024"},
    {"row":  9046, "date": "3/4/2024"},
    {"row":  9190, "date": "3/11/2024"},
    {"row":  9334, "date": "3/18/2024"},
    {"row":  9480, "date": "3/25/2024"},
    {"row":  9626, "date": "4/1/2024"},
    {"row":  9776, "date": "4/8/2024"},
    {"row":  9927, "date": "4/15/2024"},
    {"row": 10079, "date": "4/22/2024"},
    {"row": 10234, "date": "4/29/2024"},
    {"row": 10388, "date": "5/6/2024"},
    {"row": 10542, "date": "5/13/2024"},
    {"row": 10696, "date": "5/20/2024"},
    {"row": 10850, "date": "5/27/2024"},
    {"row": 11005, "date": "6/3/2024"},
    {"row": 11159, "date": "6/10/2024"},
    {"row": 11313, "date": "6/17/2024"},
    {"row": 11470, "date": "6/24/2024"},
    {"row": 11626, "date": "7/1/2024"},
    {"row": 11792, "date": "7/8/2024"},
    {"row": 11960, "date": "7/15/2024"},
    {"row": 12126, "date": "7/22/2024"},
    {"row": 12291, "date": "7/29/2024"},
    {"row": 12461, "date": "8/5/2024"},
    {"row": 12631, "date": "8/12/2024"},
    {"row": 12801, "date": "8/19/2024"},
    {"row": 12971, "date": "8/26/2024"},
    {"row": 13141, "date": "9/2/2024"},
    {"row": 13312, "date": "9/9/2024"},
    {"row": 13486, "date": "9/16/2024"},
    {"row": 13660, "date": "9/23/2024"},
    {"row": 13829, "date": "9/30/2024"},
    {"row": 14000, "date": "10/7/2024"},
    {"row": 14171, "date": "10/14/2024"},
    {"row": 14342, "date": "10/21/2024"},
    {"row": 14512, "date": "10/28/2024"},
    {"row": 14692, "date": "11/4/2024"},
    {"row": 14872, "date": "11/11/2024"},
    {"row": 15052, "date": "11/18/2024"},
    {"row": 15236, "date": "11/25/2024"},
    {"row": 15422, "date": "12/2/2024"},
    {"row": 15610, "date": "12/9/2024"},
    {"row": 15799, "date": "12/16/2024"},
    {"row": 15989, "date": "12/23/2024"},
    {"row": 16043, "date": "12/30/2024"},
    {"row": 16083, "date": "1/6/2025"},
    {"row": 16271, "date": "1/13/2025"},
    {"row": 16458, "date": "1/20/2025"},
    {"row": 16644, "date": "1/27/2025"},
    {"row": 16836, "date": "2/3/2025"},
    {"row": 17029, "date": "2/10/2025"},
    {"row": 17224, "date": "2/17/2025"},
    {"row": 17426, "date": "2/24/2025"},
    {"row": 17628, "date": "3/3/2025"},
    {"row": 17836, "date": "3/10/2025"},
    {"row": 18042, "date": "3/17/2025"},
    {"row": 18249, "date": "3/24/2025"},
    {"row": 18455, "date": "3/30/2025"},
    {"row": 18661, "date": "4/7/2025"},
    {"row": 18868, "date": "4/14/2025"},
    {"row": 19075, "date": "4/21/2025"},
    {"row": 19285, "date": "4/28/2025"},
    {"row": 19495, "date": "5/5/2025"},
    {"row": 19706, "date": "5/12/2025"},
    {"row": 19918, "date": "5/19/2025"},
    {"row": 20129, "date": "5/26/2025"},
    {"row": 20341, "date": "6/2/2025"},
    {"row": 20553, "date": "6/9/2025"},
    {"row": 20763, "date": "6/16/2025"},
    {"row": 20972, "date": "6/23/2025"},
    {"row": 21182, "date": "6/30/2025"},
    {"row": 21394, "date": "7/7/2025"},
    {"row": 21606, "date": "7/14/2025"},
    {"row": 21818, "date": "7/21/2025"},
    {"row": 22028, "date": "7/28/2025"},
    {"row": 22238, "date": "8/4/2025"},
    {"row": 22449, "date": "8/11/2025"},
    {"row": 22661, "date": "8/18/2025"},
    {"row": 22879, "date": "8/25/2025"},
    {"row": 23094, "date": "9/1/2025"},
    {"row": 23311, "date": "9/8/2025"},
    {"row": 23517, "date": "9/15/2025"},
    {"row": 23725, "date": "9/22/2025"},
    {"row": 23937, "date": "9/29/2025"},
    {"row": 24149, "date": "10/6/2025"},
    {"row": 24364, "date": "10/13/2025"},
    {"row": 24581, "date": "10/20/2025"},
]
def _cas_row_window_for_date(period_date: _date) -> tuple[int, int] | None:
    if period_date is None:
        return None
    rows = [(int(e["row"]), str(e["date"])) for e in _CAS_WEEK_ROWS]
    rows_sorted = sorted(rows, key=lambda x: x[0])
    target_idx = None
    for i, (r, dstr) in enumerate(rows_sorted):
        try:
            if _coerce_to_date(dstr) == period_date:
                target_idx = i
                break
        except Exception:
            continue
    if target_idx is None:
        return None
    start = rows_sorted[target_idx][0]
    end   = (rows_sorted[target_idx + 1][0] - 1) if target_idx + 1 < len(rows_sorted) else (start + 400)
    return (start, end)
def extract_ect_nonwip(xlsx_path: Path) -> list[dict]:
    out: list[dict] = []
    if xlsx_path.suffix.lower() not in (".xlsx", ".xlsm"):
        return out
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception:
        return out
    sh_name = _resolve_sheet_exact_or_fuzzy_openpyxl(wb, "Available WIP Hours")
    if not sh_name:
        return out
    ws = wb[sh_name]
    MIN_COL, MAX_COL = 1, 13
    current_person = None
    for row in ws.iter_rows(min_row=1, max_row=getattr(ws, "max_row", 0),
                            min_col=MIN_COL, max_col=MAX_COL, values_only=True):
        raw_name = _clean_name(row[0]) if len(row) >= 1 else ""
        if raw_name:
            current_person = raw_name
        if not current_person:
            continue
        act1 = (str(row[9]).strip() if len(row) >= 10 and row[9] is not None else "")
        hrs1 = _to_float(row[10] if len(row) >= 11 else None)
        act2 = (str(row[11]).strip() if len(row) >= 12 and row[11] is not None else "")
        hrs2 = _to_float(row[12] if len(row) >= 13 else None)
        def _clean_act(s: str) -> str:
            s = re.sub(r"^\s*\d+\.\s*", "", s)  # leading enumerations
            return s.strip()
        for act, hrs in ((act1, hrs1), (act2, hrs2)):
            if not act:
                continue
            act = _clean_act(act)
            if not act:
                continue
            if hrs is None or hrs <= 0:
                continue
            activity = "OOO" if "ooo" in act.casefold() else act
            out.append({
                "day": "Week",
                "name": current_person,
                "activity": activity,
                "hours": float(hrs),
            })
    return out
def extract_cas_activities(xlsx_path: Path, period_date: _date) -> list[dict]:
    out: list[dict] = []
    ext = xlsx_path.suffix.lower()
    window = _cas_row_window_for_date(period_date)
    if not window or ext not in (".xlsx", ".xlsm"):
        return out
    rmin, rmax = window
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception:
        return out
    from collections import defaultdict
    agg = defaultdict(lambda: {"hours": 0.0, "days": 0})
    for sh in wb.sheetnames:
        ws = wb[sh]
        try:
            if getattr(ws, "max_row", 0) < rmax:
                continue
        except Exception:
            continue
        for row in ws.iter_rows(min_row=rmin, max_row=rmax, min_col=2, max_col=3, values_only=True):
            col_b, col_c = row  # B (activity text), C (names)
            text_b = str(col_b or "").strip()
            if not text_b:
                continue
            cat = None
            lower = text_b.casefold()
            for c in NWW_CATEGORIES:
                if c in lower:
                    cat = c.upper() if c == "ooo" else c.title()
                    break
            if not cat:
                continue
            people = _split_people(col_c or "")
            if not people:
                continue
            base_hrs = _hours_for_activity(text_b)
            for person in people:
                inline_ooo = _has_inline_ooo(person)
                person_is_half = _is_half_day(person)
                nm = _clean_name(_strip_name_annotations(_strip_inline_ooo(person)))
                if not nm:
                    continue
                eff_cat = "OOO" if inline_ooo else ( "OOO" if (cat and cat.upper() == "OOO") else cat )
                if not eff_cat:
                    continue
                base_hrs = _hours_for_activity(text_b)  # 4 if half-day in B, else 8
                hrs = 4.0 if person_is_half else base_hrs
                key = (nm, "OOO" if eff_cat.upper() == "OOO" else eff_cat)
                agg[key]["hours"] += float(hrs)
                agg[key]["days"]  += 1
    for (nm, activity), vals in agg.items():
        out.append({
            "day": "Week",
            "name": nm,
            "activity": activity,
            "hours": round(vals["hours"], 2),
            "days": int(vals["days"]),   # NEW: number of days this occurred within the week
        })
    return out
TEAM_OOO_CFG = {
    "aortic":          {"sheet": "#12 Production Analysis",           "flag_col": "K"},
    "svt":             {"sheet": "#12 Production Analysis",           "flag_col": "K"},
    "crdn":            {"sheet": "#12 Production Analysis",           "flag_col": "K"},
    "ect":             {"sheet": "#12 Production Analysis",           "flag_col": "K"},
    "pvh":             {"sheet": "#12 Production Analysis",           "flag_col": "K"},
    "tct clinical":    {"sheet": "Clinical #12 Prod Analysis",        "flag_col": "L"},
    "tct commercial":  {"sheet": "Commercial #12 Prod Analysis",      "flag_col": "L"},
}
REPO_DIR = Path(r"C:\heijunka-dev")
REPO_CSV = REPO_DIR / "metrics_aggregate_dev.csv"
OUT_CSV  = REPO_DIR / "non_wip_activities.csv"
WEEKLY_HOURS_DEFAULT = 40.0
TEAM_CFG = {
    "aortic":         {"sheet_patterns": ["individual (wip non wip)"], "col": "A", "start": 1},
    "crdn":           {"sheet_patterns": ["individual (wip non wip)"], "col": "A", "start": 1},
    "ect":            {"sheet_patterns": ["individual (wip non wip)"], "col": "A", "start": 1},
    "pvh":            {"sheet_patterns": ["individual (wip non wip)"], "col": "A", "start": 1},
    "svt":            {"sheet_patterns": ["individual"],                "col": "A", "start": 1},
    "tct commercial": {"sheet_patterns": ["individual (wip non wip)"], "col": "A", "start": 1},
    "tct clinical":   {"sheet_patterns": ["individual (wip non wip)"], "col": "Z", "start": 1},
    "ph":             {"people_from": "person_hours"},
    "cas":            {"people_from": "person_hours"},
}
try:
    from pyxlsb import open_workbook as open_xlsb
except Exception:
    open_xlsb = None
def _excel_serial_to_date(n):
    try:
        return (_dt(1899, 12, 30) + timedelta(days=float(n))).date()
    except Exception:
        return None
def _norm_person(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip()).casefold()
def _coerce_to_date(v):
    if isinstance(v, _dt): return v.date()
    if isinstance(v, _date): return v
    if isinstance(v, (int, float)): return _excel_serial_to_date(v)
    s = str(v).strip()
    if not s: return None
    try:
        return dateparser.parse(s, dayfirst=False, yearfirst=False).date()
    except Exception:
        return None
_BAD_EXACT = {
    "team member",
    "tuesday",
    "4",
    "5",
    "6",
    "7",
    "commercial weeks production output",
    "clinical weeks production output",
    "0.0",
    "open",
    "0",
    "2025-10-06 00:00:00",
    "team member 1",
    "team member 2",
    "team member 3",
    "team member 4",
    "total available hours",
    "total pitches",
    "weeks production output",
    "workflow",
}
_BAD_PREFIX = ("release date", "revision", "week starting")
def _looks_like_bad_header(raw: str) -> bool:
    if raw is None:
        return True
    s = str(raw).strip()
    if not s:
        return True
    s_nf = s.rstrip(":").strip()
    k = s_nf.casefold()
    if k in _BAD_EXACT:
        return True
    for pref in _BAD_PREFIX:
        if k.startswith(pref):
            return True
    if k in {"#ref!", "nan", "0", "-", "–", "—"}:
        return True
    return False
def _clean_name(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().strip('"').strip("'")
    return "" if _looks_like_bad_header(s) else s
def _col_letter_to_index(letter: str) -> int:
    letter = str(letter).strip().upper()
    if not letter:
        return 1
    acc = 0
    for ch in letter:
        if not ("A" <= ch <= "Z"):
            continue
        acc = acc * 26 + (ord(ch) - ord("A") + 1)
    return max(1, acc)
def _norm_sheet_name(s: str) -> str:
    s = (s or "").lower()
    s = s.replace("–", "-").replace("—", "-")        # normalize dashes
    s = s.replace("(", " ").replace(")", " ")
    s = s.replace("_", " ").replace("-", " ")
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace("wip non wip", "wip non wip")
    return s
def _norm_title(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("–", "-").replace("—", "-")
    return " ".join(s.split())
def _resolve_sheet_exact_or_fuzzy_openpyxl(wb, desired_title: str) -> str | None:
    want = _norm_title(desired_title)
    m = {_norm_title(n): n for n in wb.sheetnames}
    if want in m:
        return m[want]
    for actual in wb.sheetnames:
        ns = _norm_title(actual)
        if (want in ns) or ns.startswith(want) or ns.endswith(want):
            return actual
    return None
def _resolve_sheet_exact_or_fuzzy_xlsb(wb, desired_title: str) -> str | None:
    want = _norm_title(desired_title)
    names = list(getattr(wb, "sheets", []) or [])
    m = {_norm_title(n): n for n in names}
    if want in m:
        return m[want]
    for actual in names:
        ns = _norm_title(actual)
        if (want in ns) or ns.startswith(want) or ns.endswith(want):
            return actual
    return None
def extract_ooo_per_day(xlsx_path: Path, sheet_title: str, flag_col_letter: str) -> list[dict]:
    NAME_COL_IDX = _col_letter_to_index("C")
    FLAG_COL_IDX = _col_letter_to_index(flag_col_letter)
    ext = xlsx_path.suffix.lower()
    out = []
    if ext in (".xlsx", ".xlsm"):
        try:
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        except Exception:
            return []
        sh_name = _resolve_sheet_exact_or_fuzzy_openpyxl(wb, sheet_title)
        if not sh_name:
            return []
        ws = wb[sh_name]
        for day, (rmin, rmax) in _DAY_RANGES.items():
            seen = set()
            for row in ws.iter_rows(
                min_row=rmin, max_row=rmax,
                min_col=min(NAME_COL_IDX, FLAG_COL_IDX),
                max_col=max(NAME_COL_IDX, FLAG_COL_IDX),
                values_only=True
            ):
                name = str(row[NAME_COL_IDX - min(NAME_COL_IDX, FLAG_COL_IDX)] or "").strip()
                flag = str(row[FLAG_COL_IDX - min(NAME_COL_IDX, FLAG_COL_IDX)] or "").strip().lower()
                if name and flag == "ooo":
                    key = name.casefold()
                    if key not in seen:
                        seen.add(key)
                        out.append({"day": day, "name": name, "activity": "OOO", "hours": 8.0})
        return out
    elif ext == ".xlsb" and open_xlsb is not None:
        try:
            with open_xlsb(xlsx_path) as wb:
                sh_name = _resolve_sheet_exact_or_fuzzy_xlsb(wb, sheet_title)
                if not sh_name:
                    return []
                ws = wb.get_sheet(sh_name)
                rows_by_index = {}
                for ridx, row in enumerate(ws.rows(), start=1):
                    rows_by_index[ridx] = [c.v for c in row]
                for day, (rmin, rmax) in _DAY_RANGES.items():
                    seen = set()
                    for ridx in range(rmin, rmax + 1):
                        r = rows_by_index.get(ridx) or []
                        name = str((r[NAME_COL_IDX - 1] if len(r) >= NAME_COL_IDX else "") or "").strip()
                        flag = str((r[FLAG_COL_IDX - 1] if len(r) >= FLAG_COL_IDX else "") or "").strip().lower()
                        if name and flag == "ooo":
                            key = name.casefold()
                            if key not in seen:
                                seen.add(key)
                                out.append({"day": day, "name": name, "activity": "OOO", "hours": 8.0})
            return out
        except Exception:
            return []
    return []
def _resolve_sheet_name(wb, desired_patterns: list[str]) -> str | None:
    if not desired_patterns:
        return None
    desired = {_norm_sheet_name(p) for p in desired_patterns}
    norm_to_actual = {_norm_sheet_name(n): n for n in wb.sheetnames}
    for want in desired:
        if want in norm_to_actual:
            return norm_to_actual[want]
    for actual in wb.sheetnames:
        ns = _norm_sheet_name(actual)
        for want in desired:
            if (want in ns) or ns.startswith(want) or ns.endswith(want):
                return actual
    return None
def _read_names_from_matching_sheets_row_xlsx(path: Path, sheet_patterns: list[str],
                                              row_number: int, max_cols: int = 400) -> list[str]:
    ext = path.suffix.lower()
    names = []
    want = {_norm_sheet_name(p) for p in (sheet_patterns or [])}
    if ext in (".xlsx", ".xlsm"):
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception:
            print(f"[non-wip] Could not open workbook for PH: {path}")
            return []
        for sh in wb.sheetnames:
            nsh = _norm_sheet_name(sh)
            if any(w in nsh for w in want):
                ws = wb[sh]
                for r in ws.iter_rows(min_row=row_number, max_row=row_number, min_col=1, max_col=max_cols, values_only=True):
                    for val in r:
                        nm = _clean_name(val)
                        if nm: names.append(nm)
    elif ext == ".xlsb":
        if open_xlsb is None:
            print("[non-wip] '.xlsb' requires 'pyxlsb'. Try: pip install pyxlsb")
            return []
        try:
            with open_xlsb(path) as wb:
                for sh in wb.sheets:
                    nsh = _norm_sheet_name(sh)
                    if any(w in nsh for w in want):
                        ws = wb.get_sheet(sh)
                        for ridx, row in enumerate(ws.rows(), start=1):
                            if ridx < row_number: 
                                continue
                            if ridx > row_number: 
                                break
                            for cidx, cell in enumerate(row, start=1):
                                if cidx > max_cols: 
                                    break
                                nm = _clean_name(cell.v)
                                if nm: names.append(nm)
        except Exception as e:
            print(f"[non-wip] Failed reading PH .xlsb {path.name}: {e}")
            return []
    seen, out = set(), []
    for n in names:
        k = n.casefold()
        if k not in seen:
            seen.add(k); out.append(n)
    return out
def _read_names_from_sheet_col_xlsx(path: Path, sheet_patterns: list[str], col_letter: str = "A",
                                    start_row: int = 1, max_rows: int = 400) -> list[str]:
    ext = path.suffix.lower()
    col_idx = _col_letter_to_index(col_letter)
    start_row = max(1, int(start_row))
    end_row = max(start_row, start_row + max_rows - 1)
    names = []
    if ext in (".xlsx", ".xlsm"):
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception:
            print(f"[non-wip] Could not open workbook: {path}")
            return []
        sheet_name = _resolve_sheet_name(wb, sheet_patterns)
        if not sheet_name:
            print(f"[non-wip] No sheet matched {sheet_patterns} in {path.name}")
            return []
        ws = wb[sheet_name]
        for r in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=col_idx, max_col=col_idx, values_only=True):
            nm = _clean_name(r[0])
            if nm: names.append(nm)
    elif ext == ".xlsb":
        if open_xlsb is None:
            print("[non-wip] '.xlsb' requires 'pyxlsb'. Try: pip install pyxlsb")
            return []
        try:
            with open_xlsb(path) as wb:
                sheet_name = None
                norm_to_actual = {_norm_sheet_name(n): n for n in wb.sheets}
                desired = {_norm_sheet_name(p) for p in (sheet_patterns or [])}
                for want in desired:
                    if want in norm_to_actual:
                        sheet_name = norm_to_actual[want]; break
                if sheet_name is None:
                    for actual in wb.sheets:
                        ns = _norm_sheet_name(actual)
                        if any((want in ns) or ns.startswith(want) or ns.endswith(want) for want in desired):
                            sheet_name = actual; break
                if sheet_name is None:
                    print(f"[non-wip] No sheet matched {sheet_patterns} in {path.name}")
                    return []
                sh = wb.get_sheet(sheet_name)
                for ridx, row in enumerate(sh.rows(), start=1):
                    if ridx < start_row: 
                        continue
                    if ridx > end_row: 
                        break
                    val = None
                    for cidx, cell in enumerate(row, start=1):
                        if cidx == col_idx:
                            val = cell.v; break
                    nm = _clean_name(val)
                    if nm: names.append(nm)
        except Exception as e:
            print(f"[non-wip] Failed reading .xlsb {path.name}: {e}")
            return []
    seen, out = set(), []
    for n in names:
        k = n.casefold()
        if k not in seen:
            seen.add(k); out.append(n)
    return out
def _source_file_only(s: str) -> str:
    return s.split(" :: ", 1)[0].strip()
def _parse_person_hours_cell(s: str | None) -> dict[str, float]:
    if not s:
        return {}
    try:
        obj = json.loads(s)
        out = {}
        for name, vals in (obj or {}).items():
            try:
                out[str(name).strip()] = float((vals or {}).get("actual") or 0.0)
            except Exception:
                out[str(name).strip()] = 0.0
        return out
    except Exception:
        return {}
def _lookup_actual_hours(ph_by_name: dict[str, float], person: str) -> float:
    if person in ph_by_name:
        return float(ph_by_name[person])
    want = person.casefold()
    for k, v in ph_by_name.items():
        if k.casefold() == want:
            return float(v)
    return 0.0
def _get_team_cfg(team_name: str):
    return TEAM_CFG.get(str(team_name).casefold())
def _read_people_from_file_for_team(xlsx_path: Path, team_name: str) -> list[str]:
    cfg = _get_team_cfg(team_name)
    if not cfg:
        return []
    if "row" in cfg and "sheet_patterns" in cfg:
        return _read_names_from_matching_sheets_row_xlsx(
            xlsx_path,
            sheet_patterns=cfg["sheet_patterns"],
            row_number=cfg["row"],
            max_cols=400,
        )
    patterns = cfg.get("sheet_patterns", [])
    col      = cfg.get("col", "A")
    start    = cfg.get("start", 1)
    return _read_names_from_sheet_col_xlsx(xlsx_path, sheet_patterns=patterns, col_letter=col, start_row=start)
def main():
    if not REPO_CSV.exists():
        raise FileNotFoundError(f"metrics CSV not found: {REPO_CSV}")
    df = pd.read_csv(REPO_CSV, dtype=str, keep_default_na=False)
    df["team_norm"] = df.get("team", "").astype(str).str.casefold()
    df = df[(df.get("source_file", "") != "")]
    if df.empty:
        print("[non-wip] No rows found in metrics_aggregate_dev.csv with a source file")
        return
    df["period_date"] = df["period_date"].apply(_coerce_to_date)
    df = df.dropna(subset=["period_date"])
    df["period_date"] = pd.to_datetime(df["period_date"]).dt.date
    df["source_file_only"] = df["source_file"].apply(_source_file_only)
    ph_index: dict[tuple, dict[str, float]] = {}
    for _, r in df.iterrows():
        key = (r["team_norm"], r["period_date"], r["source_file_only"])
        ph = _parse_person_hours_cell(r.get("Person Hours"))
        if key not in ph_index:
            ph_index[key] = ph
        else:
            ph_index[key].update(ph)
    unique_refs = df[["team", "team_norm", "period_date", "source_file_only"]].drop_duplicates()
    out_rows = []
    for _, row in unique_refs.iterrows():
        team        = row["team"]
        team_norm   = row["team_norm"]
        period_date = row["period_date"]
        src = row["source_file_only"]  # used for PH lookup (matches the CSV index)
        src_file = src
        if team_norm == "cas":
            src_file = r"c:\Users\wadec8\Medtronic PLC\CAS Virtual VMB - PA Board\PA Board 2.xlsx"
        ph_by_name = ph_index.get((team_norm, period_date, src), {})
        cfg = _get_team_cfg(team_norm)
        details: list[dict] = []
        p = Path(src_file)
        if team_norm == "cas" and p.exists():
            try:
                cas_extra = extract_cas_activities(p, period_date)
                if cas_extra:
                    details.extend(cas_extra)
            except Exception as e:
                print(f"[non-wip] CAS extract error for {period_date}: {e}")
        if team_norm == "ect" and p.exists():
            try:
                ect_extra = extract_ect_nonwip(p)
                if ect_extra:
                    details.extend(ect_extra)
            except Exception as e:
                print(f"[non-wip] ECT extract error for {period_date}: {e}")
        use_person_hours = bool(cfg and cfg.get("people_from") == "person_hours")
        people: list[str] = []
        if use_person_hours:
            people = [n for n in ph_by_name.keys() if _clean_name(n)]
            if team_norm == "cas" and not people and details:
                people = sorted({d["name"] for d in details if _clean_name(d.get("name"))})
            if not people:
                pass
            if not people:
                p_srcfile = Path(src_file)
                if not p_srcfile.exists():
                    print(f"[non-wip] No Person Hours names and file missing for team '{team}': {src_file}")
                    continue
                ext = p_srcfile.suffix.lower()
                if ext not in (".xlsx", ".xlsm", ".xlsb"):
                    print(f"[non-wip] No Person Hours names and unsupported file type ({ext}) for team '{team}': {src}")
                    continue
                if ext == ".xlsb" and open_xlsb is None:
                    print("[non-wip] '.xlsb' requires 'pyxlsb'. Try: pip install pyxlsb")
                    continue
                people = _read_people_from_file_for_team(p, team_norm)
        else:
            p_srcfile = Path(src_file)
            if not p_srcfile.exists():
                print(f"[non-wip] Skip missing file: {src_file}")
                continue
            ext = p_srcfile.suffix.lower()
            if ext not in (".xlsx", ".xlsm", ".xlsb"):
                print(f"[non-wip] Skip unsupported file type ({ext}): {src}")
                continue
            if ext == ".xlsb" and open_xlsb is None:
                print("[non-wip] '.xlsb' requires 'pyxlsb'. Try: pip install pyxlsb")
                continue
            people = _read_people_from_file_for_team(p, team_norm)
        if not people:
            if details:
                people = sorted({d["name"] for d in details if _clean_name(d.get("name"))})
            else:
                source_hint = "Person Hours" if use_person_hours else "workbook"
                print(f"[non-wip] No names found for team '{team}' on {period_date} from {source_hint}")
                continue
        per_person_non_wip = {}
        total_non_wip = 0.0
        total_wip_capped = 0.0
        for person in people:
            wip_actual = _lookup_actual_hours(ph_by_name, person)
            non_wip = max(0.0, WEEKLY_HOURS_DEFAULT - float(wip_actual))
            per_person_non_wip[person] = round(non_wip, 2)
            total_non_wip += non_wip
            total_wip_capped += min(float(wip_actual), WEEKLY_HOURS_DEFAULT)
        people_count = len(people)
        weekly_total_available = WEEKLY_HOURS_DEFAULT * people_count if people_count > 0 else 0.0
        pct_in_wip = (total_wip_capped / weekly_total_available * 100.0) if weekly_total_available > 0 else 0.0
        pct_in_wip = round(pct_in_wip, 2)
        row_obj = {
            "team": team,
            "period_date": period_date.isoformat(),
            "source_file": src,
            "people_count": people_count,
            "total_non_wip_hours": round(total_non_wip, 2),
            "% in WIP": pct_in_wip,
            "non_wip_by_person": json.dumps(per_person_non_wip, ensure_ascii=False),
        }
        ooo_cfg = TEAM_OOO_CFG.get(team_norm)
        if ooo_cfg and p.exists():
            try:
                ooo_details = extract_ooo_per_day(p, ooo_cfg["sheet"], ooo_cfg["flag_col"])
                if ooo_details:
                    details.extend(ooo_details)
            except Exception as e:
                print(f"[non-wip] OOO extract error for {team_norm} {period_date}: {e}")
        row_obj["non_wip_activities"] = json.dumps(details, ensure_ascii=False)
        from collections import defaultdict
        ooo_hours_by_person: dict[str, float] = defaultdict(float)
        ooo_full_days_seen = set()
        for d in details:
            act = str(d.get("activity", "")).strip().upper()
            if act != "OOO":
                continue
            nm_norm = _norm_person(d.get("name", ""))
            if not nm_norm:
                continue
            hrs = float(d.get("hours", 8.0) or 8.0)
            ooo_hours_by_person[nm_norm] += hrs
            dy = str(d.get("day", "")).strip()
            if dy in _DAY_RANGES and abs(hrs - 8.0) < 1e-6:
                ooo_full_days_seen.add((nm_norm, dy))
        canon_by_norm = {}
        for original_key in per_person_non_wip.keys():
            nk = _norm_person(original_key)
            if nk and nk not in canon_by_norm:
                canon_by_norm[nk] = original_key
        for nk, hrs in ooo_hours_by_person.items():
            orig = canon_by_norm.get(nk)
            if not orig:
                continue
            adj = float(per_person_non_wip.get(orig, 0.0)) - float(hrs)
            per_person_non_wip[orig] = round(max(0.0, adj), 2)
        total_non_wip = sum(float(v) for v in per_person_non_wip.values())
        full_week_ooo_norm = set()
        for nk in {k for k, _ in ooo_full_days_seen}:
            days = {dy for kk, dy in ooo_full_days_seen if kk == nk}
            if {"Monday","Tuesday","Wednesday","Thursday","Friday"}.issubset(days):
                full_week_ooo_norm.add(nk)
        effective_people = [pname for pname in people if _norm_person(pname) not in full_week_ooo_norm]
        people_count = len(effective_people)
        weekly_total_available = WEEKLY_HOURS_DEFAULT * people_count if people_count > 0 else 0.0
        total_wip_capped = 0.0
        for person in effective_people:
            wip_actual = _lookup_actual_hours(ph_by_name, person)
            total_wip_capped += min(float(wip_actual), WEEKLY_HOURS_DEFAULT)
        pct_in_wip = (total_wip_capped / weekly_total_available * 100.0) if weekly_total_available > 0 else 0.0
        pct_in_wip = round(pct_in_wip, 2)
        ooo_hours_total = round(sum(ooo_hours_by_person.values()), 2)
        row_obj["people_count"] = people_count
        row_obj["total_non_wip_hours"] = round(total_non_wip, 2)
        row_obj["% in WIP"] = pct_in_wip
        row_obj["non_wip_by_person"] = json.dumps(per_person_non_wip, ensure_ascii=False)
        row_obj["OOO Hours"] = ooo_hours_total
        out_rows.append(row_obj)
    if not out_rows:
        print("[non-wip] No weekly rows produced for mapped teams.")
        return
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    cols = [
        "team",
        "period_date",
        "source_file",
        "people_count",
        "total_non_wip_hours",
        "% in WIP",
        "non_wip_by_person",
        "non_wip_activities",
        "OOO Hours",
    ]
    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(out_rows)
    print(f"[non-wip] Wrote {len(out_rows)} rows to {OUT_CSV}")
if __name__ == "__main__":
    main()