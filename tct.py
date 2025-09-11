import re
import sys
import glob
from pathlib import Path
from datetime import datetime as _dt, date as _date, timedelta
from dateutil import parser as dateparser
import pandas as pd
import csv
import numpy as np
from openpyxl import load_workbook

"""
TCT-only metrics extractor
- Scans TCT Clinical and TCT Commercial workbooks (archived + current folders)
- Calculates Target/Actual UPLH, Actual HC Used, and counts HC in WIP
- Optionally joins Open Complaint Timeliness if timeliness.csv is present
- Outputs tct.csv in the current working directory
"""

# ======= Config =======
TIMELINESS_CSV = Path(r"C:\heijunka-dev") / "timeliness.csv"
OUT_CSV = Path.cwd() / "tct.csv"

# Exclusions carried over from the original script
EXCLUDED_SOURCE_FILES = {
    r"C:\Users\wadec8\Medtronic PLC\SVT PXM Team - Archived Heijunka\SVT Future Heijunka.xlsm"
}
EXCLUDED_DIRS = {
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\Weekly Heijunka Archived",
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\Clinical",
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\Commercial",
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\Remediation",
    r"C:\Users\wadec8\Medtronic PLC\TCT CQXM - 1 WIP and Schedule\WIP Blitz Power Hour",
    r"C:\Users\wadec8\Medtronic PLC\SVT PXM Team - Archived Heijunka",
}
EXCLUDED_DIRS = {s.lower().rstrip("\\").replace("/", "\\") for s in EXCLUDED_DIRS}
EXCLUDED_SOURCE_FILES = {s.lower().replace("/", "\\") for s in EXCLUDED_SOURCE_FILES}

TEAM_CONFIG = [
    {
        "name": "TCT Commercial",
        "root": r"C:\\Users\\wadec8\\Medtronic PLC\\TCT CQXM - Weekly Heijunka Archived",
        "pattern": "*.xlsb",
        "period": {"sheet": "#10 WIP Analysis", "cell": "D3"},
        "cells": {
            "Individual(WIP-Non WIP)": {
                "Total Available Hours": "I69",
                "Completed Hours": "I70",
            }
        },
        "sum_columns": {
            "Commercial #12 Prod Analysis": {
                "Target Output": "G",
                "Actual Output": "J",
            }
        },
    },
    {
        "name": "TCT Clinical",
        "root": r"C:\\Users\\wadec8\\Medtronic PLC\\TCT CQXM - Weekly Heijunka Archived",
        "pattern": "*.xlsb",
        "period": {"sheet": "#10 WIP Analysis", "cell": "D3"},
        "cells": {
            "Individual(WIP-Non WIP)": {
                "Total Available Hours": "AG69",
                "Completed Hours": "AG70",
            }
        },
        "sum_columns": {
            "Clinical #12 Prod Analysis": {
                "Target Output": "G",
                "Actual Output": "J",
            }
        },
    },
    {
        "name": "TCT Commercial",
        "root": r"C:\\Users\\wadec8\\Medtronic PLC\\TCT CQXM - 1 WIP and Schedule",
        "pattern": "*.xlsb",
        "period": {"sheet": "#10 WIP Analysis", "cell": "D3"},
        "cells": {
            "Individual(WIP-Non WIP)": {
                "Total Available Hours": "I69",
                "Completed Hours": "I70",
            }
        },
        "sum_columns": {
            "Commercial #12 Prod Analysis": {
                "Target Output": "G",
                "Actual Output": "J",
            }
        },
    },
    {
        "name": "TCT Clinical",
        "root": r"C:\\Users\\wadec8\\Medtronic PLC\\TCT CQXM - 1 WIP and Schedule",
        "pattern": "*.xlsb",
        "period": {"sheet": "#10 WIP Analysis", "cell": "D3"},
        "cells": {
            "Individual(WIP-Non WIP)": {
                "Total Available Hours": "AG69",
                "Completed Hours": "AG70",
            }
        },
        "sum_columns": {
            "Clinical #12 Prod Analysis": {
                "Target Output": "G",
                "Actual Output": "J",
            }
        },
    },
]

# ======= Helpers =======
SKIP_PATTERNS = [r"~\$", r"\.tmp$"]
DATE_REGEXES = [
    r"\b\d{4}[-_]\d{2}[-_]\d{2}\b",
    r"\b\d{8}\b",
    r"\b\d{1,2}[-_]\d{1,2}[-_]\d{2,4}\b",
    r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*[-_ ]\d{1,2}[-_, ]\d{2,4}\b",
]
USE_FILE_MTIME_IF_NO_DATE = True


def looks_like_temp(name: str) -> bool:
    return any(re.search(pat, name, flags=re.IGNORECASE) for pat in SKIP_PATTERNS)


def parse_date_from_text(text: str):
    for rx in DATE_REGEXES:
        m = re.search(rx, text, flags=re.IGNORECASE)
        if m:
            try:
                return dateparser.parse(m.group(0), dayfirst=False, yearfirst=False).date()
            except Exception:
                pass
    return None


def _excel_serial_to_date(n) -> _date | None:
    try:
        return (_dt(1899, 12, 30) + timedelta(days=float(n))).date()
    except Exception:
        return None


def col_letter_to_index(letter: str) -> int:
    letter = letter.strip().upper()
    num = 0
    for ch in letter:
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f"Invalid column letter: {letter}")
        num = num * 26 + (ord(ch) - ord('A') + 1)
    return num


def a1_to_rowcol(a1: str) -> tuple[int, int]:
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", a1.strip())
    if not m:
        raise ValueError(f"Bad cell address: {a1}")
    letters, row = m.groups()
    col = 0
    for ch in letters.upper():
        col = col * 26 + (ord(ch) - ord('A') + 1)
    return int(row), col


def read_one_cell_openpyxl(ws, a1: str):
    r, c = a1_to_rowcol(a1)
    vals = list(ws.iter_rows(min_row=r, max_row=r, min_col=c, max_col=c, values_only=True))
    return vals[0][0] if vals else None


def read_one_cell_xlsb(file_path: Path, sheet: str, a1: str):
    from pandas import read_excel
    df = read_excel(file_path, sheet_name=sheet, engine="pyxlsb", header=None)
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", a1.strip())
    if not m:
        return None
    col_letters, row_str = m.groups()
    r = int(row_str) - 1
    c = col_letter_to_index(col_letters) - 1
    try:
        return df.iat[r, c]
    except Exception:
        return None


def sum_column_openpyxl_filtered(ws, target_col: str,
                                 include_contains: dict | None = None,
                                 exclude_regex: dict | None = None,
                                 row_start: int | None = None,
                                 row_end: int | None = None,
                                 skip_hidden: bool = False) -> float | None:
    def _passes_filters(row_dict: dict) -> bool:
        if include_contains:
            for col, needle in include_contains.items():
                v = row_dict.get(col)
                if not isinstance(v, str) or needle.lower() not in v.lower():
                    return False
        if exclude_regex:
            for col, rx in exclude_regex.items():
                v = row_dict.get(col)
                if isinstance(v, str) and re.search(rx, v, flags=re.IGNORECASE):
                    return False
        return True

    t_idx = col_letter_to_index(target_col)
    need_cols = {t_idx}
    if include_contains:
        need_cols |= {col_letter_to_index(c) for c in include_contains.keys()}
    if exclude_regex:
        need_cols |= {col_letter_to_index(c) for c in exclude_regex.keys()}
    min_c, max_c = min(need_cols), max(need_cols)
    row_start = row_start or 1
    row_end = row_end or ws.max_row
    total, any_vals = 0.0, False
    for r_idx, row_vals in enumerate(ws.iter_rows(min_row=row_start, max_row=row_end,
                                                  min_col=min_c, max_col=max_c,
                                                  values_only=True), start=row_start):
        if skip_hidden and hasattr(ws, "row_dimensions"):
            rd = ws.row_dimensions.get(r_idx) if hasattr(ws.row_dimensions, "get") else None
            if rd is not None and getattr(rd, "hidden", False):
                continue
        row_map = {}
        for c_idx_off, val in enumerate(row_vals, start=min_c):
            row_map[chr(ord('A') + c_idx_off - 1)] = val
        if not _passes_filters(row_map):
            continue
        val = row_map.get(chr(ord('A') + t_idx - 1))
        try:
            if val is None:
                continue
            if isinstance(val, (int, float)):
                total += float(val); any_vals = True
            else:
                total += float(str(val).replace(",", "").strip()); any_vals = True
        except Exception:
            continue
    return total if any_vals else None


def sum_column_pyxlsb_filtered(file_path: Path, sheet: str, target_col: str,
                               include_contains: dict | None = None,
                               exclude_regex: dict | None = None,
                               row_start: int | None = None,
                               row_end: int | None = None) -> float | None:
    from pandas import read_excel
    try:
        df = read_excel(file_path, sheet_name=sheet, engine="pyxlsb", header=None)
        t = col_letter_to_index(target_col) - 1
        s = pd.to_numeric(df.iloc[:, t], errors="coerce")
        mask = pd.Series(True, index=df.index)
        if include_contains:
            for col_letter, needle in include_contains.items():
                c = col_letter_to_index(col_letter) - 1
                mask &= df.iloc[:, c].astype(str).str.contains(needle, case=False, na=False)
        if exclude_regex:
            for col_letter, rx in exclude_regex.items():
                c = col_letter_to_index(col_letter) - 1
                mask &= ~df.iloc[:, c].astype(str).str.contains(rx, case=False, na=False, regex=True)
        if row_start or row_end:
            rs = (row_start - 1) if row_start else 0
            re_ = (row_end - 1) if row_end else len(df) - 1
            idx = df.index[rs:re_ + 1]
            mask = mask.loc[idx]
        s = s.where(mask)
        tot = s.dropna().sum()
        return float(tot) if pd.notna(tot) else None
    except Exception:
        return None


def read_with_openpyxl(file_path: Path, cells_cfg: dict, sumcols_cfg: dict) -> dict:
    need_visible_rows = any(
        isinstance(spec, dict) and spec.get("skip_hidden", False)
        for mapping in (sumcols_cfg or {}).values()
        for spec in mapping.values()
    )
    wb = load_workbook(file_path, data_only=True, read_only=not need_visible_rows)
    out = {}
    for sheet_name, mapping in (cells_cfg or {}).items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        for out_name, addr in mapping.items():
            try:
                r, c = a1_to_rowcol(addr)
                vals = list(ws.iter_rows(min_row=r, max_row=r, min_col=c, max_col=c, values_only=True))
                out[out_name] = vals[0][0] if vals else None
            except Exception:
                out[out_name] = None
    for sheet_name, mapping in (sumcols_cfg or {}).items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        for out_name, spec in mapping.items():
            try:
                if isinstance(spec, str):
                    # Sum entire column
                    col = col_letter_to_index(spec)
                    total, any_vals = 0.0, False
                    for (val,) in ws.iter_rows(min_col=col, max_col=col, values_only=True):
                        try:
                            if val is None:
                                continue
                            total += float(val) if isinstance(val, (int, float)) else float(str(val).replace(",", "").strip())
                            any_vals = True
                        except Exception:
                            continue
                    out[out_name] = total if any_vals else None
                elif isinstance(spec, dict):
                    out[out_name] = sum_column_openpyxl_filtered(
                        ws,
                        spec.get("col"),
                        include_contains=spec.get("include_contains"),
                        exclude_regex=spec.get("exclude_regex"),
                        row_start=spec.get("row_start"),
                        row_end=spec.get("row_end"),
                        skip_hidden=spec.get("skip_hidden", False),
                    )
                    if out[out_name] is not None and spec.get("divide"):
                        try:
                            out[out_name] = float(out[out_name]) / float(spec["divide"])
                        except Exception:
                            pass
                else:
                    out[out_name] = None
            except Exception:
                out[out_name] = None
    return out


def read_with_pyxlsb(file_path: Path, cells_cfg: dict, sumcols_cfg: dict) -> dict:
    from pandas import read_excel
    out = {}
    for sheet_name, mapping in (cells_cfg or {}).items():
        df = read_excel(file_path, sheet_name=sheet_name, engine="pyxlsb", header=None)
        for out_name, addr in mapping.items():
            m = re.fullmatch(r"([A-Za-z]+)(\d+)", addr.strip())
            if not m:
                out[out_name] = None
                continue
            col_letters, row_str = m.groups()
            r = int(row_str) - 1
            c = col_letter_to_index(col_letters) - 1
            try:
                val = df.iat[r, c]
            except Exception:
                val = None
            out[out_name] = val
    for sheet_name, mapping in (sumcols_cfg or {}).items():
        for out_name, spec in mapping.items():
            try:
                if isinstance(spec, str):
                    out[out_name] = sum_column_pyxlsb_filtered(file_path, sheet_name, spec)
                elif isinstance(spec, dict):
                    val = sum_column_pyxlsb_filtered(
                        file_path,
                        sheet_name,
                        spec.get("col"),
                        include_contains=spec.get("include_contains"),
                        exclude_regex=spec.get("exclude_regex"),
                        row_start=spec.get("row_start"),
                        row_end=spec.get("row_end"),
                    )
                    if val is not None and spec.get("divide"):
                        try:
                            val = float(val) / float(spec["divide"])
                        except Exception:
                            pass
                    out[out_name] = val
                else:
                    out[out_name] = None
            except Exception:
                out[out_name] = None
    return out


def read_metrics_from_file(file_path: Path, cells_cfg: dict, sumcols_cfg: dict) -> dict:
    ext = file_path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return read_with_openpyxl(file_path, cells_cfg, sumcols_cfg)
    elif ext == ".xlsb":
        return read_with_pyxlsb(file_path, cells_cfg, sumcols_cfg)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _hc_in_wip_from_file(file_path: Path, sheet_name: str,
                         key_col_letter: str = "C", cond_col_letter: str = "G",
                         row_start: int = 7, row_end: int = 200) -> int | None:
    ext = file_path.suffix.lower()
    try:
        if ext in (".xlsx", ".xlsm"):
            wb = load_workbook(file_path, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                return None
            ws = wb[sheet_name]
            kc = col_letter_to_index(key_col_letter)
            cc = col_letter_to_index(cond_col_letter)
            unique_keys = set()
            for r, row in enumerate(ws.iter_rows(min_row=row_start, max_row=row_end,
                                                 min_col=min(kc, cc), max_col=max(kc, cc),
                                                 values_only=True), start=row_start):
                key_val = row[kc - min(kc, cc)]
                cond_val = row[cc - min(kc, cc)]
                try:
                    if cond_val is None:
                        continue
                    cond_num = float(str(cond_val).replace(",", "").strip())
                    if cond_num <= 0:
                        continue
                except Exception:
                    continue
                if key_val is None:
                    continue
                key_str = str(key_val).strip()
                if key_str:
                    unique_keys.add(key_str)
            return len(unique_keys)
        elif ext == ".xlsb":
            from pandas import read_excel
            df = read_excel(file_path, sheet_name=sheet_name, engine="pyxlsb", header=None)
            rs = row_start - 1
            re_ = row_end
            k = col_letter_to_index(key_col_letter) - 1
            c = col_letter_to_index(cond_col_letter) - 1
            cond_series = pd.to_numeric(df.iloc[rs:re_, c], errors="coerce")
            mask = cond_series > 0
            keys_series = df.iloc[rs:re_, k].where(mask)
            uniq = set()
            for v in keys_series.dropna().tolist():
                s = str(v).strip()
                if s and s.lower() != "nan":
                    uniq.add(s)
            return len(uniq)
        else:
            return None
    except Exception:
        return None


def infer_period_date(path: Path):
    d = parse_date_from_text(str(path))
    if d:
        return d
    if USE_FILE_MTIME_IF_NO_DATE:
        try:
            return _dt.fromtimestamp(path.stat().st_mtime).date()
        except Exception:
            return None
    return None


def normalize_period_date(df: pd.DataFrame) -> pd.DataFrame:
    if "period_date" in df.columns:
        df["period_date"] = pd.to_datetime(df["period_date"], errors="coerce").dt.normalize()
    return df


def _filter_future_periods(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "period_date" not in df.columns:
        return df
    today = pd.Timestamp.today().normalize()
    d = pd.to_datetime(df["period_date"], errors="coerce").dt.normalize()
    keep = d.isna() | (d <= today)
    return df.loc[keep].copy()


def collect_for_team(team_cfg: dict) -> list[dict]:
    root = Path(team_cfg["root"])
    pattern = team_cfg.get("pattern", "*.xlsx")
    team_name = team_cfg["name"]
    cells_cfg = team_cfg.get("cells", {})
    sumcols_cfg = team_cfg.get("sum_columns", {})
    if not root.exists():
        print(f"[WARN] Root not found for {team_name}: {root}", file=sys.stderr)
        return []
    rows = []
    for p in root.rglob(pattern):
        if p.is_dir():
            continue
        sp = str(p).lower().replace("/", "\\")
        if looks_like_temp(p.name):
            continue
        if sp in EXCLUDED_SOURCE_FILES:
            continue
        if any(sp == d or sp.startswith(d + "\\") for d in EXCLUDED_DIRS):
            continue
        try:
            # Period from sheet/cell if possible, else infer from name/mtime
            period = None
            per_cfg = team_cfg.get("period")
            if per_cfg and isinstance(per_cfg, dict):
                sheet = per_cfg.get("sheet")
                cell = per_cfg.get("cell")
                if sheet and cell:
                    ext = p.suffix.lower()
                    try:
                        if ext == ".xlsb":
                            val = read_one_cell_xlsb(p, sheet, cell)
                        else:
                            wb = load_workbook(p, data_only=True, read_only=True)
                            if sheet in wb.sheetnames:
                                ws = wb[sheet]
                                val = read_one_cell_openpyxl(ws, cell)
                            else:
                                val = None
                        if isinstance(val, (_dt, _date)):
                            period = val.date() if isinstance(val, _dt) else val
                        elif isinstance(val, (int, float)):
                            period = _excel_serial_to_date(val)
                        elif val is not None:
                            period = dateparser.parse(str(val)).date()
                    except Exception:
                        period = None
            if period is None:
                period = infer_period_date(p)

            # Skip future weeks for TCT
            if team_name.lower().startswith("tct"):
                today = _dt.today().date()
                if isinstance(period, _dt):
                    period = period.date()
                if isinstance(period, _date) and period > today:
                    print(f"[skip] TCT future period {period} -> {p}")
                    continue

            values = read_metrics_from_file(p, cells_cfg, sumcols_cfg)

            # HC in WIP from the appropriate #12 sheet
            sheet_for_hc = None
            if team_name == "TCT Commercial":
                sheet_for_hc = "Commercial #12 Prod Analysis"
            elif team_name == "TCT Clinical":
                sheet_for_hc = "Clinical #12 Prod Analysis"
            if sheet_for_hc:
                values["HC in WIP"] = _hc_in_wip_from_file(p, sheet_for_hc)

            rows.append({
                "team": team_name,
                "period_date": period,
                "source_file": str(p),
                **values,
            })
        except Exception as e:
            # Ensure columns still appear even on error
            error_cols = {}
            for mapping in cells_cfg.values():
                for out_name in mapping.keys():
                    error_cols.setdefault(out_name, None)
            for mapping in sumcols_cfg.values():
                for out_name in mapping.keys():
                    error_cols.setdefault(out_name, None)
            rows.append({
                "team": team_name,
                "period_date": infer_period_date(p),
                "source_file": str(p),
                "error": str(e),
                **error_cols,
            })
    return rows


def build_master(rows: list[dict]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df = normalize_period_date(df)
    base_cols = ["team", "period_date", "source_file"]
    metric_cols = [c for c in df.columns if c not in base_cols + ["error"]]
    cols = base_cols + metric_cols + (["error"] if "error" in df.columns else [])
    df = df.reindex(columns=cols)
    if "period_date" in df.columns:
        df = df.sort_values(["team", "period_date", "source_file"], ascending=[True, True, True])
    return df


def _parse_percentish(x):
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    try:
        if s.endswith('%'):
            return float(s[:-1])
        return float(s)
    except Exception:
        return None


def _week_start(s):
    ts = pd.to_datetime(s, errors="coerce")
    return ts.dt.to_period("W-MON").apply(lambda p: p.start_time.normalize())


def _expand_tct_rows(t: pd.DataFrame) -> pd.DataFrame:
    # If timeliness has team like "TCT", "TCT CQXM", or blank, replicate for both clinical & commercial
    if "team" not in t.columns:
        t = t.copy()
        t["team"] = "TCT"
    rows = []
    for _, r in t.iterrows():
        team = str(r.get("team", "")).strip()
        if team.lower() in {"tct", "tct cqxm", "tct total", "tct overall", ""}:
            for target in ["TCT Clinical", "TCT Commercial"]:
                nr = r.copy()
                nr["team"] = target
                rows.append(nr)
        else:
            rows.append(r)
    return pd.DataFrame(rows, columns=t.columns)


def _find_timeliness_csv() -> Path | None:
    candidates = [TIMELINESS_CSV, Path.cwd() / "timeliness.csv", Path.cwd() / "data" / "timeliness.csv"]
    for p in candidates:
        if p.exists():
            if p != TIMELINESS_CSV:
                print(f"[timeliness] Using fallback: {p}")
            return p
    print(f"[timeliness] Not found: {TIMELINESS_CSV}")
    return None


def add_open_complaint_timeliness(df: pd.DataFrame) -> pd.DataFrame:
    p = _find_timeliness_csv()
    if not p:
        return df
    try:
        t = pd.read_csv(p, dtype=str, keep_default_na=False)
    except Exception as e:
        print(f"[timeliness] Failed to read {p}: {e}")
        return df
    if t.empty:
        print("[timeliness] File empty; skipping join.")
        return df

    # Locate columns (robust)
    lower_cols = [str(c).strip().lower() for c in t.columns]

    def _first_match(names):
        for want in names:
            if want in lower_cols:
                return t.columns[lower_cols.index(want)]
        return None

    team_col = _first_match(["team", "group", "area"])  # optional
    date_col = _first_match(["period_date", "period", "week", "date", "week_start", "week ending", "week_end"]) 
    val_col = _first_match(["open complaint timeliness", "timeliness", "% timeliness", "value", "metric", "pct"]) 

    if date_col is None or val_col is None:
        cols = list(t.columns)
        if len(cols) >= 3:
            team_col, date_col, val_col = cols[0], cols[1], cols[2]
        else:
            print("[timeliness] Could not identify columns; skipping join.")
            return df

    t = t.rename(columns={
        team_col if team_col else t.columns[0]: "team",
        date_col: "period_date",
        val_col: "Open Complaint Timeliness",
    })

    # Normalize team & period
    t["team"] = t["team"].astype(str).str.strip()
    t["period_date"] = pd.to_datetime(t["period_date"], errors="coerce").dt.normalize()
    t["Open Complaint Timeliness"] = t["Open Complaint Timeliness"].apply(_parse_percentish)

    # Expand any roll-up rows to both TCT teams
    t = _expand_tct_rows(t)

    # Drop bad rows & dedupe (keep latest)
    t = t.dropna(subset=["period_date", "Open Complaint Timeliness"]).drop_duplicates(subset=["team", "period_date"], keep="last")

    # Align by Monday-of-week to tolerate slight date mismatches
    t = t.copy()
    t["week_start"] = _week_start(t["period_date"])  # Monday

    out = df.copy()
    if out.empty:
        return df
    if "team" not in out.columns or "period_date" not in out.columns:
        print("[timeliness] 'team'/'period_date' not found in metrics df; skipping join.")
        return df

    out["team"] = out["team"].astype(str).str.strip()
    out["period_date"] = pd.to_datetime(out["period_date"], errors="coerce").dt.normalize()
    out["week_start"] = _week_start(out["period_date"])  # Monday

    merged = out.merge(
        t[["team", "week_start", "Open Complaint Timeliness"]],
        on=["team", "week_start"],
        how="left",
    )

    # Secondary merge: week-only broadcast when team is absent in timeliness
    missing_mask = merged["Open Complaint Timeliness"].isna()
    if missing_mask.any():
        wk_only = t[["week_start", "Open Complaint Timeliness"]].drop_duplicates("week_start")
        merged2 = merged.loc[missing_mask].drop(columns=["Open Complaint Timeliness"]).merge(
            wk_only, on="week_start", how="left"
        )
        merged.loc[missing_mask, "Open Complaint Timeliness"] = merged2["Open Complaint Timeliness"].values

    # Cleanup
    if "week_start" in merged.columns:
        merged = merged.drop(columns=["week_start"]) 
    return merged


def _safe_div(n, d):
    try:
        n = float(str(n).replace(",", "").strip())
        d = float(str(d).replace(",", "").strip())
        return None if d == 0 else n / d
    except Exception:
        return None


def save_outputs(df: pd.DataFrame):
    if df.empty:
        print("No rows collected. Check paths/sheets/cells.")
        return

    # Ensure fallback_used and error columns exist
    for c in ("fallback_used", "error"):
        if c not in df.columns:
            df[c] = ""

    # Preferred column order
    preferred_cols = [
        "team", "period_date", "source_file",
        "Total Available Hours", "Completed Hours",
        "Target Output", "Actual Output",
        "Target UPLH", "Actual UPLH",
        "HC in WIP", "Actual HC Used",
        "Open Complaint Timeliness",
        "fallback_used", "error",
    ]
    cols = [c for c in preferred_cols if c in df.columns]
    out = df.loc[:, cols].copy()

    if "period_date" in out.columns:
        out["period_date"] = pd.to_datetime(out["period_date"], errors="coerce").dt.strftime("%Y-%m-%d")

    numeric_cols = {
        "Total Available Hours", "Completed Hours", "Target Output", "Actual Output",
        "Target UPLH", "Actual UPLH", "Actual HC Used", "HC in WIP",
    } & set(out.columns)
    for c in numeric_cols:
        out[c] = pd.to_numeric(out[c], errors="coerce")

    out = out.replace({np.nan: ""})

    out.to_csv(
        OUT_CSV,
        index=False,
        sep=",",
        encoding="utf-8",
        lineterminator="\n",
        quoting=csv.QUOTE_MINIMAL,
        date_format="%Y-%m-%d",
    )
    print(f"Saved CSV: {OUT_CSV.resolve()}")


def run_once():
    all_rows = []
    for cfg in TEAM_CONFIG:
        all_rows.extend(collect_for_team(cfg))

    df = build_master(all_rows)
    df = _filter_future_periods(df)

    # Calculations (same as original)
    df["Target UPLH"] = df.apply(lambda r: _safe_div(r.get("Target Output"), r.get("Total Available Hours")), axis=1)
    df["Actual UPLH"] = df.apply(lambda r: _safe_div(r.get("Actual Output"), r.get("Completed Hours")), axis=1)
    df["Target UPLH"] = df["Target UPLH"].round(2)
    df["Actual UPLH"] = df["Actual UPLH"].round(2)
    df["Actual HC Used"] = pd.to_numeric(df.get("Completed Hours"), errors="coerce") / 32.5
    df["Actual HC Used"] = df["Actual HC Used"].round(2)

    # Optional timeliness join
    df = add_open_complaint_timeliness(df)

    save_outputs(df)

    if not df.empty:
        with pd.option_context("display.max_columns", None, "display.width", 180):
            print("\nPreview:")
            print(df.head(12).to_string(index=False))


def main():
    run_once()


if __name__ == "__main__":
    main()
