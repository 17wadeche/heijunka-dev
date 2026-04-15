from __future__ import annotations
import argparse
import csv
import datetime as _dt
import json
import os
from typing import Any, Dict, Iterable, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
TEAM_BY_SOURCE: Dict[str, str] = {
    r"C:\Users\wadec8\Medtronic PLC\CQXM RI-Heijunka live spreadsheet shared - Documents\WIP+Non-WIP Heijunka Template CQXM  VSS 2026 03.xlsm": "VSS",
    r"C:\Users\wadec8\Medtronic PLC\Robotics Complaint Intake - Heijunka\RST(US)-Heijunka Surgical.xlsm":"Surgical Robotics",
    r"C:\Users\wadec8\Medtronic PLC\Beeman, Amy - Heijunka Dashboard Endoscopy\WIP+Non-WIP Heijunka Template.xlsm":"Endoscopy",
    r"C:\Users\wadec8\Medtronic PLC\Surgical CQXM Team - !Heijunka\AST-GST(US) - Heijunka Surgical.xlsm":"Surgical AST-GST",
    r"C:\Users\wadec8\Medtronic PLC\SI INV Backlog - Heijunka\Surgical Inv (MEIC) - Heijunka.xlsm":"Surgical AST-GST MEIC",
    r"C:\Users\wadec8\Medtronic PLC\SI INV Backlog - Heijunka\Surgical Inv (US)-Heijunka.xlsm":"Surgical AST-GST",
    r"C:\Users\wadec8\Medtronic PLC\ACM Documents - General\ACM INV (US)-Heijunka v1.0 (002).xlsm":"ACM",
    r"C:\Users\wadec8\Medtronic PLC\Surgical CQXM Team - Legal Mesh -- Shared Folder\WIP+Non-WIP Heijunka Surgical - Legal Team.xlsm":"Surgical Legal",
    r"C:\Users\wadec8\Medtronic PLC\Robotics Complaint Intake - Heijunka\RST(MEIC)-Heijunka Surgical.xlsm":"Surgical Robotics MEIC",
    r"C:\Users\wadec8\Medtronic PLC\Heijunka -VSS,ENDO,ACM - Heijunka -VSS,ENDO,ACM\Heijunka v1.0- ACM - In Use.xlsm":"ACM MEIC",
    r"C:\Users\wadec8\Medtronic PLC\Heijunka -VSS,ENDO,ACM - Heijunka -VSS,ENDO,ACM\Heijunka v1.0- VSS.xlsm":"VSS MEIC",
    r"C:\Users\wadec8\Medtronic PLC\Heijunka -VSS,ENDO,ACM - Heijunka -VSS,ENDO,ACM\Heijunka v1.0-ENDO.xlsm":"Endo MEIC",
    r"C:\Users\wadec8\Medtronic PLC\Surgical MEIC RR team - General\MIR(MEIC)- Heijunka.xlsm":"MEIC MIR",
    r"C:\Users\wadec8\Medtronic PLC\ONEFLOW - Heijunka\06 April\Heijunka Cognizant (ACM-PM).xlsm":"CTS-ACM-PM",
    r"C:\Users\wadec8\Medtronic PLC\ONEFLOW - Heijunka\06 April\Heijunka Cognizant (ACM-RI).xlsm":"CTS-ACM-RI",
    r"C:\Users\wadec8\Medtronic PLC\ONEFLOW - Heijunka\06 April\Heijunka Cognizant (GIS).xlsm":"CTS-GIS",
    r"C:\Users\wadec8\Medtronic PLC\ONEFLOW - Heijunka\06 April\Heijunka Cognizant (RR).xlsm":"CTS-RR",
    r"C:\Users\wadec8\Medtronic PLC\ONEFLOW - Heijunka\06 April\Heijunka Cognizant (SIBO).xlsm":"CTS-SIBO",
    r"C:\Users\wadec8\Medtronic PLC\ONEFLOW - Heijunka\06 April\Heijunka Cognizant (SINH).xlsm":"CTS-SINH",
    r"C:\Users\wadec8\Medtronic PLC\ONEFLOW - Heijunka\06 April\Heijunka Cognizant (Vents).xlsm":"CTS-Vents",
}
TEAM_BY_BASENAME: Dict[str, str] = {
    "WIP+Non-WIP Heijunka Template CQXM  VSS 2026 03.xlsm": "VSS",
    "RST(US)-Heijunka Surgical.xlsm":"Surgical Robotics",
    "WIP+Non-WIP Heijunka Template.xlsm":"Endoscopy",
    "AST-GST(US) - Heijunka Surgical.xlsm":"Surgical AST-GST",
    "ACM INV (US)-Heijunka v1.0 (002).xlsm":"ACM",
    "WIP+Non-WIP Heijunka Surgical - Legal Team.xlsm":"Surgical Legal",
    "Surgical Inv (MEIC) - Heijunka.xlsm":"Surgical AST-GST MEIC",
    "Surgical Inv (US)-Heijunka.xlsm":"Surgical AST-GST",
    "Heijunka v1.0- ACM - In Use.xlsm":"ACM MEIC",
    "Heijunka v1.0- VSS.xlsm":"VSS MEIC",
    "Heijunka v1.0-ENDO.xlsm":"Endo MEIC",
    "RST(MEIC)-Heijunka Surgical.xlsm":"Surgical Robotics MEIC",
    "MIR(MEIC)- Heijunka.xlsm":"MEIC MIR",
    "Heijunka Cognizant (ACM-PM).xlsm":"CTS-ACM-PM",
    "Heijunka Cognizant (ACM-RI).xlsm":"CTS-ACM-RI",
    "Heijunka Cognizant (GIS).xlsm":"CTS-GIS",
    "Heijunka Cognizant (RR).xlsm":"CTS-RR",
    "Heijunka Cognizant (SIBO).xlsm":"CTS-SIBO",
    "Heijunka Cognizant (SINH).xlsm":"CTS-SINH",
    "Heijunka Cognizant (Vents).xlsm":"CTS-Vents",
}
DEFAULT_FILES: List[str] = [
    r"C:\Users\wadec8\Medtronic PLC\CQXM RI-Heijunka live spreadsheet shared - Documents\WIP+Non-WIP Heijunka Template CQXM  VSS 2026 03.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Robotics Complaint Intake - Heijunka\RST(US)-Heijunka Surgical.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Beeman, Amy - Heijunka Dashboard Endoscopy\WIP+Non-WIP Heijunka Template.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Surgical CQXM Team - !Heijunka\AST-GST(US) - Heijunka Surgical.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\ACM Documents - General\ACM INV (US)-Heijunka v1.0 (002).xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Surgical CQXM Team - Legal Mesh -- Shared Folder\WIP+Non-WIP Heijunka Surgical - Legal Team.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\SI INV Backlog - Heijunka\Surgical Inv (MEIC) - Heijunka.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\SI INV Backlog - Heijunka\Surgical Inv (US)-Heijunka.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Robotics Complaint Intake - Heijunka\RST(MEIC)-Heijunka Surgical.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Heijunka -VSS,ENDO,ACM - Heijunka -VSS,ENDO,ACM\Heijunka v1.0- ACM - In Use.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Heijunka -VSS,ENDO,ACM - Heijunka -VSS,ENDO,ACM\Heijunka v1.0- VSS.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Heijunka -VSS,ENDO,ACM - Heijunka -VSS,ENDO,ACM\Heijunka v1.0-ENDO.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\Surgical MEIC RR team - General\MIR(MEIC)- Heijunka.xlsm",
    r"C:\Users\wadec8\Medtronic PLC\ONEFLOW - Heijunka\06 April\Heijunka Cognizant (ACM-PM).xlsm",
    r"C:\Users\wadec8\Medtronic PLC\ONEFLOW - Heijunka\06 April\Heijunka Cognizant (ACM-RI).xlsm",
    r"C:\Users\wadec8\Medtronic PLC\ONEFLOW - Heijunka\06 April\Heijunka Cognizant (GIS).xlsm",
    r"C:\Users\wadec8\Medtronic PLC\ONEFLOW - Heijunka\06 April\Heijunka Cognizant (RR).xlsm",
    r"C:\Users\wadec8\Medtronic PLC\ONEFLOW - Heijunka\06 April\Heijunka Cognizant (SIBO).xlsm",
    r"C:\Users\wadec8\Medtronic PLC\ONEFLOW - Heijunka\06 April\Heijunka Cognizant (SINH).xlsm",
    r"C:\Users\wadec8\Medtronic PLC\ONEFLOW - Heijunka\06 April\Heijunka Cognizant (Vents).xlsm",
]
NAME_COL = 12          # L
ACTIVITY_START_COL = 13  # M
ACTIVITY_END_COL = 21    # U
OOO_COL = 22             # V
NON_D2D_COL = 23  
CSV_COLUMNS = [
    "team",
    "period_date",
    "people_count",
    "total_non_wip_hours",
    "OOO Hours",
    "% in WIP",
    "non_wip_by_person",
    "non_wip_activities",
    "wip_workers",
    "wip_workers_count",
    "wip_workers_ooo_hours",
]
AVAILABILITY_SHEET = "Available WIP+Non-WIP Hours"
def _norm_path(p: str) -> str:
    return os.path.normpath(p)
def team_for_source(path: str) -> str:
    np = _norm_path(path)
    if np in TEAM_BY_SOURCE:
        return TEAM_BY_SOURCE[np]
    return TEAM_BY_BASENAME.get(os.path.basename(np), "")
def _as_text(v: Any) -> str:
    return str(v).strip() if v is not None else ""
def _cell_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None
def _norm_name(s: str) -> str:
    return " ".join((s or "").strip().split()).upper()
def _as_date(v: Any) -> Optional[_dt.date]:
    if v is None:
        return None
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%d-%b-%y", "%d-%b-%Y"):
            try:
                return _dt.datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None
def json_load_safe(v: Any) -> Any:
    if v is None:
        return {}  
    if isinstance(v, (dict, list)):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return {}
        try:
            return json.loads(s)
        except Exception:
            return {}
    return {}
def safe_float2(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s:
            return 0.0
        try:
            return float(s)
        except ValueError:
            return 0.0
    return 0.0
def _sum_simple_map(dst: dict, src: dict) -> None:
    for k, v in (src or {}).items():
        dst[k] = safe_float2(dst.get(k)) + safe_float2(v)
def _merge_unique_list_of_dicts(items: list[dict]) -> list[dict]:
    seen = set()
    out = []
    for item in items:
        if not isinstance(item, dict):
            continue
        key = (
            str(item.get("name", "")).strip(),
            str(item.get("activity", "")).strip(),
            safe_float2(item.get("hours")),
        )
        if key not in seen:
            seen.add(key)
            out.append({
                "name": key[0],
                "activity": key[1],
                "hours": key[2],
            })
    return out
def rollup_non_wip_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    team_rollup_map = {
        "Surgical AST-GST MEIC": "Surgical AST-GST",
        "Surgical INV MEIC": "Surgical AST-GST",
        "Surgical INV US": "Surgical AST-GST",
        "Surgical Robotics MEIC":"Surgical Robotics",
        "ACM MEIC":"ACM",
        "VSS MEIC":"VSS",
        "Endo MEIC":"Endoscopy",
        "CTS-ACM-PM": "ACM",
        "CTS-ACM-RI": "ACM",
        "CTS-GIS": "Endoscopy",
        "CTS-SIBO": "Surgical AST-GST",
        "CTS-SINH": "Surgical AST-GST",
        "CTS-Vents": "VSS",
    }
    buckets: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for row in rows:
        raw_team = (row.get("team", "") or "").strip()
        period_date = (row.get("period_date", "") or "").strip()
        team = team_rollup_map.get(raw_team, raw_team)
        r = dict(row)
        r["team"] = team
        key = (team, period_date)
        buckets.setdefault(key, []).append(r)
    out_rows: List[Dict[str, Any]] = []
    for (team, period_date), group in buckets.items():
        if not team or not period_date:
            if len(group) == 1:
                out_rows.extend(group)
            else:
                out_rows.append(group[0])
            continue
        people_count = 0.0
        total_non_wip_hours = 0.0
        total_ooo_hours = 0.0
        wip_workers_count = 0.0
        wip_workers_ooo_hours = 0.0
        pct_values: List[float] = []
        non_wip_by_person: Dict[str, float] = {}
        non_wip_activities_all: List[dict] = []
        wip_workers_set = set()
        for row in group:
            people_count += safe_float2(row.get("people_count"))
            total_non_wip_hours += safe_float2(row.get("total_non_wip_hours"))
            total_ooo_hours += safe_float2(row.get("OOO Hours"))
            wip_workers_count += safe_float2(row.get("wip_workers_count"))
            wip_workers_ooo_hours += safe_float2(row.get("wip_workers_ooo_hours"))
            pct = row.get("% in WIP")
            if pct not in ("", None):
                pct_values.append(safe_float2(pct))
            _sum_simple_map(non_wip_by_person, json_load_safe(row.get("non_wip_by_person")))
            nwa = json_load_safe(row.get("non_wip_activities"))
            if isinstance(nwa, list):
                non_wip_activities_all.extend(nwa)
            ww = json_load_safe(row.get("wip_workers"))
            if isinstance(ww, list):
                for person in ww:
                    name = str(person).strip()
                    if name:
                        wip_workers_set.add(name)
        pct_in_wip_avg = (
            sum(pct_values) / len(pct_values)
            if pct_values else ""
        )
        out_rows.append({
            "team": team,
            "period_date": period_date,
            "people_count": int(people_count) if float(people_count).is_integer() else people_count,
            "total_non_wip_hours": total_non_wip_hours,
            "OOO Hours": total_ooo_hours,
            "% in WIP": pct_in_wip_avg,
            "non_wip_by_person": dumps_json(non_wip_by_person),
            "non_wip_activities": dumps_json(_merge_unique_list_of_dicts(non_wip_activities_all)),
            "wip_workers": dumps_json(sorted(wip_workers_set)),
            "wip_workers_count": int(wip_workers_count) if float(wip_workers_count).is_integer() else wip_workers_count,
            "wip_workers_ooo_hours": wip_workers_ooo_hours,
        })
    out_rows.sort(key=lambda r: (
        (r.get("team", "") or "").lower(),
        r.get("period_date", "") or "9999-12-31",
    ))
    return out_rows
def parse_week_people_from_left_table(ws: Worksheet, start_row: int, end_row: int) -> List[str]:
    people: List[str] = []
    seen = set()
    current_person = ""
    for row in range(start_row + 1, end_row + 1):
        person_cell = _as_text(ws.cell(row=row, column=3).value)  # col C
        if person_cell:
            current_person = person_cell
        if not current_person:
            continue
        norm = _norm_name(current_person)
        if norm in {"X", "0", "USER1", "USER2", "USER3", "USER4", "USER10", "USER11", "User10", "User11",  "user10", "user11"}:
            continue
        if norm not in seen:
            seen.add(norm)
            people.append(current_person.strip())
    return people
def iso_date(d: Optional[_dt.date]) -> str:
    return d.isoformat() if isinstance(d, _dt.date) else ""
def dumps_json(obj: Any) -> str:
    return json.dumps(obj, ensure_ascii=False)
def safe_div(n: float, d: float) -> Optional[float]:
    if d in (0, 0.0) or d is None:
        return None
    return n / d
def iter_available_blocks(ws: Worksheet) -> Iterable[Tuple[int, _dt.date]]:
    for row in range(1, ws.max_row + 1):
        label = _as_text(ws.cell(row=row, column=1).value)
        if label.lower() == "week starting:":
            week_val = ws.cell(row=row, column=2).value
            week_date = _as_date(week_val)
            if week_date is not None:
                yield row, week_date
def find_next_week_row(starts: List[Tuple[int, _dt.date]], idx: int, max_row: int) -> int:
    if idx + 1 < len(starts):
        return starts[idx + 1][0] - 1
    return max_row
def find_actual_non_wip_header_row(ws: Worksheet, start_row: int, end_row: int) -> Optional[int]:
    for row in range(start_row, min(end_row, start_row + 25) + 1):
        row_vals = [_as_text(ws.cell(row=row, column=col).value).lower() for col in range(12, 23)]
        if "audit" in row_vals and "ooo" in row_vals:
            return row
    return None
def find_column_by_header(ws: Worksheet, header_row: int, header_name: str, start_col: int = 12, end_col: int = 22) -> Optional[int]:
    target = header_name.strip().lower()
    for col in range(start_col, end_col + 1):
        val = _as_text(ws.cell(row=header_row, column=col).value).strip().lower()
        if val == target:
            return col
    return None
def parse_available_sheet(ws: Worksheet) -> Dict[_dt.date, Dict[str, Any]]:
    results: Dict[_dt.date, Dict[str, Any]] = {}
    starts = list(iter_available_blocks(ws))
    for idx, (start_row, week_date) in enumerate(starts):
        end_row = find_next_week_row(starts, idx, ws.max_row)
        week_people = parse_week_people_from_left_table(ws, start_row, end_row)
        non_wip_by_person: Dict[str, float] = {}
        non_wip_activities: List[Dict[str, Any]] = []
        ooo_by_person: Dict[str, float] = {}
        total_non_wip_hours = 0.0
        total_ooo_hours = 0.0
        header_row = None
        for row in range(start_row, end_row + 1):
            m_val = _as_text(ws.cell(row=row, column=ACTIVITY_START_COL).value).strip().lower()
            u_val = _as_text(ws.cell(row=row, column=OOO_COL).value).strip().lower()
            v_val = _as_text(ws.cell(row=row, column=NON_D2D_COL).value).strip().lower()
            if m_val == "audit" and u_val == "ooo" and ("non" in v_val):
                header_row = row
                break
        if header_row is None:
            results[week_date] = {
                "people_count": len(week_people),
                "total_non_wip_hours": 0.0,
                "ooo_hours": 0.0,
                "non_wip_by_person": {},
                "non_wip_activities": [],
                "ooo_by_person": {},
            }
            continue
        for row in range(header_row + 1, end_row + 1):
            name = _as_text(ws.cell(row=row, column=NAME_COL).value)
            clean_name = name.strip()
            lower_name = clean_name.lower()
            if (
                not clean_name
                or lower_name in {"0", "user1", "user2", "user3", "user4", "user10", "user11"}
            ):
                continue
            lower_name = name.strip().lower()
            if lower_name == "total":
                total_non_wip_hours = (_cell_number(ws.cell(row=row, column=NON_D2D_COL).value) or 0.0) / 60.0
                total_ooo_hours = (_cell_number(ws.cell(row=row, column=OOO_COL).value) or 0.0) / 60.0
                continue
            if lower_name == "x":
                continue
            person_non_wip = (_cell_number(ws.cell(row=row, column=NON_D2D_COL).value) or 0.0) / 60.0
            person_ooo = (_cell_number(ws.cell(row=row, column=OOO_COL).value) or 0.0) / 60.0
            clean_name = name.strip()
            non_wip_by_person[clean_name] = person_non_wip
            ooo_by_person[clean_name] = person_ooo
            for col in range(ACTIVITY_START_COL, ACTIVITY_END_COL + 1):
                activity = _as_text(ws.cell(row=header_row, column=col).value).strip()
                mins = _cell_number(ws.cell(row=row, column=col).value) or 0.0
                hours = mins / 60.0
                if hours > 0:
                    non_wip_activities.append({
                        "name": clean_name,
                        "activity": activity,
                        "hours": hours,
                    })
            if person_ooo > 0:
                non_wip_activities.append({
                    "name": clean_name,
                    "activity": "OOO",
                    "hours": person_ooo,
                })
        results[week_date] = {
            "people_count": len(week_people),
            "total_non_wip_hours": total_non_wip_hours,
            "ooo_hours": total_ooo_hours,
            "non_wip_by_person": non_wip_by_person,
            "non_wip_activities": non_wip_activities,
            "ooo_by_person": ooo_by_person,
        }
    return results
def _norm_team(s: str) -> str:
    return (s or "").strip().upper()
def _norm_period_date(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    try:
        return _dt.date.fromisoformat(s).isoformat()
    except ValueError:
        return s
def script_dir() -> str:
    return os.path.dirname(os.path.abspath(__file__))
def parse_people_in_wip(raw: str) -> List[str]:
    s = (raw or "").strip()
    if not s:
        return []
    try:
        val = json.loads(s)
        if isinstance(val, list):
            return [str(x).strip() for x in val if str(x).strip()]
    except Exception:
        pass
    return []
def load_wip_lookup(path: str) -> Dict[Tuple[str, str], Dict[str, Any]]:
    lut: Dict[Tuple[str, str], Dict[str, Any]] = {}
    if not os.path.exists(path):
        return lut
    with open(path, "r", newline="", encoding="utf-8-sig") as fp:
        reader = csv.DictReader(fp)
        for row in reader:
            team = _norm_team(row.get("team", ""))
            period = _norm_period_date(row.get("period_date", ""))
            completed_hours = _cell_number(row.get("Completed Hours"))
            people_in_wip = parse_people_in_wip(row.get("People in WIP", ""))
            if team and period:
                lut[(team, period)] = {
                    "Completed Hours": float(completed_hours or 0.0),
                    "People in WIP": people_in_wip,
                }
    return lut
def find_wip_csv() -> Optional[str]:
    base = script_dir()
    candidates = [
        os.path.join(base, "MS_DATA\\MS_WIP.csv")
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return None
def blank_row_for_missing_file(path: str) -> Dict[str, Any]:
    return {
        "team": team_for_source(path),
        "period_date": "",
        "people_count": "",
        "total_non_wip_hours": "",
        "OOO Hours": "",
        "% in WIP": "",
        "non_wip_by_person": "",
        "non_wip_activities": "",
        "wip_workers": "",
        "wip_workers_count": "",
        "wip_workers_ooo_hours": "",
    }
def scrape_one_workbook(path: str, wip_lut: Dict[Tuple[str, str], Dict[str, Any]]) -> List[Dict[str, Any]]:
    team = team_for_source(path)
    wb = load_workbook(path, data_only=True)
    if AVAILABILITY_SHEET not in wb.sheetnames:
        return [{
            "team": team,
            "period_date": "",
            "people_count": "",
            "total_non_wip_hours": "",
            "OOO Hours": "",
            "% in WIP": "",
            "non_wip_by_person": "",
            "non_wip_activities": "",
            "wip_workers": "",
            "wip_workers_count": "",
            "wip_workers_ooo_hours": "",
        }]
    available_by_week = parse_available_sheet(wb[AVAILABILITY_SHEET])
    rows: List[Dict[str, Any]] = []
    for period in sorted(available_by_week.keys()):
        av = available_by_week[period]
        key = (_norm_team(team), iso_date(period))
        wip_info = wip_lut.get(key, {})
        completed_hours = float(wip_info.get("Completed Hours", 0.0) or 0.0)
        wip_workers = list(wip_info.get("People in WIP", []) or [])
        wip_workers_set = {_norm_name(x) for x in wip_workers if str(x).strip()}
        total_non_wip_hours = float(av.get("total_non_wip_hours", 0.0) or 0.0)
        pct_in_wip = safe_div(completed_hours, completed_hours + total_non_wip_hours)
        ooo_by_person = av.get("ooo_by_person", {}) or {}
        ooo_by_person_norm = {
            _norm_name(name): float(val or 0.0)
            for name, val in ooo_by_person.items()
        }
        wip_workers_ooo_hours = 0.0
        for worker in wip_workers_set:
            wip_workers_ooo_hours += ooo_by_person_norm.get(worker, 0.0)
        row = {
            "team": team,
            "period_date": iso_date(period),
            "people_count": av.get("people_count", 0),
            "total_non_wip_hours": total_non_wip_hours,
            "OOO Hours": float(av.get("ooo_hours", 0.0) or 0.0),
            "% in WIP": float(pct_in_wip) if pct_in_wip is not None else "",
            "non_wip_by_person": dumps_json(av.get("non_wip_by_person", {})),
            "non_wip_activities": dumps_json(av.get("non_wip_activities", [])),
            "wip_workers": dumps_json(wip_workers),
            "wip_workers_count": len(wip_workers),
            "wip_workers_ooo_hours": wip_workers_ooo_hours,
        }
        rows.append(row)
    return rows
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("files", nargs="*", help="Optional workbook(s) to scrape. If omitted, uses DEFAULT_FILES in the script.")
    ap.add_argument("--out", default="MS_DATA\\ms_non_wip_activities.csv", help="Output CSV path.")
    args = ap.parse_args()
    wip_csv = find_wip_csv()
    wip_lut: Dict[Tuple[str, str], Dict[str, Any]] = {}
    if wip_csv:
        wip_lut = load_wip_lookup(wip_csv)
    files = args.files or DEFAULT_FILES
    all_rows: List[Dict[str, Any]] = []
    for path in files:
        if not os.path.exists(path):
            all_rows.append(blank_row_for_missing_file(path))
            continue
        all_rows.extend(scrape_one_workbook(path, wip_lut))
    final_rows = rollup_non_wip_rows(all_rows)
    with open(args.out, "w", newline="", encoding="utf-8") as fp:
        writer = csv.DictWriter(fp, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for row in final_rows:
            writer.writerow({k: row.get(k, "") for k in CSV_COLUMNS})
    print(f"Wrote {len(final_rows)} row(s) to {args.out}")
    return 0
if __name__ == "__main__":
    raise SystemExit(main())