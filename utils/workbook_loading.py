from __future__ import annotations
import warnings
from typing import Any
from openpyxl import load_workbook
_MISSING_PIVOT_FORMULA_ERROR = (
    "openpyxl.pivot.cache.CalculatedItem'>.formula should be <class 'str'> "
    "but value is <class 'NoneType'>"
)
def load_data_only_workbook(path: str) -> Any:
    try:
        return load_workbook(path, data_only=True)
    except TypeError as exc:
        if _MISSING_PIVOT_FORMULA_ERROR not in str(exc):
            raise
        warnings.warn(
            f"{path}: malformed pivot-cache calculated item; retrying in "
            "read-only mode and ignoring pivot caches",
            RuntimeWarning,
            stacklevel=2,
        )
        return load_workbook(path, data_only=True, read_only=True, keep_links=False)