#!/usr/bin/env python3
from __future__ import annotations
import argparse, csv, json, math, os, sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from typing import Any, Dict, Iterable, List, Optional, Tuple
def _to_date(v) -> Optional[date]:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:
        n = float(s)
        return (datetime(1899, 12, 30) + timedelta(days=n)).date()
    except Exception:
        pass
    for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
        try:
            if fmt == "%Y-%m-%d":
                return datetime.fromisoformat(s).date()
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None
def _to_float(x) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, float):
        return x
    try:
        return float(str(x).replace(",", "").strip())
    except Exception:
        return None
def _clean(s: Any) -> str:
    if s is None:
        return ""
    if isinstance(s, float) and math.isnan(s):
        return ""
    return str(s).strip()
def _sheetnames_xlsb(path: str) -> List[str]:
    import pandas as pd
    with pd.ExcelFile(path, engine="pyxlsb") as xf:
        return list(xf.sheet_names)
def _rows_from_xlsb(path: str, sheet_name: str) -> Iterable[Tuple[Any, ...]]:
    import pandas as pd
    df = pd.read_excel(path, sheet_name=sheet_name, engine="pyxlsb", header=None)
    for row in df.itertuples(index=False, name=None):
        yield tuple(row)
def _sheetnames_xlsx_like(path: str) -> List[str]:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    return list(wb.sheetnames)
def _rows_from_xlsx_like_visible(path: str, sheet_name: str) -> Iterable[Tuple[Any, ...]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[sheet_name]
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            row_vals.append(ws.cell(r, c).value)
        yield tuple(row_vals)
from typing import Set 
def _hidden_rows_xlsx_like(path: str, sheet_name: str) -> Set[int]:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[sheet_name]
    max_row = ws.max_row or 0
    hidden: Set[int] = set()
    for r in range(1, max_row + 1):
        rd = ws.row_dimensions.get(r)
        if rd is None:
            continue
        is_hidden = bool(getattr(rd, "hidden", False))
        zero_h = (getattr(rd, "height", None) == 0)
        if is_hidden or zero_h:
            hidden.add(r)
    return hidden
def _hidden_rows(path: str, sheet_name: str) -> Set[int]:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _hidden_rows_xlsx_like(path, sheet_name)
    return set()
def _get_io(path: str):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsb":
        return _sheetnames_xlsb, _rows_from_xlsb
    elif ext in (".xlsx", ".xlsm"):
        return _sheetnames_xlsx_like, _rows_from_xlsx_like_visible
    else:
        raise ValueError(f"Unsupported workbook extension '{ext}'. Use .xlsb/.xlsx/.xlsm")
def _find_sheet_by_hint(sheet_names: List[str], hint: str) -> str:
    if not hint:
        raise ValueError("Empty sheet hint")
    exact = [nm for nm in sheet_names if nm.strip().lower() == hint.strip().lower()]
    if exact:
        return exact[0]
    contains = [nm for nm in sheet_names if hint.lower() in nm.lower()]
    if contains:
        return contains[0]
    raise ValueError(f"Sheet '{hint}' not found. Available: {sheet_names}")
def _week_from_row(ridx: int, anchors: List[Dict[str, Any]]) -> Optional[date]:
    if not anchors:
        return None
    parsed = []
    for a in anchors:
        try:
            r = int(a.get("row"))
            d = _to_date(a.get("date"))
            if d: parsed.append((r, d))
        except Exception:
            pass
    if not parsed:
        return None
    parsed.sort(key=lambda x: x[0])  # by row index
    wk = None
    for r, d in parsed:
        if ridx >= r:
            wk = d
        else:
            break
    return wk
def _is_non_person_label(name: Any) -> bool:
    s = _clean(name).lower()
    if not s:
        return True
    banned_exact = {
        "available hours", "available", "ooo hours", "total hours",
        "total", "people count", "hc", "hc in wip", "actual hc used",
    }
    if s in banned_exact:
        return True
    if "hour" in s or "total" in s:
        return True
    return False
def people_by_week_from_available(
    rows: Iterable[Tuple[Any, ...]],
    anchors: List[Dict[str, Any]],
    hidden_rows: Optional[Set[int]] = None,
) -> Dict[date, set]:
    PEOPLE_COL = 2  # C
    out: Dict[date, set] = defaultdict(set)
    hidden_rows = hidden_rows or set()
    for ridx, r in enumerate(rows, start=1):
        if ridx in hidden_rows:
            continue  
        r = r or tuple()
        wk = _week_from_row(ridx, anchors)
        if not wk:
            continue
        name = _clean(r[PEOPLE_COL] if len(r) > PEOPLE_COL else "")
        if not name or _is_non_person_label(name):
            continue
        out[wk].add(name)
    return out
def parse_prod_analysis(
    rows: Iterable[Tuple[Any, ...]],
    anchors: List[Dict[str, Any]],
    hidden_rows: Optional[Set[int]] = None,
) -> Dict[date, Dict[str, Any]]:
    COL_NAME, COL_FLAG, COL_MINUTES, COL_ACTIVITY = 3, 4, 8, 12
    COL_HOURS_F = 5
    nonwip_flags = {"non wip", "non-wip"}
    other_team_wip_flags = {"other team wip", "other-team-wip", "otherteamwip", "other team's wip"}
    buckets: Dict[date, Dict[str, Any]] = defaultdict(lambda: {
        "ooo_hours": 0.0,
        "ooo_by_person": defaultdict(float),
        "non_wip_by_person": defaultdict(float),
        "non_wip_activities": [],
    })
    import re
    hidden_rows = hidden_rows or set()
    for ridx, r in enumerate(rows, start=1):
        if ridx in hidden_rows:
            continue  # <-- skip hidden Excel rows
        r = r or tuple()
        wk = _week_from_row(ridx, anchors)
        if not wk:
            continue
        name = _clean(r[COL_NAME] if len(r) > COL_NAME else "")
        flag = _clean(r[COL_FLAG] if len(r) > COL_FLAG else "").lower()
        mins = _to_float(r[COL_MINUTES] if len(r) > COL_MINUTES else None) or 0.0
        act  = _clean(r[COL_ACTIVITY] if len(r) > COL_ACTIVITY else "")
        if not (flag or mins or name or act):
            continue
        b = buckets[wk]
        if flag == "ooo" and mins > 0:
            hrs = mins / 60.0
            b["ooo_hours"] += hrs
            if name:
                b["ooo_by_person"][name] += hrs
        act_key = "".join(ch for ch in act.upper() if ch.isalnum())
        is_other_team_wip = (
            flag in other_team_wip_flags
            or act_key == "OTHERTEAMWIP"
        )
        if flag in nonwip_flags or is_other_team_wip:
            hrs: float = 0.0
            if mins > 0:
                hrs = mins / 60.0
            else:
                comment = _clean(r[COL_ACTIVITY] if len(r) > COL_ACTIVITY else "")
                if comment:
                    nums = [ _to_float(m) or 0.0 for m in re.findall(r"\d+(?:\.\d+)?", comment) ]
                    total_comment_mins = sum(nums)
                else:
                    total_comment_mins = 0.0

                if total_comment_mins > 0:
                    hrs = total_comment_mins / 60.0
                else:
                    hrs_f = _to_float(r[COL_HOURS_F] if len(r) > COL_HOURS_F else None) or 0.0
                    if hrs_f > 0:
                        hrs = hrs_f
            if hrs > 0:
                if name:
                    b["non_wip_by_person"][name] += hrs
                label = "Other Team WIP" if is_other_team_wip else (act or "Non-WIP")
                b["non_wip_activities"].append({
                    "name": name,
                    "activity": label,
                    "hours": round(hrs, 2),
                })
    for wk, b in buckets.items():
        b["ooo_hours"] = round(b["ooo_hours"], 2)
        b["ooo_by_person"] = {k: round(v, 2) for k, v in b["ooo_by_person"].items()}
        b["non_wip_by_person"] = {k: round(v, 2) for k, v in b["non_wip_by_person"].items()}
    return buckets
def load_completed_hours(metrics_csv: str) -> Dict[Tuple[str, str], float]:
    out: Dict[Tuple[str, str], float] = {}
    with open(metrics_csv, "r", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            team = _clean(row.get("Team") or row.get("team") or "")
            wk = _clean(row.get("Week") or row.get("period_date") or "")
            ch = _to_float(row.get("Completed Hours") or row.get("completed_hours") or "0") or 0.0
            if team and wk:
                out[(team, wk)] = out.get((team, wk), 0.0) + ch
    return out
def weeks_for_team(metrics_csv: str, team: str) -> List[str]:
    weeks = set()
    with open(metrics_csv, "r", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            if _clean(row.get("Team")) == team:
                w = _clean(row.get("Week"))
                if w:
                    weeks.add(w)
    return sorted(weeks)
def _loads_json_maybe(s: str) -> Optional[dict]:
    if not s:
        return None
    s = s.strip()
    try:
        return json.loads(s)
    except Exception:
        try:
            return json.loads(s.replace("''", '"').replace('""', '"'))
        except Exception:
            return None
def load_person_metrics(metrics_csv: str) -> Dict[Tuple[str, str, str], Dict[str, float]]:
    out: Dict[Tuple[str, str, str], Dict[str, float]] = {}
    with open(metrics_csv, "r", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            team = _clean(row.get("Team") or row.get("team") or "")
            week = _clean(row.get("Week") or row.get("period_date") or "")
            if not (team and week):
                continue
            ph_json_raw = row.get("Person Hours") or row.get("person_hours") or ""
            ph = _loads_json_maybe(ph_json_raw) or {}
            if not isinstance(ph, dict):
                continue
            for person, payload in ph.items():
                if not person or _is_non_person_label(person):   # <-- skip labels
                    continue
                actual_val = None
                if isinstance(payload, dict):
                    actual_val = _to_float(payload.get("actual"))
                actual = float(actual_val) if actual_val is not None else 0.0
                out[(team, week, _clean(person))] = {"actual": actual}
    return out
def build_non_wip_rows(config_path: str,
                       chosen_teams: Optional[List[str]],
                       all_teams: bool,
                       metrics_csv: str) -> List[Dict[str, Any]]:
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    teams_to_run = list(cfg.keys()) if all_teams else (chosen_teams or [])
    if not teams_to_run:
        raise SystemExit("No teams specified. Use --all or --team <NAME> (repeatable).")
    completed_index = load_completed_hours(metrics_csv)
    person_metrics = load_person_metrics(metrics_csv)
    out_rows: List[Dict[str, Any]] = []
    for team in teams_to_run:
        entry = cfg.get(team) or {}
        path = entry.get("workbook")
        if not path or not os.path.exists(path):
            raise SystemExit(f"[{team}] Workbook not found: {path}")
        irl_people = {p.strip() for p in (entry.get("irl_people") or [])}
        prod_cfg = entry.get("prod_sheets") or entry.get("prod_sheet") or []
        prod_hints = prod_cfg if isinstance(prod_cfg, list) else [prod_cfg]
        avail_hint = entry.get("avail_sheet")
        if not avail_hint:
            raise SystemExit(f"[{team}] Missing 'avail_sheet' in config.")
        get_names, get_rows = _get_io(path)
        sheet_names = get_names(path)
        prod_buckets_merged: Dict[date, Dict[str, Any]] = defaultdict(lambda: {
            "ooo_hours": 0.0,
            "ooo_by_person": defaultdict(float),   # person -> OOO hours for that week
            "non_wip_by_person": defaultdict(float),
            "non_wip_activities": [],
        })
        for hint in prod_hints:
            if not hint:
                continue
            nm = _find_sheet_by_hint(sheet_names, hint)
            rows_s = list(get_rows(path, nm))
            anchors_s = (entry.get("week_anchors", {}) or {}).get(nm, [])
            hidden_prod_rows = _hidden_rows(path, nm)
            pb = parse_prod_analysis(rows_s, anchors_s, hidden_prod_rows)
            for wk, b in pb.items():
                prod_buckets_merged[wk]["ooo_hours"] += b.get("ooo_hours", 0.0)
                for person, hrs in (b.get("ooo_by_person", {}) or {}).items():
                    prod_buckets_merged[wk]["ooo_by_person"][person] += hrs
                for person, hrs in (b.get("non_wip_by_person", {}) or {}).items():
                    prod_buckets_merged[wk]["non_wip_by_person"][person] += hrs
                prod_buckets_merged[wk]["non_wip_activities"].extend(b.get("non_wip_activities", []))
        for wk, b in prod_buckets_merged.items():
            b["ooo_hours"] = round(b["ooo_hours"], 2)
            b["ooo_by_person"] = {k: round(float(v), 2) for k, v in b["ooo_by_person"].items()}
            b["non_wip_by_person"] = {k: round(float(v), 2) for k, v in b["non_wip_by_person"].items()}
        avail_name = _find_sheet_by_hint(sheet_names, avail_hint)
        avail_rows = list(get_rows(path, avail_name))
        avail_anchors = (entry.get("week_anchors", {}) or {}).get(avail_name, [])
        hidden_avail_rows = _hidden_rows(path, avail_name)
        people_by_week = people_by_week_from_available(
            avail_rows,
            avail_anchors,
            hidden_avail_rows,
        )
        if team == "Aortic":
            print("Aortic people_by_week from avail:")
            for d, names in sorted(people_by_week.items()):
                print(" ", d, "->", len(names), names)
        team_weeks = weeks_for_team(metrics_csv, team)
        for iso in team_weeks:
            wk_date = _to_date(iso)
            wk_people = people_by_week.get(wk_date, set()) if wk_date else set()
            wk_people = {p for p in wk_people if not _is_non_person_label(p)}   # <-- add this
            people_count = len(wk_people)
            completed = float(completed_index.get((team, iso), 0.0))
            bucket = prod_buckets_merged.get(wk_date, {}) if wk_date else {}
            ooo_by_person = bucket.get("ooo_by_person", {})
            per_person_nonwip: Dict[str, float] = {}
            for person in sorted(wk_people):
                ooo_p = float(ooo_by_person.get(person, 0.0))
                base_capacity = 39.0 if person in irl_people else 40.0
                capacity = max(0.0, base_capacity - ooo_p)
                pm = person_metrics.get((team, iso, person), {"available": None, "actual": 0.0})
                available = pm.get("available")  # may be None
                actual = float(pm.get("actual", 0.0))
                effective_person_hours = (
                    min(capacity, float(available))
                    if (available is not None) else
                    capacity
                )
                per_person_nonwip[person] = round(max(0.0, effective_person_hours - actual), 2)
            total_non_wip_hours = round(sum(per_person_nonwip.values()), 2)
            ooo_hours_total = float(bucket.get("ooo_hours", 0.0))
            denom = completed + total_non_wip_hours
            pct_in_wIP = round((completed / denom * 100.0), 2) if denom > 0 else None
            activities = bucket.get("non_wip_activities", [])
            out_rows.append({
                "Team": team,
                "Week": iso,
                "People Count": people_count,
                "Total Non-WIP Hours": total_non_wip_hours,
                "OOO Hours": round(ooo_hours_total, 2),
                "% in WIP": pct_in_wIP,
                "Non-WIP by Person": json.dumps(per_person_nonwip, ensure_ascii=False),
                "Non-WIP Activities": json.dumps(activities, ensure_ascii=False),
            })
        print(f"[{team}] OK: {len(team_weeks)} week(s)")
    return out_rows
def main():
    ap = argparse.ArgumentParser(description="Collect Non-WIP/OOO metrics into a new CSV.")
    ap.add_argument("--config", required=True, help="Path to teams.json")
    ap.add_argument("--metrics", required=True, help="Path to metrics.csv produced by heijunka_new_layout.py")
    ap.add_argument("--team", action="append", help="Team name from teams.json (repeatable)")
    ap.add_argument("--all", action="store_true", help="Process all teams in config")
    ap.add_argument("--out", default="non_wip.csv", help="Output CSV path (default: non_wip.csv)")
    args = ap.parse_args()
    try:
        rows = build_non_wip_rows(
            config_path=args.config,
            chosen_teams=args.team,
            all_teams=args.all,
            metrics_csv=args.metrics,
        )
    except SystemExit as e:
        print(str(e), file=sys.stderr); sys.exit(2)
    except Exception as e:
        print(f"Failed: {e}", file=sys.stderr); sys.exit(1)
    if not rows:
        print("No data produced.", file=sys.stderr); sys.exit(1)
    cols = [
        "Team", "Week",
        "People Count",
        "Total Non-WIP Hours",
        "OOO Hours",
        "% in WIP",
        "Non-WIP by Person",
        "Non-WIP Activities",
    ]
    with open(args.out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(rows)
    print(f"Wrote {len(rows)} rows -> {args.out}")
if __name__ == "__main__":
    main()