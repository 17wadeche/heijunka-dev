from __future__ import annotations
import argparse
import csv
import datetime as _dt
import json
import os
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
def _load_workbook_for_scraping(path: str):
    return load_workbook(path, data_only=True, keep_links=False, read_only=True)
MCS_DEFAULT_PATH = r"C:\Users\wadec8\Medtronic PLC\MCS COS Transformation - VMB Scheduling\Heijunka Current.xlsm"
MCS_REFRESH_MIN_PERIOD = _dt.date(2026, 6, 22)
DS_DEFAULT_DIR = r"C:\Users\wadec8\Medtronic PLC\Defibrillation Solutions - Schedule and PAB"
DS_ARCHIVE = r"C:\Users\wadec8\Medtronic PLC\Defibrillation Solutions - Schedule and PAB\Archive"
CPT_DEFAULT_DIR = r"C:\Users\wadec8\Medtronic PLC\Cardiac Pacing Therapies CQXM - Heijunka & PAB"
CPT_ARCHIVE_PAB_DIR =r"C:\Users\wadec8\Medtronic PLC\Cardiac Pacing Therapies CQXM - Heijunka & PAB\Archive\2026\4. April 2026"
CPT_ARCHIVE_PAB_DIR2 = r"C:\Users\wadec8\Medtronic PLC\Cardiac Pacing Therapies CQXM - Heijunka & PAB\Archive\2026\5. May 2026"
CPT_ARCHIVE_PAB_DIR3 = r"C:\Users\wadec8\Medtronic PLC\Cardiac Pacing Therapies CQXM - Heijunka & PAB\Archive\2026\6. Jun 2026"
CDS_DEFAULT_DIR = r"C:\Users\wadec8\Medtronic PLC\Diagnostics MDR - Heijunka and Production Analysis"
CDS_ARCHIVE_PAB_DIR = r"C:\Users\wadec8\Medtronic PLC\Diagnostics MDR - Heijunka and Production Analysis\Archived PAB"
NI_DEFAULT_DIR = r"C:\Users\wadec8\Medtronic PLC\Tier1 PXM - Non Implantables - Heijunka"
NI_ARCHIVE_APRIL_2026_DIR = r"C:\Users\wadec8\Medtronic PLC\Tier1 PXM - Non Implantables - Heijunka\Archived PAB\April 2026 - PAB"
NI_ARCHIVE = r"C:\Users\wadec8\Medtronic PLC\Tier1 PXM - Non Implantables - Heijunka\Archived PAB"
MEIC_DEFAULT_DIR = r"C:\Users\wadec8\Medtronic PLC\CRM CQXM Reports - 1.9 Heijunka Tracker"
PM_CTS_DEFAULT_DIR = r"C:\Users\wadec8\Medtronic PLC\Tier1 PXM - Non Implantables - Heijunka\PM-CTS PAB"
TEAM_BY_SOURCE: Dict[str, str] = {
    os.path.normpath(MCS_DEFAULT_PATH): "MCS",
}
TEAM_BY_BASENAME: Dict[str, str] = {
    "Heijunka Current.xlsm": "MCS",
}
EXCLUDED_FILES = {
    os.path.normpath(p)
    for p in [
        r"C:\Users\wadec8\Medtronic PLC\Cardiac Pacing Therapies CQXM - Heijunka & PAB\Archive\2026\4. April 2026\Not USED Week 20 Apr 2026 Heijunka & PAB.xlsm",
        r"C:\Users\wadec8\Medtronic PLC\Defibrillation Solutions - Schedule and PAB\Archive\Assigned DS COS PEs for 2025.xlsx",
        r"C:\Users\wadec8\Medtronic PLC\Defibrillation Solutions - Schedule and PAB\Archive\(will be archived) DS_Schedule_PAS 6.5 V1.xlsx",
        r"C:\Users\wadec8\Medtronic PLC\Defibrillation Solutions - Schedule and PAB\Archive\CPT Event Support.xlsx",
        r"C:\Users\wadec8\Medtronic PLC\Defibrillation Solutions - Schedule and PAB\Archive\DS Production Analysis Sheet and Schedule.xlsx",
        r"C:\Users\wadec8\Medtronic PLC\Tier1 PXM - Non Implantables - Heijunka\PM-CTS PAB\Revised PM-CTS Template.xlsm",
    ]
}
_AVAIL_PAT = re.compile(r"\bavailability\b", re.IGNORECASE)
_PROD_PAT = re.compile(r"\b(production|product)\s+analysis\b", re.IGNORECASE)
_MONTH_MAP = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}
CSV_COLUMNS = [
    "team",
    "period_date",
    "people_count",
    "total_non_wip_hours",
    "OOO Hours",
    "% in WIP",
    "non_wip_by_person",
    "non_wip_activities",
    "wip_workers",
    "wip_workers_count",
    "wip_workers_ooo_hours",
]
LIT_LETTERS_TEAM_NAME = "Lit & Letters"
LIT_LETTERS_PAB_SHEET = "#3 PAB"
LIT_LETTERS_PERF_WIP_SHEET = "#6 Performance WIP Time"
LIT_LETTERS_NON_WIP_TYPES = {"essential non-wip", "non-wip"}
LIT_LETTERS_PEOPLE_COUNT = 7
LIT_LETTERS_PEOPLE_COUNT_EFFECTIVE_DATE = _dt.date(2026, 6, 8)
LIT_LETTERS_PEOPLE_COUNT_FROM_EFFECTIVE_DATE = 8
def lit_letters_people_count_for_period(period: _dt.date) -> int:
    if period >= LIT_LETTERS_PEOPLE_COUNT_EFFECTIVE_DATE:
        return LIT_LETTERS_PEOPLE_COUNT_FROM_EFFECTIVE_DATE
    return LIT_LETTERS_PEOPLE_COUNT
PM_CTS_TEAM_NAME = "PM-CTS"
PM_CTS_IND_TEAM_NAME = "PM-CTS IND"
PM_CTS_IND_START = _dt.date(2026, 6, 29)
PM_CTS_REMOVAL_START = _dt.date(2026, 6, 29)
PM_CTS_PAB_SHEET = "#2 PAB"
PM_CTS_PERF_WIP_SHEET = "#3 Performance WIP Time"
PM_CTS_NON_WIP_TYPES = {"essential non-wip", "non-wip"}
PM_CTS_PEOPLE_COUNT = 14
DS_PAB_SHEET = "#2 PAB"
DS_WIP_PLAN_SHEET = "# 1 WIP plan"
DS_PERF_WIP_SHEET = "#5 Performance WIP Time"
DS_NON_WIP_TYPES = {"essential non-wip", "non-wip"}
DS_PEOPLE_COUNT = 41
DS_PEOPLE_COUNT_EFFECTIVE_DATE = _dt.date(2026, 6, 1)
DS_PEOPLE_COUNT_FROM_EFFECTIVE_DATE = 40
DS_PEOPLE_COUNT_SECOND_EFFECTIVE_DATE = _dt.date(2026, 6, 8)
DS_PEOPLE_COUNT_FROM_SECOND_EFFECTIVE_DATE = 39
DS_PEOPLE_COUNT_THIRD_EFFECTIVE_DATE = _dt.date(2026, 7, 13)
DS_PEOPLE_COUNT_FROM_THIRD_EFFECTIVE_DATE = 37
def ds_people_count_for_period(period: _dt.date) -> int:
    if period >= DS_PEOPLE_COUNT_THIRD_EFFECTIVE_DATE:
        return DS_PEOPLE_COUNT_FROM_THIRD_EFFECTIVE_DATE
    if period >= DS_PEOPLE_COUNT_SECOND_EFFECTIVE_DATE:
        return DS_PEOPLE_COUNT_FROM_SECOND_EFFECTIVE_DATE
    if period >= DS_PEOPLE_COUNT_EFFECTIVE_DATE:
        return DS_PEOPLE_COUNT_FROM_EFFECTIVE_DATE
    return DS_PEOPLE_COUNT
CPT_PAB_SHEET = "#3 PAB"
CPT_WIP_PLAN_SHEET = "# 1 WIP plan"
CPT_PERF_WIP_SHEET = "#6 Performance WIP Time"
CPT_NON_WIP_TYPES = {"essential non-wip", "non-wip"}
CPT_PEOPLE_COUNT = 42
CPT_PEOPLE_COUNT_EFFECTIVE_DATE = _dt.date(2026, 5, 4)
CPT_PEOPLE_COUNT_FROM_EFFECTIVE_DATE = 43
CPT_PEOPLE_COUNT_SECOND_EFFECTIVE_DATE = _dt.date(2026, 6, 29)
CPT_PEOPLE_COUNT_FROM_SECOND_EFFECTIVE_DATE = 42
def cpt_people_count_for_period(period: _dt.date) -> int:
    if period >= CPT_PEOPLE_COUNT_SECOND_EFFECTIVE_DATE:
        return CPT_PEOPLE_COUNT_FROM_SECOND_EFFECTIVE_DATE
    if period >= CPT_PEOPLE_COUNT_EFFECTIVE_DATE:
        return CPT_PEOPLE_COUNT_FROM_EFFECTIVE_DATE
    return CPT_PEOPLE_COUNT
CDS_PAB_SHEET = "#2 PAB"
CDS_WIP_PLAN_SHEET = "# 1 WIP plan"
CDS_PERF_METRICS_SHEET = "#4 Performance Metrics"
CDS_PERF_WIP_SHEET = "#5 Performance WIP Time"
CDS_NON_WIP_TYPES = {"essential non-wip", "non-wip"}
CDS_PEOPLE_COUNT = 6
CDS_PEOPLE_COUNT_EFFECTIVE_DATE = _dt.date(2026, 4, 24)
CDS_PEOPLE_COUNT_FROM_EFFECTIVE_DATE = 5
CDS_PEOPLE_COUNT_SECOND_EFFECTIVE_DATE = _dt.date(2026, 6, 22)
CDS_PEOPLE_COUNT_FROM_SECOND_EFFECTIVE_DATE = 6
def cds_people_count_for_period(period: _dt.date) -> int:
    if period >= CDS_PEOPLE_COUNT_SECOND_EFFECTIVE_DATE:
        return CDS_PEOPLE_COUNT_FROM_SECOND_EFFECTIVE_DATE
    if period >= CDS_PEOPLE_COUNT_EFFECTIVE_DATE:
        return CDS_PEOPLE_COUNT_FROM_EFFECTIVE_DATE
    return CDS_PEOPLE_COUNT
NI_PAB_SHEET = "#2 PAB"
NI_WIP_PLAN_SHEET = "# 1 WIP plan"
NI_PERF_METRICS_SHEET = "#4 Performance Metrics"
NI_PERF_WIP_SHEET = "#5 Performance WIP Time"
NI_NON_WIP_TYPES = {"essential non-wip", "non-wip"}
NI_PEOPLE_COUNT = 8
NI_PEOPLE_COUNT_EFFECTIVE_DATE = _dt.date(2026, 4, 17)
NI_PEOPLE_COUNT_FROM_EFFECTIVE_DATE = 7
NI_PEOPLE_COUNT_SECOND_EFFECTIVE_DATE = _dt.date(2026, 6, 29)
NI_PEOPLE_COUNT_FROM_SECOND_EFFECTIVE_DATE = 6
NI_PEOPLE_COUNT_THIRD_EFFECTIVE_DATE = _dt.date(2026, 7, 13)
NI_PEOPLE_COUNT_FROM_THIRD_EFFECTIVE_DATE = 5
def ni_people_count_for_period(period: _dt.date) -> int:
    if period >= NI_PEOPLE_COUNT_THIRD_EFFECTIVE_DATE:
            return NI_PEOPLE_COUNT_FROM_THIRD_EFFECTIVE_DATE
    if period >= NI_PEOPLE_COUNT_SECOND_EFFECTIVE_DATE:
        return NI_PEOPLE_COUNT_FROM_SECOND_EFFECTIVE_DATE
    if period >= NI_PEOPLE_COUNT_EFFECTIVE_DATE:
        return NI_PEOPLE_COUNT_FROM_EFFECTIVE_DATE
    return NI_PEOPLE_COUNT
NI_PERF_WIP_OOO_EFFECTIVE_DATE = _dt.date(2026, 5, 25)
NI_PERF_WIP_OOO_TOTAL_CELL = "Z12"
NI_PERF_WIP_OOO_NAME_COL = "A"
NI_PERF_WIP_OOO_HOURS_COL = "Z"
NI_PERF_WIP_OOO_START_ROW = 5
NI_PERF_WIP_OOO_END_ROW = 11
MEIC_PAB_SHEET = "#2 PAB"
MEIC_WIP_PLAN_SHEET = "# 1 WIP plan"
MEIC_PERF_WIP_SHEET = "#5 Performance WIP Time"
MEIC_NON_WIP_TYPES = {"essential non-wip", "non-wip"}
MEIC_PEOPLE_COUNT = 19
MEIC_TEAM_NAME = "NI & PM MEIC"
def _norm_path(p: str) -> str:
    return os.path.normpath(p)
def is_lit_letters_path(path: str) -> bool:
    base = os.path.basename(_norm_path(path)).lower()
    return (
        "pab for letters" in base
        and "lit" in base
        and "principals" in base
    )
def _is_pm_cts_ind_file(path: str) -> bool:
    base = os.path.basename(_norm_path(path)).lower()
    period = parse_period_date_from_filename(path)
    return "pm-cts ind" in base and (period is None or period >= PM_CTS_IND_START)
def _is_removed_pm_cts_file(path: str) -> bool:
    base = os.path.basename(_norm_path(path)).lower()
    if "pm-cts" not in base or "pm-cts ind" in base:
        return False
    period = parse_period_date_from_filename(path)
    return isinstance(period, _dt.date) and period >= PM_CTS_REMOVAL_START
def team_for_source(path: str) -> str:
    np = _norm_path(path)
    if is_lit_letters_path(np):
        return LIT_LETTERS_TEAM_NAME
    if _is_pm_cts_ind_file(np):
        return PM_CTS_IND_TEAM_NAME
    if _is_removed_pm_cts_file(np):
        return ""
    if np in TEAM_BY_SOURCE:
        return TEAM_BY_SOURCE[np]
    ds_root = _norm_path(DS_DEFAULT_DIR)
    if np.startswith(ds_root + os.sep) or np == ds_root:
        return "DS"
    cpt_root = _norm_path(CPT_DEFAULT_DIR)
    if np.startswith(cpt_root + os.sep) or np == cpt_root:
        return "CPT"
    cds_root = _norm_path(CDS_DEFAULT_DIR)
    if np.startswith(cds_root + os.sep) or np == cds_root:
        return "CDS"
    pm_cts_root = _norm_path(PM_CTS_DEFAULT_DIR)
    if np.startswith(pm_cts_root + os.sep) or np == pm_cts_root:
        return "PM-CTS"
    ni_root = _norm_path(NI_DEFAULT_DIR)
    if np.startswith(ni_root + os.sep) or np == ni_root:
        return "NI"
    meic_root = _norm_path(MEIC_DEFAULT_DIR)
    if np.startswith(meic_root + os.sep) or np == meic_root:
        return MEIC_TEAM_NAME
    base = os.path.basename(np)
    base_lower = base.lower()
    if base in TEAM_BY_BASENAME:
        return TEAM_BY_BASENAME[base]
    if "cpt" in base_lower and "pab" in base_lower:
        return "CPT"
    np_lower = np.lower()
    if "pab for letters" in np_lower and "lit" in np_lower and "principals" in np_lower:
        return LIT_LETTERS_TEAM_NAME
    if "defibrillation solutions" in np_lower:
        return "DS"
    if "cardiac pacing therapies" in np_lower:
        return "CPT"
    if "diagnostics mdr" in np_lower:
        return "CDS"
    if "non implantables" in np_lower:
        return "NI"
    if "pm-cts ind" in np_lower:
        return PM_CTS_IND_TEAM_NAME
    if "pm-cts" in np_lower:
        return PM_CTS_TEAM_NAME
    if "crm cqxm reports - 1.9 heijunka tracker" in np_lower:
        return MEIC_TEAM_NAME
    return ""
def _col_idx(col: str) -> int:
    n = 0
    for ch in col.upper():
        if "A" <= ch <= "Z":
            n = n * 26 + (ord(ch) - ord("A") + 1)
    return n
def _norm_text(x: str) -> str:
    return re.sub(r"\s+", " ", (x or "").strip()).lower()
def _collapse_ws(x: str) -> str:
    return re.sub(r"\s+", " ", (x or "").strip())
def normalize_person_name(name: str) -> str:
    s = _collapse_ws(name)
    if not s:
        return ""
    return s.title()
def normalize_person_key(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", _norm_text(name))
def is_excluded_person(name: str) -> bool:
    n = _norm_text(name)
    return n in {"", "do not use", "team member(s)"}
def _find_header_col(ws: Worksheet, header_text: str, *, max_header_row: int = 10) -> Optional[int]:
    want = _norm_text(header_text)
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, max_header_row),
        values_only=True,
    ):
        for c, value in enumerate(row, start=1):
            if _norm_text(str(value or "")) == want:
                return c
    return None
def _find_number_right_of_label(
    ws: Worksheet,
    label_text: str,
    *,
    lookahead: int = 10,
) -> Optional[float]:
    want = _norm_text(label_text)
    for row in ws.iter_rows(values_only=True):
        for c, value in enumerate(row):
            if want in _norm_text(str(value or "")):
                for candidate in row[c + 1:c + 1 + lookahead]:
                    n = _cell_number(candidate)
                    if n is not None:
                        return n
    return None
def _is_summary_person_row(name: str) -> bool:
    n = _norm_text(name)
    return (
        n in {
            "team tally",
            "total wip hours",
            "total non-wip hours",
            "tally",
            "total team workable hours",
            "team ooo hours",
            "team % wip",
            "team% non-wip",
        }
        or n.startswith("total ")
        or n.startswith("team ")
        or n.startswith("team%")
    )
def iter_pm_cts_non_wip_rows(
    ws_pab: Worksheet,
    start_row: int = 2,
) -> Iterable[Tuple[str, str, float]]:
    blank_run = 0
    for row in ws_pab.iter_rows(
        min_row=start_row, min_col=2, max_col=8, values_only=True
    ):
        raw_person, raw_category, raw_activity, raw_hours = row[0], row[1], row[2], row[6]
        if not any(str(value or "").strip() for value in (raw_person, raw_category, raw_activity, raw_hours)):
            blank_run += 1
            if blank_run >= MAX_TRAILING_BLANK_ROWS:
                break
            continue
        blank_run = 0
        category = _norm_text(str(raw_category or ""))
        if category not in PM_CTS_NON_WIP_TYPES:
            continue
        person = normalize_person_name(str(raw_person or ""))
        activity = _collapse_ws(str(raw_activity or "")) or category.title()
        hours = _cell_number(raw_hours)
        if not person or is_excluded_person(person) or hours is None:
            continue
        yield person, activity, float(hours)
def compute_pm_cts_total_non_wip_hours(ws_pab: Worksheet) -> float:
    return float(sum(hours for _, _, hours in iter_pm_cts_non_wip_rows(ws_pab)))
def compute_pm_cts_non_wip_by_person(ws_pab: Worksheet) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for person, _, hours in iter_pm_cts_non_wip_rows(ws_pab):
        out[person] = out.get(person, 0.0) + float(hours)
    return {person: float(total) for person, total in out.items() if total != 0}
def compute_pm_cts_total_non_wip_hours(ws_pab: Worksheet) -> float:
    return float(sum(hours for _, _, hours in iter_pm_cts_non_wip_rows(ws_pab)))
def compute_pm_cts_non_wip_by_person(ws_pab: Worksheet) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for person, _, hours in iter_pm_cts_non_wip_rows(ws_pab):
        out[person] = out.get(person, 0.0) + float(hours)
    return {person: float(total) for person, total in out.items() if total != 0}
def compute_pm_cts_ooo_by_person(ws_perf: Worksheet) -> Dict[str, float]:
    ooo_col = _find_header_col(ws_perf, "Total OOO hours") or ws_perf["S1"].column
    out: Dict[str, float] = {}
    blank_run = 0
    for row in ws_perf.iter_rows(
        min_row=3, min_col=1, max_col=ooo_col, values_only=True
    ):
        raw_name = str(row[0] or "").strip()
        raw_hours = row[ooo_col - 1]
        if not raw_name and _cell_number(raw_hours) in {None, 0}:
            blank_run += 1
            if blank_run >= MAX_TRAILING_BLANK_ROWS:
                break
            continue
        blank_run = 0
        if not raw_name:
            continue
        if is_excluded_person(raw_name) or _is_summary_person_row(raw_name):
            continue
        person = normalize_person_name(raw_name)
        hours = _cell_number(raw_hours)
        if not person or hours is None or hours == 0:
            continue
        out[person] = out.get(person, 0.0) + float(hours)
    return out
def compute_pm_cts_total_ooo_hours(ws_perf: Worksheet) -> float:
    summary_val = (
        _find_number_right_of_label(ws_perf, "Team OOO hours")
        or _find_number_right_of_label(ws_perf, "Total OOO hours")
    )
    if summary_val is not None:
        return float(summary_val)
    return float(sum(compute_pm_cts_ooo_by_person(ws_perf).values()))
def compute_pm_cts_non_wip_activities(
    ws_pab: Worksheet,
    ws_perf: Worksheet,
) -> List[Dict[str, Any]]:
    agg: Dict[Tuple[str, str], float] = {}
    for person, activity, hours in iter_pm_cts_non_wip_rows(ws_pab):
        key = (person, activity)
        agg[key] = agg.get(key, 0.0) + float(hours)
    for person, hours in compute_pm_cts_ooo_by_person(ws_perf).items():
        key = (person, "OOO")
        agg[key] = agg.get(key, 0.0) + float(hours)
    return [
        {"name": person, "activity": activity, "hours": float(hours)}
        for (person, activity), hours in sorted(
            agg.items(),
            key=lambda x: (x[0][0].lower(), x[0][1].lower()),
        )
        if hours != 0
    ]
def compute_pm_cts_wip_workers_ooo_hours(
    ws_perf: Worksheet,
    wip_workers: List[str],
) -> float:
    worker_keys = {normalize_person_key(x) for x in wip_workers if x}
    if not worker_keys:
        return 0.0
    total = 0.0
    for person, hours in compute_pm_cts_ooo_by_person(ws_perf).items():
        if normalize_person_key(person) in worker_keys:
            total += float(hours)
    return float(total)
def compute_pm_cts_people_count(ws_perf: Worksheet) -> int:
    seen = set()
    for r in range(3, ws_perf.max_row + 1):
        raw_name = str(ws_perf[f"A{r}"].value or "").strip()
        if not raw_name:
            continue
        if is_excluded_person(raw_name) or _is_summary_person_row(raw_name):
            continue
        key = normalize_person_key(raw_name)
        if key:
            seen.add(key)
    return len(seen)
def iter_lit_letters_non_wip_rows(
    ws_pab: Worksheet,
    start_row: int = 2,
) -> Iterable[Tuple[str, str, float]]:
    blank_run = 0
    for row in ws_pab.iter_rows(
        min_row=start_row, min_col=3, max_col=8, values_only=True
    ):
        raw_person, raw_area, raw_activity, raw_mins = row[0], row[1], row[2], row[5]
        if not any(str(value or "").strip() for value in (raw_person, raw_area, raw_activity, raw_mins)):
            blank_run += 1
            if blank_run >= MAX_TRAILING_BLANK_ROWS:
                break
            continue
        blank_run = 0
        area = _norm_text(str(raw_area or ""))
        if area not in LIT_LETTERS_NON_WIP_TYPES:
            continue
        person = normalize_person_name(str(raw_person or ""))
        activity = _collapse_ws(str(raw_activity or "")) or area.title()
        mins = _cell_number(raw_mins)
        if not person or is_excluded_person(person) or mins is None:
            continue
        yield person, activity, float(mins) / 60.0
def compute_lit_letters_total_non_wip_hours(ws_pab: Worksheet) -> float:
    return float(sum(hours for _, _, hours in iter_lit_letters_non_wip_rows(ws_pab)))
def compute_lit_letters_non_wip_by_person(ws_pab: Worksheet) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for person, _, hours in iter_lit_letters_non_wip_rows(ws_pab):
        out[person] = out.get(person, 0.0) + float(hours)
    return {person: float(total) for person, total in out.items() if total != 0}
def compute_lit_letters_ooo_by_person(ws_perf: Worksheet) -> Dict[str, float]:
    ooo_col = _find_header_col(ws_perf, "OOO hours") or ws_perf["AL1"].column
    out: Dict[str, float] = {}
    blank_run = 0
    for row in ws_perf.iter_rows(
        min_row=5, min_col=1, max_col=ooo_col, values_only=True
    ):
        raw_name = str(row[0] or "").strip()
        raw_hours = row[ooo_col - 1]
        if not raw_name and _cell_number(raw_hours) in {None, 0}:
            blank_run += 1
            if blank_run >= MAX_TRAILING_BLANK_ROWS:
                break
            continue
        blank_run = 0
        if not raw_name:
            continue
        if is_excluded_person(raw_name) or _is_summary_person_row(raw_name):
            continue
        person = normalize_person_name(raw_name)
        hours = _cell_number(raw_hours)
        if not person or hours is None or hours == 0:
            continue
        out[person] = out.get(person, 0.0) + float(hours)
    return out
def compute_lit_letters_total_ooo_hours(ws_perf: Worksheet) -> float:
    summary_val = (
        _find_number_right_of_label(ws_perf, "Team OOO hours")
        or _find_number_right_of_label(ws_perf, "Total OOO hours")
    )
    if summary_val is not None:
        return float(summary_val)
    return float(sum(compute_lit_letters_ooo_by_person(ws_perf).values()))
def compute_lit_letters_non_wip_activities(
    ws_pab: Worksheet,
    ws_perf: Worksheet,
) -> List[Dict[str, Any]]:
    agg: Dict[Tuple[str, str], float] = {}
    for person, activity, hours in iter_lit_letters_non_wip_rows(ws_pab):
        key = (person, activity)
        agg[key] = agg.get(key, 0.0) + float(hours)
    for person, hours in compute_lit_letters_ooo_by_person(ws_perf).items():
        key = (person, "OOO")
        agg[key] = agg.get(key, 0.0) + float(hours)
    return [
        {"name": person, "activity": activity, "hours": float(hours)}
        for (person, activity), hours in sorted(
            agg.items(),
            key=lambda x: (x[0][0].lower(), x[0][1].lower()),
        )
        if hours != 0
    ]
def compute_lit_letters_wip_workers_ooo_hours(
    ws_perf: Worksheet,
    wip_workers: List[str],
) -> float:
    worker_keys = {normalize_person_key(x) for x in wip_workers if x}
    if not worker_keys:
        return 0.0
    total = 0.0
    for person, hours in compute_lit_letters_ooo_by_person(ws_perf).items():
        if normalize_person_key(person) in worker_keys:
            total += float(hours)
    return float(total)
def compute_lit_letters_people_count(ws_perf: Worksheet) -> int:
    seen = set()
    for r in range(5, ws_perf.max_row + 1):
        raw_name = str(ws_perf[f"A{r}"].value or "").strip()
        if not raw_name:
            continue
        if is_excluded_person(raw_name) or _is_summary_person_row(raw_name):
            continue
        key = normalize_person_key(raw_name)
        if key:
            seen.add(key)
    return len(seen)
def parse_period_date_from_sheetname(sheet_name: str, *, default_year: Optional[int] = None) -> Optional[_dt.date]:
    if default_year is None:
        default_year = _dt.date.today().year
    s = sheet_name.strip()
    m = re.search(r"(\d{1,2})\s*[-/ ]\s*([A-Za-z]{3,9})(?:\s*[-/ ]\s*(\d{2,4}))?\s*$", s)
    if not m:
        return None
    day = int(m.group(1))
    mon_raw = m.group(2).strip().lower()
    year_raw = m.group(3)
    if mon_raw not in _MONTH_MAP:
        return None
    month = _MONTH_MAP[mon_raw]
    if year_raw:
        year = int(year_raw)
        if year < 100:
            year += 2000
    else:
        year = default_year
    try:
        return _dt.date(year, month, day)
    except ValueError:
        return None
def parse_period_date_from_filename(filename: str) -> Optional[_dt.date]:
    base = os.path.basename(filename)
    stem, _ = os.path.splitext(base)
    patterns = [
        r"(?<!\d)(\d{1,2})\s*[-_ ]*\s*([A-Za-z]{3,9})(?:\s*[-_ ]*\s*(\d{2,4}))?(?!\d)",
        r"(?<!\d)(\d{4})[\s\-_]+(\d{1,2})[\s\-_]+(\d{1,2})(?!\d)",
        r"(?<!\d)(\d{1,2})[\s\-_\/]+(\d{1,2})[\s\-_\/]+(\d{4})(?!\d)",
    ]
    for i, pat in enumerate(patterns):
        m = re.search(pat, stem)
        if not m:
            continue
        try:
            if i == 0:
                day = int(m.group(1))
                mon_raw = m.group(2).lower()
                year_raw = m.group(3)
                year = int(year_raw) if year_raw else 2026
                if year < 100:
                    year += 2000
                if mon_raw not in _MONTH_MAP:
                    continue
                return _dt.date(year, _MONTH_MAP[mon_raw], day)
            if i == 1:
                year = int(m.group(1))
                month = int(m.group(2))
                day = int(m.group(3))
                return _dt.date(year, month, day)
            if i == 2:
                month = int(m.group(1))
                day = int(m.group(2))
                year = int(m.group(3))
                return _dt.date(year, month, day)
        except ValueError:
            continue
    return None
def parse_period_date_from_workbook_sheetnames(wb, *, default_year: Optional[int] = None) -> Optional[_dt.date]:
    for name in wb.sheetnames:
        d = parse_period_date_from_sheetname(name, default_year=default_year)
        if d is not None:
            return d
    return None
def parse_period_date_from_perf_metrics_cell(ws: Worksheet, cell_ref: str = "B3") -> Optional[_dt.date]:
    value = ws[cell_ref].value
    if value is None or value == "":
        return None
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    if isinstance(value, (int, float)):
        try:
            return _dt.datetime.fromordinal(_dt.datetime(1899, 12, 30).toordinal() + int(value)).date()
        except Exception:
            return None
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%d-%b-%y", "%d %b %Y", "%d %B %Y"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None
def monday_of_week(d: _dt.date) -> _dt.date:
    return d - _dt.timedelta(days=d.weekday())
def iso_monday_weeks_back(today: Optional[_dt.date] = None, weeks_back: int = 3) -> List[str]:
    if today is None:
        today = _dt.date.today()
    start = monday_of_week(today)
    return [(start - _dt.timedelta(days=7 * i)).isoformat() for i in range(weeks_back + 1)]
def filter_rows_to_recent_weeks(rows: List[Dict[str, Any]], weeks_back: int = 3) -> List[Dict[str, Any]]:
    keep_weeks = set(iso_monday_weeks_back(weeks_back=weeks_back))
    return [r for r in rows if (r.get("period_date") or "").strip() in keep_weeks]
def load_existing_csv_rows(path: str) -> List[Dict[str, Any]]:
    if not path or not os.path.exists(path):
        return []
    with open(path, "r", newline="", encoding="utf-8-sig") as fp:
        return list(csv.DictReader(fp))
def _parse_iso_date(value: Any) -> Optional[_dt.date]:
    s = str(value or "").strip()
    if not s:
        return None
    try:
        return _dt.date.fromisoformat(s)
    except ValueError:
        return None
def should_refresh_existing_row(
    row: Dict[str, Any],
    refresh_weeks: set[str],
    refresh_team: Optional[str] = None,
) -> bool:
    team = (row.get("team") or "").strip()
    if refresh_team is not None and team != refresh_team:
        return False
    period_iso = (row.get("period_date") or "").strip()
    if period_iso not in refresh_weeks:
        return False
    period = _parse_iso_date(period_iso)
    if team == "MCS" and period is not None and period < MCS_REFRESH_MIN_PERIOD:
        return False
    return True
def should_include_recent_row(row: Dict[str, Any]) -> bool:
    team = (row.get("team") or "").strip()
    period = _parse_iso_date(row.get("period_date"))
    if team == "MCS" and period is not None and period < MCS_REFRESH_MIN_PERIOD:
        return False
    return True
def merge_existing_with_recent_rows(
    existing_rows: List[Dict[str, Any]],
    recent_rows: List[Dict[str, Any]],
    weeks_back: int = 3,
    refresh_team: Optional[str] = None,
) -> List[Dict[str, Any]]:
    refresh_weeks = set(iso_monday_weeks_back(weeks_back=weeks_back))
    frozen_rows = [
        r for r in existing_rows
        if not should_refresh_existing_row(r, refresh_weeks, refresh_team=refresh_team)
    ]
    return frozen_rows + recent_rows
def iso_date(d: Optional[_dt.date]) -> str:
    return d.isoformat() if isinstance(d, _dt.date) else ""
def _cell_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        vs = v.strip().replace(",", "")
        if vs == "":
            return None
        if vs.endswith("%"):
            try:
                return float(vs[:-1]) / 100.0
            except ValueError:
                return None
        try:
            return float(vs)
        except ValueError:
            return None
    return None
def sum_range(ws: Worksheet, cell1: str, cell2: str) -> float:
    total = 0.0
    for row in ws[cell1:cell2]:
        for c in row:
            n = _cell_number(c.value)
            if n is not None:
                total += n
    return total
def read_merged_value(ws: Worksheet, top_left_cell: str) -> str:
    v = ws[top_left_cell].value
    return str(v).strip() if v is not None else ""
def find_sheets_by_period(wb, *, kind: str) -> Dict[_dt.date, str]:
    out: Dict[_dt.date, str] = {}
    for name in wb.sheetnames:
        if kind == "availability":
            if not _AVAIL_PAT.search(name):
                continue
        elif kind == "production":
            if not _PROD_PAT.search(name):
                continue
        else:
            raise ValueError("kind must be availability or production")
        d = parse_period_date_from_sheetname(name)
        if d:
            out[d] = name
    return out
def iter_prod_rows(ws_prod: Worksheet, start_row: int = 7, end_row: int = 406) -> Iterable[Tuple[int, str, str, Optional[float], Optional[float]]]:
    maxr = min(ws_prod.max_row, end_row)
    for r, row in enumerate(
        ws_prod.iter_rows(
            min_row=start_row, max_row=maxr, min_col=4, max_col=7, values_only=True
        ),
        start=start_row,
    ):
        person, station = row[0], row[1]
        target = _cell_number(row[2])
        output = _cell_number(row[3])
        p = str(person).strip() if person is not None else ""
        s = str(station).strip() if station is not None else ""
        if p == "" and s == "" and target is None and output is None:
            continue
        yield (r, p, s, target, output)
def wip_workers_from_prod(ws_prod: Worksheet) -> List[str]:
    seen = set()
    for _, person, station, target, output in iter_prod_rows(ws_prod, start_row=7):
        if output is None:
            continue
        if is_excluded_person(person):
            continue
        norm_person = normalize_person_name(person)
        if not norm_person:
            continue
        seen.add(norm_person)
    return sorted(seen)
NON_WIP_SPECS = [
    ("A13", ("B18", "F21"), "A", ("B17", "F17")),
    ("A23", ("B28", "F31"), "A", ("B27", "F27")),
    ("H3",  ("I8", "M11"), "H", ("I7", "M7")),
    ("H23", ("I28", "M31"), "H", ("I27", "M27")),
    ("H33", ("I38", "M41"), "H", ("I37", "M37")),
    ("O3",  ("P8", "T11"), "O", ("P7", "T7")),
    ("O13", ("P18", "T21"), "O", ("P17", "T17")),
    ("O23", ("P28", "T31"), "O", ("P27", "T27")),
    ("O43", ("P48", "T51"), "O", ("P47", "T47")),
    ("O53", ("P58", "T61"), "O", ("P57", "T57")),
]
def compute_total_non_wip_hours(ws_av: Worksheet) -> float:
    total = 0.0
    for _, (c1, c2), _, _ in NON_WIP_SPECS:
        total += sum_range(ws_av, c1, c2)
    return total
def compute_total_ooo_hours(ws_av: Worksheet) -> float:
    total = 0.0
    for _, _, _, (c1, c2) in NON_WIP_SPECS:
        total += sum_range(ws_av, c1, c2)
    return total
def compute_non_wip_by_person(ws_av: Worksheet) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for name_cell, (c1, c2), _, _ in NON_WIP_SPECS:
        raw_name = read_merged_value(ws_av, name_cell)
        if not raw_name or is_excluded_person(raw_name):
            continue
        name = normalize_person_name(raw_name)
        if not name:
            continue
        out[name] = out.get(name, 0.0) + sum_range(ws_av, c1, c2)
    return out
def compute_non_wip_activities(ws_av: Worksheet) -> List[Dict[str, Any]]:
    agg: Dict[Tuple[str, str], float] = {}
    for name_cell, (c1, c2), label_col, (ooo_c1, ooo_c2) in NON_WIP_SPECS:
        raw_name = read_merged_value(ws_av, name_cell)
        if not raw_name or is_excluded_person(raw_name):
            continue
        name = normalize_person_name(raw_name)
        if not name:
            continue
        min_row = ws_av[c1].row
        max_row = ws_av[c2].row
        min_col = ws_av[c1].column
        max_col = ws_av[c2].column
        for r in range(min_row, max_row + 1):
            activity_label = ws_av[f"{label_col}{r}"].value
            activity = str(activity_label).strip() if activity_label is not None else ""
            if activity == "":
                continue
            if _norm_text(activity) in {"weekday", "hours available", "available for cos1", "wip block?"}:
                continue
            hours_sum = 0.0
            for c in range(min_col, max_col + 1):
                val = _cell_number(ws_av.cell(row=r, column=c).value)
                if val is None or val == 0:
                    continue
                hours_sum += float(val)
            if hours_sum:
                key = (name, activity)
                agg[key] = agg.get(key, 0.0) + hours_sum
        ooo_hours = sum_range(ws_av, ooo_c1, ooo_c2)
        if ooo_hours:
            key = (name, "OOO")
            agg[key] = agg.get(key, 0.0) + float(ooo_hours)
    return [
        {"name": n, "activity": a, "hours": float(h)}
        for (n, a), h in sorted(agg.items(), key=lambda x: (x[0][0].lower(), x[0][1].lower()))
    ]
def compute_ooo_by_person(ws_av: Worksheet) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for name_cell, _, _, (c1, c2) in NON_WIP_SPECS:
        raw_name = read_merged_value(ws_av, name_cell)
        if not raw_name or is_excluded_person(raw_name):
            continue
        name = normalize_person_name(raw_name)
        if not name:
            continue
        out[name] = out.get(name, 0.0) + sum_range(ws_av, c1, c2)
    return out
def load_completed_hours_from_crm_wip(crm_wip_csv: str) -> Dict[Tuple[str, str], float]:
    out: Dict[Tuple[str, str], float] = {}
    if not os.path.exists(crm_wip_csv):
        return out
    with open(crm_wip_csv, "r", encoding="utf-8-sig", newline="") as fp:
        r = csv.DictReader(fp)
        for row in r:
            team = (row.get("team") or "").strip()
            period = (row.get("period_date") or "").strip()
            ch = row.get("Completed Hours")
            try:
                val = float(ch) if ch not in (None, "") else None
            except ValueError:
                val = None
            if team and period and val is not None:
                out[(team, period)] = val
    return out
def parse_people_in_wip_value(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, list):
        raw_items = value
    else:
        s = str(value).strip()
        if not s:
            return []
        try:
            parsed = json.loads(s)
            raw_items = parsed if isinstance(parsed, list) else [parsed]
        except Exception:
            raw_items = re.split(r"[;,|\n]+", s)
    out: List[str] = []
    seen = set()
    for item in raw_items:
        name = normalize_person_name(str(item))
        if not name:
            continue
        key = normalize_person_key(name)
        if key in seen:
            continue
        seen.add(key)
        out.append(name)
    return out
def compute_ds_ooo_by_person(ws_wip_plan: Worksheet) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for r in range(3, ws_wip_plan.max_row + 1):
        person = normalize_person_name(str(ws_wip_plan[f"DL{r}"].value or ""))
        if not person or is_excluded_person(person):
            continue
        hours = _cell_number(ws_wip_plan[f"EF{r}"].value)
        if hours is None or hours == 0:
            continue
        out[person] = out.get(person, 0.0) + float(hours)
    return out
def compute_ds_non_wip_activities(ws_pab: Worksheet, ws_wip_plan: Worksheet) -> List[Dict[str, Any]]:
    agg: Dict[Tuple[str, str], float] = {}
    for person, activity, hours in iter_ds_non_wip_rows(ws_pab):
        key = (person, activity)
        agg[key] = agg.get(key, 0.0) + float(hours)
    for person, hours in compute_ds_ooo_by_person(ws_wip_plan).items():
        key = (person, "OOO")
        agg[key] = agg.get(key, 0.0) + float(hours)
    return [
        {"name": person, "activity": activity, "hours": float(hours)}
        for (person, activity), hours in sorted(agg.items(), key=lambda x: (x[0][0].lower(), x[0][1].lower()))
        if hours != 0
    ]
def compute_cpt_ooo_by_person(ws_wip_plan: Worksheet) -> Dict[str, float]:
    return compute_ooo_by_person_from_wip_plan(ws_wip_plan, "DB", "DT")
def compute_cpt_non_wip_activities(ws_pab: Worksheet, ws_wip_plan: Worksheet) -> List[Dict[str, Any]]:
    agg: Dict[Tuple[str, str], float] = {}
    for person, activity, hours in iter_cpt_non_wip_rows(ws_pab):
        key = (person, activity)
        agg[key] = agg.get(key, 0.0) + float(hours)
    for person, hours in compute_cpt_ooo_by_person(ws_wip_plan).items():
        key = (person, "OOO")
        agg[key] = agg.get(key, 0.0) + float(hours)
    return [
        {"name": person, "activity": activity, "hours": float(hours)}
        for (person, activity), hours in sorted(agg.items(), key=lambda x: (x[0][0].lower(), x[0][1].lower()))
        if hours != 0
    ]
def load_people_in_wip_from_crm_wip(crm_wip_csv: str) -> Dict[Tuple[str, str], List[str]]:
    out: Dict[Tuple[str, str], List[str]] = {}
    if not os.path.exists(crm_wip_csv):
        return out
    with open(crm_wip_csv, "r", encoding="utf-8-sig", newline="") as fp:
        r = csv.DictReader(fp)
        for row in r:
            team = (row.get("team") or "").strip()
            period = (row.get("period_date") or "").strip()
            if not team or not period:
                continue
            people = parse_people_in_wip_value(row.get("People in WIP"))
            if not people:
                continue
            key = (team, period)
            existing = out.get(key, [])
            seen = {normalize_person_key(x) for x in existing}
            for person in people:
                pkey = normalize_person_key(person)
                if pkey not in seen:
                    existing.append(person)
                    seen.add(pkey)
            out[key] = existing
    return out
def safe_div(n: float, d: float) -> Optional[float]:
    if d == 0:
        return None
    return n / d
def scrape_one_lit_letters_workbook(
    path: str,
    people_in_wip_lookup: Dict[Tuple[str, str], List[str]],
) -> List[Dict[str, Any]]:
    team = LIT_LETTERS_TEAM_NAME
    wb = _load_workbook_for_scraping(path)
    period = parse_period_date_from_filename(path)
    if period is None:
        period = parse_period_date_from_workbook_sheetnames(wb)
    if period is None:
        return []
    period_iso = iso_date(period)
    ws_pab = _get_required_sheet(wb, LIT_LETTERS_PAB_SHEET)
    ws_perf = _get_required_sheet(wb, LIT_LETTERS_PERF_WIP_SHEET)
    pab_rows = list(iter_lit_letters_non_wip_rows(ws_pab))
    total_non_wip_hours = float(sum(hours for _, _, hours in pab_rows))
    non_wip_by_person: Dict[str, float] = {}
    activity_agg: Dict[Tuple[str, str], float] = {}
    for person, activity, hours in pab_rows:
        non_wip_by_person[person] = non_wip_by_person.get(person, 0.0) + hours
        activity_agg[(person, activity)] = activity_agg.get((person, activity), 0.0) + hours
    ooo_by_person = compute_lit_letters_ooo_by_person(ws_perf)
    ooo_hours = float(
        _find_number_right_of_label(ws_perf, "Team OOO hours")
        or _find_number_right_of_label(ws_perf, "Total OOO hours")
        or sum(ooo_by_person.values())
    )
    for person, hours in ooo_by_person.items():
        activity_agg[(person, "OOO")] = activity_agg.get((person, "OOO"), 0.0) + hours
    non_wip_activities = [
        {"name": person, "activity": activity, "hours": float(hours)}
        for (person, activity), hours in sorted(
            activity_agg.items(), key=lambda item: (item[0][0].lower(), item[0][1].lower())
        )
        if hours != 0
    ]
    pct_in_wip = _find_number_right_of_label(ws_perf, "Team % WIP")
    if pct_in_wip is None:
        total_wip_hours = _find_number_right_of_label(ws_perf, "Total WIP Hours")
        total_workable_hours = _find_number_right_of_label(ws_perf, "Total Team workable hours")
        if total_wip_hours is not None and total_workable_hours is not None:
            pct_in_wip = safe_div(float(total_wip_hours), float(total_workable_hours))
    wip_workers = people_in_wip_lookup.get((team, period_iso), [])
    wip_workers_count = len({normalize_person_key(x) for x in wip_workers if x})
    wip_workers_ooo_hours = compute_wip_workers_ooo_hours_from_ooo_by_person(ooo_by_person, wip_workers)
    row = {
        "team": team,
        "period_date": period_iso,
        "people_count": lit_letters_people_count_for_period(period),
        "total_non_wip_hours": float(total_non_wip_hours),
        "OOO Hours": float(ooo_hours),
        "% in WIP": float(pct_in_wip) if pct_in_wip is not None else "",
        "non_wip_by_person": json.dumps(non_wip_by_person, ensure_ascii=False),
        "non_wip_activities": json.dumps(non_wip_activities, ensure_ascii=False),
        "wip_workers": json.dumps(wip_workers, ensure_ascii=False),
        "wip_workers_count": int(wip_workers_count),
        "wip_workers_ooo_hours": float(wip_workers_ooo_hours),
    }
    return [row]
def scrape_one_pm_cts_workbook(
    path: str,
    people_in_wip_lookup: Dict[Tuple[str, str], List[str]],
) -> List[Dict[str, Any]]:
    team = team_for_source(path) or PM_CTS_TEAM_NAME
    wb = _load_workbook_for_scraping(path)
    period = parse_period_date_from_filename(path)
    if period is None:
        period = parse_period_date_from_workbook_sheetnames(wb)
    if period is None:
        return []
    period_iso = iso_date(period)
    ws_pab = _get_required_sheet(wb, PM_CTS_PAB_SHEET)
    ws_perf = _get_required_sheet(wb, PM_CTS_PERF_WIP_SHEET)
    pab_rows = list(iter_pm_cts_non_wip_rows(ws_pab))
    total_non_wip_hours = float(sum(hours for _, _, hours in pab_rows))
    non_wip_by_person: Dict[str, float] = {}
    activity_agg: Dict[Tuple[str, str], float] = {}
    for person, activity, hours in pab_rows:
        non_wip_by_person[person] = non_wip_by_person.get(person, 0.0) + hours
        activity_agg[(person, activity)] = activity_agg.get((person, activity), 0.0) + hours
    ooo_by_person = compute_pm_cts_ooo_by_person(ws_perf)
    ooo_hours = float(
        _find_number_right_of_label(ws_perf, "Team OOO hours")
        or _find_number_right_of_label(ws_perf, "Total OOO hours")
        or sum(ooo_by_person.values())
    )
    for person, hours in ooo_by_person.items():
        activity_agg[(person, "OOO")] = activity_agg.get((person, "OOO"), 0.0) + hours
    non_wip_activities = [
        {"name": person, "activity": activity, "hours": float(hours)}
        for (person, activity), hours in sorted(
            activity_agg.items(), key=lambda item: (item[0][0].lower(), item[0][1].lower())
        )
        if hours != 0
    ]
    pct_in_wip = _find_number_right_of_label(ws_perf, "Team % WIP")
    if pct_in_wip is None:
        total_wip_hours = _find_number_right_of_label(ws_perf, "Total WIP Hours")
        total_workable_hours = _find_number_right_of_label(ws_perf, "Total Team workable hours")
        if total_wip_hours is not None and total_workable_hours is not None:
            pct_in_wip = safe_div(float(total_wip_hours), float(total_workable_hours))
    wip_workers = people_in_wip_lookup.get((team, period_iso), [])
    wip_workers_count = len({normalize_person_key(x) for x in wip_workers if x})
    wip_workers_ooo_hours = compute_wip_workers_ooo_hours_from_ooo_by_person(ooo_by_person, wip_workers)
    row = {
        "team": team,
        "period_date": period_iso,
        "people_count": PM_CTS_PEOPLE_COUNT,
        "total_non_wip_hours": float(total_non_wip_hours),
        "OOO Hours": float(ooo_hours),
        "% in WIP": float(pct_in_wip) if pct_in_wip is not None else "",
        "non_wip_by_person": json.dumps(non_wip_by_person, ensure_ascii=False),
        "non_wip_activities": json.dumps(non_wip_activities, ensure_ascii=False),
        "wip_workers": json.dumps(wip_workers, ensure_ascii=False),
        "wip_workers_count": int(wip_workers_count),
        "wip_workers_ooo_hours": float(wip_workers_ooo_hours),
    }
    return [row]
def scrape_one_workbook(path: str, completed_hours_lookup: Dict[Tuple[str, str], float]) -> List[Dict[str, Any]]:
    team = team_for_source(path)
    wb = _load_workbook_for_scraping(path)
    avail_sheets = find_sheets_by_period(wb, kind="availability")
    prod_sheets = find_sheets_by_period(wb, kind="production")
    periods = sorted(set(avail_sheets.keys()) | set(prod_sheets.keys()))
    rows: List[Dict[str, Any]] = []
    for period in periods:
        ws_av = wb[avail_sheets[period]] if period in avail_sheets else None
        ws_prod = wb[prod_sheets[period]] if period in prod_sheets else None
        if ws_av is None:
            continue
        period_iso = iso_date(period)
        total_non_wip_hours = compute_total_non_wip_hours(ws_av)
        ooo_hours = compute_total_ooo_hours(ws_av)
        non_wip_by_person = compute_non_wip_by_person(ws_av)
        non_wip_activities = compute_non_wip_activities(ws_av)
        wip_workers: List[str] = wip_workers_from_prod(ws_prod) if ws_prod is not None else []
        wip_workers_count = len(wip_workers)
        ooo_by_person = compute_ooo_by_person(ws_av)
        wip_workers_ooo_hours = float(sum(float(ooo_by_person.get(p, 0.0)) for p in wip_workers))
        completed_hours = completed_hours_lookup.get((team, period_iso), 0.0)
        pct_in_wip = safe_div(float(completed_hours), float(completed_hours) + float(total_non_wip_hours))
        row = {
            "team": team,
            "period_date": period_iso,
            "people_count": 10,
            "total_non_wip_hours": float(total_non_wip_hours),
            "OOO Hours": float(ooo_hours),
            "% in WIP": float(pct_in_wip) if pct_in_wip is not None else "",
            "non_wip_by_person": json.dumps(non_wip_by_person, ensure_ascii=False),
            "non_wip_activities": json.dumps(non_wip_activities, ensure_ascii=False),
            "wip_workers": json.dumps(wip_workers, ensure_ascii=False),
            "wip_workers_count": int(wip_workers_count),
            "wip_workers_ooo_hours": wip_workers_ooo_hours,
        }
        rows.append(row)
    return rows
def _get_required_sheet(wb, sheet_name: str) -> Worksheet:
    wanted = sheet_name.strip().lower()
    for actual in wb.sheetnames:
        if actual.strip().lower() == wanted:
            return wb[actual]
    raise KeyError(f"Missing required sheet: {sheet_name}. Found: {wb.sheetnames}")
def iter_ds_non_wip_rows(ws_pab: Worksheet, start_row: int = 2) -> Iterable[Tuple[str, str, float]]:
    for r in range(start_row, ws_pab.max_row + 1):
        category = _norm_text(str(ws_pab[f"D{r}"].value or ""))
        if category not in DS_NON_WIP_TYPES:
            continue
        person = normalize_person_name(str(ws_pab[f"C{r}"].value or ""))
        activity = _collapse_ws(str(ws_pab[f"E{r}"].value or ""))
        hours = _cell_number(ws_pab[f"I{r}"].value)
        if not person or hours is None:
            continue
        yield person, activity, float(hours)
def compute_ds_total_non_wip_hours(ws_pab: Worksheet) -> float:
    return float(sum(hours for _, _, hours in iter_ds_non_wip_rows(ws_pab)))
def compute_ds_non_wip_by_person(ws_pab: Worksheet) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for person, _, hours in iter_ds_non_wip_rows(ws_pab):
        out[person] = out.get(person, 0.0) + float(hours)
    return {person: float(total) for person, total in out.items() if total != 0}
def compute_ds_non_wip_activities(ws_pab: Worksheet, ws_wip_plan: Worksheet) -> List[Dict[str, Any]]:
    agg: Dict[Tuple[str, str], float] = {}
    for person, activity, hours in iter_ds_non_wip_rows(ws_pab):
        key = (person, activity)
        agg[key] = agg.get(key, 0.0) + float(hours)
    for person, hours in compute_ds_ooo_by_person(ws_wip_plan).items():
        key = (person, "OOO")
        agg[key] = agg.get(key, 0.0) + float(hours)
    return [
        {"name": person, "activity": activity, "hours": float(hours)}
        for (person, activity), hours in sorted(
            agg.items(), key=lambda x: (x[0][0].lower(), x[0][1].lower())
        )
        if hours != 0
    ]
def compute_ds_wip_workers_ooo_hours(ws_wip_plan: Worksheet, wip_workers: List[str]) -> float:
    worker_keys = {normalize_person_key(x) for x in wip_workers if x}
    if not worker_keys:
        return 0.0
    total = 0.0
    for r in range(1, ws_wip_plan.max_row + 1):
        name = str(ws_wip_plan[f"DL{r}"].value or "").strip()
        if not name:
            continue
        if normalize_person_key(name) not in worker_keys:
            continue
        val = _cell_number(ws_wip_plan[f"EF{r}"].value)
        if val is not None:
            total += float(val)
    return float(total)
def iter_cpt_non_wip_rows(ws_pab: Worksheet, start_row: int = 2) -> Iterable[Tuple[str, str, float]]:
    yield from iter_team_non_wip_rows(ws_pab, CPT_NON_WIP_TYPES, start_row=start_row)
def compute_cpt_total_non_wip_hours(ws_pab: Worksheet) -> float:
    return float(sum(hours for _, _, hours in iter_cpt_non_wip_rows(ws_pab)))
def compute_cpt_non_wip_by_person(ws_pab: Worksheet) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for person, _, hours in iter_cpt_non_wip_rows(ws_pab):
        out[person] = out.get(person, 0.0) + float(hours)
    return {person: float(total) for person, total in out.items() if total != 0}
def compute_cpt_non_wip_activities(ws_pab: Worksheet, ws_wip_plan: Worksheet) -> List[Dict[str, Any]]:
    agg: Dict[Tuple[str, str], float] = {}
    for person, activity, hours in iter_cpt_non_wip_rows(ws_pab):
        key = (person, activity)
        agg[key] = agg.get(key, 0.0) + float(hours)
    for person, hours in compute_cpt_ooo_by_person(ws_wip_plan).items():
        key = (person, "OOO")
        agg[key] = agg.get(key, 0.0) + float(hours)
    return [
        {"name": person, "activity": activity, "hours": float(hours)}
        for (person, activity), hours in sorted(
            agg.items(), key=lambda x: (x[0][0].lower(), x[0][1].lower())
        )
        if hours != 0
    ]
def compute_cpt_wip_workers_ooo_hours(ws_wip_plan: Worksheet, wip_workers: List[str]) -> float:
    return compute_wip_workers_ooo_hours_from_ooo_by_person(
        compute_cpt_ooo_by_person(ws_wip_plan),
        wip_workers,
    )
MAX_TRAILING_BLANK_ROWS = 50

def iter_team_non_wip_rows(
    ws_pab: Worksheet,
    non_wip_types: set[str],
    start_row: int = 2,
    *,
    max_trailing_blank_rows: int = MAX_TRAILING_BLANK_ROWS,
) -> Iterable[Tuple[str, str, float]]:
    blank_run = 0
    for row in ws_pab.iter_rows(
        min_row=start_row,
        min_col=3,   # C: person
        max_col=9,   # I: hours
        values_only=True,
    ):
        raw_person = row[0] if len(row) > 0 else None
        raw_category = row[1] if len(row) > 1 else None
        raw_activity = row[2] if len(row) > 2 else None
        raw_hours = row[6] if len(row) > 6 else None
        hours = _cell_number(raw_hours)

        if (
            not str(raw_person or "").strip()
            and not str(raw_category or "").strip()
            and not str(raw_activity or "").strip()
            and (hours is None or hours == 0)
        ):
            blank_run += 1
            if blank_run >= max_trailing_blank_rows:
                break
            continue

        blank_run = 0
        category = _norm_text(str(raw_category or ""))
        if category not in non_wip_types:
            continue
        person = normalize_person_name(str(raw_person or ""))
        activity = _collapse_ws(str(raw_activity or ""))
        if not person or is_excluded_person(person) or _is_summary_person_row(person) or hours is None:
            continue
        yield person, activity, float(hours)
def compute_ooo_by_person_from_wip_plan(
    ws_wip_plan: Worksheet,
    name_col: str,
    hours_col: str,
    start_row: int = 3,
    *,
    max_trailing_blank_rows: int = MAX_TRAILING_BLANK_ROWS,
) -> Dict[str, float]:
    out: Dict[str, float] = {}
    name_idx = _col_idx(name_col)
    hours_idx = _col_idx(hours_col)
    min_col = min(name_idx, hours_idx)
    max_col = max(name_idx, hours_idx)
    name_offset = name_idx - min_col
    hours_offset = hours_idx - min_col
    blank_run = 0
    for row in ws_wip_plan.iter_rows(
        min_row=start_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        raw_name = row[name_offset] if len(row) > name_offset else None
        hours = _cell_number(row[hours_offset] if len(row) > hours_offset else None)
        if not str(raw_name or "").strip() and (hours is None or hours == 0):
            blank_run += 1
            if blank_run >= max_trailing_blank_rows:
                break
            continue

        blank_run = 0
        person = normalize_person_name(str(raw_name or ""))
        if not person or is_excluded_person(person) or _is_summary_person_row(person):
            continue
        if hours is None or hours == 0:
            continue
        out[person] = out.get(person, 0.0) + float(hours)
    return out
def compute_ooo_by_person_from_fixed_rows(
    ws: Worksheet,
    *,
    name_col: str,
    hours_col: str,
    start_row: int,
    end_row: int,
) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for r in range(start_row, min(ws.max_row, end_row) + 1):
        person = normalize_person_name(str(ws[f"{name_col}{r}"].value or ""))
        if not person or is_excluded_person(person) or _is_summary_person_row(person):
            continue
        hours = _cell_number(ws[f"{hours_col}{r}"].value)
        if hours is None or hours == 0:
            continue
        out[person] = out.get(person, 0.0) + float(hours)
    return out
def compute_team_non_wip_activities_from_ooo_by_person(
    ws_pab: Worksheet,
    non_wip_types: set[str],
    ooo_by_person: Dict[str, float],
) -> List[Dict[str, Any]]:
    agg: Dict[Tuple[str, str], float] = {}
    for person, activity, hours in iter_team_non_wip_rows(ws_pab, non_wip_types):
        key = (person, activity)
        agg[key] = agg.get(key, 0.0) + float(hours)
    for person, hours in ooo_by_person.items():
        key = (person, "OOO")
        agg[key] = agg.get(key, 0.0) + float(hours)
    return [
        {"name": person, "activity": activity, "hours": float(hours)}
        for (person, activity), hours in sorted(
            agg.items(), key=lambda x: (x[0][0].lower(), x[0][1].lower())
        )
        if hours != 0
    ]
def _person_match_keys(name: str) -> set[str]:
    raw = str(name or "")
    without_trailing_parenthetical = re.sub(r"\s*\([^)]*\)\s*$", "", raw)
    return {
        key
        for key in (
            normalize_person_key(raw),
            normalize_person_key(without_trailing_parenthetical),
        )
        if key
    }
def compute_wip_workers_ooo_hours_from_ooo_by_person(
    ooo_by_person: Dict[str, float],
    wip_workers: List[str],
) -> float:
    worker_keys: set[str] = set()
    for worker in wip_workers:
        worker_keys.update(_person_match_keys(worker))
    if not worker_keys:
        return 0.0
    total = 0.0
    for person, hours in ooo_by_person.items():
        if _person_match_keys(person) & worker_keys:
            total += float(hours)
    return float(total)
def compute_team_total_non_wip_hours(ws_pab: Worksheet, non_wip_types: set[str]) -> float:
    return float(sum(hours for _, _, hours in iter_team_non_wip_rows(ws_pab, non_wip_types)))
def compute_team_non_wip_by_person(ws_pab: Worksheet, non_wip_types: set[str]) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for person, _, hours in iter_team_non_wip_rows(ws_pab, non_wip_types):
        out[person] = out.get(person, 0.0) + float(hours)
    return {person: float(total) for person, total in out.items() if total != 0}
def compute_team_non_wip_activities(ws_pab: Worksheet, ws_wip_plan: Worksheet, non_wip_types: set[str], ooo_name_col: str, ooo_hours_col: str) -> List[Dict[str, Any]]:
    return compute_team_non_wip_activities_from_ooo_by_person(
        ws_pab,
        non_wip_types,
        compute_ooo_by_person_from_wip_plan(ws_wip_plan, ooo_name_col, ooo_hours_col),
    )
def compute_team_wip_workers_ooo_hours(ws_wip_plan: Worksheet, wip_workers: List[str], name_col: str, hours_col: str) -> float:
    return compute_wip_workers_ooo_hours_from_ooo_by_person(
        compute_ooo_by_person_from_wip_plan(ws_wip_plan, name_col, hours_col, start_row=1),
        wip_workers,
    )
def scrape_one_mapped_workbook(
    path: str,
    people_in_wip_lookup: Dict[Tuple[str, str], List[str]],
    *,
    team: str,
    pab_sheet: str,
    wip_plan_sheet: str,
    perf_metrics_sheet: Optional[str],
    perf_metrics_date_cell: Optional[str],
    perf_wip_sheet: str,
    pct_in_wip_cell: str,
    ooo_total_cell: str,
    ooo_name_col: str,
    ooo_hours_col: str,
    people_count: int | Any,
    non_wip_types: set[str],
    period_date_offset_days: int = 0,
) -> List[Dict[str, Any]]:
    wb = _load_workbook_for_scraping(path)
    period: Optional[_dt.date] = None
    if perf_metrics_sheet and perf_metrics_date_cell:
        ws_perf_metrics = _get_required_sheet(wb, perf_metrics_sheet)
        period = parse_period_date_from_perf_metrics_cell(ws_perf_metrics, perf_metrics_date_cell)
        if period is not None and period_date_offset_days:
            period = period + _dt.timedelta(days=period_date_offset_days)
    if period is None:
        period = parse_period_date_from_filename(path)
    if period is None:
        period = parse_period_date_from_workbook_sheetnames(wb)
    if period is None:
        return []
    period_iso = iso_date(period)
    ws_pab = _get_required_sheet(wb, pab_sheet)
    ws_wip_plan = _get_required_sheet(wb, wip_plan_sheet)
    ws_perf = _get_required_sheet(wb, perf_wip_sheet)
    pab_rows = list(iter_team_non_wip_rows(ws_pab, non_wip_types))
    total_non_wip_hours = float(sum(hours for _, _, hours in pab_rows))
    ooo_hours = float(_cell_number(ws_wip_plan[ooo_total_cell].value) or 0.0)
    pct_in_wip = _cell_number(ws_perf[pct_in_wip_cell].value)

    non_wip_by_person: Dict[str, float] = {}
    activity_agg: Dict[Tuple[str, str], float] = {}
    for person, activity, hours in pab_rows:
        non_wip_by_person[person] = non_wip_by_person.get(person, 0.0) + float(hours)
        key = (person, activity)
        activity_agg[key] = activity_agg.get(key, 0.0) + float(hours)
    non_wip_by_person = {person: float(total) for person, total in non_wip_by_person.items() if total != 0}

    ooo_by_person = compute_ooo_by_person_from_wip_plan(ws_wip_plan, ooo_name_col, ooo_hours_col)
    for person, hours in ooo_by_person.items():
        key = (person, "OOO")
        activity_agg[key] = activity_agg.get(key, 0.0) + float(hours)
    non_wip_activities = [
        {"name": person, "activity": activity, "hours": float(hours)}
        for (person, activity), hours in sorted(
            activity_agg.items(), key=lambda x: (x[0][0].lower(), x[0][1].lower())
        )
        if hours != 0
    ]

    wip_workers = people_in_wip_lookup.get((team, period_iso), [])
    wip_workers_count = len({normalize_person_key(x) for x in wip_workers if x})
    wip_workers_ooo_hours = compute_wip_workers_ooo_hours_from_ooo_by_person(ooo_by_person, wip_workers)
    effective_people_count = (
        people_count(period)
        if callable(people_count)
        else people_count
    )
    row = {
        "team": team,
        "period_date": period_iso,
        "people_count": effective_people_count,
        "total_non_wip_hours": float(total_non_wip_hours),
        "OOO Hours": float(ooo_hours),
        "% in WIP": float(pct_in_wip) if pct_in_wip is not None else "",
        "non_wip_by_person": json.dumps(non_wip_by_person, ensure_ascii=False),
        "non_wip_activities": json.dumps(non_wip_activities, ensure_ascii=False),
        "wip_workers": json.dumps(wip_workers, ensure_ascii=False),
        "wip_workers_count": int(wip_workers_count),
        "wip_workers_ooo_hours": float(wip_workers_ooo_hours),
    }
    return [row]
def scrape_one_ds_workbook(path: str, people_in_wip_lookup: Dict[Tuple[str, str], List[str]]) -> List[Dict[str, Any]]:
    return scrape_one_mapped_workbook(
        path,
        people_in_wip_lookup,
        team="DS",
        pab_sheet=DS_PAB_SHEET,
        wip_plan_sheet=DS_WIP_PLAN_SHEET,
        perf_metrics_sheet=None,
        perf_metrics_date_cell=None,
        perf_wip_sheet=DS_PERF_WIP_SHEET,
        pct_in_wip_cell="J2",
        ooo_total_cell="EF2",
        ooo_name_col="DL",
        ooo_hours_col="EF",
        people_count=ds_people_count_for_period,
        non_wip_types=DS_NON_WIP_TYPES,
    )
def scrape_one_cpt_workbook(path: str, people_in_wip_lookup: Dict[Tuple[str, str], List[str]]) -> List[Dict[str, Any]]:
    return scrape_one_mapped_workbook(
        path,
        people_in_wip_lookup,
        team="CPT",
        pab_sheet=CPT_PAB_SHEET,
        wip_plan_sheet=CPT_WIP_PLAN_SHEET,
        perf_metrics_sheet=None,
        perf_metrics_date_cell=None,
        perf_wip_sheet=CPT_PERF_WIP_SHEET,
        pct_in_wip_cell="G2",
        ooo_total_cell="DT2",
        ooo_name_col="DB",
        ooo_hours_col="DT",
        people_count=cpt_people_count_for_period,
        non_wip_types=CPT_NON_WIP_TYPES,
    )
def scrape_one_cds_workbook(path: str, people_in_wip_lookup: Dict[Tuple[str, str], List[str]]) -> List[Dict[str, Any]]:
    return scrape_one_mapped_workbook(
        path,
        people_in_wip_lookup,
        team="CDS",
        pab_sheet=CDS_PAB_SHEET,
        wip_plan_sheet=CDS_WIP_PLAN_SHEET,
        perf_metrics_sheet=CDS_PERF_METRICS_SHEET,
        perf_metrics_date_cell="B3",
        perf_wip_sheet=CDS_PERF_WIP_SHEET,
        pct_in_wip_cell="J2",
        ooo_total_cell="CV2",
        ooo_name_col="CH",
        ooo_hours_col="CV",
        people_count=cds_people_count_for_period,
        non_wip_types=CDS_NON_WIP_TYPES,
        period_date_offset_days=-4,
    )
def scrape_one_ni_workbook(path: str, people_in_wip_lookup: Dict[Tuple[str, str], List[str]]) -> List[Dict[str, Any]]:
    team = "NI"
    wb = _load_workbook_for_scraping(path)
    period: Optional[_dt.date] = None
    ws_perf_metrics = _get_required_sheet(wb, NI_PERF_METRICS_SHEET)
    period = parse_period_date_from_perf_metrics_cell(ws_perf_metrics, "B3")
    if period is not None:
        period = period + _dt.timedelta(days=-4)
    if period is None:
        period = parse_period_date_from_filename(path)
    if period is None:
        period = parse_period_date_from_workbook_sheetnames(wb)
    if period is None:
        return []
    period_iso = iso_date(period)
    ws_pab = _get_required_sheet(wb, NI_PAB_SHEET)
    ws_wip_plan = _get_required_sheet(wb, NI_WIP_PLAN_SHEET)
    ws_perf = _get_required_sheet(wb, NI_PERF_WIP_SHEET)
    total_non_wip_hours = compute_team_total_non_wip_hours(ws_pab, NI_NON_WIP_TYPES)
    if period >= NI_PERF_WIP_OOO_EFFECTIVE_DATE:
        ooo_hours = float(_cell_number(ws_perf[NI_PERF_WIP_OOO_TOTAL_CELL].value) or 0.0)
        ooo_by_person = compute_ooo_by_person_from_fixed_rows(
            ws_perf,
            name_col=NI_PERF_WIP_OOO_NAME_COL,
            hours_col=NI_PERF_WIP_OOO_HOURS_COL,
            start_row=NI_PERF_WIP_OOO_START_ROW,
            end_row=NI_PERF_WIP_OOO_END_ROW,
        )
    else:
        ooo_hours = float(_cell_number(ws_wip_plan["BR2"].value) or 0.0)
        ooo_by_person = compute_ooo_by_person_from_wip_plan(ws_wip_plan, "BI", "BR")
    pct_in_wip = _cell_number(ws_perf["J2"].value)
    non_wip_by_person = compute_team_non_wip_by_person(ws_pab, NI_NON_WIP_TYPES)
    non_wip_activities = compute_team_non_wip_activities_from_ooo_by_person(
        ws_pab,
        NI_NON_WIP_TYPES,
        ooo_by_person,
    )
    wip_workers = people_in_wip_lookup.get((team, period_iso), [])
    wip_workers_count = len({normalize_person_key(x) for x in wip_workers if x})
    wip_workers_ooo_hours = compute_wip_workers_ooo_hours_from_ooo_by_person(ooo_by_person, wip_workers)
    row = {
        "team": team,
        "period_date": period_iso,
        "people_count": ni_people_count_for_period(period),
        "total_non_wip_hours": float(total_non_wip_hours),
        "OOO Hours": float(ooo_hours),
        "% in WIP": float(pct_in_wip) if pct_in_wip is not None else "",
        "non_wip_by_person": json.dumps(non_wip_by_person, ensure_ascii=False),
        "non_wip_activities": json.dumps(non_wip_activities, ensure_ascii=False),
        "wip_workers": json.dumps(wip_workers, ensure_ascii=False),
        "wip_workers_count": int(wip_workers_count),
        "wip_workers_ooo_hours": float(wip_workers_ooo_hours),
    }
    return [row]
def scrape_one_meic_workbook(path: str, people_in_wip_lookup: Dict[Tuple[str, str], List[str]]) -> List[Dict[str, Any]]:
    return scrape_one_mapped_workbook(
        path,
        people_in_wip_lookup,
        team=MEIC_TEAM_NAME,
        pab_sheet=MEIC_PAB_SHEET,
        wip_plan_sheet=MEIC_WIP_PLAN_SHEET,
        perf_metrics_sheet=None,
        perf_metrics_date_cell=None,
        perf_wip_sheet=MEIC_PERF_WIP_SHEET,
        pct_in_wip_cell="J2",
        ooo_total_cell="DT2",
        ooo_name_col="DB",
        ooo_hours_col="DT",
        people_count=MEIC_PEOPLE_COUNT,
        non_wip_types=MEIC_NON_WIP_TYPES,
    )

TEAM_ALIASES: Dict[str, str] = {
    "mcs": "MCS",
    "ds": "DS",
    "defibrillation solutions": "DS",
    "cpt": "CPT",
    "cardiac pacing therapies": "CPT",
    "cds": "CDS",
    "diagnostics": "CDS",
    "diagnostics mdr": "CDS",
    "ni": "NI",
    "non implantables": "NI",
    "non-implantables": "NI",
    "pm-cts": PM_CTS_TEAM_NAME,
    "pm cts": PM_CTS_TEAM_NAME,
    "pm_cts": PM_CTS_TEAM_NAME,
    "pm-cts ind": PM_CTS_IND_TEAM_NAME,
    "pm cts ind": PM_CTS_IND_TEAM_NAME,
    "pm_cts_ind": PM_CTS_IND_TEAM_NAME,
    "meic": MEIC_TEAM_NAME,
    "ni & pm meic": MEIC_TEAM_NAME,
    "ni and pm meic": MEIC_TEAM_NAME,
    "lit": LIT_LETTERS_TEAM_NAME,
    "lit & letters": LIT_LETTERS_TEAM_NAME,
    "lit and letters": LIT_LETTERS_TEAM_NAME,
}

TEAM_DEFAULT_INPUTS: Dict[str, List[str]] = {
    "MCS": [MCS_DEFAULT_PATH],
    "DS": [DS_DEFAULT_DIR, DS_ARCHIVE],
    "CPT": [CPT_DEFAULT_DIR, CPT_ARCHIVE_PAB_DIR, CPT_ARCHIVE_PAB_DIR2, CPT_ARCHIVE_PAB_DIR3],
    "CDS": [CDS_DEFAULT_DIR, CDS_ARCHIVE_PAB_DIR],
    "NI": [NI_DEFAULT_DIR, NI_ARCHIVE_APRIL_2026_DIR, NI_ARCHIVE],
    MEIC_TEAM_NAME: [MEIC_DEFAULT_DIR],
    PM_CTS_TEAM_NAME: [PM_CTS_DEFAULT_DIR],
    PM_CTS_IND_TEAM_NAME: [PM_CTS_DEFAULT_DIR],
    LIT_LETTERS_TEAM_NAME: [],
}


def normalize_team_arg(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    key = _norm_text(value).replace("_", " ")
    team = TEAM_ALIASES.get(key)
    if team:
        return team
    valid = sorted(TEAM_DEFAULT_INPUTS)
    raise ValueError(f"Unknown team {value!r}. Valid teams: {', '.join(valid)}")


def default_inputs_for_team(team: Optional[str]) -> List[str]:
    if team:
        return list(TEAM_DEFAULT_INPUTS.get(team, []))
    inputs: List[str] = []
    for team_inputs in TEAM_DEFAULT_INPUTS.values():
        inputs.extend(team_inputs)
    return inputs


def recent_week_starts(weeks_back: int) -> set[str]:
    return set(iso_monday_weeks_back(weeks_back=weeks_back))


def recent_months(weeks_back: int) -> set[Tuple[int, int]]:
    months: set[Tuple[int, int]] = set()
    for week_start_iso in recent_week_starts(weeks_back):
        week_start = _dt.date.fromisoformat(week_start_iso)
        for day_offset in range(7):
            d = week_start + _dt.timedelta(days=day_offset)
            months.add((d.year, d.month))
    return months


def parse_month_year_from_path(path: str) -> Optional[Tuple[int, int]]:
    """Find folder names like '4. April 2026' or 'June 2026'."""
    parts = re.split(r"[\\/]+", _norm_path(path))
    for part in reversed(parts):
        s = part.strip().lower()
        m = re.search(r"(?:\d+\.\s*)?([a-z]{3,9})\s+(20\d{2})", s)
        if not m:
            continue
        mon_raw = m.group(1)
        if mon_raw not in _MONTH_MAP:
            continue
        return int(m.group(2)), _MONTH_MAP[mon_raw]
    return None
def file_looks_recent_enough(path: str, *, weeks_back: int) -> bool:
    period = parse_period_date_from_filename(path)
    if period is not None:
        return monday_of_week(period).isoformat() in recent_week_starts(weeks_back)
    month_year = parse_month_year_from_path(path)
    if month_year is not None:
        return month_year in recent_months(weeks_back)
    return True
def filter_files_to_recent_weeks(files: List[str], *, weeks_back: int) -> List[str]:
    return [f for f in files if file_looks_recent_enough(f, weeks_back=weeks_back)]
def lit_letters_search_roots() -> List[str]:
    roots = [
        NI_DEFAULT_DIR,
        PM_CTS_DEFAULT_DIR,
        MEIC_DEFAULT_DIR,
        DS_DEFAULT_DIR,
        CPT_DEFAULT_DIR,
        CDS_DEFAULT_DIR,
    ]
    out: List[str] = []
    seen = set()
    for root in roots:
        nr = _norm_path(root)
        if nr in seen:
            continue
        seen.add(nr)
        out.append(nr)
    return out


def discover_lit_letters_files(*, weeks_back: int) -> List[str]:
    """Find Lit & Letters workbooks by filename under the likely team folders."""
    out: List[str] = []
    seen = set()
    for root in lit_letters_search_roots():
        if not os.path.isdir(root):
            continue
        for dirpath, dirnames, filenames in os.walk(root):
            # Avoid Office/OneDrive temporary folders and hidden/system folders.
            dirnames[:] = [
                d for d in dirnames
                if not d.startswith("~") and not d.startswith(".")
            ]
            for name in filenames:
                if name.startswith("~$"):
                    continue
                ext = os.path.splitext(name)[1].lower()
                if ext not in {".xlsx", ".xlsm"}:
                    continue
                fp = _norm_path(os.path.join(dirpath, name))
                if fp in seen or fp in EXCLUDED_FILES:
                    continue
                if not is_lit_letters_path(fp):
                    continue
                if not file_looks_recent_enough(fp, weeks_back=weeks_back):
                    continue
                seen.add(fp)
                out.append(fp)
    return sorted(out)
def expand_input_paths(paths: List[str]) -> List[str]:
    out: List[str] = []
    seen = set()
    def add_file(fp: str) -> None:
        np = _norm_path(fp)
        if np in EXCLUDED_FILES:
            return
        if _is_removed_pm_cts_file(np):
            return
        if np in seen:
            return
        if not os.path.isfile(np):
            return
        ext = os.path.splitext(np)[1].lower()
        if ext not in {".xlsx", ".xlsm"}:
            return
        team = team_for_source(np)
        base = os.path.basename(np)
        base_lower = base.lower()
        if base.startswith("~$"):
            return
        if team in {"CDS", "NI"} and "pab" not in base_lower:
            return
        if team == "CPT" and not any(token in base_lower for token in ("pab", "heijunka")):
            return
        seen.add(np)
        out.append(np)
    for p in paths:
        np = _norm_path(p)
        if os.path.isdir(np):
            for name in sorted(os.listdir(np)):
                add_file(os.path.join(np, name))
        else:
            add_file(np)
    return out
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "files",
        nargs="*",
        help="Excel workbook(s) or directory/directories to scrape (.xlsx/.xlsm).",
    )
    ap.add_argument("--crm_wip", default="CRM_DATA\\CRM_WIP.csv", help="Path to CRM_WIP.csv (default: CRM_WIP.csv).")
    ap.add_argument("--out", default="CRM_DATA\\crm_non_wip_activities.csv", help="Output CSV path.")
    ap.add_argument(
        "--weeks-back",
        type=int,
        default=3,
        help="Number of prior weeks to include in addition to the current week (default: 3).",
    )
    ap.add_argument(
        "--team",
        default=None,
        help=(
            "Optional team to scrape. Examples: MCS, DS, CPT, CDS, NI, PM-CTS, PM-CTS IND, "
            "MEIC, 'Lit & Letters'. If omitted, all configured teams are scraped."
        ),
    )
    args = ap.parse_args()
    try:
        selected_team = normalize_team_arg(args.team)
    except ValueError as exc:
        ap.error(str(exc))
    if args.files:
        inputs = args.files
        files = expand_input_paths(inputs)
    elif selected_team == LIT_LETTERS_TEAM_NAME:
        inputs = []
        files = discover_lit_letters_files(weeks_back=args.weeks_back)
        if not files:
            roots = "; ".join(lit_letters_search_roots())
            ap.error(
                "No Lit & Letters workbook was found in the default search roots. "
                "Pass the workbook or folder path after the script name. "
                f"Searched: {roots}"
            )
    else:
        inputs = default_inputs_for_team(selected_team)
        if selected_team and not inputs:
            ap.error(
                f"No default input path is configured for {selected_team!r}. "
                "Pass the workbook or folder path after the script name."
            )
        files = expand_input_paths(inputs)

    if selected_team:
        files = [f for f in files if team_for_source(f) == selected_team]

    before_recent_filter = len(files)
    files = filter_files_to_recent_weeks(files, weeks_back=args.weeks_back)
    skipped_by_date = before_recent_filter - len(files)

    if selected_team:
        print(
            f"Scraping team {selected_team}: {len(files)} file(s) found"
            + (f" ({skipped_by_date} older/newer dated file(s) skipped)" if skipped_by_date else "")
        )
    else:
        print(
            f"Scraping all configured teams: {len(files)} file(s) found"
            + (f" ({skipped_by_date} older/newer dated file(s) skipped)" if skipped_by_date else "")
        )
    completed_hours_lookup = load_completed_hours_from_crm_wip(args.crm_wip)
    people_in_wip_lookup = load_people_in_wip_from_crm_wip(args.crm_wip)
    all_rows: List[Dict[str, Any]] = []
    for f in files:
        team = team_for_source(f)
        try:
            if team == LIT_LETTERS_TEAM_NAME:
                all_rows.extend(scrape_one_lit_letters_workbook(f, people_in_wip_lookup))
            elif team in {PM_CTS_TEAM_NAME, PM_CTS_IND_TEAM_NAME}:
                all_rows.extend(scrape_one_pm_cts_workbook(f, people_in_wip_lookup))
            elif team == "DS":
                all_rows.extend(scrape_one_ds_workbook(f, people_in_wip_lookup))
            elif team == "CPT":
                all_rows.extend(scrape_one_cpt_workbook(f, people_in_wip_lookup))
            elif team == "CDS":
                all_rows.extend(scrape_one_cds_workbook(f, people_in_wip_lookup))
            elif team == "NI":
                all_rows.extend(scrape_one_ni_workbook(f, people_in_wip_lookup))
            elif team == MEIC_TEAM_NAME:
                all_rows.extend(scrape_one_meic_workbook(f, people_in_wip_lookup))
            else:
                all_rows.extend(scrape_one_workbook(f, completed_hours_lookup))
        except Exception as exc:
            print(f"Skipping {f}: {exc}")
    recent_rows = [
        r for r in filter_rows_to_recent_weeks(all_rows, weeks_back=args.weeks_back)
        if should_include_recent_row(r)
    ]
    all_rows = merge_existing_with_recent_rows(
        load_existing_csv_rows(args.out),
        recent_rows,
        weeks_back=args.weeks_back,
        refresh_team=selected_team,
    )
    all_rows.sort(key=lambda r: (str(r.get("team", "")), str(r.get("period_date", ""))))
    with open(args.out, "w", newline="", encoding="utf-8") as fp:
        w = csv.DictWriter(fp, fieldnames=CSV_COLUMNS)
        w.writeheader()
        for r in all_rows:
            w.writerow({k: r.get(k, "") for k in CSV_COLUMNS})
    print(f"Wrote {len(all_rows)} row(s) to {args.out}")
    return 0
if __name__ == "__main__":
    raise SystemExit(main())