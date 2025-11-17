# heijunka_new_layout.py
from __future__ import annotations
import argparse, json, os, sys, csv, math
from dataclasses import dataclass, field
from datetime import datetime, date, timedelta
from typing import Dict, List, Tuple, Optional, Any, Iterable
from collections import defaultdict
import shutil, tempfile, uuid
def _excel_path_for_open(p: str) -> str:
    return os.path.abspath(p)
def _copy_to_temp_if_needed(src_path: str) -> Optional[str]:
    try:
        tmp_dir = tempfile.gettempdir()
        tmp_name = f"heijunka_{uuid.uuid4().hex}.xlsb"
        tmp_path = os.path.join(tmp_dir, tmp_name)
        shutil.copy2(src_path, tmp_path)
        return tmp_path
    except Exception:
        return None
try:
    import pythoncom
    class _ComMessageFilter:
        def HandleInComingCall(self, dwCallType, hTaskCaller, dwTickCount, lpInterfaceInfo):
            return 0  # SERVERCALL_ISHANDLED
        def RetryRejectedCall(self, hTaskCallee, dwTickCount, dwRejectType):
            if dwRejectType == 2:
                return 100  # retry in 100ms
            return -1
        def MessagePending(self, hTaskCallee, dwTickCount, dwPendingType):
            return 2  # PENDINGMSG_WAITDEFPROCESS
except Exception:
    _ComMessageFilter = None
EXCLUDED_CELLS = {
    "OOO",
    "NON WIP", "NON-WIP", "NONWIP",
    "OTHER TEAM WIP", "OTHER TEAM'S WIP", "OTHER_TEAM_WIP", "OTHERTEAMWIP"
}
def _is_excluded_cell(cell: str) -> bool:
    return (cell or "").strip().upper() in EXCLUDED_CELLS
def _week_from_row(ridx: int, anchors: List[Dict[str, Any]]) -> Optional[date]:
    if not anchors:
        return None
    try:
        parsed = [(int(a["row"]), _to_date(a["date"])) for a in anchors if "row" in a and "date" in a]
        parsed = [(r, d) for (r, d) in parsed if d is not None]
        if not parsed:
            return None
        parsed.sort(key=lambda x: x[0])  # by row
        wk = None
        for r, d in parsed:
            if ridx >= r:
                wk = d
            else:
                break
        return wk
    except Exception:
        return None
def _to_date(v) -> Optional[date]:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        d = _excel_serial_to_date(v)
        if d:
            return d
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y",
                "%d-%b-%Y", "%d-%b-%y", "%d %b %Y", "%d %b %y",
                "%d-%B-%Y", "%d-%B-%y", "%d %B %Y", "%d %B %y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None
def _excel_serial_to_date(n) -> Optional[date]:
    try:
        return (datetime(1899, 12, 30) + timedelta(days=float(n))).date()
    except Exception:
        return None
def _to_float(x) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, float):
        return x
    try:
        return float(str(x).replace(",", "").strip())
    except Exception:
        return None
def _clean_name(s: Any) -> str:
    if s is None:
        return ""
    if isinstance(s, float) and math.isnan(s):
        return ""
    return str(s).strip()
@dataclass
class WeekRollup:
    week: date
    total_available_hours: float = 0.0
    completed_hours: float = 0.0
    target_output: float = 0.0
    actual_output: float = 0.0
    target_uplh: Optional[float] = None
    actual_uplh: Optional[float] = None
    hc_in_wip: int = 0
    actual_hc_used: Optional[float] = None
    person_hours: Dict[str, Dict[str, float]] = field(default_factory=dict)
    outputs_by_person: Dict[str, Dict[str, float]] = field(default_factory=dict)
    outputs_by_cell: Dict[str, Dict[str, float]] = field(default_factory=dict)
    cell_hours: Dict[str, float] = field(default_factory=dict)
    hours_by_cell_by_person: Dict[str, Dict[str, float]] = field(default_factory=dict)
    outputs_by_cell_by_person: Dict[str, Dict[str, Dict[str, float]]] = field(default_factory=dict)
    uplh_by_cell_by_person: Dict[str, Dict[str, Dict[str, float]]] = field(default_factory=dict)
    def finalize(self):
        self.total_available_hours = round(float(self.total_available_hours or 0.0), 2)
        self.completed_hours = round(float(self.completed_hours or 0.0), 2)
        self.target_output = round(float(self.target_output or 0.0), 2)
        self.actual_output = round(float(self.actual_output or 0.0), 2)
        self.target_uplh = round(self.target_output / self.completed_hours, 2) if self.completed_hours else None
        self.actual_uplh = round(self.actual_output / self.completed_hours, 2) if self.completed_hours else None
        self.hc_in_wip = sum(1 for v in self.person_hours.values() if (v.get("actual", 0) or 0) > 0)
        self.actual_hc_used = round(self.completed_hours / 30.0, 2) if self.completed_hours else None
        uplh: Dict[str, Dict[str, Dict[str, float]]] = {}
        for cell, per_map in self.hours_by_cell_by_person.items():
            for person, hrs in per_map.items():
                h = float(hrs or 0.0)
                if h <= 0:
                    continue
                outs = self.outputs_by_cell_by_person.get(cell, {}).get(person, {})
                a = float(outs.get("output", 0) or 0.0)
                t = float(outs.get("target", 0) or 0.0)
                uplh.setdefault(cell, {})[person] = {
                    "actual": round(a / h, 2) if h else None,
                    "target": round(t / h, 2) if h else None,
                }
        self.uplh_by_cell_by_person = uplh
def _rows_from_xlsx_visible(path: str, sheet_name: str):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=False)  # need row_dimensions
    ws = wb[sheet_name]
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        rd = ws.row_dimensions.get(r)
        hidden = bool(getattr(rd, "hidden", False))
        zero_h = (getattr(rd, "height", None) == 0)
        if hidden or zero_h:
            continue
        vals = []
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            vals.append(cell.value)
        yield r, tuple(vals)
def _rows_from_xlsb_visible(path: str, sheet_name: str):
    try:
        import time, os as _os
        import pythoncom
        import win32com.client as win32  # type: ignore
        import pywintypes  # type: ignore
    except Exception:
        import pandas as pd
        df = pd.read_excel(path, sheet_name=sheet_name, engine="pyxlsb", header=None)
        for i, row in enumerate(df.itertuples(index=False, name=None), start=1):
            yield i, tuple(row)
        return
    norm_path = _os.path.abspath(path)
    xl = wb = None
    created_app = False
    opened_here = False
    mf_cookie = None
    def _same_file(a: str, b: str) -> bool:
        try:
            return _os.path.samefile(a, b)
        except Exception:
            return _os.path.normcase(_os.path.abspath(a)) == _os.path.normcase(_os.path.abspath(b))
    def _get_ws(_wb, name: str):
        try:
            ws = _wb.Worksheets(name)
        except Exception:
            ws = None
            try:
                for s in _wb.Worksheets:
                    if str(s.Name).strip().lower() == name.strip().lower():
                        ws = s
                        break
            except Exception:
                pass
        if ws is None:
            try:
                available = [str(s.Name) for s in _wb.Worksheets]
            except Exception:
                available = []
            raise RuntimeError(f"Worksheet '{name}' not found. Available: {available}")
        try:
            if int(getattr(ws, "Type", -4167)) != -4167:  # -4167 == xlWorksheet
                raise RuntimeError(f"'{name}' is not a worksheet.")
        except Exception:
            pass
        return ws
    def _safe_used_range(ws):
        last_err = None
        for _ in range(8):
            try:
                ur = ws.UsedRange
                r1 = int(ur.Row)
                c1 = int(ur.Column)
                rc = int(ur.Rows.Count)
                cc = int(ur.Columns.Count)
                if rc <= 0 or cc <= 0:
                    return None
                return (r1, c1, rc, cc)
            except Exception as e:
                last_err = e
                time.sleep(0.15)
        if last_err:
            raise last_err
        return None
    try:
        pythoncom.CoInitialize()
        if _ComMessageFilter is not None:
            mf_cookie = pythoncom.CoRegisterMessageFilter(_ComMessageFilter(), None)
    except Exception:
        pass
    try:
        try:
            xl = win32.GetActiveObject("Excel.Application")
        except Exception:
            xl = win32.DispatchEx("Excel.Application")
            created_app = True
        if created_app:
            try:
                xl.Visible = False
                xl.DisplayAlerts = False
            except Exception:
                pass
        for w in xl.Workbooks:
            try:
                if _same_file(str(w.FullName), norm_path):
                    wb = w
                    break
            except Exception:
                continue
        if wb is None:
            OPEN_ARGS = dict(
                UpdateLinks=0, ReadOnly=True, AddToMru=False,
                IgnoreReadOnlyRecommended=True, Notify=False
            )
            max_tries = 15
            src_path = _excel_path_for_open(norm_path)
            last_exc = None
            for attempt in range(1, max_tries + 1):
                try:
                    wb = xl.Workbooks.Open(src_path, **OPEN_ARGS)
                    opened_here = True
                    break
                except pywintypes.com_error as e:
                    last_exc = e
                    code = e.args[0] if (e.args and isinstance(e.args[0], int)) else None
                    msg  = str(e).lower()
                    if code in (-2147418111, -2147417848):  # call rejected / disconnected
                        try:
                            pythoncom.PumpWaitingMessages()
                        except Exception:
                            pass
                        time.sleep(min(0.2 * attempt, 2.5))
                        continue
                    if "cannot access the file" in msg or "same name as a currently open workbook" in msg:
                        tmp_copy = _copy_to_temp_if_needed(src_path)
                        if tmp_copy:
                            try:
                                wb = xl.Workbooks.Open(tmp_copy, **OPEN_ARGS)
                                opened_here = True
                                break
                            except Exception as _e2:
                                last_exc = _e2
                        time.sleep(min(0.3 * attempt, 3.0))
                        continue
                    if "permission denied" in msg:
                        time.sleep(min(0.3 * attempt, 4.0))
                        continue
                    time.sleep(min(0.2 * attempt, 1.5))
                    continue
                except PermissionError as e:
                    last_exc = e
                    time.sleep(min(0.3 * attempt, 4.0))
                    continue
            if wb is None:
                raise RuntimeError(f"Failed to open workbook read-only (it may be exclusively locked or name-colliding): {norm_path} | last={last_exc}")
        ws = _get_ws(wb, sheet_name)
        meta = _safe_used_range(ws)
        if meta is None:
            return  # empty sheet
        first_row, first_col, rows, cols = meta
        for attempt in range(6):
            try:
                last_row = first_row + rows - 1
                last_col = first_col + cols - 1
                rng = ws.Range(ws.Cells(first_row, first_col), ws.Cells(last_row, last_col))
                data = rng.Value  # safe 2D array/tuple
                break
            except Exception:
                time.sleep(0.1)
        else:
            data = ws.UsedRange.Value
        if rows == 1 and cols == 1:
            data = ((data,),)
        elif rows == 1:
            data = (tuple(data),)
        elif cols == 1:
            data = tuple((d,) for d in data)
        for i in range(rows):
            ridx = first_row + i
            hidden = False
            for _ in range(3):
                try:
                    cell = ws.Cells(ridx, 1)
                    hidden = bool(cell.EntireRow.Hidden) or float(cell.RowHeight or 0) == 0.0
                    break
                except Exception:
                    time.sleep(0.05)
            if hidden:
                continue
            row_vals = data[i] if i < len(data) else tuple()
            if row_vals is None:
                row_vals = tuple(None for _ in range(cols))
            elif isinstance(row_vals, tuple):
                if len(row_vals) < cols:
                    row_vals = row_vals + tuple(None for _ in range(cols - len(row_vals)))
            else:
                row_vals = (row_vals,) + tuple(None for _ in range(cols - 1))
            yield ridx, tuple(row_vals)
    finally:
        try:
            if opened_here and wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        if xl is not None and created_app:
            try:
                xl.Quit()
            except Exception:
                pass
        try:
            if mf_cookie is not None:
                pythoncom.CoRegisterMessageFilter(None, None)
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
def parse_available_rows(rows_with_idx: Iterable[Tuple[int, Tuple[Any, ...]]],
                         anchors: Optional[List[Dict[str, Any]]] = None) -> Dict[date, Dict[str, float]]:
    day_idxs: Optional[List[int]] = None
    default_day_idxs = [4, 5, 6, 7, 8]
    avail_per_week: Dict[date, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    current_person: Optional[str] = None
    def _detect_day_idxs(r: Tuple[Any, ...]) -> Optional[List[int]]:
        labels = ["monday", "tuesday", "wednesday", "thursday", "friday"]
        idxs = {}
        for i, v in enumerate(r):
            s = _clean_name(v).lower()
            if s in labels:
                idxs[s] = i
        if all(k in idxs for k in labels):
            return [idxs["monday"], idxs["tuesday"], idxs["wednesday"], idxs["thursday"], idxs["friday"]]
        return None
    for ridx, r in rows_with_idx:
        r = r or tuple()
        if day_idxs is None:
            di = _detect_day_idxs(r)
            if di:
                day_idxs = di
        name = _clean_name(r[2] if len(r) > 2 else "")
        if name:
            current_person = name
        flag = _clean_name(r[3] if len(r) > 3 else "").strip().lower()
        if not current_person or "available wip" not in flag:
            continue
        wk = _week_from_row(ridx, anchors or [])
        if wk is None:
            continue
        cols = day_idxs or default_day_idxs
        s = 0.0
        for c in cols:
            v = r[c] if len(r) > c else None
            fv = _to_float(v)
            if fv is not None:
                s += fv
        if s:
            avail_per_week[wk][current_person] += s
    return avail_per_week
def parse_prod_rows(rows_with_idx: Iterable[Tuple[int, Tuple[Any, ...]]],
                    anchors: Optional[List[Dict[str, Any]]] = None) -> Dict[date, Dict[str, Any]]:
    COL_DATE, COL_NAME, COL_CELL, COL_TARGET, COL_MINUTES, COL_OUTPUT = 0, 3, 4, 7, 8, 10
    buckets: Dict[date, Dict[str, Any]] = defaultdict(lambda: {
        "completed_hours_by_person": defaultdict(float),
        "outputs_by_person": defaultdict(lambda: {"target": 0.0, "output": 0.0}),
        "outputs_by_cell": defaultdict(lambda: {"target": 0.0, "output": 0.0}),
        "hours_by_cell_by_person": defaultdict(lambda: defaultdict(float)),
        "outputs_by_cell_by_person": defaultdict(lambda: defaultdict(lambda: {"target": 0.0, "output": 0.0})),
        "target_output_total": 0.0,
        "actual_output_total": 0.0,
        "names_in_wip": set(),
    })
    for ridx, r in rows_with_idx:
        r = r or tuple()
        wk = _week_from_row(ridx, anchors or [])
        if wk is None:
            continue
        name = _clean_name(r[COL_NAME] if len(r) > COL_NAME else "")
        cell = _clean_name(r[COL_CELL] if len(r) > COL_CELL else "")
        tgt  = _to_float(r[COL_TARGET] if len(r) > COL_TARGET else None) or 0.0
        mins = _to_float(r[COL_MINUTES] if len(r) > COL_MINUTES else None) or 0.0
        outp = _to_float(r[COL_OUTPUT] if len(r) > COL_OUTPUT else None) or 0.0
        if not (name or cell or tgt or mins or outp):
            continue
        is_excluded = _is_excluded_cell(cell)
        b = buckets[wk]
        if not is_excluded and name and cell and mins > 0:
            hrs = mins / 60.0
            b["completed_hours_by_person"][name] += hrs
            b["hours_by_cell_by_person"][cell][name] += hrs
            b["names_in_wip"].add(name)
        if name and (tgt or outp):
            b["outputs_by_person"][name]["target"] += tgt
            b["outputs_by_person"][name]["output"] += outp
        if not is_excluded and cell and (tgt or outp):
            b["outputs_by_cell"][cell]["target"] += tgt
            b["outputs_by_cell"][cell]["output"] += outp
        if not is_excluded and cell and name and (tgt or outp):
            b["outputs_by_cell_by_person"][cell][name]["target"] += tgt
            b["outputs_by_cell_by_person"][cell][name]["output"] += outp
        b["target_output_total"] += tgt
        b["actual_output_total"] += outp
    out: Dict[date, Dict[str, Any]] = {}
    for wk, b in buckets.items():
        out[wk] = {
            "completed_hours_by_person": dict(b["completed_hours_by_person"]),
            "outputs_by_person": {k: dict(v) for k, v in b["outputs_by_person"].items()},
            "outputs_by_cell": {k: dict(v) for k, v in b["outputs_by_cell"].items()},
            "hours_by_cell_by_person": {k: dict(v) for k, v in b["hours_by_cell_by_person"].items()},
            "outputs_by_cell_by_person": {cell: {p: dict(vals) for p, vals in per.items()}
                                          for cell, per in b["outputs_by_cell_by_person"].items()},
            "target_output_total": b["target_output_total"],
            "actual_output_total": b["actual_output_total"],
            "hc_in_wip": len(b["names_in_wip"]),
        }
    return out
def _sheetnames_xlsx_like(path: str) -> List[str]:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    return list(wb.sheetnames)
def _rows_from_xlsx_like(path: str, sheet_name: str) -> Iterable[Tuple[Any, ...]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    for r in ws.iter_rows(values_only=True):
        yield tuple(r)
def _sheetnames_xlsb(path: str) -> List[str]:
    import pandas as pd
    with pd.ExcelFile(path, engine="pyxlsb") as xf:
        return list(xf.sheet_names)
def _rows_from_xlsb(path: str, sheet_name: str) -> Iterable[Tuple[Any, ...]]:
    import pandas as pd
    df = pd.read_excel(path, sheet_name=sheet_name, engine="pyxlsb", header=None)
    for row in df.itertuples(index=False, name=None):
        yield tuple(row)
def _get_visible_rows_reader(path: str):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsb":
        return _rows_from_xlsb_visible  # keep only if you still have any .xlsb left
    elif ext in (".xlsx", ".xlsm"):
        return _rows_from_xlsx_visible
    else:
        raise ValueError(f"Unsupported workbook extension '{ext}'.")
def _get_all_rows_reader(path: str):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsb":
        return _rows_from_xlsb
    elif ext in (".xlsx", ".xlsm"):
        return _rows_from_xlsx_like
    else:
        raise ValueError(f"Unsupported workbook extension '{ext}'.")
def _sort_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    def _parse_week(r: Dict[str, Any]) -> date:
        w = str(r.get("Week", "")).strip()
        try:
            return datetime.fromisoformat(w).date()
        except Exception:
            d = _to_date(w)
            return d if d else date.min
    return sorted(rows, key=lambda r: (str(r.get("Team","")), _parse_week(r)))
def _find_sheet_by_hint(sheet_names: List[str], hint: str) -> str:
    exact = [nm for nm in sheet_names if nm.strip().lower() == hint.strip().lower()]
    if exact:
        return exact[0]
    contains = [nm for nm in sheet_names if hint.lower() in nm.lower()]
    if contains:
        return contains[0]
    raise ValueError(f"Sheet '{hint}' not found. Available: {sheet_names}")
def build_weekly_metrics_from_file(path: str, prod_hints: List[str], avail_hint: str,
                                   week_anchors_by_sheet: Optional[Dict[str, List[Dict[str, Any]]]] = None) -> List[Dict[str, Any]]:
    week_anchors_by_sheet = week_anchors_by_sheet or {}
    ext = os.path.splitext(path)[1].lower()
    sheet_names = (_sheetnames_xlsb(path) if ext == ".xlsb" else _sheetnames_xlsx_like(path))
    read_visible_rows = _get_visible_rows_reader(path)
    read_all_rows     = _get_all_rows_reader(path)
    avail_name = _find_sheet_by_hint(sheet_names, avail_hint)
    avail_rows_with_idx = list(read_visible_rows(path, avail_name))
    avail = parse_available_rows(avail_rows_with_idx, anchors=week_anchors_by_sheet.get(avail_name, []))
    prod_merged: Dict[date, Dict[str, Any]] = {}
    for hint in prod_hints:
        prod_name = _find_sheet_by_hint(sheet_names, hint)
        prod_rows_with_idx = list(
            (i + 1, tuple(r))
            for i, r in enumerate(read_all_rows(path, prod_name))
        )
        anchors = week_anchors_by_sheet.get(prod_name, [])
        prod_part = parse_prod_rows(prod_rows_with_idx, anchors=anchors)
        for wk, b in prod_part.items():
            if wk not in prod_merged:
                prod_merged[wk] = {
                    "completed_hours_by_person": defaultdict(float),
                    "outputs_by_person": defaultdict(lambda: {"target": 0.0, "output": 0.0}),
                    "outputs_by_cell": defaultdict(lambda: {"target": 0.0, "output": 0.0}),
                    "hours_by_cell_by_person": defaultdict(lambda: defaultdict(float)),
                    "outputs_by_cell_by_person": defaultdict(lambda: defaultdict(lambda: {"target": 0.0, "output": 0.0})),
                    "target_output_total": 0.0,
                    "actual_output_total": 0.0,
                    "hc_in_wip": 0,
                    "_names_in_wip": set(),
                }
            tgt = prod_merged[wk]
            for name, hrs in b["completed_hours_by_person"].items():
                tgt["completed_hours_by_person"][name] += hrs
                tgt["_names_in_wip"].add(name)
            for name, vv in b["outputs_by_person"].items():
                tgt["outputs_by_person"][name]["target"] += vv.get("target", 0.0)
                tgt["outputs_by_person"][name]["output"] += vv.get("output", 0.0)
            for cell, vv in b["outputs_by_cell"].items():
                tgt["outputs_by_cell"][cell]["target"] += vv.get("target", 0.0)
                tgt["outputs_by_cell"][cell]["output"] += vv.get("output", 0.0)
            for cell, per in b["hours_by_cell_by_person"].items():
                for name, hrs in per.items():
                    tgt["hours_by_cell_by_person"][cell][name] += hrs
            for cell, per in b["outputs_by_cell_by_person"].items():
                for name, vv in per.items():
                    tgt["outputs_by_cell_by_person"][cell][name]["target"] += vv.get("target", 0.0)
                    tgt["outputs_by_cell_by_person"][cell][name]["output"] += vv.get("output", 0.0)
            tgt["target_output_total"] += b.get("target_output_total", 0.0)
            tgt["actual_output_total"] += b.get("actual_output_total", 0.0)
    prod: Dict[date, Dict[str, Any]] = {}
    for wk, b in prod_merged.items():
        prod[wk] = {
            "completed_hours_by_person": dict(b["completed_hours_by_person"]),
            "outputs_by_person": {k: dict(v) for k, v in b["outputs_by_person"].items()},
            "outputs_by_cell": {k: dict(v) for k, v in b["outputs_by_cell"].items()},
            "hours_by_cell_by_person": {k: dict(v) for k, v in b["hours_by_cell_by_person"].items()},
            "outputs_by_cell_by_person": {
                cell: {p: dict(vals) for p, vals in per.items()}
                for cell, per in b["outputs_by_cell_by_person"].items()
            },
            "target_output_total": b["target_output_total"],
            "actual_output_total": b["actual_output_total"],
            "hc_in_wip": len(b["_names_in_wip"]),
        }
    allowed_weeks = set()
    for sh, a in week_anchors_by_sheet.items():
        for it in a or []:
            d = _to_date(it.get("date"))
            if d:
                allowed_weeks.add(d)
    all_weeks = sorted((set(avail.keys()) | set(prod.keys())) & allowed_weeks)
    rows: List[Dict[str, Any]] = []
    for wk in all_weeks:
        roll = WeekRollup(week=wk)
        per_available = avail.get(wk, {})
        per_actual    = prod.get(wk, {}).get("completed_hours_by_person", {})
        people = set(per_available.keys()) | set(per_actual.keys())
        for name in sorted(people):
            a_av = float(per_available.get(name, 0.0) or 0.0)
            a_ac = float(per_actual.get(name, 0.0) or 0.0)
            roll.person_hours[name] = {"available": round(a_av, 2), "actual": round(a_ac, 2)}
        roll.total_available_hours = round(sum(v.get("available", 0.0) for v in roll.person_hours.values()), 2)
        roll.completed_hours       = round(sum(v.get("actual", 0.0) for v in roll.person_hours.values()), 2)
        roll.outputs_by_person         = prod.get(wk, {}).get("outputs_by_person", {})
        roll.outputs_by_cell           = prod.get(wk, {}).get("outputs_by_cell", {})
        roll.cell_hours                = {cell: round(sum(per.values()), 2)
                                          for cell, per in prod.get(wk, {}).get("hours_by_cell_by_person", {}).items()}
        roll.hours_by_cell_by_person   = prod.get(wk, {}).get("hours_by_cell_by_person", {})
        roll.outputs_by_cell_by_person = prod.get(wk, {}).get("outputs_by_cell_by_person", {})
        roll.target_output             = prod.get(wk, {}).get("target_output_total", 0.0)
        roll.actual_output             = prod.get(wk, {}).get("actual_output_total", 0.0)
        roll.hc_in_wip                 = int(prod.get(wk, {}).get("hc_in_wip", 0))
        roll.finalize()
        rows.append({
            "Week": roll.week.isoformat(),
            "Total Available Hours": roll.total_available_hours,
            "Completed Hours": roll.completed_hours,
            "Target Output": roll.target_output,
            "Actual Output": roll.actual_output,
            "Target UPLH": roll.target_uplh,
            "Actual UPLH": roll.actual_uplh,
            "HC in WIP": roll.hc_in_wip,
            "Actual HC Used": roll.actual_hc_used,
            "Person Hours": json.dumps(roll.person_hours, ensure_ascii=False),
            "Outputs by Person": json.dumps(roll.outputs_by_person, ensure_ascii=False),
            "Outputs by Cell/Station": json.dumps(roll.outputs_by_cell, ensure_ascii=False),
            "Cell/Station Hours": json.dumps(roll.cell_hours, ensure_ascii=False),
            "Hours by Cell/Station - by person": json.dumps(roll.hours_by_cell_by_person, ensure_ascii=False),
            "Output by Cell/Station - by person": json.dumps(roll.outputs_by_cell_by_person, ensure_ascii=False),
            "UPLH by Cell/Station - by person": json.dumps(roll.uplh_by_cell_by_person, ensure_ascii=False),
        })
    return rows
def _norm_week_str(s: str) -> str:
    s = (s or "").strip()
    try:
        return datetime.fromisoformat(s).date().isoformat()
    except Exception:
        d = _to_date(s)
        return d.isoformat() if d else s
def _read_value_by_team_week(path: str,
                             value_col: Optional[str],
                             value_hint: Optional[str]) -> Tuple[Dict[Tuple[str,str], Any], str]:
    if not path or not os.path.exists(path):
        return {}, value_col or (value_hint or "Value")
    week_aliases = {"week", "period_date", "period", "period_start", "date"}
    with open(path, "r", newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        headers = [h for h in (r.fieldnames or [])]
        team_col = next((h for h in headers if h and h.strip().lower() == "team"), None)
        if not team_col:
            team_col = next((h for h in headers if h and "team" == h.strip().lower()), None)
        lower_map = {h: (h or "").strip().lower() for h in headers if h}
        week_col = None
        for h, low in lower_map.items():
            if low in week_aliases:
                week_col = h
                break
        if not team_col or not week_col:
            raise RuntimeError(f"{path}: must contain Team and a week/period column "
                               f"(accepted: {sorted(week_aliases)}). Got headers: {headers}")
        val_col = value_col
        if not val_col:
            if value_hint:
                needle = value_hint.strip().lower()
                val_col = next((h for h in headers
                                if h not in (team_col, week_col)
                                and needle in (h or "").strip().lower()), None)
            if not val_col:
                val_col = next((h for h in headers if h not in (team_col, week_col)), None)
        if not val_col:
            raise RuntimeError(f"{path}: no value column found (looked for hint '{value_hint}')")
        out: Dict[Tuple[str, str], Any] = {}
        for row in r:
            team_raw = row.get(team_col, "")
            week_raw = row.get(week_col, "")
            team = str(team_raw).strip()
            week = _norm_week_str(str(week_raw))
            if not team or not week:
                continue
            out[(team, week)] = row.get(val_col, "")
        return out, val_col
def _apply_aux_column(rows: List[Dict[str, Any]], merged_cols: List[str], aux: Dict[Tuple[str,str], Any], colname: str):
    if colname not in merged_cols:
        merged_cols.append(colname)
    for r in rows:
        key = (str(r.get("Team","")).strip(), _norm_week_str(str(r.get("Week","")).strip()))
        if key in aux:
            r[colname] = aux[key]
        elif colname not in r:
            r[colname] = ""
def _read_csv_if_exists(path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    if not os.path.exists(path):
        return [], []
    rows: List[Dict[str, Any]] = []
    with open(path, "r", newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            rows.append(dict(row))
        cols = list(r.fieldnames or [])
    return rows, cols
def _key(row: Dict[str, Any]) -> Tuple[str, str]:
    return (str(row.get("Team", "")).strip(), str(row.get("Week", "")).strip())
def _merge_rows(existing: List[Dict[str, Any]], new_rows: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[str]]:
    by_key: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for r in existing:
        by_key[_key(r)] = r
    for r in new_rows:
        by_key[_key(r)] = r
    existing_keys = [_key(r) for r in existing]
    new_keys = [_key(r) for r in new_rows]
    appended = [k for k in new_keys if k not in existing_keys]
    merged_order = [k for k in existing_keys if k in by_key] + appended
    merged_rows = [by_key[k] for k in merged_order]
    existing_cols = set(existing[0].keys()) if existing else set()
    all_cols = list(existing[0].keys()) if existing else []
    for r in new_rows:
        for c in r.keys():
            if c not in existing_cols:
                all_cols.append(c)
                existing_cols.add(c)
    if "Team" in all_cols:
        all_cols = ["Team"] + [c for c in all_cols if c != "Team"]
    if "Week" in all_cols:
        all_cols = ["Team", "Week"] + [c for c in all_cols if c not in ("Team", "Week")]
    return merged_rows, all_cols
def main():
    p = argparse.ArgumentParser(description="Aggregate metrics from unified Heijunka workbooks (.xlsb/.xlsx/.xlsm).")
    p.add_argument("--team", action="append", help="Team key to load from config JSON (repeatable)")
    p.add_argument("--all", action="store_true", help="Process all teams in config")
    p.add_argument("--config", help="Path to teams.json that maps team to workbook and sheet names")
    p.add_argument("--prod-sheet", action="append",
               help="Override: one or more 'Prod Analysis' sheet names (repeatable)")
    p.add_argument("--avail-sheet", help="Override: exact/partial name of the 'Available WIP+Non-WIP Hours' sheet")
    p.add_argument("workbook", nargs="?", help="Direct path to a single workbook (if not using --team/--config)")
    p.add_argument("--team-name", help="Team label to use when processing a single workbook without config")
    p.add_argument("--closures", help="Path to closures.csv to append as 'Closures'")
    p.add_argument("--timeliness", help="Path to timeliness.csv to append as 'Open Complaint Timeliness'")
    p.add_argument("--out", help="CSV output path", default="metrics.csv")
    args = p.parse_args()
    if not args.config and os.path.exists("teams.json"):
        args.config = "teams.json"
    if not args.out:
        args.out = "metrics.csv"
    if not getattr(args, "closures", None) and os.path.exists("closures.csv"):
        args.closures = "closures.csv"
    if not getattr(args, "timeliness", None) and os.path.exists("timeliness.csv"):
        args.timeliness = "timeliness.csv"
    if args.team and not args.config:
        print("Provide --config or place teams.json in the current folder.", file=sys.stderr)
        sys.exit(2)
    jobs: List[Tuple[str, str, List[str], str, Dict[str, List[Dict[str, Any]]]]] = []
    default_prod = "Prod Analysis"
    default_avail = "Available WIP+Non-WIP Hours"
    if args.config and (args.all or args.team):
        try:
            with open(args.config, "r", encoding="utf-8") as f:
                cfg = json.load(f)
        except Exception as e:
            print(f"Failed to read config '{args.config}': {e}", file=sys.stderr)
            sys.exit(2)
        chosen = list(cfg.keys()) if args.all else args.team
        for team in chosen:
            if team not in cfg:
                print(f"Team '{team}' not found in {args.config}. Available: {', '.join(cfg.keys())}", file=sys.stderr)
                sys.exit(2)
            entry = cfg[team]
            wb = entry.get("workbook")
            if not wb:
                print(f"Config for team '{team}' must include 'workbook'.", file=sys.stderr)
                sys.exit(2)
            if args.prod_sheet:
                prod_hints = list(args.prod_sheet)                     # from repeated --prod-sheet
            else:
                prod_cfg = entry.get("prod_sheets") or entry.get("prod_sheet") or default_prod
                prod_hints = prod_cfg if isinstance(prod_cfg, list) else [prod_cfg]
            avail_hint = args.avail_sheet or entry.get("avail_sheet") or default_avail
            jobs.append((team, wb, prod_hints, avail_hint, entry.get("week_anchors", {})))
    elif args.workbook:
        team_label = args.team_name or "Unnamed"
        prod_cfg = args.prod_sheet or default_prod
        prod_hints = prod_cfg if isinstance(prod_cfg, list) else ([prod_cfg] if args.prod_sheet else [default_prod])
        jobs.append((team_label, args.workbook, prod_hints, args.avail_sheet or default_avail, {}))
    else:
        print("Provide either --config with --all/--team, or a positional WORKBOOK path.", file=sys.stderr)
        sys.exit(2)
    all_rows: List[Dict[str, Any]] = []
    for team_label, path, prod_hints, avail_hint, week_anchors in jobs:
        if not os.path.exists(path):
            print(f"[{team_label}] Workbook not found: {path}", file=sys.stderr)
            sys.exit(2)
        try:
            rows = build_weekly_metrics_from_file(
                path,
                prod_hints=prod_hints,
                avail_hint=avail_hint,
                week_anchors_by_sheet=week_anchors
            )
            for r in rows:
                r_with_team = {"Team": team_label, **r}
                all_rows.append(r_with_team)
            print(f"[{team_label}] OK: {len(rows)} weekly rows (prod sheets: {prod_hints})")
        except Exception as e:
            print(f"[{team_label}] Error while building metrics: {e}", file=sys.stderr)
            sys.exit(1)
            for r in rows:
                r_with_team = {"Team": team_label, **r}
                all_rows.append(r_with_team)
            print(f"[{team_label}] OK: {len(rows)} weekly rows")
        except Exception as e:
            print(f"[{team_label}] Error while building metrics: {e}", file=sys.stderr)
            sys.exit(1)
    if not all_rows:
        print("No data found (check sheet names/column placements).", file=sys.stderr)
        sys.exit(1)
    existing_rows, existing_cols = _read_csv_if_exists(args.out)
    merged_rows, merged_cols = _merge_rows(existing_rows, all_rows)
    merged_rows = _sort_rows(merged_rows)
    closures_map = {}
    opened_map   = {}
    if getattr(args, "closures", None):
        closures_map, _ = _read_value_by_team_week(
            args.closures,
            value_col="Closures",
            value_hint="closures",
        )
        opened_map, _ = _read_value_by_team_week(
            args.closures,
            value_col="Opened",
            value_hint="opened",
        )
    if closures_map:
        _apply_aux_column(merged_rows, merged_cols, closures_map, "Closures")  # force name
    if opened_map:
        _apply_aux_column(merged_rows, merged_cols, opened_map, "Opened")      # NEW
    time_map, time_col = _read_value_by_team_week(
        args.timeliness,
        value_col=None,
        value_hint="timeliness",
    ) if getattr(args, "timeliness", None) else ({}, "Open Complaint Timeliness")
    if time_map:
        _apply_aux_column(merged_rows, merged_cols, time_map, "Open Complaint Timeliness")
    with open(args.out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=merged_cols)
        w.writeheader()
        for r in merged_rows:
            w.writerow({k: r.get(k, "") for k in merged_cols})
    updated_keys = {(_key(r)) for r in all_rows}
    print(f"Wrote {len(merged_rows)} total rows "
          f"(updated {len(updated_keys)} Team-Week rows) -> {args.out}")
if __name__ == "__main__":
    main()